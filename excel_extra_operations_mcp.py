#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel Advanced Operations MCP (Multi-purpose Connector for Python with Excel)
-------------------------------------------------------
Biblioteca para operaciones avanzadas en Excel:
- Validación de Datos y Formato Condicional
- Filtros, Ordenación y Agrupación
- Imágenes y Objetos
- Otras Operaciones Avanzadas

Author: MCP Team
Version: 1.0
"""

import os
import sys
import json
import logging
import tempfile
from pathlib import Path
from typing import List, Dict, Union, Optional, Tuple, Any, Callable

# Configuración de logging
logger = logging.getLogger("excel_advanced_operations_mcp")
logger.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stderr)
handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
logger.addHandler(handler)

# Importar MCP
try:
    from mcp.server.fastmcp import FastMCP
    HAS_MCP = True
except ImportError:
    logger.warning("No se pudo importar FastMCP. Las funcionalidades de servidor MCP no estarán disponibles.")
    HAS_MCP = False

# Intentar importar las bibliotecas necesarias
try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, CellIsRule, Rule
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
    from openpyxl.worksheet.filters import AutoFilter, SortState, SortCondition
    from openpyxl.drawing.image import Image
    HAS_OPENPYXL = True
except ImportError as e:
    logger.warning(f"Error al importar bibliotecas esenciales: {e}")
    logger.warning("Es posible que algunas funcionalidades no estén disponibles")
    HAS_OPENPYXL = False

# Intentar importar bibliotecas adicionales
try:
    from io import BytesIO
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    logger.warning("No se pudo importar PIL. La manipulación avanzada de imágenes estará limitada.")
    HAS_PIL = False

# Intentar importar pywin32 para uso alternativo con Excel
try:
    import win32com.client
    HAS_PYWIN32 = True
except ImportError:
    logger.warning("No se pudo importar win32com.client. La funcionalidad alternativa de Excel COM no estará disponible.")
    HAS_PYWIN32 = False

# Excepciones personalizadas
class AdvancedOperationsError(Exception):
    """Excepción base para todos los errores de Excel Advanced Operations MCP."""
    pass

class ValidationError(AdvancedOperationsError):
    """Se lanza cuando hay un problema con la validación de datos."""
    pass

class FormattingError(AdvancedOperationsError):
    """Se lanza cuando hay un problema con el formato condicional."""
    pass

class FilterError(AdvancedOperationsError):
    """Se lanza cuando hay un problema con filtros o ordenación."""
    pass

class GroupingError(AdvancedOperationsError):
    """Se lanza cuando hay un problema con la agrupación."""
    pass

class ImageError(AdvancedOperationsError):
    """Se lanza cuando hay un problema con imágenes u objetos."""
    pass

class OperationError(AdvancedOperationsError):
    """Se lanza cuando hay un problema con operaciones avanzadas."""
    pass

class SheetNotFoundError(AdvancedOperationsError):
    """Se lanza cuando no se encuentra una hoja en el archivo Excel."""
    pass

class RangeError(AdvancedOperationsError):
    """Se lanza cuando hay un problema con un rango de celdas."""
    pass

# Clase auxiliar para gestionar rangos de Excel
class ExcelRange:
    """
    Clase para manipular y convertir rangos de Excel.
    
    Esta clase proporciona métodos para convertir entre notación de Excel (A1:B5)
    y coordenadas de Python (0-based), además de validar rangos.
    """
    
    @staticmethod
    def parse_cell_ref(cell_ref: str) -> Tuple[int, int]:
        """
        Convierte una referencia de celda en estilo A1 a coordenadas (fila, columna) 0-based.
        
        Args:
            cell_ref: Referencia de celda en formato Excel (ej: 'A1', 'B5')
            
        Returns:
            Tupla (fila, columna) con índices base 0
            
        Raises:
            ValueError: Si la referencia de celda no es válida
        """
        if not cell_ref or not isinstance(cell_ref, str):
            raise ValueError(f"Referencia de celda inválida: {cell_ref}")
        
        # Extraer la parte de columna (letras)
        col_str = ''.join(c for c in cell_ref if c.isalpha())
        # Extraer la parte de fila (números)
        row_str = ''.join(c for c in cell_ref if c.isdigit())
        
        if not col_str or not row_str:
            raise ValueError(f"Formato de celda inválido: {cell_ref}")
        
        # Convertir columna a índice (A->0, B->1, etc.)
        col_idx = 0
        for c in col_str.upper():
            col_idx = col_idx * 26 + (ord(c) - ord('A') + 1)
        col_idx -= 1  # Ajustar a base 0
        
        # Convertir fila a índice (base 0)
        row_idx = int(row_str) - 1
        
        return row_idx, col_idx
    
    @staticmethod
    def parse_range(range_str: str) -> Tuple[int, int, int, int]:
        """
        Convierte un rango en estilo A1:B5 a coordenadas (row1, col1, row2, col2) 0-based.
        
        Args:
            range_str: Rango en formato Excel (ej: 'A1:B5')
            
        Returns:
            Tupla (fila_inicio, col_inicio, fila_fin, col_fin) con índices base 0
            
        Raises:
            ValueError: Si el rango no es válido
        """
        if not range_str or not isinstance(range_str, str):
            raise ValueError(f"Rango inválido: {range_str}")
        
        # Manejar rangos con referencia a hoja
        if '!' in range_str:
            parts = range_str.split('!')
            if len(parts) != 2:
                raise ValueError(f"Formato de rango con hoja inválido: {range_str}")
            range_str = parts[1]  # Usar solo la parte del rango
        
        # Dividir el rango en celdas de inicio y fin
        if ':' in range_str:
            start_cell, end_cell = range_str.split(':')
            start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
            end_row, end_col = ExcelRange.parse_cell_ref(end_cell)
        else:
            # Si es una sola celda, inicio y fin son iguales
            start_row, start_col = ExcelRange.parse_cell_ref(range_str)
            end_row, end_col = start_row, start_col
        
        return start_row, start_col, end_row, end_col
    
    @staticmethod
    def cell_to_a1(row: int, col: int) -> str:
        """
        Convierte coordenadas (fila, columna) 0-based a referencia de celda A1.
        
        Args:
            row: Índice de fila (base 0)
            col: Índice de columna (base 0)
            
        Returns:
            Referencia de celda en formato A1
        """
        if row < 0 or col < 0:
            raise ValueError(f"Índices negativos no válidos: fila={row}, columna={col}")
        
        # Convertir columna a letras
        col_str = ""
        col_val = col + 1  # Convertir a base 1 para cálculo
        
        while col_val > 0:
            remainder = (col_val - 1) % 26
            col_str = chr(65 + remainder) + col_str
            col_val = (col_val - 1) // 26
        
        # Convertir fila a número (base 1 para Excel)
        row_val = row + 1
        
        return f"{col_str}{row_val}"
    
    @staticmethod
    def range_to_a1(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
        """
        Convierte coordenadas de rango 0-based a rango A1:B5.
        
        Args:
            start_row: Fila inicial (base 0)
            start_col: Columna inicial (base 0)
            end_row: Fila final (base 0)
            end_col: Columna final (base 0)
            
        Returns:
            Rango en formato A1:B5
        """
        start_cell = ExcelRange.cell_to_a1(start_row, start_col)
        end_cell = ExcelRange.cell_to_a1(end_row, end_col)
        
        return f"{start_cell}:{end_cell}"

# Función auxiliar para obtener una hoja de trabajo
def get_sheet(wb, sheet_name) -> Any:
    """
    Obtiene una hoja de Excel por nombre.
    
    Args:
        wb: Objeto workbook de openpyxl
        sheet_name: Nombre de la hoja
        
    Returns:
        Objeto worksheet
        
    Raises:
        SheetNotFoundError: Si la hoja no existe
    """
    if not wb:
        raise AdvancedOperationsError("El workbook no puede ser None")
    
    if sheet_name not in wb.sheetnames:
        raise SheetNotFoundError(f"La hoja '{sheet_name}' no existe en el workbook. Hojas disponibles: {wb.sheetnames}")
    
    return wb[sheet_name]


# 10. Validación de Datos y Formato Condicional
def add_data_validation(ws, cell_range, validation_type, criteria):
    """
    Añade validación de datos a un rango de celdas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell_range (str): Rango de celdas en formato A1:B5
        validation_type (str): Tipo de validación ('whole', 'decimal', 'list', 'date', 'time', 'textLength', 'custom')
        criteria (dict): Diccionario con criterios de validación:
            - operator: Operador para la comparación
            - value1: Primer valor de comparación
            - value2: Segundo valor para rangos
            - formula1: Primera fórmula para validación personalizada
            - formula2: Segunda fórmula para rangos personalizados
            - allow_blank: Permitir celdas vacías
            - show_error: Mostrar mensaje de error
            - error_title: Título del mensaje de error
            - error_message: Mensaje de error completo
            - show_input: Mostrar mensaje de entrada
            - input_title: Título del mensaje de entrada
            - input_message: Mensaje de entrada completo
            - dropdown: Mostrar lista desplegable (para tipo 'list')
            
    Returns:
        Objeto DataValidation configurado
        
    Raises:
        ValidationError: Si hay un problema con la validación de datos
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Crear objeto de validación según el tipo
        valid_types = ['whole', 'decimal', 'list', 'date', 'time', 'textLength', 'custom']
        if validation_type not in valid_types:
            raise ValidationError(f"Tipo de validación '{validation_type}' no válido. Tipos válidos: {valid_types}")
        
        # Extraer parámetros del diccionario criteria
        operator = criteria.get('operator', '')
        formula1 = criteria.get('formula1', criteria.get('value1', ''))
        formula2 = criteria.get('formula2', criteria.get('value2', ''))
        allow_blank = criteria.get('allow_blank', True)
        show_error = criteria.get('show_error', True)
        error_title = criteria.get('error_title', '')
        error_message = criteria.get('error_message', '')
        show_input = criteria.get('show_input', True)
        input_title = criteria.get('input_title', '')
        input_message = criteria.get('input_message', '')
        dropdown = criteria.get('dropdown', True)
        
        # Convertir value1/value2 a formula1/formula2 si es necesario
        if 'value1' in criteria and not formula1:
            value1 = criteria['value1']
            if isinstance(value1, list):
                # Para validación de lista, convertir a string separado por comas
                formula1 = ",".join(str(v) for v in value1)
            else:
                formula1 = str(value1)
        
        if 'value2' in criteria and not formula2:
            formula2 = str(criteria['value2'])
        
        # Crear validación de datos
        dv = DataValidation(
            type=validation_type,
            operator=operator,
            formula1=formula1,
            formula2=formula2,
            allow_blank=allow_blank,
            showErrorMessage=show_error,
            errorTitle=error_title,
            error=error_message,
            showInputMessage=show_input,
            promptTitle=input_title,
            prompt=input_message,
            showDropDown=not dropdown if validation_type == 'list' else None
        )
        
        # Añadir el rango a la validación
        dv.add(cell_range)
        
        # Añadir validación a la hoja
        ws.add_data_validation(dv)
        
        return dv
    
    except Exception as e:
        if "Unknown operator" in str(e):
            raise ValidationError(f"Operador '{operator}' no válido para validación de tipo '{validation_type}'")
        elif "Invalid formula" in str(e):
            raise ValidationError(f"Fórmula inválida: '{formula1}' o '{formula2}'")
        else:
            raise ValidationError(f"Error al añadir validación de datos: {e}")

def remove_data_validation(ws, cell_range):
    """
    Elimina la validación de datos de un rango de celdas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell_range (str): Rango de celdas en formato A1:B5
        
    Raises:
        ValidationError: Si hay un problema al eliminar la validación
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Verificar que hay validaciones en la hoja
        if not hasattr(ws, 'data_validations') or not ws.data_validations:
            logger.warning(f"No hay validaciones de datos en la hoja")
            return
        
        # Parsear el rango para encontrar todas las celdas afectadas
        start_row, start_col, end_row, end_col = ExcelRange.parse_range(cell_range)
        
        # Ajustar a base 1 para Excel
        start_row += 1
        start_col += 1
        end_row += 1
        end_col += 1
        
        # Crear conjunto de celdas en el rango
        cells_to_remove = set()
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell_ref = f"{get_column_letter(col)}{row}"
                cells_to_remove.add(cell_ref)
        
        # Verificar cada validación y eliminar las celdas del rango
        validations_to_remove = []
        for validation in ws.data_validations.dataValidation:
            # Crear nuevo conjunto de sqref sin las celdas a eliminar
            new_cells = []
            cells_removed = False
            
            # Procesar cada rango en la validación
            if hasattr(validation, 'sqref') and validation.sqref:
                for sq in validation.sqref.split():
                    # Verificar si este rango está completamente dentro del rango a eliminar
                    if ':' in sq:
                        # Rango múltiple como A1:B5
                        sq_start, sq_end = sq.split(':')
                        sq_start_row, sq_start_col = ExcelRange.parse_cell_ref(sq_start)
                        sq_end_row, sq_end_col = ExcelRange.parse_cell_ref(sq_end)
                        
                        # Ajustar a base 1 para Excel
                        sq_start_row += 1
                        sq_start_col += 1
                        sq_end_row += 1
                        sq_end_col += 1
                        
                        # Verificar si hay intersección con el rango a eliminar
                        if (sq_end_row < start_row or sq_start_row > end_row or 
                            sq_end_col < start_col or sq_start_col > end_col):
                            # No hay intersección, mantener este rango
                            new_cells.append(sq)
                        else:
                            # Hay intersección, manejar por celda individual
                            cells_removed = True
                            for r in range(sq_start_row, sq_end_row + 1):
                                for c in range(sq_start_col, sq_end_col + 1):
                                    cell_ref = f"{get_column_letter(c)}{r}"
                                    if cell_ref not in cells_to_remove:
                                        new_cells.append(cell_ref)
                    else:
                        # Celda individual como A1
                        if sq not in cells_to_remove:
                            new_cells.append(sq)
                        else:
                            cells_removed = True
            
            # Actualizar la validación o marcarla para eliminar
            if new_cells:
                validation.sqref = " ".join(new_cells)
            elif cells_removed:
                validations_to_remove.append(validation)
        
        # Eliminar validaciones vacías
        for validation in validations_to_remove:
            ws.data_validations.dataValidation.remove(validation)
        
    except Exception as e:
        raise ValidationError(f"Error al eliminar validación de datos: {e}")

def add_conditional_formatting(ws, cell_range, rule_type, rule_params, style=None):
    """
    Añade formato condicional a un rango de celdas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell_range (str): Rango de celdas en formato A1:B5
        rule_type (str): Tipo de regla ('cellIs', 'expression', 'colorScale', 'dataBar', 'iconSet', 'containsText', 'duplicateValues', etc.)
        rule_params (dict): Parámetros específicos para el tipo de regla
        style (dict, opcional): Diccionario con estilos a aplicar (font, fill, etc.)
        
    Returns:
        ID de la regla añadida
        
    Raises:
        FormattingError: Si hay un problema con el formato condicional
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Crear regla según el tipo
        rule = None
        
        if rule_type == 'cellIs':
            # Regla basada en el valor de la celda
            operator = rule_params.get('operator', 'equal')
            formula = rule_params.get('formula', '')
            value = rule_params.get('value', '')
            
            # Si se proporciona valor, usarlo para la fórmula
            if value and not formula:
                if isinstance(value, str) and not value.startswith('='):
                    # Para strings, añadir comillas
                    formula = f'"{value}"'
                else:
                    formula = str(value)
            
            # Crear regla
            rule = CellIsRule(
                operator=operator,
                formula=[formula] if formula else [],
                stopIfTrue=rule_params.get('stopIfTrue', False)
            )
            
        elif rule_type == 'expression':
            # Regla basada en fórmula
            formula = rule_params.get('formula', '')
            if not formula:
                raise FormattingError("Debe proporcionar una fórmula para reglas de tipo 'expression'")
            
            # Asegurarse de que la fórmula empiece con =
            if not formula.startswith('='):
                formula = f"={formula}"
                
            rule = FormulaRule(
                formula=[formula],
                stopIfTrue=rule_params.get('stopIfTrue', False)
            )
            
        elif rule_type == 'colorScale':
            # Escala de colores
            cfvo1 = rule_params.get('min_value', {'type': 'min', 'val': 0})
            cfvo2 = rule_params.get('mid_value', None)
            cfvo3 = rule_params.get('max_value', {'type': 'max', 'val': 0})
            
            # Colores para la escala
            color1 = rule_params.get('min_color', 'FF638EC6')  # Azul claro
            color2 = rule_params.get('mid_color', None)
            color3 = rule_params.get('max_color', 'FF990000')  # Rojo
            
            if cfvo2 and color2:
                # Escala de 3 colores
                rule = ColorScaleRule(
                    start_type=cfvo1.get('type', 'min'),
                    start_value=cfvo1.get('val', 0),
                    start_color=color1,
                    mid_type=cfvo2.get('type', 'percentile'),
                    mid_value=cfvo2.get('val', 50),
                    mid_color=color2,
                    end_type=cfvo3.get('type', 'max'),
                    end_value=cfvo3.get('val', 0),
                    end_color=color3
                )
            else:
                # Escala de 2 colores
                rule = ColorScaleRule(
                    start_type=cfvo1.get('type', 'min'),
                    start_value=cfvo1.get('val', 0),
                    start_color=color1,
                    end_type=cfvo3.get('type', 'max'),
                    end_value=cfvo3.get('val', 0),
                    end_color=color3
                )
                
        elif rule_type == 'containsText':
            # Contiene texto
            text = rule_params.get('text', '')
            if not text:
                raise FormattingError("Debe proporcionar el texto a buscar para reglas de tipo 'containsText'")
            
            formula = f'NOT(ISERROR(SEARCH("{text}",A1)))'
            rule = FormulaRule(
                formula=[formula],
                stopIfTrue=rule_params.get('stopIfTrue', False)
            )
            
        elif rule_type == 'duplicateValues':
            # Valores duplicados
            from openpyxl.formatting.rule import DuplicateRule
            rule = DuplicateRule(
                stopIfTrue=rule_params.get('stopIfTrue', False)
            )
            
        else:
            raise FormattingError(f"Tipo de regla no soportado: '{rule_type}'")
        
        # Aplicar estilos si se proporcionan
        if style and hasattr(rule, 'dxf') and rule.dxf:
            # Font (fuente)
            if 'font' in style:
                font_style = style['font']
                rule.dxf.font = Font(
                    name=font_style.get('name'),
                    size=font_style.get('size'),
                    bold=font_style.get('bold'),
                    italic=font_style.get('italic'),
                    color=font_style.get('color')
                )
            
            # Fill (relleno)
            if 'fill' in style:
                fill_style = style['fill']
                rule.dxf.fill = PatternFill(
                    fill_type=fill_style.get('type', 'solid'),
                    fgColor=fill_style.get('color')
                )
            
            # Border (borde)
            if 'border' in style:
                border_style = style['border']
                rule.dxf.border = Border(
                    left=Side(style=border_style.get('style'), color=border_style.get('color')),
                    right=Side(style=border_style.get('style'), color=border_style.get('color')),
                    top=Side(style=border_style.get('style'), color=border_style.get('color')),
                    bottom=Side(style=border_style.get('style'), color=border_style.get('color'))
                )
        
        # Si el código ha llegado hasta aquí y rule sigue siendo None, es un error
        if rule is None:
            raise FormattingError(f"No se pudo crear una regla de formato condicional para tipo: '{rule_type}'")
        
        # Añadir regla a la hoja
        ws.conditional_formatting.add(cell_range, rule)
        
        # Identificar la regla añadida (índice en la lista de formatos condicionales)
        rule_id = len(ws.conditional_formatting.cf_rules) - 1
        
        return rule_id
    
    except Exception as e:
        raise FormattingError(f"Error al añadir formato condicional: {e}")

def remove_conditional_formatting(ws, cell_range, rule_id=None):
    """
    Elimina reglas de formato condicional de un rango de celdas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell_range (str): Rango de celdas en formato A1:B5
        rule_id (int, opcional): ID específico de regla a eliminar, o None para todas las reglas del rango
        
    Raises:
        FormattingError: Si hay un problema al eliminar el formato condicional
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Si no hay formatos condicionales, no hay nada que hacer
        if not ws.conditional_formatting:
            logger.warning(f"No hay formatos condicionales en la hoja")
            return
            
        if rule_id is not None:
            # Eliminar una regla específica
            rules_to_remove = []
            for idx, (cf_range, cf_rules) in enumerate(ws.conditional_formatting.cf_rules.items()):
                if idx == rule_id:
                    rules_to_remove.append((cf_range, cf_rules))
            
            for cf_range, cf_rules in rules_to_remove:
                del ws.conditional_formatting.cf_rules[cf_range]
        else:
            # Eliminar todas las reglas que afecten al rango especificado
            rules_to_remove = []
            
            for cf_range, cf_rules in list(ws.conditional_formatting.cf_rules.items()):
                # Comprobar si el rango afecta a cell_range
                if cf_range == cell_range:
                    # Eliminar completamente
                    rules_to_remove.append((cf_range, cf_rules))
                elif ':' in cf_range and ':' in cell_range:
                    # Comprobar si hay intersección
                    cf_min_row, cf_min_col, cf_max_row, cf_max_col = ExcelRange.parse_range(cf_range)
                    cell_min_row, cell_min_col, cell_max_row, cell_max_col = ExcelRange.parse_range(cell_range)
                    
                    # Verificar intersección
                    if (cf_max_row >= cell_min_row and cf_min_row <= cell_max_row and
                        cf_max_col >= cell_min_col and cf_min_col <= cell_max_col):
                        # Hay intersección, eliminar
                        rules_to_remove.append((cf_range, cf_rules))
            
            # Eliminar las reglas identificadas
            for cf_range, cf_rules in rules_to_remove:
                del ws.conditional_formatting.cf_rules[cf_range]
    
    except Exception as e:
        raise FormattingError(f"Error al eliminar formato condicional: {e}")


# 11. Filtros, Ordenación y Agrupación
def set_auto_filter(ws, cell_range):
    """
    Activa el filtro automático en un rango de celdas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell_range (str): Rango de celdas en formato A1:B5
        
    Returns:
        Objeto AutoFilter configurado
        
    Raises:
        FilterError: Si hay un problema con el filtro
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Configurar filtro automático
        ws.auto_filter.ref = cell_range
        return ws.auto_filter
    
    except Exception as e:
        raise FilterError(f"Error al configurar filtro automático: {e}")

def sort_range(ws, cell_range, sort_by, ascending=True):
    """
    Ordena un rango de celdas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell_range (str): Rango de celdas en formato A1:B5
        sort_by (str or list): Columna por la que ordenar (ej: "A") o lista de tuplas [(columna, ascendente), ...]
        ascending (bool): Si es True, ordena ascendente; si es False, descendente
        
    Raises:
        FilterError: Si hay un problema con la ordenación
        
    Note:
        Esta función tiene limitaciones en openpyxl que no permite ordenación directa.
        Para una ordenación completa, se recomienda usar Excel COM si está disponible.
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Parsear el rango
        start_row, start_col, end_row, end_col = ExcelRange.parse_range(cell_range)
        
        # Ajustar a base 1 para Excel
        start_row += 1
        start_col += 1
        end_row += 1
        end_col += 1
        
        # Verificar si se puede usar Excel COM
        if HAS_PYWIN32:
            # Obtener el archivo original
            if hasattr(ws, 'parent') and hasattr(ws.parent, 'path'):
                file_path = ws.parent.path
                
                # Información para registro
                logger.info(f"Usando Excel COM para ordenar rango {cell_range} en {file_path}")
                
                # Crear una aplicación Excel invisible
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                
                try:
                    # Abrir el archivo
                    wb = excel.Workbooks.Open(file_path)
                    excel_ws = wb.Worksheets(ws.title)
                    
                    # Obtener el rango a ordenar
                    excel_range = excel_ws.Range(cell_range)
                    
                    # Configurar la ordenación
                    if isinstance(sort_by, list):
                        # Ordenación por múltiples columnas
                        # Crear SortFields para cada columna
                        for i, (col, asc) in enumerate(sort_by):
                            # Convertir letra de columna a índice relativo dentro del rango
                            col_idx = column_index_from_string(col) - start_col + 1
                            
                            # Añadir campo de ordenación
                            if i == 0:
                                # Primera columna con método Sort
                                excel_range.Sort(
                                    Key1=excel_ws.Cells(start_row, column_index_from_string(col)),
                                    Order1=1 if asc else 2,  # 1=ascendente, 2=descendente
                                    Header=1,  # Incluye encabezado
                                    Orientation=1,  # 1=ordenar por columnas
                                )
                            else:
                                # Columnas adicionales con SortFields
                                field_index = i + 1
                                excel_range.Sort.SortFields.Add(
                                    Key=excel_ws.Cells(start_row, column_index_from_string(col)),
                                    SortOn=0,  # 0=ordenar por valores
                                    Order=1 if asc else 2,
                                    CustomOrder="",
                                    DataOption=0  # 0=ordenar por datos
                                )
                    else:
                        # Ordenación por una sola columna
                        col = sort_by
                        excel_range.Sort(
                            Key1=excel_ws.Cells(start_row, column_index_from_string(col)),
                            Order1=1 if ascending else 2,
                            Header=1,
                            Orientation=1
                        )
                    
                    # Guardar y cerrar
                    wb.Save()
                    wb.Close()
                    
                finally:
                    # Cerrar Excel
                    excel.Quit()
                
                return {
                    'method': 'excel_com',
                    'message': 'Ordenación realizada usando Excel COM'
                }
            else:
                logger.warning("No se puede usar Excel COM porque no se conoce la ruta del archivo")
        
        # Si no se puede usar Excel COM, intentar con openpyxl (limitado)
        logger.warning("La ordenación directa en openpyxl es limitada y puede no funcionar correctamente")
        logger.warning("Se implementará un algoritmo básico de ordenación")
        
        # Extraer datos del rango
        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                row_data.append(ws.cell(row=row, column=col).value)
            data.append(row_data)
        
        # Determinar la columna por la que ordenar
        if isinstance(sort_by, list):
            # Ordenar por múltiples columnas
            sort_indices = []
            for col, asc in sort_by:
                idx = column_index_from_string(col) - start_col
                sort_indices.append((idx, asc))
            
            # Función de ordenación para múltiples columnas
            def custom_sort_key(row):
                return tuple((row[idx] if asc else -row[idx]) for idx, asc in sort_indices)
            
            # Ordenar los datos
            header = data[0]  # Guardar encabezado
            rows = data[1:]  # Filas a ordenar
            rows.sort(key=custom_sort_key)
            data = [header] + rows  # Restaurar encabezado
        else:
            # Ordenar por una sola columna
            sort_idx = column_index_from_string(sort_by) - start_col
            
            # Ordenar los datos
            header = data[0]  # Guardar encabezado
            rows = data[1:]  # Filas a ordenar
            rows.sort(key=lambda row: row[sort_idx], reverse=not ascending)
            data = [header] + rows  # Restaurar encabezado
        
        # Escribir datos ordenados de vuelta al rango
        for i, row_data in enumerate(data):
            for j, cell_value in enumerate(row_data):
                ws.cell(row=start_row + i, column=start_col + j, value=cell_value)
        
        return {
            'method': 'openpyxl',
            'message': 'Ordenación básica realizada. Para mejores resultados, use Excel directamente.'
        }
    
    except Exception as e:
        raise FilterError(f"Error al ordenar rango: {e}")

def group_rows(ws, start_row, end_row):
    """
    Agrupa filas en un rango.
    
    Args:
        ws: Objeto worksheet de openpyxl
        start_row (int): Fila inicial a agrupar
        end_row (int): Fila final a agrupar
        
    Raises:
        GroupingError: Si hay un problema con la agrupación
        
    Note:
        Esta función tiene limitaciones en openpyxl.
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Asegurarse de que los índices son enteros
        start_row = int(start_row)
        end_row = int(end_row)
        
        # Validar rango
        if start_row < 1 or end_row < start_row:
            raise GroupingError(f"Rango de filas inválido: {start_row}-{end_row}")
        
        # Intentar usar openpyxl directamente
        if hasattr(ws, 'row_dimensions'):
            for i in range(start_row, end_row + 1):
                if i in ws.row_dimensions:
                    ws.row_dimensions[i].outlineLevel = 1
            
            # Configurar el resumen abajo (como en Excel)
            ws.sheet_properties.outlinePr.summaryBelow = True
            
            return {
                'method': 'openpyxl',
                'start_row': start_row,
                'end_row': end_row,
                'outline_level': 1
            }
        else:
            raise GroupingError("Esta hoja no soporta agrupación de filas con openpyxl")
    
    except Exception as e:
        raise GroupingError(f"Error al agrupar filas: {e}")

def ungroup_rows(ws, start_row, end_row):
    """
    Desagrupa filas en un rango.
    
    Args:
        ws: Objeto worksheet de openpyxl
        start_row (int): Fila inicial a desagrupar
        end_row (int): Fila final a desagrupar
        
    Raises:
        GroupingError: Si hay un problema con la desagrupación
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Asegurarse de que los índices son enteros
        start_row = int(start_row)
        end_row = int(end_row)
        
        # Validar rango
        if start_row < 1 or end_row < start_row:
            raise GroupingError(f"Rango de filas inválido: {start_row}-{end_row}")
        
        # Intentar usar openpyxl directamente
        if hasattr(ws, 'row_dimensions'):
            for i in range(start_row, end_row + 1):
                if i in ws.row_dimensions:
                    ws.row_dimensions[i].outlineLevel = 0
            
            return {
                'method': 'openpyxl',
                'start_row': start_row,
                'end_row': end_row,
                'outline_level': 0
            }
        else:
            raise GroupingError("Esta hoja no soporta desagrupación de filas con openpyxl")
    
    except Exception as e:
        raise GroupingError(f"Error al desagrupar filas: {e}")

def group_columns(ws, start_col, end_col):
    """
    Agrupa columnas en un rango.
    
    Args:
        ws: Objeto worksheet de openpyxl
        start_col (str): Columna inicial a agrupar (ej: "A")
        end_col (str): Columna final a agrupar (ej: "D")
        
    Raises:
        GroupingError: Si hay un problema con la agrupación
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Convertir letras de columna a índices
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        # Validar rango
        if start_col_idx < 1 or end_col_idx < start_col_idx:
            raise GroupingError(f"Rango de columnas inválido: {start_col}-{end_col}")
        
        # Intentar usar openpyxl directamente
        if hasattr(ws, 'column_dimensions'):
            for i in range(start_col_idx, end_col_idx + 1):
                col_letter = get_column_letter(i)
                if col_letter in ws.column_dimensions:
                    ws.column_dimensions[col_letter].outlineLevel = 1
            
            # Configurar el resumen a la derecha (como en Excel)
            ws.sheet_properties.outlinePr.summaryRight = True
            
            return {
                'method': 'openpyxl',
                'start_col': start_col,
                'end_col': end_col,
                'outline_level': 1
            }
        else:
            raise GroupingError("Esta hoja no soporta agrupación de columnas con openpyxl")
    
    except Exception as e:
        raise GroupingError(f"Error al agrupar columnas: {e}")

def ungroup_columns(ws, start_col, end_col):
    """
    Desagrupa columnas en un rango.
    
    Args:
        ws: Objeto worksheet de openpyxl
        start_col (str): Columna inicial a desagrupar (ej: "A")
        end_col (str): Columna final a desagrupar (ej: "D")
        
    Raises:
        GroupingError: Si hay un problema con la desagrupación
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Convertir letras de columna a índices
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        # Validar rango
        if start_col_idx < 1 or end_col_idx < start_col_idx:
            raise GroupingError(f"Rango de columnas inválido: {start_col}-{end_col}")
        
        # Intentar usar openpyxl directamente
        if hasattr(ws, 'column_dimensions'):
            for i in range(start_col_idx, end_col_idx + 1):
                col_letter = get_column_letter(i)
                if col_letter in ws.column_dimensions:
                    ws.column_dimensions[col_letter].outlineLevel = 0
            
            return {
                'method': 'openpyxl',
                'start_col': start_col,
                'end_col': end_col,
                'outline_level': 0
            }
        else:
            raise GroupingError("Esta hoja no soporta desagrupación de columnas con openpyxl")
    
    except Exception as e:
        raise GroupingError(f"Error al desagrupar columnas: {e}")


# 12. Imágenes y Objetos
def add_image(ws, image_path_or_bytes, cell, width=None, height=None):
    """
    Añade una imagen a la hoja de Excel.
    
    Args:
        ws: Objeto worksheet de openpyxl
        image_path_or_bytes (str or bytes): Ruta al archivo de imagen o bytes de la imagen
        cell (str): Celda de anclaje (ej: "A1")
        width (int, opcional): Ancho de la imagen en píxeles
        height (int, opcional): Alto de la imagen en píxeles
        
    Returns:
        ID de la imagen insertada
        
    Raises:
        ImageError: Si hay un problema con la imagen
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Verificar si se proporciona ruta o bytes
        if isinstance(image_path_or_bytes, str):
            # Es una ruta de archivo
            # Verificar que el archivo existe
            if not os.path.exists(image_path_or_bytes):
                raise ImageError(f"El archivo de imagen no existe: {image_path_or_bytes}")
            
            # Cargar imagen
            img = Image(image_path_or_bytes)
        else:
            # Es un objeto de bytes
            from io import BytesIO
            img = Image(BytesIO(image_path_or_bytes))
        
        # Ajustar tamaño si se proporciona
        if width and height:
            # Calcular la relación de aspecto original
            aspect_ratio = img.width / img.height
            
            # Si solo se proporciona uno, calcular el otro manteniendo aspecto
            if width and not height:
                height = width / aspect_ratio
            elif height and not width:
                width = height * aspect_ratio
            
            # Ajustar tamaño
            img.width = width
            img.height = height
        
        # Añadir la imagen a la hoja
        ws.add_image(img, cell)
        
        # Determinar el ID de la imagen (basado en su posición en la lista)
        image_id = len(ws._images) - 1 if hasattr(ws, '_images') else 0
        
        return image_id
    
    except Exception as e:
        raise ImageError(f"Error al añadir imagen: {e}")

def delete_image(ws, image_id):
    """
    Elimina una imagen de la hoja.
    
    Args:
        ws: Objeto worksheet de openpyxl
        image_id (int): ID de la imagen a eliminar
        
    Raises:
        ImageError: Si hay un problema al eliminar la imagen
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Verificar que hay imágenes en la hoja
        if not hasattr(ws, '_images') or not ws._images:
            raise ImageError("No hay imágenes en la hoja")
        
        # Verificar que el ID es válido
        if image_id < 0 or image_id >= len(ws._images):
            raise ImageError(f"No existe una imagen con ID {image_id}")
        
        # Eliminar la imagen
        del ws._images[image_id]
    
    except Exception as e:
        raise ImageError(f"Error al eliminar imagen: {e}")

def move_object(ws, object_id, new_cell):
    """
    Mueve un objeto (imagen, gráfico) a otra celda.
    
    Args:
        ws: Objeto worksheet de openpyxl
        object_id (str): ID del objeto (formato: "image-0", "chart-2", etc.)
        new_cell (str): Nueva celda de anclaje (ej: "B5")
        
    Raises:
        ImageError: Si hay un problema al mover el objeto
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Parsear el ID del objeto para identificar tipo e índice
        object_parts = object_id.split('-')
        if len(object_parts) != 2:
            raise ImageError(f"Formato de ID de objeto inválido: '{object_id}'. Use formato 'tipo-índice' (ej: 'image-0', 'chart-2')")
        
        object_type = object_parts[0].lower()
        object_index = int(object_parts[1])
        
        if object_type == 'image':
            # Verificar que hay imágenes en la hoja
            if not hasattr(ws, '_images') or not ws._images:
                raise ImageError("No hay imágenes en la hoja")
            
            # Verificar que el índice es válido
            if object_index < 0 or object_index >= len(ws._images):
                raise ImageError(f"No existe una imagen con índice {object_index}")
            
            # Cambiar la posición de la imagen
            ws._images[object_index].anchor = new_cell
            
        elif object_type == 'chart':
            # Verificar que hay gráficos en la hoja
            if not hasattr(ws, '_charts') or not ws._charts:
                raise ImageError("No hay gráficos en la hoja")
            
            # Verificar que el índice es válido
            if object_index < 0 or object_index >= len(ws._charts):
                raise ImageError(f"No existe un gráfico con índice {object_index}")
            
            # Cambiar la posición del gráfico (depende de la estructura interna)
            chart_tuple = ws._charts[object_index]
            if len(chart_tuple) > 1:
                # En algunos casos es una tupla (chart, position)
                chart, position = chart_tuple
                ws._charts[object_index] = (chart, new_cell)
            else:
                # En otros casos el anchor está en el objeto
                chart = chart_tuple[0]
                if hasattr(chart, 'anchor'):
                    chart.anchor = new_cell
                else:
                    raise ImageError("Este gráfico no tiene un anchor que pueda modificarse")
        
        else:
            raise ImageError(f"Tipo de objeto no soportado: '{object_type}'. Tipos soportados: 'image', 'chart'")
    
    except Exception as e:
        raise ImageError(f"Error al mover objeto: {e}")

def list_objects(ws):
    """
    Lista todos los objetos (imágenes, gráficos, tablas) en una hoja.
    
    Args:
        ws: Objeto worksheet de openpyxl
        
    Returns:
        Lista de diccionarios con información de los objetos:
        [{'id': 'image-0', 'type': 'image', 'anchor': 'A1'}, ...]
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    objects_info = []
    
    try:
        # Listar imágenes
        if hasattr(ws, '_images'):
            for i, img in enumerate(ws._images):
                anchor = img.anchor if hasattr(img, 'anchor') else "desconocido"
                objects_info.append({
                    'id': f"image-{i}",
                    'type': 'image',
                    'anchor': anchor,
                    'width': img.width if hasattr(img, 'width') else None,
                    'height': img.height if hasattr(img, 'height') else None
                })
        
        # Listar gráficos
        if hasattr(ws, '_charts'):
            for i, chart_tuple in enumerate(ws._charts):
                # Extraer gráfico y posición
                if len(chart_tuple) > 1:
                    chart, position = chart_tuple
                    anchor = position
                else:
                    chart = chart_tuple[0]
                    anchor = chart.anchor if hasattr(chart, 'anchor') else "desconocido"
                
                # Determinar tipo de gráfico
                from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart
                
                chart_type = "desconocido"
                if isinstance(chart, BarChart):
                    chart_type = "bar" if getattr(chart, "type", None) == "bar" else "column"
                elif isinstance(chart, LineChart):
                    chart_type = "line"
                elif isinstance(chart, PieChart):
                    chart_type = "pie"
                elif isinstance(chart, ScatterChart):
                    chart_type = "scatter"
                
                objects_info.append({
                    'id': f"chart-{i}",
                    'type': 'chart',
                    'chart_type': chart_type,
                    'anchor': anchor,
                    'title': chart.title if hasattr(chart, 'title') and chart.title else f"Chart {i}"
                })
        
        # Listar tablas
        if hasattr(ws, 'tables') and ws.tables:
            for i, (table_name, table) in enumerate(ws.tables.items()):
                objects_info.append({
                    'id': f"table-{i}",
                    'type': 'table',
                    'name': table_name,
                    'ref': table.ref if hasattr(table, 'ref') else "desconocido",
                    'display_name': table.displayName if hasattr(table, 'displayName') else table_name
                })
    
    except Exception as e:
        logger.warning(f"Error al listar objetos: {e}")
    
    return objects_info


# 15. Otras Operaciones Avanzadas
def find_and_replace(ws, find_text, replace_text, range_str=None, case_sensitive=False):
    """
    Busca y reemplaza texto en una hoja.
    
    Args:
        ws: Objeto worksheet de openpyxl
        find_text (str): Texto a buscar
        replace_text (str): Texto de reemplazo
        range_str (str, opcional): Rango en formato A1:B5, o None para toda la hoja
        case_sensitive (bool): Si es True, la búsqueda distingue mayúsculas y minúsculas
        
    Returns:
        Número de reemplazos realizados
        
    Raises:
        OperationError: Si hay un problema con la operación
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Determinar el rango a procesar
        if range_str:
            start_row, start_col, end_row, end_col = ExcelRange.parse_range(range_str)
            
            # Ajustar a base 1 para Excel
            start_row += 1
            start_col += 1
            end_row += 1
            end_col += 1
        else:
            # Toda la hoja
            start_row = 1
            start_col = 1
            end_row = ws.max_row
            end_col = ws.max_column
        
        # Contar reemplazos
        replacements = 0
        
        # Procesar cada celda en el rango
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                
                # Solo procesar celdas con valores de texto
                if cell.value and isinstance(cell.value, str):
                    original_value = cell.value
                    
                    # Realizar el reemplazo según sensibilidad a mayúsculas
                    if case_sensitive:
                        new_value = original_value.replace(find_text, replace_text)
                    else:
                        # Reemplazar ignorando mayúsculas y minúsculas
                        # Para esto, implementamos nuestra propia lógica
                        value_lower = original_value.lower()
                        find_lower = find_text.lower()
                        
                        # Buscar todas las ocurrencias y reemplazar
                        current_pos = 0
                        new_value = original_value
                        
                        while True:
                            pos = value_lower.find(find_lower, current_pos)
                            if pos == -1:
                                break
                            
                            # Reemplazar la instancia encontrada
                            new_value = new_value[:pos] + replace_text + new_value[pos + len(find_text):]
                            
                            # Actualizar las cadenas de búsqueda para la siguiente iteración
                            value_lower = new_value.lower()
                            
                            # Avanzar posición
                            current_pos = pos + len(replace_text)
                    
                    # Verificar si se realizó algún cambio
                    if new_value != original_value:
                        cell.value = new_value
                        replacements += 1
        
        return replacements
    
    except Exception as e:
        raise OperationError(f"Error en búsqueda y reemplazo: {e}")

def set_page_setup(ws, orientation, paper_size, margins):
    """
    Configura opciones de página (orientación, tamaño, márgenes).
    
    Args:
        ws: Objeto worksheet de openpyxl
        orientation (str): Orientación ('portrait' o 'landscape')
        paper_size (str): Tamaño de papel ('letter', 'legal', 'a4', etc.)
        margins (dict): Diccionario con márgenes en pulgadas (top, right, bottom, left, header, footer)
        
    Raises:
        OperationError: Si hay un problema con la configuración
    """
    if not ws:
        raise AdvancedOperationsError("El worksheet no puede ser None")
    
    try:
        # Verificar que la hoja tiene atributos de configuración de página
        if not hasattr(ws, 'page_setup') or not hasattr(ws, 'page_margins'):
            raise OperationError("Esta hoja no tiene atributos de configuración de página")
        
        # Configurar orientación
        if orientation.lower() == 'portrait':
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        elif orientation.lower() == 'landscape':
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        else:
            raise OperationError(f"Orientación inválida: '{orientation}'. Use 'portrait' o 'landscape'")
        
        # Configurar tamaño de papel
        paper_sizes = {
            'letter': ws.PAPERSIZE_LETTER,
            'legal': ws.PAPERSIZE_LEGAL,
            'a4': ws.PAPERSIZE_A4,
            'a3': ws.PAPERSIZE_A3,
            'a5': ws.PAPERSIZE_A5
        }
        
        if paper_size.lower() in paper_sizes:
            ws.page_setup.paperSize = paper_sizes[paper_size.lower()]
        else:
            raise OperationError(f"Tamaño de papel inválido: '{paper_size}'. Tamaños válidos: {list(paper_sizes.keys())}")
        
        # Configurar márgenes
        if isinstance(margins, dict):
            for margin_name, margin_value in margins.items():
                if hasattr(ws.page_margins, margin_name):
                    setattr(ws.page_margins, margin_name, float(margin_value))
                else:
                    logger.warning(f"Margen desconocido: '{margin_name}'")
        else:
            raise OperationError("Los márgenes deben proporcionarse como un diccionario")
    
    except Exception as e:
        raise OperationError(f"Error al configurar opciones de página: {e}")

def add_macro(wb, vba_code):
    """
    Añade código VBA (macro) al libro Excel.
    
    Args:
        wb: Objeto workbook de openpyxl
        vba_code (str): Código VBA a añadir
        
    Raises:
        OperationError: Si hay un problema al añadir la macro o no es soportado
        
    Note:
        Esta función tiene limitaciones en openpyxl.
    """
    if not wb:
        raise AdvancedOperationsError("El workbook no puede ser None")
    
    try:
        # Verificar si el libro ya está en formato XLSM
        if not hasattr(wb, 'vba_archive'):
            logger.warning("Este libro no está en formato XLSM. Se requiere guardar como .xlsm para soportar macros.")
            
            # Intentar habilitar macros si es posible
            wb.is_template = False
            
            # Configurar para que se guarde con extensión xlsm
            if hasattr(wb, '_write_vba'):
                wb._write_vba = True
            else:
                logger.warning("No se pudo habilitar el soporte de VBA en este objeto workbook")
        
        # Verificar si podemos usar Excel COM
        if HAS_PYWIN32:
            # Posible implementación con Excel COM
            logger.info("Se intentará añadir la macro usando Excel COM (VBA)")
            
            # Obtener el archivo original
            if hasattr(wb, 'path'):
                file_path = wb.path
                
                # Comprobar que la extensión es .xlsm
                if not file_path.lower().endswith('.xlsm'):
                    new_path = os.path.splitext(file_path)[0] + '.xlsm'
                    logger.warning(f"El archivo debe guardarse como .xlsm para soportar macros. Se sugiere guardar como: {new_path}")
                    
                    # Intentar guardar con nueva extensión
                    wb.save(new_path)
                    file_path = new_path
                
                # Crear una aplicación Excel invisible
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                
                try:
                    # Abrir el archivo
                    workbook = excel.Workbooks.Open(file_path)
                    
                    # Añadir módulo VBA
                    vbcomponent = workbook.VBProject.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
                    vbcomponent.CodeModule.AddFromString(vba_code)
                    
                    # Guardar y cerrar
                    workbook.Save()
                    workbook.Close()
                    
                    return {
                        'method': 'excel_com',
                        'message': 'Macro añadida correctamente usando Excel COM'
                    }
                
                finally:
                    # Cerrar Excel
                    excel.Quit()
            else:
                logger.warning("No se puede usar Excel COM porque no se conoce la ruta del archivo")
        
        # Si no se puede usar Excel COM, intentar con openpyxl (muy limitado)
        logger.warning("Añadir macros con openpyxl tiene funcionalidad muy limitada")
        logger.warning("Se recomienda añadir macros directamente desde Excel")
        
        return {
            'method': 'unsupported',
            'message': 'Añadir macros con openpyxl no está completamente soportado. Use Excel directamente.'
        }
    
    except Exception as e:
        raise OperationError(f"Error al añadir macro: {e}")

def save_as_pdf(wb, filename, sheet_names=None):
    """
    Guarda hojas de Excel como PDF.
    
    Args:
        wb: Objeto workbook de openpyxl
        filename (str): Nombre del archivo PDF a crear
        sheet_names (list, opcional): Lista de nombres de hojas a incluir, o None para todas
        
    Raises:
        OperationError: Si hay un problema al guardar como PDF
        
    Note:
        Esta función requiere Excel (COM) instalado para funcionar.
    """
    if not wb:
        raise AdvancedOperationsError("El workbook no puede ser None")
    
    try:
        # Verificar que tenemos Excel COM disponible
        if not HAS_PYWIN32:
            raise OperationError("Se requiere Excel COM (pywin32) para exportar a PDF")
        
        # Obtener el archivo original
        if not hasattr(wb, 'path'):
            # Si el workbook no tiene path, debemos guardarlo primero
            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            temp_file.close()
            wb.save(temp_file.name)
            file_path = temp_file.name
        else:
            file_path = wb.path
        
        # Crear una aplicación Excel invisible
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        
        try:
            # Abrir el archivo
            workbook = excel.Workbooks.Open(file_path)
            
            # Determinar qué hojas exportar
            if sheet_names:
                # Comprobar que las hojas existen
                for sheet_name in sheet_names:
                    try:
                        workbook.Worksheets(sheet_name)
                    except:
                        raise OperationError(f"Hoja no encontrada: '{sheet_name}'")
                
                # Exportar solo las hojas seleccionadas
                sheets_to_export = []
                for sheet_name in sheet_names:
                    sheets_to_export.append(workbook.Worksheets(sheet_name).Index)
            else:
                # Exportar todas las hojas
                sheets_to_export = None
            
            # Exportar a PDF
            if sheets_to_export:
                workbook.Worksheets(sheets_to_export).Select()
            else:
                workbook.Worksheets.Select()
            
            # Guardar como PDF
            workbook.ActiveSheet.ExportAsFixedFormat(
                Type=0,  # 0 = PDF
                Filename=filename,
                Quality=0,  # 0 = Standard
                IncludeDocProperties=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False
            )
            
            # Cerrar sin guardar cambios
            workbook.Close(SaveChanges=False)
            
            # Limpiar archivo temporal si se creó
            if not hasattr(wb, 'path'):
                os.unlink(file_path)
            
            return {
                'method': 'excel_com',
                'filename': filename,
                'sheets': sheet_names if sheet_names else 'all',
                'message': 'Archivo PDF creado correctamente'
            }
        
        finally:
            # Cerrar Excel
            excel.Quit()
    
    except Exception as e:
        # Limpiar archivo temporal si se creó y hubo error
        if not hasattr(wb, 'path') and 'file_path' in locals() and os.path.exists(file_path):
            os.unlink(file_path)
        
        raise OperationError(f"Error al guardar como PDF: {e}")


# Crear el servidor MCP como variable global
mcp = None
if HAS_MCP:
    # Esta es la variable global que el sistema MCP busca
    mcp = FastMCP("Excel Advanced Operations MCP", 
                 dependencies=["openpyxl", "Pillow"])
    logger.info("Servidor MCP iniciado correctamente")
    
    # 10. Validación de Datos y Formato Condicional - Herramientas MCP
    @mcp.tool(description="Añade validación de datos a un rango de celdas")
    def add_data_validation_tool(file_path, sheet_name, cell_range, validation_type, criteria):
        """Añade validación de datos a un rango de celdas"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Convertir criterios de JSON a diccionario si es necesario
            if isinstance(criteria, str):
                try:
                    criteria = json.loads(criteria)
                except:
                    criteria = {'value1': criteria}
            
            # Añadir validación
            validation = add_data_validation(ws, cell_range, validation_type, criteria)
            
            # Guardar cambios
            wb.save(file_path)
            
            # Preparar mensaje según el tipo de validación
            if validation_type == 'list':
                values = criteria.get('value1', []) if isinstance(criteria.get('value1'), list) else criteria.get('formula1', '')
                message = f"Validación de lista añadida al rango {cell_range} con valores: {values}"
            elif validation_type in ['whole', 'decimal']:
                operator = criteria.get('operator', '')
                value1 = criteria.get('value1', criteria.get('formula1', ''))
                value2 = criteria.get('value2', criteria.get('formula2', ''))
                message = f"Validación numérica ({validation_type}) añadida al rango {cell_range}: {operator} {value1}"
                if value2:
                    message += f" y {value2}"
            else:
                message = f"Validación de tipo '{validation_type}' añadida correctamente al rango {cell_range}"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "validation_type": validation_type,
                "message": message
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al añadir validación de datos: {e}"
            }
    
    @mcp.tool(description="Elimina la validación de datos de un rango de celdas")
    def remove_data_validation_tool(file_path, sheet_name, cell_range):
        """Elimina la validación de datos de un rango de celdas"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Eliminar validación
            remove_data_validation(ws, cell_range)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "message": f"Validación de datos eliminada del rango {cell_range}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al eliminar validación de datos: {e}"
            }
    
    @mcp.tool(description="Añade formato condicional a un rango de celdas")
    def add_conditional_formatting_tool(file_path, sheet_name, cell_range, rule_type, rule_params, style=None):
        """Añade formato condicional a un rango de celdas"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Convertir params y style de JSON a diccionario si es necesario
            if isinstance(rule_params, str):
                try:
                    rule_params = json.loads(rule_params)
                except:
                    # Si no es JSON válido, intentar un valor simple
                    rule_params = {'value': rule_params}
            
            if isinstance(style, str):
                try:
                    style = json.loads(style)
                except:
                    # Si no es JSON válido, intentar un color simple
                    style = {'fill': {'color': style, 'type': 'solid'}}
            
            # Añadir formato condicional
            rule_id = add_conditional_formatting(ws, cell_range, rule_type, rule_params, style)
            
            # Guardar cambios
            wb.save(file_path)
            
            # Preparar mensaje según el tipo de regla
            if rule_type == 'cellIs':
                operator = rule_params.get('operator', '')
                value = rule_params.get('value', rule_params.get('formula', ''))
                message = f"Formato condicional 'cellIs' añadido al rango {cell_range}: {operator} {value}"
            elif rule_type == 'expression':
                formula = rule_params.get('formula', '')
                message = f"Formato condicional basado en fórmula añadido al rango {cell_range}: {formula}"
            elif rule_type == 'colorScale':
                message = f"Escala de colores añadida al rango {cell_range}"
            else:
                message = f"Formato condicional de tipo '{rule_type}' añadido correctamente al rango {cell_range}"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "rule_type": rule_type,
                "rule_id": rule_id,
                "message": message
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al añadir formato condicional: {e}"
            }
    
    @mcp.tool(description="Elimina reglas de formato condicional de un rango de celdas")
    def remove_conditional_formatting_tool(file_path, sheet_name, cell_range, rule_id=None):
        """Elimina reglas de formato condicional de un rango de celdas"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Convertir rule_id a entero si es posible y no es None
            if rule_id is not None:
                try:
                    rule_id = int(rule_id)
                except:
                    # Si no es un entero válido, generar error
                    raise FormattingError(f"ID de regla inválido: '{rule_id}'. Debe ser un número entero")
            
            # Eliminar formato condicional
            remove_conditional_formatting(ws, cell_range, rule_id)
            
            # Guardar cambios
            wb.save(file_path)
            
            # Preparar mensaje según si se eliminó una regla específica o todas
            if rule_id is not None:
                message = f"Regla de formato condicional con ID {rule_id} eliminada"
            else:
                message = f"Todas las reglas de formato condicional eliminadas del rango {cell_range}"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "rule_id": rule_id,
                "message": message
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al eliminar formato condicional: {e}"
            }
    
    # 11. Filtros, Ordenación y Agrupación - Herramientas MCP
    @mcp.tool(description="Activa el filtro automático en un rango de celdas")
    def set_auto_filter_tool(file_path, sheet_name, cell_range):
        """Activa el filtro automático en un rango de celdas"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Configurar filtro automático
            auto_filter = set_auto_filter(ws, cell_range)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "message": f"Filtro automático activado en el rango {cell_range}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al activar filtro automático: {e}"
            }
    
    @mcp.tool(description="Ordena un rango de celdas")
    def sort_range_tool(file_path, sheet_name, cell_range, sort_by, ascending=True):
        """Ordena un rango de celdas"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Convertir sort_by de JSON a lista si es necesario
            if isinstance(sort_by, str):
                try:
                    # Intentar como JSON
                    sort_by_data = json.loads(sort_by)
                    # Comprobar si es una lista o un valor simple
                    if isinstance(sort_by_data, list):
                        sort_by = sort_by_data
                    else:
                        # Es un valor simple (probablemente una columna)
                        sort_by = sort_by
                except:
                    # Si no es JSON válido, es una columna
                    sort_by = sort_by
            
            # Convertir ascending a booleano
            if isinstance(ascending, str):
                ascending = ascending.lower() in ['true', 't', 'yes', 'y', '1']
            
            # Ordenar rango
            result = sort_range(ws, cell_range, sort_by, ascending)
            
            # Guardar cambios
            wb.save(file_path)
            
            # Preparar mensaje según el resultado
            if isinstance(sort_by, list):
                columns_info = ", ".join([f"{col} ({'asc' if asc else 'desc'})" for col, asc in sort_by])
                message = f"Rango {cell_range} ordenado por múltiples columnas: {columns_info}"
            else:
                message = f"Rango {cell_range} ordenado por columna {sort_by} en orden {'ascendente' if ascending else 'descendente'}"
            
            # Añadir detalles del método usado
            if isinstance(result, dict) and 'method' in result:
                message += f" (método: {result['method']})"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "sort_by": sort_by,
                "ascending": ascending,
                "method": result.get('method') if isinstance(result, dict) else 'default',
                "message": message
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al ordenar rango: {e}"
            }
    
    @mcp.tool(description="Agrupa filas en un rango")
    def group_rows_tool(file_path, sheet_name, start_row, end_row):
        """Agrupa filas en un rango"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Convertir a enteros
            start_row = int(start_row)
            end_row = int(end_row)
            
            # Agrupar filas
            result = group_rows(ws, start_row, end_row)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "start_row": start_row,
                "end_row": end_row,
                "method": result.get('method') if isinstance(result, dict) else 'default',
                "message": f"Filas {start_row}-{end_row} agrupadas correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al agrupar filas: {e}"
            }
    
    @mcp.tool(description="Desagrupa filas en un rango")
    def ungroup_rows_tool(file_path, sheet_name, start_row, end_row):
        """Desagrupa filas en un rango"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Convertir a enteros
            start_row = int(start_row)
            end_row = int(end_row)
            
            # Desagrupar filas
            result = ungroup_rows(ws, start_row, end_row)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "start_row": start_row,
                "end_row": end_row,
                "method": result.get('method') if isinstance(result, dict) else 'default',
                "message": f"Filas {start_row}-{end_row} desagrupadas correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al desagrupar filas: {e}"
            }
    
    @mcp.tool(description="Agrupa columnas en un rango")
    def group_columns_tool(file_path, sheet_name, start_col, end_col):
        """Agrupa columnas en un rango"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Verificar que son letras de columna válidas
            if not all(c.isalpha() for c in start_col + end_col):
                raise GroupingError(f"Las columnas deben especificarse como letras (ej: 'A', 'BC')")
            
            # Agrupar columnas
            result = group_columns(ws, start_col, end_col)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "start_col": start_col,
                "end_col": end_col,
                "method": result.get('method') if isinstance(result, dict) else 'default',
                "message": f"Columnas {start_col}-{end_col} agrupadas correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al agrupar columnas: {e}"
            }
    
    @mcp.tool(description="Desagrupa columnas en un rango")
    def ungroup_columns_tool(file_path, sheet_name, start_col, end_col):
        """Desagrupa columnas en un rango"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Verificar que son letras de columna válidas
            if not all(c.isalpha() for c in start_col + end_col):
                raise GroupingError(f"Las columnas deben especificarse como letras (ej: 'A', 'BC')")
            
            # Desagrupar columnas
            result = ungroup_columns(ws, start_col, end_col)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "start_col": start_col,
                "end_col": end_col,
                "method": result.get('method') if isinstance(result, dict) else 'default',
                "message": f"Columnas {start_col}-{end_col} desagrupadas correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al desagrupar columnas: {e}"
            }
    
    # 12. Imágenes y Objetos - Herramientas MCP
    @mcp.tool(description="Añade una imagen a la hoja de Excel")
    def add_image_tool(file_path, sheet_name, image_path_or_bytes, cell, width=None, height=None):
        """Añade una imagen a la hoja de Excel"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Convertir width y height a enteros si no son None
            if width is not None:
                width = int(width)
            
            if height is not None:
                height = int(height)
            
            # Añadir imagen
            image_id = add_image(ws, image_path_or_bytes, cell, width, height)
            
            # Guardar cambios
            wb.save(file_path)
            
            # Preparar mensaje con detalles
            size_info = ""
            if width and height:
                size_info = f" con tamaño {width}x{height} píxeles"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell": cell,
                "image_id": image_id,
                "width": width,
                "height": height,
                "message": f"Imagen añadida correctamente en la celda {cell}{size_info} (ID: image-{image_id})"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al añadir imagen: {e}"
            }
    
    @mcp.tool(description="Elimina una imagen de la hoja")
    def delete_image_tool(file_path, sheet_name, image_id):
        """Elimina una imagen de la hoja"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Extraer el índice si se proporciona en formato 'image-N'
            if isinstance(image_id, str) and image_id.startswith('image-'):
                try:
                    image_id = int(image_id.split('-')[1])
                except:
                    raise ImageError(f"ID de imagen inválido: '{image_id}'. Formato esperado: 'image-N' o un número entero")
            else:
                # Intentar convertir a entero
                image_id = int(image_id)
            
            # Eliminar imagen
            delete_image(ws, image_id)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "image_id": image_id,
                "message": f"Imagen con ID {image_id} eliminada correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al eliminar imagen: {e}"
            }
    
    @mcp.tool(description="Mueve un objeto (imagen, gráfico) a otra celda")
    def move_object_tool(file_path, sheet_name, object_id, new_cell):
        """Mueve un objeto (imagen, gráfico) a otra celda"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Mover objeto
            move_object(ws, object_id, new_cell)
            
            # Guardar cambios
            wb.save(file_path)
            
            # Extraer tipo y nombre para el mensaje
            object_parts = object_id.split('-')
            object_type = object_parts[0] if len(object_parts) > 0 else "objeto"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "object_id": object_id,
                "new_cell": new_cell,
                "message": f"{object_type.capitalize()} con ID {object_id} movido a la celda {new_cell}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al mover objeto: {e}"
            }
    
    @mcp.tool(description="Lista todos los objetos (imágenes, gráficos, tablas) en una hoja")
    def list_objects_tool(file_path, sheet_name):
        """Lista todos los objetos (imágenes, gráficos, tablas) en una hoja"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Listar objetos
            objects = list_objects(ws)
            
            # Contar por tipo
            count_by_type = {}
            for obj in objects:
                obj_type = obj['type']
                count_by_type[obj_type] = count_by_type.get(obj_type, 0) + 1
            
            # Preparar mensaje de resumen
            type_summary = ", ".join([f"{count} {obj_type}{'s' if count > 1 else ''}" for obj_type, count in count_by_type.items()])
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "objects": objects,
                "count": len(objects),
                "count_by_type": count_by_type,
                "message": f"Se encontraron {len(objects)} objetos en la hoja {sheet_name}: {type_summary}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al listar objetos: {e}"
            }
    
    # 15. Otras Operaciones Avanzadas - Herramientas MCP
    @mcp.tool(description="Busca y reemplaza texto en una hoja")
    def find_and_replace_tool(file_path, sheet_name, find_text, replace_text, range_str=None, case_sensitive=False):
        """Busca y reemplaza texto en una hoja"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Convertir case_sensitive a booleano
            if isinstance(case_sensitive, str):
                case_sensitive = case_sensitive.lower() in ['true', 't', 'yes', 'y', '1']
            
            # Buscar y reemplazar
            replacements = find_and_replace(ws, find_text, replace_text, range_str, case_sensitive)
            
            # Guardar cambios
            wb.save(file_path)
            
            # Preparar mensaje con detalles
            range_info = f" en el rango {range_str}" if range_str else " en toda la hoja"
            case_info = " (distinguiendo mayúsculas y minúsculas)" if case_sensitive else ""
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "find_text": find_text,
                "replace_text": replace_text,
                "range": range_str,
                "case_sensitive": case_sensitive,
                "replacements": replacements,
                "message": f"Se realizaron {replacements} reemplazos de '{find_text}' por '{replace_text}'{range_info}{case_info}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error en búsqueda y reemplazo: {e}"
            }
    
    @mcp.tool(description="Configura opciones de página (orientación, tamaño, márgenes)")
    def set_page_setup_tool(file_path, sheet_name, orientation, paper_size, margins):
        """Configura opciones de página (orientación, tamaño, márgenes)"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Convertir margins de JSON a diccionario si es necesario
            if isinstance(margins, str):
                try:
                    margins = json.loads(margins)
                except:
                    raise OperationError(f"Formato de márgenes inválido: '{margins}'. Debe ser un objeto JSON con valores numéricos")
            
            # Configurar opciones de página
            set_page_setup(ws, orientation, paper_size, margins)
            
            # Guardar cambios
            wb.save(file_path)
            
            # Preparar mensaje con detalles
            margins_info = ", ".join([f"{k}: {v}" for k, v in margins.items()])
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "orientation": orientation,
                "paper_size": paper_size,
                "margins": margins,
                "message": f"Opciones de página configuradas: orientación '{orientation}', tamaño '{paper_size}', márgenes ({margins_info})"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al configurar opciones de página: {e}"
            }
    
    @mcp.tool(description="Añade código VBA (macro) al libro Excel")
    def add_macro_tool(file_path, vba_code):
        """Añade código VBA (macro) al libro Excel"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Añadir macro
            result = add_macro(wb, vba_code)
            
            # Guardar cambios si no se usó Excel COM
            if result.get('method') != 'excel_com':
                wb.save(file_path)
            
            # Verificar extensión
            is_xlsm = file_path.lower().endswith('.xlsm')
            
            return {
                "success": True,
                "file_path": file_path,
                "method": result.get('method', 'unknown'),
                "message": result.get('message', 'Macro añadida correctamente'),
                "warning": None if is_xlsm else "El archivo debe guardarse con extensión .xlsm para soportar macros"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al añadir macro: {e}"
            }
    
    @mcp.tool(description="Guarda hojas de Excel como PDF")
    def save_as_pdf_tool(file_path, filename, sheet_names=None):
        """Guarda hojas de Excel como PDF"""
        try:
            # Convertir sheet_names de JSON a lista si es necesario
            if isinstance(sheet_names, str):
                try:
                    # Intentar como JSON
                    sheet_names = json.loads(sheet_names)
                except:
                    # Si no es JSON válido, intentar como una lista separada por comas
                    sheet_names = [s.strip() for s in sheet_names.split(',')]
            
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Guardar como PDF
            result = save_as_pdf(wb, filename, sheet_names)
            
            return {
                "success": True,
                "file_path": file_path,
                "pdf_file": filename,
                "sheets": sheet_names if sheet_names else "all",
                "method": result.get('method', 'unknown'),
                "message": result.get('message', f"Archivo PDF '{filename}' creado correctamente")
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al guardar como PDF: {e}"
            }

if __name__ == "__main__":
    # Código de ejemplo de uso
    logger.info("Excel Advanced Operations MCP - Ejemplo de uso")
    
    # Verificar argumentos
    if len(sys.argv) > 1:
        comando = sys.argv[1].lower()
        
        if comando == "validacion" and len(sys.argv) > 4:
            archivo, hoja, rango = sys.argv[2], sys.argv[3], sys.argv[4]
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb[hoja]
                
                # Añadir validación
                criteria = {'operator': 'between', 'formula1': '1', 'formula2': '100'}
                add_data_validation(ws, rango, 'whole', criteria)
                
                wb.save(archivo)
                logger.info(f"Validación añadida correctamente al rango {rango}")
            except Exception as e:
                logger.error(f"Error: {e}")
                
        elif comando == "formato" and len(sys.argv) > 4:
            archivo, hoja, rango = sys.argv[2], sys.argv[3], sys.argv[4]
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb[hoja]
                
                # Añadir formato condicional
                rule_params = {'operator': 'greaterThan', 'formula': '50'}
                style = {'fill': {'color': 'FF6D9EEB', 'type': 'solid'}}
                add_conditional_formatting(ws, rango, 'cellIs', rule_params, style)
                
                wb.save(archivo)
                logger.info(f"Formato condicional añadido correctamente al rango {rango}")
            except Exception as e:
                logger.error(f"Error: {e}")
                
        elif comando == "filtro" and len(sys.argv) > 4:
            archivo, hoja, rango = sys.argv[2], sys.argv[3], sys.argv[4]
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb[hoja]
                
                # Configurar filtro automático
                set_auto_filter(ws, rango)
                
                wb.save(archivo)
                logger.info(f"Filtro automático configurado correctamente en el rango {rango}")
            except Exception as e:
                logger.error(f"Error: {e}")
                
        elif comando == "imagen" and len(sys.argv) > 4:
            archivo, hoja, imagen, celda = sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5]
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb[hoja]
                
                # Añadir imagen
                add_image(ws, imagen, celda)
                
                wb.save(archivo)
                logger.info(f"Imagen añadida correctamente en la celda {celda}")
            except Exception as e:
                logger.error(f"Error: {e}")
                
        else:
            logger.info("Comando no reconocido o faltan argumentos")
            logger.info("Uso: python excel_advanced_operations_mcp.py [validacion|formato|filtro|imagen] archivo.xlsx hoja [args adicionales]")
    else:
        logger.info("Uso: python excel_advanced_operations_mcp.py [validacion|formato|filtro|imagen] archivo.xlsx hoja [args adicionales]")
