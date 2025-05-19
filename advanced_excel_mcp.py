# -*- coding: utf-8 -*-

"""
Advanced Excel MCP (Multi-purpose Connector for Python with Excel)
-------------------------------------------------------
Biblioteca para operaciones avanzadas en Excel:
- Tablas (Excel Tables)
- Fórmulas y Cálculos
- Gráficos (Charts)
- Tablas Dinámicas (Pivot Tables)

Author: MCP Team
Version: 1.0
"""

import os
import sys
import json
import logging
from typing import List, Dict, Union, Optional, Tuple, Any, Callable

# Configuración de logging
logger = logging.getLogger("advanced_excel_mcp")
logger.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stderr)
handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
logger.addHandler(handler)

# Importar MCP
try:
    from mcp.server.fastmcp import FastMCP
    HAS_MCP = True
except ImportError:
    logger.warning("No se pudo importar FastMCP. Las funcionalidades de servidor MCP no estarán disponibles.")
    HAS_MCP = False

# Intentar importar las bibliotecas necesarias
try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, ScatterChart, AreaChart,
        Reference, Series
    )
    from openpyxl.pivot.table import PivotTable, PivotField
    from openpyxl.pivot.cache import PivotCache
    HAS_OPENPYXL = True
except ImportError as e:
    logger.warning(f"Error al importar bibliotecas esenciales: {e}")
    logger.warning("Es posible que algunas funcionalidades no estén disponibles")
    HAS_OPENPYXL = False

# Intentar importar pywin32 para uso alternativo con Excel
try:
    import win32com.client
    HAS_PYWIN32 = True
except ImportError:
    logger.warning("No se pudo importar win32com.client. La funcionalidad alternativa de Excel COM no estará disponible.")
    HAS_PYWIN32 = False

# Excepciones personalizadas
class AdvancedExcelError(Exception):
    """Excepción base para todos los errores de Advanced Excel MCP."""
    pass

class TableError(AdvancedExcelError):
    """Se lanza cuando hay un problema con una tabla de Excel."""
    pass

class ChartError(AdvancedExcelError):
    """Se lanza cuando hay un problema con un gráfico."""
    pass

class FormulaError(AdvancedExcelError):
    """Se lanza cuando hay un problema con una fórmula."""
    pass

class PivotTableError(AdvancedExcelError):
    """Se lanza cuando hay un problema con una tabla dinámica."""
    pass

class RangeError(AdvancedExcelError):
    """Se lanza cuando hay un problema con un rango de celdas."""
    pass

class SheetNotFoundError(AdvancedExcelError):
    """Se lanza cuando no se encuentra una hoja en el archivo Excel."""
    pass


# Mapeo de nombres de estilo a números de estilo de Excel
CHART_STYLE_NAMES = {
    # Estilos claros
    'light-1': 1, 'light-2': 2, 'light-3': 3, 'light-4': 4, 'light-5': 5, 'light-6': 6,
    'office-1': 1, 'office-2': 2, 'office-3': 3, 'office-4': 4, 'office-5': 5, 'office-6': 6,
    'white': 1, 'minimal': 2, 'soft': 3, 'gradient': 4, 'muted': 5, 'outlined': 6,
    
    # Estilos oscuros
    'dark-1': 7, 'dark-2': 8, 'dark-3': 9, 'dark-4': 10, 'dark-5': 11, 'dark-6': 12, 
    'dark-blue': 7, 'dark-gray': 8, 'dark-green': 9, 'dark-red': 10, 'dark-purple': 11, 'dark-orange': 12,
    'navy': 7, 'charcoal': 8, 'forest': 9, 'burgundy': 10, 'indigo': 11, 'rust': 12,
    
    # Estilos coloridos
    'colorful-1': 13, 'colorful-2': 14, 'colorful-3': 15, 'colorful-4': 16, 
    'colorful-5': 17, 'colorful-6': 18, 'colorful-7': 19, 'colorful-8': 20,
    'bright': 13, 'vivid': 14, 'rainbow': 15, 'multi': 16, 'contrast': 17, 'vibrant': 18,
    
    # Temas de Office
    'ion-1': 21, 'ion-2': 22, 'ion-3': 23, 'ion-4': 24,
    'wisp-1': 25, 'wisp-2': 26, 'wisp-3': 27, 'wisp-4': 28,
    'aspect-1': 29, 'aspect-2': 30, 'aspect-3': 31, 'aspect-4': 32,
    'badge-1': 33, 'badge-2': 34, 'badge-3': 35, 'badge-4': 36,
    'gallery-1': 37, 'gallery-2': 38, 'gallery-3': 39, 'gallery-4': 40,
    'median-1': 41, 'median-2': 42, 'median-3': 43, 'median-4': 44,
    
    # Estilos para tipos específicos de gráficos
    'column-default': 1, 'column-dark': 7, 'column-colorful': 13, 
    'bar-default': 1, 'bar-dark': 7, 'bar-colorful': 13,
    'line-default': 1, 'line-dark': 7, 'line-markers': 3, 'line-dash': 5,
    'pie-default': 1, 'pie-dark': 7, 'pie-explosion': 4, 'pie-3d': 10,
    'area-default': 1, 'area-dark': 7, 'area-transparent': 5, 'area-stacked': 9,
    'scatter-default': 1, 'scatter-dark': 7, 'scatter-bubble': 4, 'scatter-smooth': 9,
}

# Mapeo entre estilos y paletas de colores recomendadas
STYLE_TO_PALETTE = {
    # Estilos claros (1-6)
    1: 'office', 2: 'office', 3: 'colorful', 4: 'colorful', 5: 'pastel', 6: 'pastel',
    # Estilos oscuros (7-12)
    7: 'dark-blue', 8: 'dark-gray', 9: 'dark-green', 10: 'dark-red', 11: 'dark-purple', 12: 'dark-orange',
    # Estilos coloridos (13-20)
    13: 'colorful', 14: 'colorful', 15: 'colorful', 16: 'colorful', 
    17: 'colorful', 18: 'colorful', 19: 'colorful', 20: 'colorful',
    # Temas de Office (21-48)
    21: 'ion', 22: 'ion', 23: 'ion', 24: 'ion',
    25: 'wisp', 26: 'wisp', 27: 'wisp', 28: 'wisp',
    29: 'aspect', 30: 'aspect', 31: 'aspect', 32: 'aspect',
    33: 'badge', 34: 'badge', 35: 'badge', 36: 'badge',
    37: 'gallery', 38: 'gallery', 39: 'gallery', 40: 'gallery',
    41: 'office', 42: 'office', 43: 'office', 44: 'office',
    45: 'default', 46: 'default', 47: 'default', 48: 'default'
}

# Mapeo de temas de colores de Excel a sus valores hexadecimales
EXCEL_COLOR_THEMES = {
    # Tema Office (predeterminado)
    'office': {
        'text1': '000000', 'text2': 'FFFFFF',
        'accent1': '4472C4', 'accent2': 'ED7D31', 'accent3': 'A5A5A5',
        'accent4': 'FFC000', 'accent5': '5B9BD5', 'accent6': '70AD47',
        'hyperlink': '0563C1', 'followed_hyperlink': '954F72'
    },
    # Tema Ion
    'ion': {
        'text1': '1F3853', 'text2': 'FFFFFF',
        'accent1': '2996CC', 'accent2': 'A7BFDE', 'accent3': 'C8D4E1',
        'accent4': 'FFE699', 'accent5': 'FF9900', 'accent6': 'CCFF99',
        'hyperlink': '045A8D', 'followed_hyperlink': '98333F'
    },
    # Tema Wisp
    'wisp': {
        'text1': '4C4C4C', 'text2': 'FFFFFF',
        'accent1': 'A0CBE8', 'accent2': 'F4B183', 'accent3': 'C5E0B3',
        'accent4': 'FFE699', 'accent5': 'B4A7D6', 'accent6': 'FA8072',
        'hyperlink': '1155CC', 'followed_hyperlink': '6699CC'
    },
    # Tema Aspect
    'aspect': {
        'text1': '002060', 'text2': 'FFFFFF',
        'accent1': '008BD5', 'accent2': 'FFC000', 'accent3': '92CDDC',
        'accent4': 'FF9C00', 'accent5': 'C1DFF0', 'accent6': '8FBC8F',
        'hyperlink': '054186', 'followed_hyperlink': '954F72'
    },
    # Tema Badge
    'badge': {
        'text1': '252525', 'text2': 'FFFFFF',
        'accent1': 'FF8C00', 'accent2': 'DA3A2A', 'accent3': '65A1D1',
        'accent4': '007A33', 'accent5': 'FFFF00', 'accent6': '4169E1',
        'hyperlink': '0000EE', 'followed_hyperlink': 'EE82EE'
    },
    # Tema Gallery
    'gallery': {
        'text1': '404040', 'text2': 'FFFFFF',
        'accent1': '5170B0', 'accent2': 'A65D1F', 'accent3': '7F7F7F',
        'accent4': 'E1692C', 'accent5': '648FA6', 'accent6': '417808',
        'hyperlink': '045FB4', 'followed_hyperlink': '5F04B4'
    },
    # Tema Median
    'median': {
        'text1': '0B243B', 'text2': 'FFFFFF',
        'accent1': '00B0F0', 'accent2': 'FFC000', 'accent3': '92D050',
        'accent4': 'FF0000', 'accent5': 'C000C0', 'accent6': 'FF9900',
        'hyperlink': '002060', 'followed_hyperlink': '954F72'
    }
}

# Paletas de colores más específicas para diferentes estilos de gráficos
CHART_COLOR_SCHEMES = {
    # Estilos de color estándar
    'default': ['4472C4', 'ED7D31', 'A5A5A5', 'FFC000', '5B9BD5', '70AD47', '8549BA', 'C55A11'],
    'colorful': ['5B9BD5', 'ED7D31', 'A5A5A5', 'FFC000', '4472C4', '70AD47', '264478', '9E480E'],
    'monochrome': ['000000', '404040', '808080', 'BFBFBF', 'FFFFFF', 'D9D9D9', '595959', '262626'],
    'pastel': ['9DC3E6', 'FFD966', 'C5E0B3', 'F4B183', 'B4A7D6', '8FBCDB', 'D89595', 'B7B7B7'],
    
    # Estilos basados en los temas
    'office': [EXCEL_COLOR_THEMES['office']['accent1'], EXCEL_COLOR_THEMES['office']['accent2'], 
              EXCEL_COLOR_THEMES['office']['accent3'], EXCEL_COLOR_THEMES['office']['accent4'],
              EXCEL_COLOR_THEMES['office']['accent5'], EXCEL_COLOR_THEMES['office']['accent6']],
              
    'ion': [EXCEL_COLOR_THEMES['ion']['accent1'], EXCEL_COLOR_THEMES['ion']['accent2'],
           EXCEL_COLOR_THEMES['ion']['accent3'], EXCEL_COLOR_THEMES['ion']['accent4'],
           EXCEL_COLOR_THEMES['ion']['accent5'], EXCEL_COLOR_THEMES['ion']['accent6']],
           
    'wisp': [EXCEL_COLOR_THEMES['wisp']['accent1'], EXCEL_COLOR_THEMES['wisp']['accent2'],
            EXCEL_COLOR_THEMES['wisp']['accent3'], EXCEL_COLOR_THEMES['wisp']['accent4'],
            EXCEL_COLOR_THEMES['wisp']['accent5'], EXCEL_COLOR_THEMES['wisp']['accent6']],
            
    'aspect': [EXCEL_COLOR_THEMES['aspect']['accent1'], EXCEL_COLOR_THEMES['aspect']['accent2'],
              EXCEL_COLOR_THEMES['aspect']['accent3'], EXCEL_COLOR_THEMES['aspect']['accent4'],
              EXCEL_COLOR_THEMES['aspect']['accent5'], EXCEL_COLOR_THEMES['aspect']['accent6']],
              
    'badge': [EXCEL_COLOR_THEMES['badge']['accent1'], EXCEL_COLOR_THEMES['badge']['accent2'],
             EXCEL_COLOR_THEMES['badge']['accent3'], EXCEL_COLOR_THEMES['badge']['accent4'],
             EXCEL_COLOR_THEMES['badge']['accent5'], EXCEL_COLOR_THEMES['badge']['accent6']],
    
    # Estilos oscuros (dark)
    'dark-blue': ['2F5597', '1F3864', '4472C4', '5B9BD5', '8FAADC', '2E75B5', '255E91', '1C4587'],
    'dark-red': ['952213', 'C0504D', 'FF8B6B', 'EA6B66', 'DA3903', 'FF4500', 'B22222', '8B0000'],
    'dark-green': ['1E6C41', '375623', '548235', '70AD47', '9BC169', '006400', '228B22', '3CB371'],
    'dark-purple': ['5C3292', '7030A0', '8064A2', '9A7FBA', 'B3A2C7', '800080', '9400D3', '8B008B'],
    'dark-orange': ['C55A11', 'ED7D31', 'F4B183', 'FFC000', 'FFD966', 'FF8C00', 'FF7F50', 'FF4500'],
    'dark-teal': ['008080', '31869B', '4BACC6', '92CDDC', '93CDDD', '0D8591', '0F6A77', '115E67'],
    
    # Estilos neutros
    'greyscale': ['000000', '333333', '666666', '999999', 'CCCCCC', 'EEEEEE', 'F3F3F3', 'FFFFFF'],
    'sepia': ['5F4B32', '74624B', '8A7964', 'A0917E', 'B6A997', 'CCC2B1', 'E2DBCA', 'F8F4E3']
}

# Función auxiliar para convertir estilo a número
def parse_chart_style(style):
    """
    Convierte diferentes formatos de estilo a un número de estilo de Excel (1-48).
    
    Args:
        style: Estilo en formato int, str numérico, 'styleN', o nombre descriptivo
        
    Returns:
        Número de estilo entre 1-48, o None si no es un estilo válido
    """
    if isinstance(style, int) and 1 <= style <= 48:
        return style
        
    if isinstance(style, str):
        # Caso 1: String numérico '5'
        if style.isdigit():
            style_num = int(style)
            if 1 <= style_num <= 48:
                return style_num
                
        # Caso 2: Formato 'styleN' o 'Style N'
        style_lower = style.lower()
        if style_lower.startswith('style'):
            try:
                # Extraer el número después de 'style'
                num_part = ''.join(c for c in style_lower[5:] if c.isdigit())
                if num_part:
                    style_num = int(num_part)
                    if 1 <= style_num <= 48:
                        return style_num
            except (ValueError, IndexError):
                pass
                
        # Caso 3: Nombre descriptivo ('dark-blue', etc.)
        if style_lower in CHART_STYLE_NAMES:
            return CHART_STYLE_NAMES[style_lower]
            
    return None

# Función auxiliar para aplicar estilos a gráficos y colores a la vez
def apply_chart_style(chart, style):
    """
    Aplica un estilo predefinido al gráfico, incluyendo paleta de colores adecuada.
    
    Args:
        chart: Objeto de gráfico openpyxl
        style: Estilo en cualquier formato soportado (número, nombre, etc.)
    
    Returns:
        True si se aplicó correctamente, False en caso contrario
    """
    # Convertir a número de estilo si no lo es ya
    style_number = parse_chart_style(style)
    
    if style_number is None:
        style_str = str(style) if style else "None"
        logger.warning(f"Estilo de gráfico inválido: '{style_str}'. Debe ser un número entre 1-48 o un nombre de estilo válido.")
        logger.info("Nombres de estilo válidos incluyen: 'dark-blue', 'light-1', 'colorful-3', etc.")
        return False
        
    if not (1 <= style_number <= 48):
        logger.warning(f"Estilo de gráfico inválido: {style_number}. Debe estar entre 1 y 48.")
        return False
    
    # Paso 1: Aplicar número de estilo a atributos del estilo nativo de Excel
    try:
        # La propiedad style en openpyxl se corresponde con el número de estilo de Excel
        chart.style = style_number
        logger.info(f"Aplicado estilo nativo {style_number} al gráfico")
        
        # Configurar opciones adicionales basadas en el tipo de gráfico y estilo
        chart_type = "unknown"
        if hasattr(chart, "type"):
            chart_type = chart.type
        elif isinstance(chart, BarChart):
            chart_type = "bar" if getattr(chart, "type", None) == "bar" else "column"
        elif isinstance(chart, LineChart):
            chart_type = "line"
        elif isinstance(chart, PieChart):
            chart_type = "pie"
        elif isinstance(chart, ScatterChart):
            chart_type = "scatter"
        elif isinstance(chart, AreaChart):
            chart_type = "area"
        
        # Configuraciones específicas según el tipo de gráfico y estilo
        if chart_type == "column" or chart_type == "bar":
            # Ajustes específicos para gráficos de columnas/barras
            if hasattr(chart, "gapWidth"):
                # Estilos más densos tienen menos espacio entre barras
                if style_number in [1, 7, 13, 19, 25, 31, 37, 43]:
                    chart.gapWidth = 50  # Barras más anchas
                elif style_number in [3, 9, 15, 21, 27, 33, 39, 45]:
                    chart.gapWidth = 150  # Espacio estándar
                elif style_number in [5, 11, 17, 23, 29, 35, 41, 47]:
                    chart.gapWidth = 200  # Barras más delgadas
                
        elif chart_type == "line":
            # Ajustes específicos para gráficos de línea
            for i, series in enumerate(chart.series):
                if hasattr(series, "marker"):
                    # Los estilos 3, 9, 15, etc. tienen marcadores
                    if style_number % 6 == 3:
                        series.marker.symbol = "circle"
                        series.marker.size = 5
                    # Los estilos 5, 11, 17, etc. tienen líneas discontinuas
                    elif style_number % 6 == 5:
                        if hasattr(series.graphicalProperties, "line"):
                            series.graphicalProperties.line.dashStyle = "dash"
                        
        elif chart_type == "pie":
            # Ajustes específicos para gráficos de tarta
            # Los estilos 4, 10, 16, etc. tienen separación entre porciones
            if style_number % 6 == 4 and hasattr(chart, "explosion"):
                chart.explosion = 10  # Separación de porciones
            
        elif chart_type == "scatter":
            # Ajustes específicos para gráficos de dispersión
            # Estilo 9, 15, etc. tendrían líneas suaves entre puntos
            if style_number % 6 == 3 and hasattr(chart, "smoothed"):
                chart.smoothed = True
            
        # Configurar opciones comunes del área de trazado
        if hasattr(chart, "plotArea"):
            # Ajustar cuadrícula para ciertos estilos
            if style_number in [4, 5, 12, 15, 31, 34, 37, 40, 43, 46]:
                # Estilos de tipo pastel suelen tener líneas de cuadrícula más suaves
                if hasattr(chart.plotArea, "gridLines"):
                    chart.plotArea.gridLines.visible = False
            
            # Configurar borde según el estilo
            if style_number in [6, 24]:
                # Estilos monocromáticos suelen tener bordes más fuertes
                if hasattr(chart, "border"):
                    chart.border.width = 2
    except Exception as e:
        logger.warning(f"Error al aplicar estilo {style_number}: {e}")
    
    # Paso 2: Aplicar paleta de colores asociada según el tema correspondiente al estilo
    palette_name = STYLE_TO_PALETTE.get(style_number, 'default')
    colors = CHART_COLOR_SCHEMES.get(palette_name, CHART_COLOR_SCHEMES['default'])
    
    # Aplicar colores a las series
    try:
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.fill import ColorChoice
        
        for i, series in enumerate(chart.series):
            if i < len(colors):
                # Asegurarse de que existen propiedades gráficas
                if not hasattr(series, 'graphicalProperties') or series.graphicalProperties is None:
                    series.graphicalProperties = GraphicalProperties()
                    
                # Asignar color usando ColorChoice para mejor compatibilidad
                color = colors[i % len(colors)]
                if isinstance(color, str) and color.startswith('#'):
                    color = color[1:]
                    
                series.graphicalProperties.solidFill = ColorChoice(srgbClr=color)
                
        logger.info(f"Aplicado estilo {style_number} con paleta '{palette_name}' al gráfico")
        return True
        
    except Exception as e:
        logger.warning(f"Error al aplicar colores para estilo {style_number}: {e}")
        return False

# Clase auxiliar para gestionar rangos de Excel
class ExcelRange:
    
    @staticmethod
    def parse_cell_ref(cell_ref: str) -> Tuple[int, int]:
        """
        Convierte una referencia de celda en estilo A1 a coordenadas (fila, columna) 0-based.
        
        Args:
            cell_ref: Referencia de celda en formato Excel (ej: 'A1', 'B5')
            
        Returns:
            Tupla (fila, columna) con índices base 0
            
        Raises:
            ValueError: Si la referencia de celda no es válida
        """
        if not cell_ref or not isinstance(cell_ref, str):
            raise ValueError(f"Referencia de celda inválida: {cell_ref}")
        
        # Extraer la parte de columna (letras)
        col_str = ''.join(c for c in cell_ref if c.isalpha())
        # Extraer la parte de fila (números)
        row_str = ''.join(c for c in cell_ref if c.isdigit())
        
        if not col_str or not row_str:
            raise ValueError(f"Formato de celda inválido: {cell_ref}")
        
        # Convertir columna a índice (A->0, B->1, etc.)
        col_idx = 0
        for c in col_str.upper():
            col_idx = col_idx * 26 + (ord(c) - ord('A') + 1)
        col_idx -= 1  # Ajustar a base 0
        
        # Convertir fila a índice (base 0)
        row_idx = int(row_str) - 1
        
        return row_idx, col_idx
    
    @staticmethod
    def parse_range(range_str: str) -> Tuple[int, int, int, int]:
        """
        Convierte un rango en estilo A1:B5 a coordenadas (row1, col1, row2, col2) 0-based.
        
        Args:
            range_str: Rango en formato Excel (ej: 'A1:B5')
            
        Returns:
            Tupla (fila_inicio, col_inicio, fila_fin, col_fin) con índices base 0
            
        Raises:
            ValueError: Si el rango no es válido
        """
        if not range_str or not isinstance(range_str, str):
            raise ValueError(f"Rango inválido: {range_str}")
        
        # Manejar rangos con referencia a hoja
        if '!' in range_str:
            parts = range_str.split('!')
            if len(parts) != 2:
                raise ValueError(f"Formato de rango con hoja inválido: {range_str}")
            range_str = parts[1]  # Usar solo la parte del rango
        
        # Dividir el rango en celdas de inicio y fin
        if ':' in range_str:
            start_cell, end_cell = range_str.split(':')
            start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
            end_row, end_col = ExcelRange.parse_cell_ref(end_cell)
        else:
            # Si es una sola celda, inicio y fin son iguales
            start_row, start_col = ExcelRange.parse_cell_ref(range_str)
            end_row, end_col = start_row, start_col
        
        return start_row, start_col, end_row, end_col
    
    @staticmethod
    def cell_to_a1(row: int, col: int) -> str:
        """
        Convierte coordenadas (fila, columna) 0-based a referencia de celda A1.
        
        Args:
            row: Índice de fila (base 0)
            col: Índice de columna (base 0)
            
        Returns:
            Referencia de celda en formato A1
        """
        if row < 0 or col < 0:
            raise ValueError(f"Índices negativos no válidos: fila={row}, columna={col}")
        
        # Convertir columna a letras
        col_str = ""
        col_val = col + 1  # Convertir a base 1 para cálculo
        
        while col_val > 0:
            remainder = (col_val - 1) % 26
            col_str = chr(65 + remainder) + col_str
            col_val = (col_val - 1) // 26
        
        # Convertir fila a número (base 1 para Excel)
        row_val = row + 1
        
        return f"{col_str}{row_val}"
    
    @staticmethod
    def range_to_a1(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
        """
        Convierte coordenadas de rango 0-based a rango A1:B5.
        
        Args:
            start_row: Fila inicial (base 0)
            start_col: Columna inicial (base 0)
            end_row: Fila final (base 0)
            end_col: Columna final (base 0)
            
        Returns:
            Rango en formato A1:B5
        """
        start_cell = ExcelRange.cell_to_a1(start_row, start_col)
        end_cell = ExcelRange.cell_to_a1(end_row, end_col)
        
        return f"{start_cell}:{end_cell}"

# Función auxiliar para obtener una hoja de trabajo
def get_sheet(wb, sheet_name) -> Any:
    """
    Obtiene una hoja de Excel por nombre.
    
    Args:
        wb: Objeto workbook de openpyxl
        sheet_name: Nombre de la hoja
        
    Returns:
        Objeto worksheet
        
    Raises:
        SheetNotFoundError: Si la hoja no existe
    """
    if not wb:
        raise AdvancedExcelError("El workbook no puede ser None")
    
    if sheet_name not in wb.sheetnames:
        raise SheetNotFoundError(f"La hoja '{sheet_name}' no existe en el workbook. Hojas disponibles: {wb.sheetnames}")
    
    return wb[sheet_name]


# 6. Tablas (Excel Tables)
def add_table(ws, table_name, cell_range, style=None):
    """
    Define un rango como Tabla con estilo.
    
    Args:
        ws: Objeto worksheet de openpyxl
        table_name (str): Nombre único para la tabla
        cell_range (str): Rango en formato A1:B5
        style (str, opcional): Nombre de estilo predefinido o dict personalizado
        
    Returns:
        Objeto Table creado
        
    Raises:
        TableError: Si hay un problema con la tabla (ej. nombre duplicado)
    """
    if not ws:
        raise AdvancedExcelError("El worksheet no puede ser None")
    
    try:
        # Verificar si ya existe una tabla con ese nombre
        if hasattr(ws, 'tables') and table_name in ws.tables:
            raise TableError(f"Ya existe una tabla con el nombre '{table_name}'")
        
        # Crear objeto de tabla
        table = Table(displayName=table_name, ref=cell_range)
        
        # Aplicar estilo
        if style:
            if isinstance(style, dict):
                # Estilo personalizado
                style_info = TableStyleInfo(**style)
            else:
                # Estilo predefinido
                style_info = TableStyleInfo(
                    name=style,
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
            table.tableStyleInfo = style_info
        
        # Añadir tabla a la hoja
        ws.add_table(table)
        return table
    
    except Exception as e:
        if "Duplicate name" in str(e):
            raise TableError(f"Ya existe una tabla con el nombre '{table_name}'")
        elif "Invalid coordinate" in str(e) or "Invalid cell" in str(e):
            raise RangeError(f"Rango inválido: '{cell_range}'")
        else:
            raise TableError(f"Error al añadir tabla: {e}")

def remove_table(ws, table_name):
    """
    Elimina una tabla existente.
    
    Args:
        ws: Objeto worksheet de openpyxl
        table_name (str): Nombre de la tabla a eliminar
        
    Raises:
        TableError: Si la tabla no existe
    """
    if not ws:
        raise AdvancedExcelError("El worksheet no puede ser None")
    
    try:
        # Verificar si existe la tabla
        if not hasattr(ws, 'tables') or table_name not in ws.tables:
            raise TableError(f"No existe una tabla con el nombre '{table_name}'")
        
        # Eliminar la tabla
        del ws.tables[table_name]
    
    except TableError:
        raise
    except Exception as e:
        raise TableError(f"Error al eliminar tabla: {e}")

def refresh_table(ws, table_name, new_range=None):
    """
    Actualiza el rango de una tabla o recalcula fórmulas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        table_name (str): Nombre de la tabla a actualizar
        new_range (str, opcional): Nuevo rango en formato A1:B5
        
    Returns:
        Objeto Table actualizado
        
    Raises:
        TableError: Si la tabla no existe o hay problemas al actualizar
    """
    if not ws:
        raise AdvancedExcelError("El worksheet no puede ser None")
    
    try:
        # Verificar si existe la tabla
        if not hasattr(ws, 'tables') or table_name not in ws.tables:
            raise TableError(f"No existe una tabla con el nombre '{table_name}'")
        
        # Obtener la tabla
        table = ws.tables[table_name]
        
        # Actualizar el rango si se proporciona uno nuevo
        if new_range:
            old_range = table.ref
            table.ref = new_range
            return table, {"old_range": old_range, "new_range": new_range}
        
        return table
    
    except TableError:
        raise
    except Exception as e:
        raise TableError(f"Error al actualizar tabla: {e}")


# 7. Fórmulas y Cálculos
def set_formula(ws, cell, formula):
    """
    Establece una fórmula en una celda.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell (str): Referencia de celda (ej. "A1")
        formula (str): Fórmula Excel, con o sin signo =
        
    Returns:
        Celda actualizada
        
    Raises:
        FormulaError: Si hay un problema con la fórmula
    """
    if not ws:
        raise AdvancedExcelError("El worksheet no puede ser None")
    
    try:
        # Añadir signo = si no está presente
        if formula and not formula.startswith('='):
            formula = f"={formula}"
        
        # Establecer la fórmula
        ws[cell] = formula
        return ws[cell]
    
    except KeyError:
        raise RangeError(f"Celda inválida: '{cell}'")
    except Exception as e:
        raise FormulaError(f"Error al establecer fórmula: {e}")

def get_formula(ws, cell):
    """
    Obtiene la fórmula de una celda.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell (str): Referencia de celda (ej. "A1")
        
    Returns:
        Fórmula como cadena, o None si no hay fórmula
        
    Raises:
        RangeError: Si la celda no existe
    """
    if not ws:
        raise AdvancedExcelError("El worksheet no puede ser None")
    
    try:
        # Obtener la celda
        cell_obj = ws[cell]
        
        # Verificar si tiene fórmula
        if cell_obj.data_type == 'f':
            return cell_obj.value
        else:
            return None
    
    except KeyError:
        raise RangeError(f"Celda inválida: '{cell}'")
    except Exception as e:
        raise FormulaError(f"Error al obtener fórmula: {e}")

def evaluate_formula(ws, cell):
    """
    Evalúa la fórmula en una celda (carga un valor calculado).
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell (str): Referencia de celda (ej. "A1")
        
    Returns:
        Valor calculado, o None si no hay fórmula
        
    Raises:
        FormulaError: Si hay un problema al evaluar
        
    Note:
        Esta función es limitada en openpyxl, ya que generalmente
        solo evalúa al cargar con data_only=True.
    """
    if not ws:
        raise AdvancedExcelError("El worksheet no puede ser None")
    
    try:
        # Obtener la celda
        cell_obj = ws[cell]
        
        # Verificar si tiene fórmula
        if cell_obj.data_type == 'f':
            formula = cell_obj.value
            
            # Intentar forzar el cálculo (limitado en openpyxl)
            # Solo un mensaje informativo
            logger.warning("La evaluación de fórmulas es limitada en openpyxl. Por favor, guarde y recargue el archivo con data_only=True para cálculos precisos.")
            
            # Devolver la fórmula original
            return {
                "formula": formula,
                "calculated_value": None,
                "warning": "La evaluación de fórmulas es limitada en openpyxl."
            }
        else:
            # Si no es una fórmula, devolver el valor actual
            return {
                "formula": None,
                "calculated_value": cell_obj.value
            }
    
    except KeyError:
        raise RangeError(f"Celda inválida: '{cell}'")
    except Exception as e:
        raise FormulaError(f"Error al evaluar fórmula: {e}")


# 8. Gráficos (Charts)
def add_chart(wb, sheet_name, chart_type, data_range, title=None, position=None, style=None, theme=None, custom_palette=None):
    """
    Inserta gráfico nativo en la hoja.
    
    Args:
        wb: Objeto workbook de openpyxl
        sheet_name (str): Nombre de la hoja donde insertar el gráfico
        chart_type (str): Tipo de gráfico ('column', 'bar', 'line', 'pie', 'doughnut', 'scatter', 'area', 'histogram', 'box', 'combo', etc.)
        data_range (str): Rango de datos en formato A1:B5
        title (str, opcional): Título del gráfico
        position (str, opcional): Celda de anclaje (ej. "E5")
        style: Estilo del gráfico que puede ser:
            - Número entero (1-48) para estilos predefinidos de Excel
            - String numérico ('5')
            - Formato 'styleN' ('style7')
            - Nombre descriptivo ('dark-blue', 'light-1', etc.)
            - Dict con opciones personalizadas (colors, width, height, etc.)
        theme (str, opcional): Nombre del tema de color ('Office', 'Ion', 'Wisp', 'Aspect', etc.)
        custom_palette (list, opcional): Lista de colores personalizados en formato hexadecimal (['#E81123', '#2D7D9A', etc.])
        
    Returns:
        Id del gráfico y objeto chart
        
    Raises:
        ChartError: Si hay un problema con el gráfico
    """
    if not wb:
        raise AdvancedExcelError("El workbook no puede ser None")
    
    try:
        # Obtener la hoja
        ws = get_sheet(wb, sheet_name)
        
        # Validar y normalizar la posición
        if position:
            # Si la posición es un rango, tomar solo la primera celda
            if ':' in position:
                position = position.split(':')[0]
            
            # Validar que es una referencia de celda válida
            try:
                # Intentar parsear para verificar que es una referencia válida
                ExcelRange.parse_cell_ref(position)
            except ValueError:
                raise ValueError(f"Posición inválida '{position}'. Debe ser una referencia de celda (ej: 'E4')")
        
        # Crear objeto de gráfico según el tipo
        chart = None
        if chart_type.lower() == 'column':
            chart = BarChart()
            chart.type = "col"
        elif chart_type.lower() == 'bar':
            chart = BarChart()
            chart.type = "bar"
        elif chart_type.lower() == 'line':
            chart = LineChart()
        elif chart_type.lower() == 'pie':
            chart = PieChart()
        elif chart_type.lower() == 'doughnut':
            chart = PieChart()
            # Configurar como doughnut si es posible
            if hasattr(chart, 'doughnutHole'):
                chart.doughnutHole = 50  # Porcentaje del agujero (0-100)
        elif chart_type.lower() == 'scatter':
            chart = ScatterChart()
        elif chart_type.lower() == 'area':
            chart = AreaChart()
        elif chart_type.lower() == 'histogram':
            # En versiones modernas de openpyxl
            try:
                from openpyxl.chart.histogram_chart import HistogramChart
                chart = HistogramChart()
            except ImportError:
                # Fallback si no está disponible
                logger.warning("Histogram chart no disponible en esta versión de openpyxl. Usando ColumnChart como alternativa.")
                chart = BarChart()
                chart.type = "col"
        elif chart_type.lower() == 'box':
            # En versiones modernas de openpyxl
            try:
                from openpyxl.chart.box_chart import BoxChart
                chart = BoxChart()
            except ImportError:
                # Fallback si no está disponible
                logger.warning("Box chart no disponible en esta versión de openpyxl. Usando ColumnChart como alternativa.")
                chart = BarChart()
                chart.type = "col"
        elif chart_type.lower() == 'combo':
            # Para gráficos combinados, usamos LineChart como base y luego añadimos otras series
            chart = LineChart()
            # Marcar que es un combo chart para procesamiento especial posterior
            chart._is_combo = True
        else:
            raise ChartError(f"Tipo de gráfico no soportado: '{chart_type}'")
        
        # Configurar título si se proporciona
        if title:
            chart.title = title
            
        # Procesar tema y paleta personalizada si se proporcionan
        if theme or custom_palette:
            # Crear o modificar estilo si es necesario
            if style is None:
                style = {}
            elif not isinstance(style, dict):
                # Convertir a dict para poder agregar theme/palette
                style_number = parse_chart_style(style)
                style = {'chart_style': style_number}
            
            # Procesar tema si se proporciona
            if theme:
                theme_lower = theme.lower()
                # Convertir a un esquema de colores
                for theme_name in EXCEL_COLOR_THEMES:
                    if theme_name.lower() == theme_lower:
                        # Asignar el tema como esquema de colores
                        style['color_scheme'] = theme_name
                        logger.info(f"Aplicando tema de colores '{theme_name}' al gráfico")
                        break
                else:
                    logger.warning(f"Tema de color no reconocido: '{theme}'. Usando tema por defecto.")
            
            # Procesar paleta personalizada si se proporciona
            if custom_palette:
                # Asignar directamente la paleta al estilo
                style['colors'] = custom_palette
                logger.info(f"Aplicando paleta personalizada con {len(custom_palette)} colores")
        
        # Asegurarnos de que el rango tiene el formato correcto (no incluye nombre de hoja)
        if '!' in data_range:
            # Si ya tiene formato con nombre de hoja, lo dejamos como está
            range_string = data_range
            # Extraer solo la parte del rango después de !
            clean_range = data_range.split('!')[1]
        else:
            clean_range = data_range
            # Añadir el nombre de la hoja al rango para cumplir con el formato requerido
            # Escapar nombres de hojas con espacios o caracteres especiales
            if ' ' in sheet_name or any(c in sheet_name for c in "![]{}?"):
                sheet_prefix = f"'{sheet_name}'!"
            else:
                sheet_prefix = f"{sheet_name}!"
            range_string = f"{sheet_prefix}{data_range}"
        
        # Parsear rango de datos
        try:
            # Parsear el rango para obtener los límites
            min_row, min_col, max_row, max_col = ExcelRange.parse_range(clean_range)
            
            # Ajustar a base 1 para Reference
            min_row += 1
            min_col += 1
            max_row += 1
            max_col += 1
            
            # Determinar si los datos están organizados en columnas o filas
            is_column_oriented = (max_row - min_row) > (max_col - min_col)
            
            # Para gráficos que necesitan categorías (la mayoría excepto scatter)
            if chart_type.lower() != 'scatter':
                if is_column_oriented:
                    # Datos organizados en columnas
                    categories = Reference(ws, min_row=min_row, max_row=max_row, min_col=min_col, max_col=min_col)
                    data = Reference(ws, min_row=min_row, max_row=max_row, min_col=min_col+1, max_col=max_col)
                    # En Python 2 openpyxl puede no soportar el parámetro titles_from_data
                    try:
                        chart.add_data(data, titles_from_data=True)
                    except TypeError:
                        # Fallback para versiones antiguas de openpyxl
                        chart.add_data(data)
                    chart.set_categories(categories)
                else:
                    # Datos organizados en filas
                    categories = Reference(ws, min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col)
                    data = Reference(ws, min_row=min_row+1, max_row=max_row, min_col=min_col, max_col=max_col)
                    # En Python 2 openpyxl puede no soportar el parámetro titles_from_data
                    try:
                        chart.add_data(data, titles_from_data=True)
                    except TypeError:
                        # Fallback para versiones antiguas de openpyxl
                        chart.add_data(data)
                    chart.set_categories(categories)
            else:
                # Para gráficos de dispersión
                data_ref = Reference(ws, min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col)
                chart.add_data(data_ref)
                
            # Aplicar estilos guardados después de añadir los datos
            if hasattr(chart, '_style_number'):
                style_number = chart._style_number
                logger.info(f"Aplicando estilo guardado {style_number} después de añadir datos")
                apply_chart_style(chart, style_number)
                # Eliminar atributo temporal
                delattr(chart, '_style_number')
                
            # Aplicar colores personalizados después de añadir los datos
            # Este paso es crucial porque ahora las series existen
            if hasattr(chart, '_custom_colors') and chart._custom_colors:
                from openpyxl.chart.shapes import GraphicalProperties
                from openpyxl.drawing.fill import ColorChoice
                
                for i, series in enumerate(chart.series):
                    if i < len(chart._custom_colors):
                        if not hasattr(series, 'graphicalProperties') or series.graphicalProperties is None:
                            series.graphicalProperties = GraphicalProperties()
                        
                        color = chart._custom_colors[i]
                        if isinstance(color, str) and color.startswith('#'):
                            color = color[1:]
                            
                        series.graphicalProperties.solidFill = ColorChoice(srgbClr=color)
                
                # Limpiar los colores guardados para evitar aplicarlos dos veces
                del chart._custom_colors
            
            # Aplicar mejoras adicionales según el tipo de gráfico
            if chart_type.lower() == 'column' or chart_type.lower() == 'bar':
                # Configurar espaciado entre columnas/barras para mejor aspecto visual
                for s in chart.series:
                    if hasattr(s, 'overlap'):
                        s.overlap = 0
                    if hasattr(s, 'gapWidth'):
                        s.gapWidth = 150
            
            elif chart_type.lower() == 'line':
                # Mejorar la visualización de líneas
                for s in chart.series:
                    if hasattr(s, 'marker'):
                        # Activar marcadores
                        s.marker.symbol = 'circle'
                        s.marker.size = 5
            
            elif chart_type.lower() == 'pie':
                # Configurar opciones específicas para gráficos de sectores
                if hasattr(chart, 'firstSliceAng'):
                    chart.firstSliceAng = 0  # Empezar desde arriba
                # Mostrar etiquetas de porcentaje
                chart.dataLabels = True
        
        except Exception as e:
            raise RangeError(f"Error al procesar rango de datos '{data_range}': {e}")
        
        # Aplicar estilos
        if style is not None:
            if isinstance(style, dict):
                # Procesar opciones de estilo como diccionario (este bloque se maneja después)
                pass
            else:
                # Procesar estilo especificado (número, nombre, etc.)
                style_number = parse_chart_style(style)
                if style_number is not None:
                    # Verificar si podemos aplicar el estilo ahora o después
                    if hasattr(chart, 'series') and len(chart.series) > 0:
                        # Aplicar el estilo inmediatamente si ya tenemos series
                        logger.info(f"Aplicando estilo {style_number} (original: '{style}')") 
                        apply_chart_style(chart, style_number)
                    else:
                        # Guardar el estilo para aplicarlo después de añadir datos
                        chart._style_number = style_number
                        logger.info(f"El estilo {style_number} (original: '{style}') se aplicará después de añadir datos")
                else:
                    logger.warning(f"Estilo de gráfico inválido: '{style}'. Debe ser un número entre 1-48 o un nombre de estilo válido.")
                    logger.info("Estilos soportados: números 1-48, 'styleN', 'dark-blue', 'light-1', 'colorful-3', etc.")
        
        if style is not None and isinstance(style, dict):
            # Procesar opciones de estilo como diccionario
            # Tamaño del gráfico
            if 'width' in style and 'height' in style:
                chart.width = style['width']
                chart.height = style['height']
                
            # Asegurar que el Chart tenga sus series antes de intentar modificarlas
            # Esto es crucial y resuelve el problema principal
            if hasattr(chart, 'series') and len(chart.series) > 0:
                # Colores personalizados para series
                if 'colors' in style and isinstance(style['colors'], list):
                    for i, series in enumerate(chart.series):
                        if i < len(style['colors']):
                            # Importar si no está disponible
                            from openpyxl.chart.shapes import GraphicalProperties
                            from openpyxl.drawing.fill import ColorChoice
                            
                            # Asegurarse de que existan las propiedades gráficas
                            if not hasattr(series, 'graphicalProperties'):
                                series.graphicalProperties = GraphicalProperties()
                            elif series.graphicalProperties is None:
                                series.graphicalProperties = GraphicalProperties()
                            
                            # Asignar color asegurándonos de que no tiene el prefijo #
                            color = style['colors'][i]
                            if isinstance(color, str) and color.startswith('#'):
                                color = color[1:]
                            
                            # Aplicar el color de forma explícita
                            series.graphicalProperties.solidFill = ColorChoice(srgbClr=color)
            else:
                logger.warning("No se pueden aplicar colores personalizados porque el gráfico no tiene series aún.")
                logger.warning("Los colores se aplicarán después de añadir los datos.")
                # Guardar los colores para aplicarlos después de añadir datos
                chart._custom_colors = style.get('colors', [])
            
            # Estilo predefinido (tiene prioridad)
            if 'chart_style' in style:
                try:
                    style_number = int(style['chart_style'])
                    # Usar función para aplicar estilo y paleta de colores
                    apply_chart_style(chart, style_number)
                except (ValueError, TypeError):
                    logger.warning(f"Estilo de gráfico inválido: {style['chart_style']}. Debe ser un número entre 1 y 48.")
                
            # Esquema de colores predefinido
            if 'color_scheme' in style:
                scheme = style['color_scheme']
                # Obtener colores del esquema seleccionado
                colors = CHART_COLOR_SCHEMES.get(scheme, CHART_COLOR_SCHEMES['default'])
                
                # Guardar los colores para aplicarlos después de añadir datos si no hay series todavía
                if not hasattr(chart, 'series') or len(chart.series) == 0:
                    chart._custom_colors = colors
                    logger.info(f"Se aplicará el esquema de colores '{scheme}' después de añadir datos.")
                else:
                    # Aplicar colores a las series existentes
                    from openpyxl.chart.shapes import GraphicalProperties
                    from openpyxl.drawing.fill import ColorChoice
                    
                    for i, series in enumerate(chart.series):
                        if i < len(colors):
                            # Asegurarse de que existen propiedades gráficas
                            if not hasattr(series, 'graphicalProperties') or series.graphicalProperties is None:
                                series.graphicalProperties = GraphicalProperties()
                            
                            # Aplicar color
                            color = colors[i % len(colors)]
                            if isinstance(color, str) and color.startswith('#'):
                                color = color[1:]
                            series.graphicalProperties.solidFill = ColorChoice(srgbClr=color)
        
        # Posicionar el gráfico en la hoja
        if position:
            ws.add_chart(chart, position)
        else:
            ws.add_chart(chart)
        
        # Determinar el ID del gráfico (basado en su posición en la lista)
        chart_id = len(ws._charts) - 1
        
        return chart_id, chart
    
    except SheetNotFoundError:
        raise
    except ChartError:
        raise
    except RangeError:
        raise
    except Exception as e:
        raise ChartError(f"Error al crear gráfico: {e}")

def delete_chart(wb, sheet_name, chart_id):
    """
    Elimina un gráfico de la hoja.
    
    Args:
        wb: Objeto workbook de openpyxl
        sheet_name (str): Nombre de la hoja
        chart_id (int): ID del gráfico a eliminar
        
    Raises:
        ChartError: Si el gráfico no existe
    """
    if not wb:
        raise AdvancedExcelError("El workbook no puede ser None")
    
    try:
        # Obtener la hoja
        ws = get_sheet(wb, sheet_name)
        
        # Verificar que el chart_id sea válido
        if not hasattr(ws, '_charts') or chart_id >= len(ws._charts) or chart_id < 0:
            raise ChartError(f"No existe un gráfico con ID {chart_id} en la hoja '{sheet_name}'")
        
        # Eliminar el gráfico
        del ws._charts[chart_id]
    
    except SheetNotFoundError:
        raise
    except ChartError:
        raise
    except Exception as e:
        raise ChartError(f"Error al eliminar gráfico: {e}")

def list_charts(ws):
    """
    Lista todos los gráficos en una hoja.
    
    Args:
        ws: Objeto worksheet de openpyxl
        
    Returns:
        Lista de diccionarios con información de los gráficos:
        [{'id': 0, 'type': 'bar', 'title': 'Ventas por región'}, ...]
    """
    if not ws:
        raise AdvancedExcelError("El worksheet no puede ser None")
    
    charts_info = []
    
    try:
        if hasattr(ws, '_charts'):
            for i, chart_rel in enumerate(ws._charts):
                chart = chart_rel[0]  # El elemento 0 es el objeto chart, el 1 es position
                
                # Determinar el tipo de gráfico
                chart_type = "desconocido"
                if isinstance(chart, BarChart):
                    chart_type = "bar" if chart.type == "bar" else "column"
                elif isinstance(chart, LineChart):
                    chart_type = "line"
                elif isinstance(chart, PieChart):
                    chart_type = "pie"
                elif isinstance(chart, ScatterChart):
                    chart_type = "scatter"
                elif isinstance(chart, AreaChart):
                    chart_type = "area"
                
                # Recopilar información del gráfico
                chart_info = {
                    'id': i,
                    'type': chart_type,
                    'title': chart.title if hasattr(chart, 'title') and chart.title else f"Chart {i}",
                    'position': chart_rel[1] if len(chart_rel) > 1 else None,
                    'series_count': len(chart.series) if hasattr(chart, 'series') else 0
                }
                
                charts_info.append(chart_info)
    
    except Exception as e:
        logger.warning(f"Error al listar gráficos: {e}")
    
    return charts_info


# 9. Pivot Tables (Tablas Dinámicas)
def add_pivot_table(wb, source_sheet, source_range, target_sheet, target_cell, rows, cols, data_fields):
    """
    Crea una tabla dinámica.
    
    Args:
        wb: Objeto workbook de openpyxl
        source_sheet (str): Hoja con datos fuente
        source_range (str): Rango de datos fuente (A1:E10)
        target_sheet (str): Hoja donde crear la tabla dinámica
        target_cell (str): Celda de anclaje (ej. "A1")
        rows (list): Campos para filas
        cols (list): Campos para columnas
        data_fields (list): Campos para valores y funciones
        
    Returns:
        Objeto PivotTable creado
        
    Raises:
        PivotTableError: Si hay problemas con la tabla dinámica
    """
    if not wb:
        raise AdvancedExcelError("El workbook no puede ser None")
    
    try:
        # Obtener hoja de origen
        source_ws = get_sheet(wb, source_sheet)
        
        # Obtener hoja de destino
        target_ws = get_sheet(wb, target_sheet)
        
        # Crear caché de pivot (requerido por openpyxl)
        # Nota: Esta implementación es básica debido a limitaciones de openpyxl con tablas dinámicas
        
        logger.warning("Las tablas dinámicas en openpyxl tienen funcionalidad limitada y pueden no funcionar como se espera.")
        logger.warning("Se crea una estructura básica que necesitará ser complementada en Excel.")
        
        # Intentar crear la caché de datos (este es un paso necesario)
        try:
            # Parsear el rango
            min_row, min_col, max_row, max_col = ExcelRange.parse_range(source_range)
            
            # Ajustar a base 1 para Reference
            min_row += 1
            min_col += 1
            max_row += 1
            max_col += 1
            
            # Crear referencia de datos para la caché
            data_reference = Reference(source_ws, min_row=min_row, min_col=min_col,
                                     max_row=max_row, max_col=max_col)
            
            # Crear caché de pivot
            pivot_cache = PivotCache(cacheSource=data_reference, cacheDefinition={'refreshOnLoad': True})
            
            # Generar un ID único para la tabla dinámica
            pivot_name = f"PivotTable{len(wb._pivots) + 1 if hasattr(wb, '_pivots') else 1}"
            
            # Crear la tabla dinámica
            pivot_table = PivotTable(name=pivot_name, cache=pivot_cache,
                                    location=target_cell, rowGrandTotals=True, colGrandTotals=True)
            
            # Añadir campos de fila
            for row_field in rows:
                pivot_table.rowFields.append(PivotField(data=row_field))
            
            # Añadir campos de columna
            for col_field in cols:
                pivot_table.colFields.append(PivotField(data=col_field))
            
            # Añadir campos de datos
            for data_field in data_fields:
                pivot_table.dataFields.append(PivotField(data=data_field))
            
            # Añadir la tabla dinámica a la hoja de destino
            target_ws.add_pivot_table(pivot_table)
            
            return pivot_table
            
        except Exception as pivot_error:
            logger.error(f"Error al crear tabla dinámica: {pivot_error}")
            raise PivotTableError(f"Error al crear tabla dinámica: {pivot_error}")
    
    except SheetNotFoundError:
        raise
    except PivotTableError:
        raise
    except Exception as e:
        raise PivotTableError(f"Error al crear tabla dinámica: {e}")

def delete_pivot_table(wb, sheet_name, pivot_name):
    """
    Elimina una tabla dinámica.
    
    Args:
        wb: Objeto workbook de openpyxl
        sheet_name (str): Nombre de la hoja con la tabla dinámica
        pivot_name (str): Nombre de la tabla dinámica
        
    Raises:
        PivotTableError: Si la tabla no existe
    """
    if not wb:
        raise AdvancedExcelError("El workbook no puede ser None")
    
    try:
        # Obtener la hoja
        ws = get_sheet(wb, sheet_name)
        
        # Verificar si existe la tabla dinámica
        if not hasattr(ws, '_pivots'):
            raise PivotTableError(f"No hay tablas dinámicas en la hoja '{sheet_name}'")
        
        # Buscar la tabla dinámica por nombre y eliminarla
        pivot_found = False
        for i, pt in enumerate(ws._pivots):
            if pt.name == pivot_name:
                del ws._pivots[i]
                pivot_found = True
                break
        
        if not pivot_found:
            raise PivotTableError(f"No existe una tabla dinámica con el nombre '{pivot_name}' en la hoja '{sheet_name}'")
    
    except SheetNotFoundError:
        raise
    except PivotTableError:
        raise
    except Exception as e:
        raise PivotTableError(f"Error al eliminar tabla dinámica: {e}")

def list_pivot_tables(wb):
    """
    Lista todas las tablas dinámicas en el workbook.
    
    Args:
        wb: Objeto workbook de openpyxl
        
    Returns:
        Dict con tablas dinámicas por hoja
    """
    if not wb:
        raise AdvancedExcelError("El workbook no puede ser None")
    
    pivot_tables = {}
    
    try:
        # Recorrer todas las hojas
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Verificar si la hoja tiene tablas dinámicas
            if hasattr(ws, '_pivots') and ws._pivots:
                sheet_pivots = []
                
                for i, pt in enumerate(ws._pivots):
                    pivot_info = {
                        'name': pt.name,
                        'id': i,
                        'location': pt.location if hasattr(pt, 'location') else None
                    }
                    sheet_pivots.append(pivot_info)
                
                if sheet_pivots:
                    pivot_tables[sheet_name] = sheet_pivots
    
    except Exception as e:
        logger.warning(f"Error al listar tablas dinámicas: {e}")
    
    return pivot_tables


# Crear el servidor MCP como variable global
mcp = None
if HAS_MCP:
    # Esta es la variable global que el sistema MCP busca
    mcp = FastMCP("Advanced Excel MCP", 
                 dependencies=["openpyxl"])
    logger.info("Servidor MCP iniciado correctamente")
    
    # 6. Tablas (Excel Tables) - Herramientas MCP
    @mcp.tool(description="Define un rango como Tabla con estilo")
    def add_table_tool(file_path, sheet_name, table_name, cell_range, style=None):
        """Define un rango como Tabla con estilo"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Convertir style de JSON a dict si es necesario
            if style and isinstance(style, str):
                try:
                    style = json.loads(style)
                except:
                    pass
            
            # Añadir la tabla
            table = add_table(ws, table_name, cell_range, style)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "table_name": table_name,
                "range": cell_range,
                "style": style,
                "message": f"Tabla '{table_name}' creada correctamente en el rango {cell_range}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al crear tabla: {e}"
            }
    
    @mcp.tool(description="Elimina una tabla existente")
    def remove_table_tool(file_path, sheet_name, table_name):
        """Elimina una tabla existente"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Eliminar la tabla
            remove_table(ws, table_name)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "table_name": table_name,
                "message": f"Tabla '{table_name}' eliminada correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al eliminar tabla: {e}"
            }
    
    @mcp.tool(description="Actualiza el rango de una tabla o recalcula fórmulas")
    def refresh_table_tool(file_path, sheet_name, table_name, new_range=None):
        """Actualiza el rango de una tabla o recalcula fórmulas"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Actualizar la tabla
            result = refresh_table(ws, table_name, new_range)
            
            # Guardar cambios
            wb.save(file_path)
            
            # Preparar mensaje según si se actualizó el rango o no
            if new_range:
                message = f"Rango de tabla '{table_name}' actualizado a {new_range}"
                if isinstance(result, tuple) and len(result) > 1:
                    table, info = result
                    old_range = info.get('old_range', 'desconocido')
                    message = f"Rango de tabla '{table_name}' actualizado de {old_range} a {new_range}"
            else:
                message = f"Tabla '{table_name}' actualizada correctamente"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "table_name": table_name,
                "new_range": new_range,
                "message": message
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al actualizar tabla: {e}"
            }
    
    # 7. Fórmulas y Cálculos - Herramientas MCP
    @mcp.tool(description="Establece una fórmula en una celda")
    def set_formula_tool(file_path, sheet_name, cell, formula):
        """Establece una fórmula en una celda"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Establecer la fórmula
            cell_obj = set_formula(ws, cell, formula)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell": cell,
                "formula": formula,
                "message": f"Fórmula establecida correctamente en la celda {cell}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al establecer fórmula: {e}"
            }
    
    @mcp.tool(description="Obtiene la fórmula de una celda")
    def get_formula_tool(file_path, sheet_name, cell):
        """Obtiene la fórmula de una celda"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Obtener la fórmula
            formula = get_formula(ws, cell)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell": cell,
                "formula": formula,
                "has_formula": formula is not None,
                "message": f"Fórmula obtenida correctamente de la celda {cell}" if formula else f"La celda {cell} no contiene una fórmula"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al obtener fórmula: {e}"
            }
    
    @mcp.tool(description="Evalúa la fórmula en una celda (carga un valor calculado)")
    def evaluate_formula_tool(file_path, sheet_name, cell):
        """Evalúa la fórmula en una celda (carga un valor calculado)"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Evaluar la fórmula
            result = evaluate_formula(ws, cell)
            
            # Preparar mensaje según si hay fórmula o no
            if result.get('formula'):
                message = f"Fórmula en la celda {cell}: {result['formula']}"
                if 'warning' in result:
                    message += f" (Advertencia: {result['warning']})"
            else:
                message = f"La celda {cell} contiene el valor: {result['calculated_value']}"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell": cell,
                "formula": result.get('formula'),
                "calculated_value": result.get('calculated_value'),
                "has_formula": result.get('formula') is not None,
                "warning": result.get('warning'),
                "message": message
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al evaluar fórmula: {e}"
            }
    
    # 8. Gráficos (Charts) - Herramientas MCP
    @mcp.tool(description="Inserta gráfico nativo en la hoja")
    def add_chart_tool(file_path, sheet_name, chart_type, data_range, title=None, position=None, style=None, theme=None, custom_palette=None):
        """Inserta gráfico nativo en la hoja.
        
        Los estilos pueden especificarse como:
        - Números entre 1-48 (estilos predefinidos de Excel)
        - Strings como 'style1', 'Style 5', etc.
        - Nombres descriptivos como 'dark-blue', 'light-1', 'colorful-3', etc.
        - Esquemas de color específicos: 'office', 'ion', 'wisp', 'aspect', 'dark-blue', etc.
        - Para gráficos específicos: 'column-dark', 'line-markers', 'pie-explosion', etc.
        - Dictionary con opciones avanzadas: {'width': 800, 'height': 600, 'colors': ['#FF0000', '#00FF00']}
        
        Tambien se puede especificar un tema con el parámetro theme ('Office', 'Ion', 'Wisp', 'Aspect', etc.)
        o una paleta personalizada con custom_palette (['#E81123', '#2D7D9A', '#5C2D91'])
        """
        try:
            # Intentar primero con openpyxl
            try:
                # Abrir el archivo
                wb = openpyxl.load_workbook(file_path)
                
                # Convertir style de JSON a dict o int o string especial (styleN)
                if style and isinstance(style, str):
                    try:
                        # Intentar primero como JSON
                        style = json.loads(style)
                    except (json.JSONDecodeError, ValueError, TypeError):
                        # Si no es JSON válido, comprobar si es "styleN"
                        if style.lower().startswith("style") and style[5:].isdigit():
                            # Es un formato "styleN", extraer N como número
                            try:
                                style_number = int(style[5:])
                                style = style_number  # Guardar como entero para procesarlo luego
                                logger.info(f"Convertido '{style}' a estilo numérico {style_number}")
                            except (ValueError, IndexError):
                                logger.warning(f"No se pudo extraer número de estilo de '{style}'")
                        # Si es un número como string, convertirlo a entero
                        elif style.isdigit():
                            style = int(style)
                        # Si no es ni JSON, ni "styleN", ni número, dejarlo como está
                        
                # Asegurarnos de que el objeto style tenga el formato correcto
                # para facilitar el manejo dentro de la función add_chart
                if isinstance(style, dict):
                    # Si hay colores como cadenas, convertirlos a una lista utilizable
                    if 'colors' in style and isinstance(style['colors'], str):
                        try:
                            style['colors'] = json.loads(style['colors'])
                        except:
                            # Si no es JSON, podría ser una lista separada por comas
                            style['colors'] = [c.strip() for c in style['colors'].split(',')]
                    
                    # Si hay un esquema de colores como string, asegurarse de que es accesible
                    if 'color_scheme' in style and isinstance(style['color_scheme'], str):
                        scheme_name = style['color_scheme'].strip()
                        if scheme_name in CHART_COLOR_SCHEMES:
                            style['color_scheme'] = scheme_name
                        else:
                            # Si el esquema no existe, usar default
                            logger.warning(f"Esquema de colores no reconocido: '{scheme_name}'. Usando esquema 'default'.")
                            style['color_scheme'] = 'default'
                
                # Validar que la hoja existe
                if sheet_name not in wb.sheetnames:
                    raise SheetNotFoundError(f"La hoja '{sheet_name}' no existe en el archivo")
                    
                # Validar el rango de datos
                if not data_range or not isinstance(data_range, str):
                    raise RangeError(f"Rango de datos inválido: {data_range}")
                    
                # Asegurarnos de que el rango no incluye el nombre de la hoja
                if '!' in data_range:
                    # Extraer solo la parte del rango después de !
                    sheet_part, range_part = data_range.split('!')
                    # Verificar que la hoja mencionada coincide
                    if sheet_part.strip("'") != sheet_name:
                        logger.warning(f"La hoja en el rango ({sheet_part}) no coincide con la hoja especificada ({sheet_name})")
                    data_range = range_part
            
                # Validar y normalizar la posición
                if position:
                    # Si la posición es un rango, tomar solo la primera celda
                    if ':' in position:
                        position = position.split(':')[0]
                    
                    # Validar que es una referencia de celda válida
                    try:
                        # Intentar parsear para verificar que es una referencia válida
                        ExcelRange.parse_cell_ref(position)
                    except ValueError:
                        raise ValueError(f"Posición inválida '{position}'. Debe ser una referencia de celda (ej: 'E4')")
                
                # Procesar paleta personalizada si viene como string
                if custom_palette and isinstance(custom_palette, str):
                    try:
                        # Intentar como JSON
                        custom_palette = json.loads(custom_palette)
                    except:
                        # Intentar como lista separada por comas
                        custom_palette = [c.strip() for c in custom_palette.split(',')]
                
                # Crear gráfico
                chart_id, chart = add_chart(wb, sheet_name, chart_type, data_range, title, position, style, theme, custom_palette)
                
                # Guardar cambios
                wb.save(file_path)
                
                # Extraer tipo de gráfico para mejor mensaje de respuesta
                chart_type_display = chart_type
                if chart_type.lower() == "col":
                    chart_type_display = "column"
                
                # Procesar info de estilo para respuesta
                style_info = ""
                if isinstance(style, dict):
                    style_parts = []
                    if 'width' in style:
                        style_parts.append(f"ancho:{style['width']}")
                    if 'height' in style:
                        style_parts.append(f"alto:{style['height']}")
                    if 'colors' in style:
                        num_colors = len(style['colors']) if isinstance(style['colors'], list) else "personalizado"
                        style_parts.append(f"colores:{num_colors}")
                    if 'chart_style' in style:
                        style_parts.append(f"estilo:{style['chart_style']}")
                    if 'color_scheme' in style:
                        style_parts.append(f"esquema:{style['color_scheme']}")
                    style_info = f"Configuración personalizada: {', '.join(style_parts)}"
                else:
                    # Obtener info del estilo si es válido
                    style_number = parse_chart_style(style)
                    if style_number is not None:
                        scheme_name = STYLE_TO_PALETTE.get(style_number, 'default')
                        style_info = f"Estilo {style_number} ('{style}') con esquema de colores '{scheme_name}'"
                    else:
                        style_info = "Estilo estándar (no se especificó un estilo válido)"
                
                return {
                    "success": True,
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "chart_id": chart_id,
                    "chart_type": chart_type_display,
                    "data_range": data_range,
                    "title": title,
                    "position": position,
                    "style": style_info,
                    "message": f"Gráfico '{chart_type_display}' creado correctamente con ID {chart_id}"
                }
            
            except Exception as openpyxl_error:
                logger.warning(f"Error con openpyxl: {openpyxl_error}. Intentando con COM Excel/pywin32...")
                
                # Si openpyxl falla y tenemos pywin32, intentar con Excel COM
                if HAS_PYWIN32:
                    try:
                        # Mapeo de tipos de gráficos a constantes de Excel
                        excel_chart_types = {
                            'column': -4100,  # xlColumnClustered
                            'bar': -4120,     # xlBarClustered
                            'line': 4,        # xlLine
                            'pie': 5,         # xlPie
                            'scatter': -4169,  # xlXYScatterSmooth
                            'area': 1         # xlArea
                        }
                        
                        # Obtener el tipo de gráfico para Excel
                        excel_chart_type = excel_chart_types.get(chart_type.lower(), -4100)  # Default a column
                        
                        # Iniciar Excel
                        excel = win32com.client.Dispatch("Excel.Application")
                        excel.Visible = False  # Evitar mostrar la aplicación
                        
                        # Abrir el archivo
                        wb = excel.Workbooks.Open(file_path)
                        ws = wb.Sheets(sheet_name)
                        
                        # Convertir posición si está especificada
                        left, top, width, height = 100, 100, 400, 300  # Valores predeterminados
                        if position:
                            # Convertir posición de celda a coordenadas
                            cell = ws.Range(position)
                            left = cell.Left
                            top = cell.Top
                        
                        # Añadir gráfico
                        chart = ws.Shapes.AddChart2(-1, excel_chart_type).Chart
                        
                        # Configurar el rango de datos
                        chart.SetSourceData(ws.Range(data_range))
                        
                        # Configurar título si se proporciona
                        if title:
                            chart.HasTitle = True
                            chart.ChartTitle.Text = title
                            
                        # Procesar cualquier formato de estilo
                        style_number = parse_chart_style(style)
                        if style_number is not None:
                            # Excel COM aplica estilo directamente con estilo numérico
                            logger.info(f"Detectado estilo predefinido: {style_number} (original: '{style}')")
                            try:
                                chart.Chart.ChartStyle = style_number
                            except Exception as e:
                                logger.warning(f"Error al aplicar estilo {style_number}: {e}")
                        elif isinstance(style, str):
                            logger.warning(f"Estilo no reconocido: '{style}'. Se aplicará un estilo estándar.")
                            logger.info("Estilos soportados: números 1-48, 'styleN', 'dark-blue', 'light-1', etc.")
                        
                        # Aplicar estilos adicionales si se proporcionan
                        if style and isinstance(style, dict):
                            if 'width' in style and 'height' in style:
                                chart.Parent.Width = style['width']
                                chart.Parent.Height = style['height']
                        
                        # Posicionar el gráfico
                        chart.Parent.Left = left
                        chart.Parent.Top = top
                        chart.Parent.Width = width
                        chart.Parent.Height = height
                        
                        # Guardar y cerrar
                        wb.Save()
                        wb.Close()
                        excel.Quit()
                        
                        return {
                            "success": True,
                            "file_path": file_path,
                            "sheet_name": sheet_name,
                            "chart_type": chart_type,
                            "data_range": data_range,
                            "title": title,
                            "position": position,
                            "message": f"Gráfico '{chart_type}' creado correctamente usando Excel COM",
                            "note": "Se utilizó la API COM de Excel con pywin32 para crear el gráfico."
                        }
                    except Exception as com_error:
                        # Si también falla el método COM, relanzar el error original
                        logger.error(f"También falló el método COM: {com_error}")
                        raise openpyxl_error
                else:
                    # Si no tenemos pywin32, relanzar el error original
                    raise openpyxl_error
        except Exception as e:
            error_msg = f"Error al crear gráfico: {e}"
            
            # Añadir sugerencia para instalar pywin32 si es relevante
            if "'NoneType' object has no attribute" in str(e) and not HAS_PYWIN32:
                error_msg += " Se recomienda instalar pywin32 para una funcionalidad alternativa: pip install pywin32"
            
            return {
                "success": False,
                "error": str(e),
                "message": error_msg
            }
    
    @mcp.tool(description="Elimina un gráfico de la hoja")
    def delete_chart_tool(file_path, sheet_name, chart_id):
        """Elimina un gráfico de la hoja"""
        try:
            # Convertir chart_id a entero
            chart_id = int(chart_id)
            
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Eliminar el gráfico
            delete_chart(wb, sheet_name, chart_id)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "chart_id": chart_id,
                "message": f"Gráfico con ID {chart_id} eliminado correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al eliminar gráfico: {e}"
            }
    
    @mcp.tool(description="Lista todos los gráficos en una hoja")
    def list_charts_tool(file_path, sheet_name):
        """Lista todos los gráficos en una hoja"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Obtener la hoja
            ws = get_sheet(wb, sheet_name)
            
            # Listar los gráficos
            charts = list_charts(ws)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "charts": charts,
                "count": len(charts),
                "message": f"Se encontraron {len(charts)} gráficos en la hoja {sheet_name}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al listar gráficos: {e}"
            }
    
    # 9. Pivot Tables (Tablas Dinámicas) - Herramientas MCP
    @mcp.tool(description="Crea una tabla dinámica")
    def add_pivot_table_tool(file_path, source_sheet, source_range, target_sheet, target_cell, rows, cols, data_fields):
        """Crea una tabla dinámica"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Convertir listas de JSON a listas si es necesario
            for param_name, param_value in [('rows', rows), ('cols', cols), ('data_fields', data_fields)]:
                if isinstance(param_value, str):
                    try:
                        locals()[param_name] = json.loads(param_value)
                    except:
                        # Si no es JSON válido, asumir que es una sola cadena
                        locals()[param_name] = [param_value]
                elif not isinstance(param_value, list):
                    locals()[param_name] = [param_value]
            
            # Crear la tabla dinámica
            pivot_table = add_pivot_table(wb, source_sheet, source_range, target_sheet, target_cell, rows, cols, data_fields)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "source_sheet": source_sheet,
                "source_range": source_range,
                "target_sheet": target_sheet,
                "target_cell": target_cell,
                "pivot_name": pivot_table.name if hasattr(pivot_table, 'name') else "PivotTable",
                "message": f"Tabla dinámica creada correctamente en {target_sheet} desde {source_sheet}",
                "warning": "Las tablas dinámicas en openpyxl tienen funcionalidad limitada y pueden necesitar ajustes en Excel"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al crear tabla dinámica: {e}"
            }
    
    @mcp.tool(description="Elimina una tabla dinámica")
    def delete_pivot_table_tool(file_path, sheet_name, pivot_name):
        """Elimina una tabla dinámica"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Eliminar la tabla dinámica
            delete_pivot_table(wb, sheet_name, pivot_name)
            
            # Guardar cambios
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "pivot_name": pivot_name,
                "message": f"Tabla dinámica '{pivot_name}' eliminada correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al eliminar tabla dinámica: {e}"
            }
    
    @mcp.tool(description="Lista todas las tablas dinámicas en el workbook")
    def list_pivot_tables_tool(file_path):
        """Lista todas las tablas dinámicas en el workbook"""
        try:
            # Abrir el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Listar las tablas dinámicas
            pivot_tables = list_pivot_tables(wb)
            
            # Contar el total
            total_count = sum(len(pivots) for pivots in pivot_tables.values())
            
            return {
                "success": True,
                "file_path": file_path,
                "pivot_tables": pivot_tables,
                "count": total_count,
                "message": f"Se encontraron {total_count} tablas dinámicas en el workbook"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al listar tablas dinámicas: {e}"
            }


if __name__ == "__main__":
    # Código de ejemplo de uso
    logger.info("Advanced Excel MCP - Ejemplo de uso")
    
    # Verificar argumentos
    if len(sys.argv) > 1:
        comando = sys.argv[1].lower()
        
        if comando == "tabla" and len(sys.argv) > 4:
            archivo, hoja, tabla, rango = sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5]
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb[hoja]
                
                # Crear tabla
                add_table(ws, tabla, rango, "TableStyleMedium9")
                
                wb.save(archivo)
                logger.info(f"Tabla '{tabla}' creada correctamente en rango {rango}")
            except Exception as e:
                logger.error(f"Error: {e}")
                
        elif comando == "formula" and len(sys.argv) > 4:
            archivo, hoja, celda, formula = sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5]
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb[hoja]
                
                # Establecer fórmula
                set_formula(ws, celda, formula)
                
                wb.save(archivo)
                logger.info(f"Fórmula establecida correctamente en celda {celda}")
            except Exception as e:
                logger.error(f"Error: {e}")
                
        elif comando == "grafico" and len(sys.argv) > 4:
            archivo, hoja, tipo, rango = sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5]
            try:
                wb = openpyxl.load_workbook(archivo)
                
                # Crear gráfico
                chart_id, _ = add_chart(wb, hoja, tipo, rango, f"Gráfico de {tipo}", "A10")
                
                wb.save(archivo)
                logger.info(f"Gráfico creado correctamente con ID {chart_id}")
            except Exception as e:
                logger.error(f"Error: {e}")
                
        else:
            logger.info("Comando no reconocido o faltan argumentos")
            logger.info("Uso: python advanced_excel_mcp.py [tabla|formula|grafico] archivo.xlsx hoja [args adicionales]")
    else:
        logger.info("Uso: python advanced_excel_mcp.py [tabla|formula|grafico] archivo.xlsx hoja [args adicionales]")
