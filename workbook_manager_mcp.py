#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Workbook Manager MCP (Multi-purpose Connector for Python)
-------------------------------------------------------
Biblioteca para gestionar workbooks y hojas de Excel:
- Creación y manejo de workbooks
- Gestión de propiedades
- Manipulación de hojas (añadir, eliminar, renombrar, copiar)
- Configuración de propiedades de hojas

Author: MCP Team
Version: 1.0
"""

import os
import sys
import logging
from typing import List, Dict, Union, Optional, Tuple, Any

# Configuración de logging
logger = logging.getLogger("workbook_manager_mcp")
logger.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stderr)
handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
logger.addHandler(handler)

# Importar MCP
try:
    from mcp.server.fastmcp import FastMCP
    HAS_MCP = True
except ImportError:
    logger.warning("No se pudo importar FastMCP. Las funcionalidades de servidor MCP no estarán disponibles.")
    HAS_MCP = False

# Intentar importar las bibliotecas necesarias
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.filters import AutoFilter
    HAS_OPENPYXL = True
except ImportError as e:
    logger.warning(f"Error al importar bibliotecas esenciales: {e}")
    logger.warning("Es posible que algunas funcionalidades no estén disponibles")
    HAS_OPENPYXL = False

# Excepciones personalizadas
class WorkbookManagerError(Exception):
    """Excepción base para todos los errores del Workbook Manager."""
    pass

class FileExistsError(WorkbookManagerError):
    """Se lanza cuando se intenta crear un archivo que ya existe."""
    pass

class FileNotFoundError(WorkbookManagerError):
    """Se lanza cuando no se encuentra un archivo Excel."""
    pass

class SheetExistsError(WorkbookManagerError):
    """Se lanza cuando se intenta crear una hoja que ya existe."""
    pass

class SheetNotFoundError(WorkbookManagerError):
    """Se lanza cuando no se encuentra una hoja en el archivo Excel."""
    pass


# 1. Gestión de Workbooks

def create_workbook(filename: str, overwrite: bool = False) -> Any:
    """
    Crea un nuevo fichero Excel vacío.
    
    Args:
        filename (str): Ruta y nombre del archivo a crear.
        overwrite (bool, opcional): Si es True, sobreescribe archivo existente.
        
    Returns:
        Objeto Workbook.
        
    Raises:
        FileExistsError: Si el archivo existe y overwrite es False.
    """
    if os.path.exists(filename) and not overwrite:
        raise FileExistsError(f"El archivo '{filename}' ya existe. Use overwrite=True para sobreescribir.")
    
    wb = openpyxl.Workbook()
    if overwrite:
        save_workbook(wb, filename)
    return wb

def open_workbook(filename: str) -> Any:
    """
    Abre un fichero Excel existente.
    
    Args:
        filename (str): Ruta del archivo.
        
    Returns:
        Objeto Workbook.
        
    Raises:
        FileNotFoundError: Si el archivo no existe.
    """
    if not os.path.exists(filename):
        raise FileNotFoundError(f"El archivo '{filename}' no existe.")
    
    try:
        wb = openpyxl.load_workbook(filename)
        return wb
    except Exception as e:
        logger.error(f"Error al abrir el archivo '{filename}': {e}")
        raise WorkbookManagerError(f"Error al abrir el archivo: {e}")

def save_workbook(wb: Any, filename: Optional[str] = None) -> str:
    """
    Guarda el Workbook en disco.
    
    Args:
        wb: Objeto Workbook.
        filename (str, opcional): Si se indica, guarda con otro nombre.
        
    Returns:
        Ruta del fichero guardado.
        
    Raises:
        WorkbookManagerError: Si hay error al guardar.
    """
    if not wb:
        raise WorkbookManagerError("El workbook no puede ser None")
    
    try:
        # Si no se proporciona filename, usar el filename original si existe
        if not filename and hasattr(wb, 'path'):
            filename = wb.path
        elif not filename:
            raise WorkbookManagerError("Debe proporcionar un nombre de archivo")
        
        wb.save(filename)
        return filename
    except Exception as e:
        logger.error(f"Error al guardar el workbook en '{filename}': {e}")
        raise WorkbookManagerError(f"Error al guardar el workbook: {e}")

def close_workbook(wb: Any) -> None:
    """
    Cierra el Workbook en memoria.
    
    Args:
        wb: Objeto Workbook.
        
    Returns:
        Ninguno.
    """
    if not wb:
        return
    
    try:
        # Openpyxl realmente no tiene un método close(),
        # pero podemos eliminar referencias para ayudar al GC
        if hasattr(wb, "_archive"):
            wb._archive.close()
    except Exception as e:
        logger.warning(f"Advertencia al cerrar workbook: {e}")

def get_properties(wb: Any) -> Dict[str, str]:
    """
    Lee metadatos del fichero (autor, título, empresa, fecha).
    
    Args:
        wb: Objeto Workbook.
        
    Returns:
        Diccionario con propiedades.
    """
    if not wb:
        raise WorkbookManagerError("El workbook no puede ser None")
    
    props = {}
    
    # Acceder a las propiedades del documento
    if hasattr(wb, 'properties'):
        # Extraer propiedades principales
        props_obj = wb.properties
        
        # Propiedades principales
        if hasattr(props_obj, 'creator'):
            props['author'] = props_obj.creator
        if hasattr(props_obj, 'title'):
            props['title'] = props_obj.title
        if hasattr(props_obj, 'description'):
            props['description'] = props_obj.description
        if hasattr(props_obj, 'subject'):
            props['subject'] = props_obj.subject
        if hasattr(props_obj, 'keywords'):
            props['keywords'] = props_obj.keywords
        if hasattr(props_obj, 'category'):
            props['category'] = props_obj.category
        
        # Opcionalmente extraer propiedades de la empresa
        if hasattr(props_obj, 'company'):
            props['company'] = props_obj.company
        elif hasattr(props_obj, 'lastModifiedBy'):
            props['lastModifiedBy'] = props_obj.lastModifiedBy
        
        # Fechas (convertidas a string)
        if hasattr(props_obj, 'created') and props_obj.created:
            props['created'] = str(props_obj.created)
        if hasattr(props_obj, 'modified') and props_obj.modified:
            props['modified'] = str(props_obj.modified)
    
    return props

def set_properties(wb: Any, **props) -> None:
    """
    Escribe metadatos del fichero (autor, título, empresa, fecha).
    
    Args:
        wb: Objeto Workbook.
        **props: Propiedades a establecer (author, title, company, etc.)
        
    Returns:
        Ninguno.
    """
    if not wb:
        raise WorkbookManagerError("El workbook no puede ser None")
    
    if not hasattr(wb, 'properties'):
        logger.warning("Este workbook no tiene propiedades accesibles")
        return
    
    # Mapeo de nombres de propiedades
    prop_mapping = {
        'author': 'creator',
        'creator': 'creator',
        'title': 'title',
        'description': 'description',
        'subject': 'subject',
        'keywords': 'keywords',
        'category': 'category',
        'company': 'company',
        'lastModifiedBy': 'lastModifiedBy'
    }
    
    # Establecer propiedades
    for prop_name, prop_value in props.items():
        if prop_name in prop_mapping:
            openpyxl_prop_name = prop_mapping[prop_name]
            if hasattr(wb.properties, openpyxl_prop_name):
                setattr(wb.properties, openpyxl_prop_name, prop_value)
            else:
                logger.warning(f"La propiedad '{openpyxl_prop_name}' no es accesible en este workbook")
        else:
            logger.warning(f"Propiedad no reconocida: '{prop_name}'")

def list_sheets(wb: Any) -> List[str]:
    """
    Devuelve lista de nombres de hojas.
    
    Args:
        wb: Objeto Workbook.
        
    Returns:
        List[str]: Lista de nombres de hojas.
    """
    if not wb:
        raise WorkbookManagerError("El workbook no puede ser None")
    
    if hasattr(wb, 'sheetnames'):
        return wb.sheetnames
    
    # Alternativa si no se puede acceder a sheetnames
    sheet_names = []
    for sheet in wb.worksheets:
        if hasattr(sheet, 'title'):
            sheet_names.append(sheet.title)
    
    return sheet_names


# 2. Gestión de Hojas (Sheets)

def add_sheet(wb: Any, sheet_name: str, index: Optional[int] = None) -> Any:
    """
    Añade una nueva hoja vacía.
    
    Args:
        wb: Objeto Workbook.
        sheet_name (str): Nombre de la hoja.
        index (int, opcional): Posición en pestañas.
        
    Returns:
        Hoja creada.
        
    Raises:
        SheetExistsError: Si ya existe una hoja con ese nombre.
    """
    if not wb:
        raise WorkbookManagerError("El workbook no puede ser None")
    
    # Verificar si ya existe una hoja con ese nombre
    if sheet_name in list_sheets(wb):
        raise SheetExistsError(f"Ya existe una hoja con el nombre '{sheet_name}'")
    
    # Crear nueva hoja
    if index is not None:
        ws = wb.create_sheet(sheet_name, index)
    else:
        ws = wb.create_sheet(sheet_name)
    
    return ws

def delete_sheet(wb: Any, sheet_name: str) -> None:
    """
    Elimina la hoja indicada.
    
    Args:
        wb: Objeto Workbook.
        sheet_name (str): Nombre de la hoja a eliminar.
        
    Raises:
        SheetNotFoundError: Si la hoja no existe.
    """
    if not wb:
        raise WorkbookManagerError("El workbook no puede ser None")
    
    # Verificar que la hoja existe
    if sheet_name not in list_sheets(wb):
        raise SheetNotFoundError(f"La hoja '{sheet_name}' no existe en el workbook")
    
    # Eliminar la hoja
    try:
        del wb[sheet_name]
    except Exception as e:
        logger.error(f"Error al eliminar la hoja '{sheet_name}': {e}")
        raise WorkbookManagerError(f"Error al eliminar la hoja: {e}")

def rename_sheet(wb: Any, old_name: str, new_name: str) -> None:
    """
    Renombra una hoja.
    
    Args:
        wb: Objeto Workbook.
        old_name (str): Nombre actual de la hoja.
        new_name (str): Nuevo nombre para la hoja.
        
    Raises:
        SheetNotFoundError: Si la hoja original no existe.
        SheetExistsError: Si ya existe una hoja con el nuevo nombre.
    """
    if not wb:
        raise WorkbookManagerError("El workbook no puede ser None")
    
    # Verificar que la hoja original existe
    if old_name not in list_sheets(wb):
        raise SheetNotFoundError(f"La hoja '{old_name}' no existe en el workbook")
    
    # Verificar que no exista una hoja con el nuevo nombre
    if new_name in list_sheets(wb) and old_name != new_name:
        raise SheetExistsError(f"Ya existe una hoja con el nombre '{new_name}'")
    
    # Renombrar la hoja
    try:
        wb[old_name].title = new_name
    except Exception as e:
        logger.error(f"Error al renombrar la hoja '{old_name}' a '{new_name}': {e}")
        raise WorkbookManagerError(f"Error al renombrar la hoja: {e}")

def copy_sheet(wb: Any, source_name: str, target_name: str) -> Any:
    """
    Duplica una hoja completa.
    
    Args:
        wb: Objeto Workbook.
        source_name (str): Nombre de la hoja origen.
        target_name (str): Nombre de la nueva hoja.
        
    Returns:
        Hoja creada.
        
    Raises:
        SheetNotFoundError: Si la hoja origen no existe.
        SheetExistsError: Si ya existe una hoja con el nombre destino.
    """
    if not wb:
        raise WorkbookManagerError("El workbook no puede ser None")
    
    # Verificar que la hoja origen existe
    if source_name not in list_sheets(wb):
        raise SheetNotFoundError(f"La hoja origen '{source_name}' no existe en el workbook")
    
    # Verificar que no exista una hoja con el nombre destino
    if target_name in list_sheets(wb):
        raise SheetExistsError(f"Ya existe una hoja con el nombre '{target_name}'")
    
    # Duplicar la hoja
    try:
        # Openpyxl no tiene un método directo para copiar hojas,
        # así que debemos crear una nueva y copiar el contenido
        source_sheet = wb[source_name]
        target_sheet = wb.create_sheet(target_name)
        
        # Copiar dimensiones y propiedades
        target_sheet.sheet_properties = source_sheet.sheet_properties
        
        # Copiar contenido (celdas)
        for row in source_sheet.rows:
            for cell in row:
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                target_cell.value = cell.value
                if cell.has_style:
                    target_cell.font = cell.font.copy()
                    target_cell.border = cell.border.copy()
                    target_cell.fill = cell.fill.copy()
                    target_cell.number_format = cell.number_format
                    target_cell.protection = cell.protection.copy()
                    target_cell.alignment = cell.alignment.copy()
        
        # Copiar fusiones de celdas
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))
        
        # Copiar dimensiones de columnas
        for col_idx, column in enumerate(source_sheet.columns, 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in source_sheet.column_dimensions:
                source_col_dim = source_sheet.column_dimensions[col_letter]
                target_col_dim = target_sheet.column_dimensions[col_letter]
                target_col_dim.width = source_col_dim.width
                target_col_dim.hidden = source_col_dim.hidden
        
        # Copiar dimensiones de filas
        for row_idx, row in enumerate(source_sheet.rows, 1):
            if row_idx in source_sheet.row_dimensions:
                source_row_dim = source_sheet.row_dimensions[row_idx]
                target_row_dim = target_sheet.row_dimensions[row_idx]
                target_row_dim.height = source_row_dim.height
                target_row_dim.hidden = source_row_dim.hidden
        
        return target_sheet
    except Exception as e:
        logger.error(f"Error al copiar la hoja '{source_name}' a '{target_name}': {e}")
        raise WorkbookManagerError(f"Error al copiar la hoja: {e}")

def set_sheet_properties(ws: Any, **options) -> None:
    """
    Configura pestaña: color de pestaña, visibilidad, freeze_panes, filtros.
    
    Args:
        ws: Objeto worksheet.
        **options: Opciones a establecer:
            - tab_color: Color de la pestaña (formato hex: "FF0000")
            - visible: Booleano indicando si la hoja es visible
            - freeze_panes: Celda donde congelar (ej: "B2")
            - auto_filter_range: Rango para filtro automático (ej: "A1:D10")
            
    Returns:
        Ninguno.
    """
    if not ws:
        raise WorkbookManagerError("El worksheet no puede ser None")
    
    try:
        # Establecer color de pestaña
        if 'tab_color' in options:
            ws.sheet_properties.tabColor = options['tab_color']
        
        # Establecer visibilidad
        if 'visible' in options:
            if not options['visible']:
                ws.sheet_state = 'hidden'
            else:
                ws.sheet_state = 'visible'
        
        # Establecer freeze_panes
        if 'freeze_panes' in options:
            freeze_cell = options['freeze_panes']
            ws.freeze_panes = ws[freeze_cell]
        
        # Establecer filtro automático
        if 'auto_filter_range' in options:
            filter_range = options['auto_filter_range']
            ws.auto_filter.ref = filter_range
    
    except Exception as e:
        logger.error(f"Error al establecer propiedades de hoja: {e}")
        raise WorkbookManagerError(f"Error al establecer propiedades de hoja: {e}")


# Crear el servidor MCP como variable global
mcp = None
if HAS_MCP:
    # Esta es la variable global que el sistema MCP busca
    mcp = FastMCP("Workbook Manager MCP", 
                 dependencies=["openpyxl"])
    logger.info("Servidor MCP iniciado correctamente")
    
    # 1. Gestión de Workbooks - Herramientas MCP
    @mcp.tool(description="Crea un nuevo fichero Excel vacío")
    def create_workbook_tool(filename, overwrite=False):
        """Crea un nuevo fichero Excel vacío"""
        try:
            wb = create_workbook(filename, overwrite)
            save_workbook(wb, filename)
            
            return {
                "success": True,
                "file_path": filename,
                "message": f"Archivo Excel creado correctamente: {filename}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al crear archivo Excel: {e}"
            }
    
    @mcp.tool(description="Abre un fichero Excel existente")
    def open_workbook_tool(filename):
        """Abre un fichero Excel existente"""
        try:
            wb = open_workbook(filename)
            sheet_names = list_sheets(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "sheets": sheet_names,
                "sheet_count": len(sheet_names),
                "message": f"Archivo Excel abierto correctamente: {filename}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al abrir archivo Excel: {e}"
            }
    
    @mcp.tool(description="Guarda el Workbook en disco")
    def save_workbook_tool(filename, new_filename=None):
        """Guarda el Workbook en disco"""
        try:
            wb = open_workbook(filename)
            saved_path = save_workbook(wb, new_filename or filename)
            close_workbook(wb)
            
            return {
                "success": True,
                "original_file": filename,
                "saved_file": saved_path,
                "message": f"Archivo Excel guardado correctamente: {saved_path}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al guardar archivo Excel: {e}"
            }
    
    @mcp.tool(description="Obtiene las propiedades de un archivo Excel")
    def get_properties_tool(filename):
        """Obtiene las propiedades de un archivo Excel"""
        try:
            wb = open_workbook(filename)
            props = get_properties(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "properties": props,
                "message": f"Propiedades obtenidas correctamente de: {filename}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al obtener propiedades: {e}"
            }
    
    @mcp.tool(description="Establece las propiedades de un archivo Excel")
    def set_properties_tool(filename, **props):
        """Establece las propiedades de un archivo Excel"""
        try:
            wb = open_workbook(filename)
            set_properties(wb, **props)
            save_workbook(wb, filename)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "properties_set": props,
                "message": f"Propiedades establecidas correctamente en: {filename}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al establecer propiedades: {e}"
            }
    
    @mcp.tool(description="Lista las hojas disponibles en un archivo Excel")
    def list_sheets_tool(filename):
        """Lista las hojas disponibles en un archivo Excel"""
        try:
            wb = open_workbook(filename)
            sheets = list_sheets(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "sheets": sheets,
                "count": len(sheets),
                "message": f"Se encontraron {len(sheets)} hojas en el archivo Excel"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al listar hojas: {e}"
            }
    
    # 2. Gestión de Hojas - Herramientas MCP
    @mcp.tool(description="Añade una nueva hoja vacía")
    def add_sheet_tool(filename, sheet_name, index=None):
        """Añade una nueva hoja vacía"""
        try:
            wb = open_workbook(filename)
            ws = add_sheet(wb, sheet_name, index)
            save_workbook(wb, filename)
            
            sheets = list_sheets(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "sheet_name": sheet_name,
                "sheet_index": sheets.index(sheet_name),
                "all_sheets": sheets,
                "message": f"Hoja '{sheet_name}' añadida correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al añadir hoja: {e}"
            }
    
    @mcp.tool(description="Elimina la hoja indicada")
    def delete_sheet_tool(filename, sheet_name):
        """Elimina la hoja indicada"""
        try:
            wb = open_workbook(filename)
            delete_sheet(wb, sheet_name)
            save_workbook(wb, filename)
            
            remaining_sheets = list_sheets(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "deleted_sheet": sheet_name,
                "remaining_sheets": remaining_sheets,
                "remaining_count": len(remaining_sheets),
                "message": f"Hoja '{sheet_name}' eliminada correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al eliminar hoja: {e}"
            }
    
    @mcp.tool(description="Renombra una hoja")
    def rename_sheet_tool(filename, old_name, new_name):
        """Renombra una hoja"""
        try:
            wb = open_workbook(filename)
            rename_sheet(wb, old_name, new_name)
            save_workbook(wb, filename)
            
            sheets = list_sheets(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "old_name": old_name,
                "new_name": new_name,
                "all_sheets": sheets,
                "message": f"Hoja renombrada de '{old_name}' a '{new_name}'"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al renombrar hoja: {e}"
            }
    
    @mcp.tool(description="Duplica una hoja completa")
    def copy_sheet_tool(filename, source_name, target_name):
        """Duplica una hoja completa"""
        try:
            wb = open_workbook(filename)
            copy_sheet(wb, source_name, target_name)
            save_workbook(wb, filename)
            
            sheets = list_sheets(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "source_sheet": source_name,
                "target_sheet": target_name,
                "all_sheets": sheets,
                "message": f"Hoja '{source_name}' copiada a '{target_name}'"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al copiar hoja: {e}"
            }
    
    @mcp.tool(description="Configura propiedades de una hoja")
    def set_sheet_properties_tool(filename, sheet_name, **options):
        """Configura propiedades de una hoja"""
        try:
            wb = open_workbook(filename)
            
            # Verificar que la hoja existe
            if sheet_name not in list_sheets(wb):
                raise SheetNotFoundError(f"La hoja '{sheet_name}' no existe en el workbook")
                
            ws = wb[sheet_name]
            set_sheet_properties(ws, **options)
            save_workbook(wb, filename)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "sheet_name": sheet_name,
                "options_set": options,
                "message": f"Propiedades de hoja '{sheet_name}' configuradas correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al configurar propiedades de hoja: {e}"
            }

if __name__ == "__main__":
    # Código de ejemplo de uso
    logger.info("Workbook Manager MCP - Ejemplo de uso")
    
    # Verificar argumentos
    if len(sys.argv) > 1:
        comando = sys.argv[1].lower()
        
        if comando == "crear" and len(sys.argv) > 2:
            archivo = sys.argv[2]
            try:
                wb = create_workbook(archivo, overwrite=True)
                save_workbook(wb, archivo)
                logger.info(f"Archivo creado correctamente: {archivo}")
                close_workbook(wb)
            except Exception as e:
                logger.error(f"Error al crear archivo: {e}")
                
        elif comando == "listar" and len(sys.argv) > 2:
            archivo = sys.argv[2]
            try:
                wb = open_workbook(archivo)
                hojas = list_sheets(wb)
                logger.info(f"Hojas en el archivo {archivo}: {hojas}")
                close_workbook(wb)
            except Exception as e:
                logger.error(f"Error al listar hojas: {e}")
                
        elif comando == "añadir" and len(sys.argv) > 3:
            archivo, hoja = sys.argv[2], sys.argv[3]
            try:
                wb = open_workbook(archivo)
                ws = add_sheet(wb, hoja)
                save_workbook(wb, archivo)
                logger.info(f"Hoja '{hoja}' añadida correctamente a {archivo}")
                close_workbook(wb)
            except Exception as e:
                logger.error(f"Error al añadir hoja: {e}")
                
        else:
            logger.info("Comando no reconocido o faltan argumentos")
            logger.info("Uso: python workbook_manager_mcp.py [crear|listar|añadir] archivo.xlsx [nombre_hoja]")
    else:
        logger.info("Uso: python workbook_manager_mcp.py [crear|listar|añadir] archivo.xlsx [nombre_hoja]")
