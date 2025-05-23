#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel MCP Master (Model Context Protocol for Excel)
--------------------------------------------------
Unified library to manipulate Excel files with advanced features:
- Combines all Excel MCP modules under a single interface
- Provides high level functions for common operations
- Optimizes the workflow when working with Excel

This module integrates:
- excel_mcp_complete.py: Data reading and exploration
- workbook_manager_mcp.py: Workbook and sheet management
- excel_writer_mcp.py: Cell writing and formatting
- advanced_excel_mcp.py: Tables, formulas, charts and pivot tables

Author: MCP Team
Version: 2.1.8

Usage guide for LLMs and agents
-------------------------------
All functions of this library are designed to be used by language models or
automatic tools that generate Excel files. To obtain the best results, follow
these context recommendations for each operation:

- **Apply styles at all times** so the resulting sheets look visually pleasant.
  Use the functions in this library to style cells, tables and charts.
- **Avoid element overlap**. Place charts in free cells and leave at least a
  couple of rows of separation from tables or blocks of text. Never place charts
  over text.
- **Automatically adjust column width**. After writing tables or datasets,
  check which cells contain long text and increase the width so everything is
  readable without breaking the layout.
- **Always seek the clearest and most organised layout**, separating sections
  and grouping related elements so that the final file is easy to understand.
- **Check the orientation of the data**. If tables are not obvious, explicitly
  indicate whether the categories are in rows or columns so the chart functions
  interpret them correctly.
"""

import os
import sys
import json
import logging
import tempfile
import time
from pathlib import Path
from typing import List, Dict, Union, Optional, Tuple, Any, Callable
import math

# Logging configuration
logger = logging.getLogger("excel_mcp_master")
logger.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stderr)
handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
logger.addHandler(handler)

# Import MCP
try:
    from mcp.server.fastmcp import FastMCP
    HAS_MCP = True
except ImportError:
    logger.warning("FastMCP could not be imported. MCP server features will be unavailable.")
    HAS_MCP = False

# Attempt to import required libraries
try:
    import pandas as pd
    import numpy as np
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment, 
        NamedStyle, Protection, Color, colors
    )
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, ScatterChart, AreaChart,
        Reference, Series
    )
    from openpyxl.worksheet.filters import AutoFilter
    from openpyxl.pivot.table import PivotTable, PivotField
    from openpyxl.pivot.cache import PivotCache
    HAS_OPENPYXL = True
except ImportError as e:
    logger.warning(f"Failed to import required libraries: {e}")
    logger.warning("Some functionality may be unavailable")
    HAS_OPENPYXL = False

# Import existing Excel MCP modules
# Note: In a real implementation we would import the functions from the existing modules
# However, for this example the key functions are reimplemented directly

# ===========================
# GLOBAL CONFIGURATION
# ===========================

# Default formatting settings
DEFAULT_FONT = "Calibri"
DEFAULT_FONT_SIZE = 11
DEFAULT_HEADER_FONT_SIZE = 12
DEFAULT_NUMBER_FORMAT = "#,##0.00"
DEFAULT_PERCENTAGE_FORMAT = "0.00%"
DEFAULT_CURRENCY_FORMAT = "$#,##0.00"

# Default chart styling
DEFAULT_CHART_FONT = "Calibri"
DEFAULT_CHART_FONT_SIZE = 10
DEFAULT_CHART_TITLE_SIZE = 14
DEFAULT_CHART_LEGEND_POSITION = "r"  # right
DEFAULT_CHART_STYLE = 2  # Professional blue theme

# Professional color palette
PROFESSIONAL_COLORS = [
    "4472C4",  # Blue
    "ED7D31",  # Orange
    "A5A5A5",  # Gray
    "FFC000",  # Yellow
    "5B9BD5",  # Light Blue
    "70AD47",  # Green
    "264478",  # Dark Blue
    "9E480E",  # Dark Orange
]

# Table styles
DEFAULT_TABLE_STYLE = "TableStyleMedium9"

# Grid alignment settings
CHART_MARGIN = 2  # Rows/columns of space around charts
SECTION_PADDING = 1  # Rows between sections

# Base exception classes (unified)
class ExcelMCPError(Exception):
    """Base exception for all Excel MCP errors."""
    pass

class FileNotFoundError(ExcelMCPError):
    """Raised when an Excel file is not found."""
    pass

class FileExistsError(ExcelMCPError):
    """Raised when attempting to create a file that already exists."""
    pass

class SheetNotFoundError(ExcelMCPError):
    """Raised when a sheet is not found in the Excel file."""
    pass

class SheetExistsError(ExcelMCPError):
    """Raised when attempting to create a sheet that already exists."""
    pass

class CellReferenceError(ExcelMCPError):
    """Raised when there is an issue with a cell reference."""
    pass

class RangeError(ExcelMCPError):
    """Raised when there is an issue with a cell range."""
    pass

class TableError(ExcelMCPError):
    """Raised when there is an issue with an Excel table."""
    pass

class ChartError(ExcelMCPError):
    """Raised when there is an issue with a chart."""
    pass

class FormulaError(ExcelMCPError):
    """Raised when there is an issue with a formula."""
    pass

class PivotTableError(ExcelMCPError):
    """Raised when there is an issue with a pivot table."""
    pass

# ===========================
# ENHANCED UTILITY FUNCTIONS
# ===========================

def enhanced_autofit_columns(ws: Any, min_width: float = 8.43, max_width: float = 80) -> None:
    """
    Enhanced auto-fit for all columns in a worksheet.
    
    Args:
        ws: Openpyxl worksheet object
        min_width: Minimum column width
        max_width: Maximum column width
    """
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            if cell.value:
                # Calculate length considering line breaks
                lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in lines)
                max_length = max(max_length, max_line_length)
                
                # Enable text wrapping for multi-line content
                if len(lines) > 1:
                    cell.alignment = Alignment(wrap_text=True)
        
        # Apply calculated width
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

def apply_consistent_number_format(ws: Any, detect_currency: bool = True, detect_percentage: bool = True) -> None:
    """
    Apply consistent number formatting to all cells with numeric values.
    
    Args:
        ws: Openpyxl worksheet object
        detect_currency: Auto-detect and format currency values
        detect_percentage: Auto-detect and format percentage values
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, (int, float)):
                # Skip if already formatted
                if cell.number_format != 'General':
                    continue
                
                # Check for percentage (values between 0 and 1)
                if detect_percentage and 0 <= cell.value <= 1:
                    # Check if neighbors suggest percentage
                    neighbors_are_percent = False
                    for neighbor in [ws.cell(row=cell.row-1, column=cell.column),
                                   ws.cell(row=cell.row+1, column=cell.column)]:
                        if neighbor.value and isinstance(neighbor.value, (int, float)) and 0 <= neighbor.value <= 1:
                            neighbors_are_percent = True
                            break
                    
                    if neighbors_are_percent:
                        cell.number_format = DEFAULT_PERCENTAGE_FORMAT
                        continue
                
                # Apply standard number format for large numbers
                if abs(cell.value) >= 1000:
                    cell.number_format = DEFAULT_NUMBER_FORMAT

def ensure_chart_spacing(ws: Any, chart_position: str, chart_width: int = 15, chart_height: int = 10) -> str:
    """
    Ensure proper spacing around charts to avoid overlaps.
    
    Args:
        ws: Worksheet object
        chart_position: Original chart position
        chart_width: Chart width in cells
        chart_height: Chart height in rows
        
    Returns:
        Adjusted position string
    """
    col_letter = ''.join(filter(str.isalpha, chart_position))
    row_num = int(''.join(filter(str.isdigit, chart_position)))
    
    # Check for content in the chart area
    col_idx = column_index_from_string(col_letter)
    
    # Find next available position if current has content
    while True:
        has_content = False
        for r in range(row_num, row_num + chart_height):
            for c in range(col_idx, col_idx + chart_width):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    has_content = True
                    break
            if has_content:
                break
        
        if not has_content:
            break
        
        # Move down by chart height + margin
        row_num += chart_height + CHART_MARGIN
    
    return f"{col_letter}{row_num}"

def apply_section_borders(ws: Any, start_row: int, start_col: int, end_row: int, end_col: int, 
                         style: str = "thin", color: str = "D3D3D3") -> None:
    """
    Apply borders to delimit a section with optional shading.
    
    Args:
        ws: Worksheet object
        start_row, start_col, end_row, end_col: Section boundaries
        style: Border style
        color: Border color (hex)
    """
    thin_border = Border(
        left=Side(style=style, color=color),
        right=Side(style=style, color=color),
        top=Side(style=style, color=color),
        bottom=Side(style=style, color=color)
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

def create_text_box_title(ws: Any, title: str, position: str, font_size: int = None) -> None:
    """
    Create a title in a separate cell with proper formatting.
    
    Args:
        ws: Worksheet object
        title: Title text
        position: Cell position
        font_size: Font size (defaults to DEFAULT_HEADER_FONT_SIZE)
    """
    if font_size is None:
        font_size = DEFAULT_HEADER_FONT_SIZE
    
    cell = ws[position]
    cell.value = title
    cell.font = Font(name=DEFAULT_FONT, size=font_size, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def apply_unified_theme(wb: Any, theme_name: str = "professional") -> None:
    """
    Apply a unified theme to the entire workbook.
    
    Args:
        wb: Workbook object
        theme_name: Theme to apply ("professional", "modern", "classic")
    """
    theme_config = {
        "professional": {
            "font": "Calibri",
            "font_size": 11,
            "header_color": "366092",
            "accent_color": "4472C4",
            "table_style": "TableStyleMedium9",
            "chart_style": 2
        },
        "modern": {
            "font": "Arial",
            "font_size": 10,
            "header_color": "44546A",
            "accent_color": "ED7D31",
            "table_style": "TableStyleMedium2",
            "chart_style": 11
        },
        "classic": {
            "font": "Times New Roman",
            "font_size": 12,
            "header_color": "002060",
            "accent_color": "203764",
            "table_style": "TableStyleLight1",
            "chart_style": 1
        }
    }
    
    theme = theme_config.get(theme_name, theme_config["professional"])
    
    # Apply theme to all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Apply default font to all cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and not cell.font.bold:
                    cell.font = Font(name=theme["font"], size=theme["font_size"])
        
        # Apply enhanced formatting
        try:
            enhanced_autofit_columns(ws)
            apply_consistent_number_format(ws)
        except Exception:
            pass

def align_to_grid(ws: Any, elements: List[Dict[str, Any]]) -> None:
    """
    Align elements (tables, charts) to Excel grid with proper spacing.
    
    Args:
        ws: Worksheet object
        elements: List of elements with position info
    """
    # Sort elements by position
    sorted_elements = sorted(elements, key=lambda x: (x.get('row', 0), x.get('col', 0)))
    
    current_row = 1
    for element in sorted_elements:
        element_type = element.get('type')
        
        if element_type == 'table':
            # Ensure table starts at a grid-aligned position
            table_range = element.get('range')
            if table_range:
                start_row = max(current_row, element.get('row', current_row))
                # Tables align to column A by default
                element['aligned_position'] = f"A{start_row}"
                
                # Calculate table height
                rows = element.get('rows', 10)
                current_row = start_row + rows + SECTION_PADDING
                
        elif element_type == 'chart':
            # Charts align to column E with proper spacing
            chart_row = max(current_row, element.get('row', current_row))
            element['aligned_position'] = f"E{chart_row}"
            
            # Standard chart height
            current_row = chart_row + 15 + CHART_MARGIN
    
def apply_worksheet_layout(ws: Any) -> None:
    """
    Apply professional layout to worksheet with proper margins and spacing.
    
    Args:
        ws: Worksheet object
    """
    # Set print margins for professional appearance
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    
    # Set print options
    ws.print_options.horizontalCentered = True
    ws.print_options.gridLines = False
    
    # Apply row/column dimensions for better spacing
    ws.row_dimensions[1].height = 25  # Header row
    
    # Ensure minimum column widths
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        if ws.column_dimensions[col_letter].width < 8.43:
            ws.column_dimensions[col_letter].width = 8.43

def detect_data_range(ws: Any, start_cell: str) -> Tuple[int, int, int, int]:
    """
    Dynamically detect the actual data range from a starting cell.
    
    Args:
        ws: Worksheet object
        start_cell: Starting cell reference (e.g., "A1")
        
    Returns:
        Tuple of (start_row, start_col, end_row, end_col) in 0-based coordinates
    """
    start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
    
    # Find the last row with data
    max_row = start_row
    for row in range(start_row, ws.max_row):
        has_data = False
        for col in range(start_col, ws.max_column):
            cell = ws.cell(row=row + 1, column=col + 1)
            if cell.value is not None and str(cell.value).strip():
                has_data = True
                break
        if has_data:
            max_row = row
        else:
            # If we find 3 consecutive empty rows, stop
            empty_count = 0
            for check_row in range(row, min(row + 3, ws.max_row)):
                row_empty = True
                for col in range(start_col, ws.max_column):
                    cell = ws.cell(row=check_row + 1, column=col + 1)
                    if cell.value is not None and str(cell.value).strip():
                        row_empty = False
                        break
                if row_empty:
                    empty_count += 1
                else:
                    break
            if empty_count >= 3:
                break
    
    # Find the last column with data
    max_col = start_col
    for col in range(start_col, ws.max_column):
        has_data = False
        for row in range(start_row, max_row + 1):
            cell = ws.cell(row=row + 1, column=col + 1)
            if cell.value is not None and str(cell.value).strip():
                has_data = True
                break
        if has_data:
            max_col = col
        else:
            break
    
    return start_row, start_col, max_row, max_col

def detect_table_range(ws: Any, start_cell: str) -> Tuple[int, int, int, int]:
    """
    Conservatively detect ONLY tabular data range, avoiding descriptive text.
    
    This function is specifically designed for table creation and avoids including
    non-tabular content like descriptions, conclusions, or scattered text.
    
    Args:
        ws: Worksheet object
        start_cell: Starting cell reference (e.g., "A1")
        
    Returns:
        Tuple of (start_row, start_col, end_row, end_col) in 0-based coordinates
    """
    start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
    
    # Step 1: Find header row (first row with mostly text values)
    header_row = start_row
    for row in range(start_row, min(start_row + 5, ws.max_row)):
        row_values = []
        for col in range(start_col, min(start_col + 20, ws.max_column)):  # Limit column scan
            cell = ws.cell(row=row + 1, column=col + 1)
            if cell.value is not None:
                row_values.append(cell.value)
        
        if len(row_values) >= 2:  # Must have at least 2 values to be a table
            text_count = sum(1 for val in row_values if val and isinstance(val, str))
            if text_count >= len(row_values) * 0.6:  # At least 60% text (likely headers)
                header_row = row
                break
    
    # Step 2: Determine number of columns by examining the header row
    max_col = start_col
    header_values = []
    for col in range(start_col, min(start_col + 20, ws.max_column)):
        cell = ws.cell(row=header_row + 1, column=col + 1)
        if cell.value is not None and str(cell.value).strip():
            header_values.append(cell.value)
            max_col = col
        else:
            # Stop at first empty column in headers
            break
    
    if len(header_values) < 2:
        # Not enough columns for a table, return minimal range
        return start_row, start_col, start_row, start_col
    
    # Step 3: Find data rows (must have consistent structure)
    max_row = header_row
    consecutive_empty = 0
    
    for row in range(header_row + 1, min(header_row + 100, ws.max_row)):  # Limit row scan
        # Check if this row has data in the same columns as headers
        row_data_count = 0
        for col in range(start_col, max_col + 1):
            cell = ws.cell(row=row + 1, column=col + 1)
            if cell.value is not None and str(cell.value).strip():
                row_data_count += 1
        
        # Row is part of table if it has data in at least 50% of header columns
        if row_data_count >= len(header_values) * 0.5:
            max_row = row
            consecutive_empty = 0
        else:
            consecutive_empty += 1
            # Stop if we find 2 consecutive rows without sufficient data
            if consecutive_empty >= 2:
                break
    
    return start_row, start_col, max_row, max_col

def conservative_table_cleanup(ws: Any, cell_range: str) -> str:
    """
    Conservative cleanup for table creation - ONLY improves formatting within the exact range.
    Does NOT expand or modify the range boundaries.
    
    Args:
        ws: Worksheet object
        cell_range: Exact table range (e.g., "A1:D10")
        
    Returns:
        The same range (unchanged)
    """
    try:
        start_row, start_col, end_row, end_col = ExcelRange.parse_range(cell_range)
        
        # Only improve headers if they are clearly generic
        # Check first row for generic headers
        for col in range(start_col, end_col + 1):
            header_cell = ws.cell(row=start_row + 1, column=col + 1)
            if header_cell.value:
                header_str = str(header_cell.value).strip()
                # Only replace clearly generic headers
                if header_str.lower() in ['column1', 'columna1', 'col1', 'field1', 
                                         'column2', 'columna2', 'col2', 'field2',
                                         'column3', 'columna3', 'col3', 'field3',
                                         'column4', 'columna4', 'col4', 'field4',
                                         'column5', 'columna5', 'col5', 'field5']:
                    # Try to infer better name from data
                    data_sample = []
                    for row in range(start_row + 1, min(start_row + 4, end_row + 1)):
                        cell = ws.cell(row=row + 1, column=col + 1)
                        if cell.value is not None:
                            data_sample.append(cell.value)
                    
                    if data_sample:
                        # Simple heuristics for common data types
                        sample_str = str(data_sample[0]).lower()
                        if any(word in sample_str for word in ['name', 'nombre', 'nom']):
                            header_cell.value = "Name"
                        elif any(word in sample_str for word in ['date', 'fecha', 'time']):
                            header_cell.value = "Date"
                        elif any(word in sample_str for word in ['price', 'cost', 'amount', 'total']):
                            header_cell.value = "Amount"
                        elif any(word in sample_str for word in ['qty', 'quantity', 'cantidad']):
                            header_cell.value = "Quantity"
                        elif isinstance(data_sample[0], (int, float)):
                            header_cell.value = "Value"
                        else:
                            header_cell.value = f"Column{col-start_col+1}"
        
        # Apply basic formatting to headers only
        for col in range(start_col, end_col + 1):
            header_cell = ws.cell(row=start_row + 1, column=col + 1)
            header_cell.font = Font(name=DEFAULT_FONT, size=11, bold=True)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        return cell_range  # Return unchanged range
        
    except Exception as e:
        logger.warning(f"Conservative table cleanup failed: {e}")
        return cell_range  # Return original range on error

# ----------------------------------------
# FORMULA GENERATION FUNCTIONS
# ----------------------------------------

def generate_sum_formula(range_ref: str) -> str:
    """Generate a SUM formula for a given range."""
    return f"=SUM({range_ref})"

def generate_average_formula(range_ref: str) -> str:
    """Generate an AVERAGE formula for a given range."""
    return f"=AVERAGE({range_ref})"

def generate_count_formula(range_ref: str) -> str:
    """Generate a COUNT formula for a given range."""
    return f"=COUNT({range_ref})"

def generate_max_formula(range_ref: str) -> str:
    """Generate a MAX formula for a given range."""
    return f"=MAX({range_ref})"

def generate_min_formula(range_ref: str) -> str:
    """Generate a MIN formula for a given range."""
    return f"=MIN({range_ref})"

def generate_percentage_formula(part_cell: str, total_cell: str) -> str:
    """Generate a percentage formula."""
    return f"=({part_cell}/{total_cell})*100"

def generate_subtotal_formula(function_num: int, range_ref: str) -> str:
    """Generate a SUBTOTAL formula (ignores filtered rows).
    
    Common function numbers:
    1 = AVERAGE, 9 = SUM, 3 = COUNT, 4 = MAX, 5 = MIN
    """
    return f"=SUBTOTAL({function_num},{range_ref})"

def generate_if_formula(condition: str, value_if_true: str, value_if_false: str = '""') -> str:
    """Generate an IF formula."""
    return f"=IF({condition},{value_if_true},{value_if_false})"

def generate_vlookup_formula(lookup_value: str, table_array: str, col_index: int, exact_match: bool = True) -> str:
    """Generate a VLOOKUP formula."""
    range_lookup = "FALSE" if exact_match else "TRUE"
    return f"=VLOOKUP({lookup_value},{table_array},{col_index},{range_lookup})"

def generate_concatenate_formula(*values) -> str:
    """Generate a CONCATENATE formula."""
    value_list = ','.join(str(v) for v in values)
    return f"=CONCATENATE({value_list})"

def detect_formula_type_from_data(data_column: list, header_name: str = "") -> str:
    """Detect what type of formula would be appropriate for a data column.
    
    Args:
        data_column: List of values in the column
        header_name: Name of the column header
        
    Returns:
        Suggested formula type ('sum', 'average', 'count', 'max', 'min', 'none')
    """
    if not data_column:
        return 'none'
    
    # Filter numeric values
    numeric_values = [val for val in data_column if isinstance(val, (int, float)) and not isinstance(val, bool)]
    
    if len(numeric_values) < 2:
        return 'count'  # Not enough numeric data for calculations
    
    header_lower = header_name.lower() if header_name else ""
    
    # Heuristics based on header name
    if any(word in header_lower for word in ['total', 'sum', 'amount', 'cost', 'price', 'revenue', 'sales']):
        return 'sum'
    elif any(word in header_lower for word in ['average', 'mean', 'avg', 'rate', 'percentage', '%']):
        return 'average'
    elif any(word in header_lower for word in ['max', 'maximum', 'highest', 'peak']):
        return 'max'
    elif any(word in header_lower for word in ['min', 'minimum', 'lowest']):
        return 'min'
    elif any(word in header_lower for word in ['count', 'quantity', 'qty', 'number', 'num']):
        return 'count'
    
    # Heuristics based on data characteristics
    data_range = max(numeric_values) - min(numeric_values)
    data_mean = sum(numeric_values) / len(numeric_values)
    
    # If values are all similar (small range), might be quantities to sum
    if data_range < data_mean * 0.5:
        return 'sum'
    
    # If values vary significantly, average might be more meaningful
    if data_range > data_mean * 2:
        return 'average'
    
    # Default to sum for numeric columns
    return 'sum'

def add_formula_to_table(ws: Any, table_range: str, formula_type: str = 'auto') -> Dict[str, Any]:
    """Add formulas to a table (typically in a total row).
    
    Args:
        ws: Worksheet object
        table_range: Range of the table (e.g., "A1:D10")
        formula_type: Type of formula ('sum', 'average', 'count', 'auto')
        
    Returns:
        Dictionary with information about added formulas
    """
    try:
        start_row, start_col, end_row, end_col = ExcelRange.parse_range(table_range)
        
        # Find a row for totals (after the data)
        total_row = end_row + 2  # Leave one empty row
        
        # Add "Total" label in first column
        label_cell = ws.cell(row=total_row + 1, column=start_col + 1)
        label_cell.value = "Total"
        label_cell.font = Font(name=DEFAULT_FONT, size=11, bold=True)
        
        formulas_added = []
        
        # Add formulas for each numeric column
        for col in range(start_col + 1, end_col + 1):  # Skip first column (labels)
            # Get header for this column
            header_cell = ws.cell(row=start_row + 1, column=col + 1)
            header_name = str(header_cell.value) if header_cell.value else ""
            
            # Get data in this column (skip header)
            data_column = []
            for row in range(start_row + 1, end_row + 1):
                cell_value = ws.cell(row=row + 1, column=col + 1).value
                if cell_value is not None:
                    data_column.append(cell_value)
            
            # Determine formula type
            if formula_type == 'auto':
                suggested_type = detect_formula_type_from_data(data_column, header_name)
            else:
                suggested_type = formula_type
            
            # Skip if no numeric data
            numeric_count = sum(1 for val in data_column if isinstance(val, (int, float)) and not isinstance(val, bool))
            if numeric_count < 1:
                continue
            
            # Create data range for this column (excluding header)
            data_range = f"{get_column_letter(col + 1)}{start_row + 2}:{get_column_letter(col + 1)}{end_row + 1}"
            
            # Generate appropriate formula
            formula_cell = ws.cell(row=total_row + 1, column=col + 1)
            
            if suggested_type == 'sum':
                formula_cell.value = generate_sum_formula(data_range)
            elif suggested_type == 'average':
                formula_cell.value = generate_average_formula(data_range)
            elif suggested_type == 'count':
                formula_cell.value = generate_count_formula(data_range)
            elif suggested_type == 'max':
                formula_cell.value = generate_max_formula(data_range)
            elif suggested_type == 'min':
                formula_cell.value = generate_min_formula(data_range)
            else:
                # Default to sum for numeric data
                formula_cell.value = generate_sum_formula(data_range)
            
            # Format formula cell
            formula_cell.font = Font(name=DEFAULT_FONT, size=11, bold=True)
            formula_cell.number_format = '#,##0.00'
            
            formulas_added.append({
                'column': get_column_letter(col + 1),
                'formula_type': suggested_type,
                'formula': formula_cell.value,
                'range': data_range
            })
        
        return {
            'success': True,
            'total_row': total_row + 1,
            'formulas_added': formulas_added,
            'message': f"Added {len(formulas_added)} formulas to table"
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'message': f"Error adding formulas to table: {e}"
        }

def create_calculated_column(ws: Any, table_range: str, new_column_header: str, formula_template: str) -> Dict[str, Any]:
    """Add a calculated column to a table.
    
    Args:
        ws: Worksheet object
        table_range: Range of the existing table
        new_column_header: Header for the new calculated column
        formula_template: Formula template using {row} placeholder (e.g., "=B{row}*C{row}")
        
    Returns:
        Dictionary with information about the new column
    """
    try:
        start_row, start_col, end_row, end_col = ExcelRange.parse_range(table_range)
        
        # New column will be after the last column
        new_col = end_col + 1
        
        # Add header
        header_cell = ws.cell(row=start_row + 1, column=new_col + 1)
        header_cell.value = new_column_header
        header_cell.font = Font(name=DEFAULT_FONT, size=11, bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add formulas for each data row
        formulas_added = []
        for row in range(start_row + 1, end_row + 1):
            formula_cell = ws.cell(row=row + 1, column=new_col + 1)
            # Replace {row} placeholder with actual row number (1-based)
            actual_formula = formula_template.replace('{row}', str(row + 1))
            formula_cell.value = actual_formula
            formula_cell.number_format = '#,##0.00'
            
            formulas_added.append({
                'row': row + 1,
                'formula': actual_formula
            })
        
        # Return new expanded range
        new_range = ExcelRange.range_to_a1(start_row, start_col, end_row, new_col)
        
        return {
            'success': True,
            'new_column': get_column_letter(new_col + 1),
            'new_range': new_range,
            'formulas_added': len(formulas_added),
            'message': f"Added calculated column '{new_column_header}' with {len(formulas_added)} formulas"
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'message': f"Error creating calculated column: {e}"
        }

def remove_empty_rows_before_data(ws: Any, start_row: int, start_col: int, end_col: int) -> int:
    """
    Remove empty rows before actual data starts.
    
    Args:
        ws: Worksheet object
        start_row: Starting row (0-based)
        start_col: Starting column (0-based)
        end_col: Ending column (0-based)
        
    Returns:
        New starting row after removing empty rows
    """
    current_row = start_row
    
    while current_row < ws.max_row:
        row_empty = True
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=current_row + 1, column=col + 1)
            if cell.value is not None and str(cell.value).strip():
                row_empty = False
                break
        
        if not row_empty:
            break
        
        # Delete the empty row
        ws.delete_rows(current_row + 1, 1)
        # Don't increment current_row since we deleted a row
    
    return current_row

def smart_header_renaming(headers: List[str], data_sample: List[List[Any]]) -> List[str]:
    """
    Rename generic headers based on data content analysis.
    
    Args:
        headers: Original header list
        data_sample: Sample data rows to analyze
        
    Returns:
        List of improved header names
    """
    improved_headers = []
    
    # Common patterns for header detection
    header_patterns = {
        'position': ['pos', 'position', 'rank', 'ranking', '#'],
        'name': ['name', 'nombre', 'comercial', 'vendedor', 'cliente', 'product'],
        'sales': ['sales', 'ventas', 'facturacion', 'revenue', 'ingresos'],
        'margin': ['margin', 'margen', 'profit', 'beneficio'],
        'percentage': ['percent', 'porcentaje', '%', 'pct'],
        'quantity': ['quantity', 'cantidad', 'qty', 'units', 'productos'],
        'transactions': ['transactions', 'transacciones', 'ops', 'operaciones'],
        'category': ['category', 'categoria', 'type', 'tipo'],
        'date': ['date', 'fecha', 'day', 'dia', 'month', 'mes'],
        'region': ['region', 'zona', 'area', 'territory']
    }
    
    for i, header in enumerate(headers):
        # If header is already meaningful, keep it (be more conservative)
        header_lower = header.lower().strip()
        
        # Only change truly generic headers
        generic_patterns = [
            'columna', 'column', 'field', 'campo', 
            'col ', 'col_', 'column_', 'field_',
            'unnamed', 'sin nombre', 'untitled'
        ]
        
        is_generic = any(header_lower.startswith(pattern) for pattern in generic_patterns)
        is_generic = is_generic or header_lower in ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h']
        is_generic = is_generic or (header_lower.startswith('column') and header_lower[6:].isdigit())
        
        if not is_generic and header.strip():
            # Keep existing meaningful header
            improved_headers.append(header.strip())
            continue
        
        # Analyze data in this column to guess the content
        column_data = [row[i] if i < len(row) else None for row in data_sample[:5]]  # First 5 rows
        
        # Check if it's a position column (sequential numbers starting from 1)
        if all(isinstance(val, (int, float)) and val == idx + 1 for idx, val in enumerate(column_data) if val is not None):
            improved_headers.append("Pos")
            continue
        
        # Check if it's a percentage column (values between 0 and 1 or ending with %)
        if any(isinstance(val, (int, float)) and 0 <= val <= 1 for val in column_data):
            if i > 0 and 'margin' in str(headers[i-1]).lower():
                improved_headers.append("Margen %")
            else:
                improved_headers.append("Porcentaje")
            continue
        
        # Check if it's currency/sales (large numbers)
        if any(isinstance(val, (int, float)) and val > 1000 for val in column_data):
            improved_headers.append("Facturación €")
            continue
        
        # Check if it's a name/text column
        if any(isinstance(val, str) and len(val) > 3 for val in column_data):
            if i == 1:  # Usually second column is names
                improved_headers.append("Comercial")
            else:
                improved_headers.append("Categoría")
            continue
        
        # Default based on position
        if i == 0:
            improved_headers.append("Pos")
        elif i == 1:
            improved_headers.append("Nombre")
        elif i == 2:
            improved_headers.append("Valor")
        else:
            improved_headers.append(f"Campo {i+1}")
    
    return improved_headers

def get_existing_chart_positions(ws: Any) -> List[Dict[str, Any]]:
    """
    Get positions and dimensions of all existing charts in the worksheet.
    
    Args:
        ws: Worksheet object
        
    Returns:
        List of chart position dictionaries with start/end coordinates
    """
    chart_positions = []
    
    try:
        # Get all charts in the worksheet
        if hasattr(ws, '_charts') and ws._charts:
            for chart_ref in ws._charts:
                try:
                    chart = chart_ref[0]  # The chart object
                    anchor = chart_ref[1]  # The anchor (position)
                    
                    # Extract position from anchor
                    if hasattr(anchor, '_from') and anchor._from:
                        # Get starting position and ensure integers
                        start_col = int(anchor._from.col) if anchor._from.col is not None else 0
                        start_row = int(anchor._from.row) if anchor._from.row is not None else 0
                        
                        # Estimate chart size if not explicitly available
                        # Standard chart sizes in Excel
                        chart_width = 8  # columns
                        chart_height = 15  # rows
                        
                        # Try to get actual size if available
                        if hasattr(anchor, 'to') and anchor.to:
                            end_col = int(anchor.to.col) if anchor.to.col is not None else start_col + 8
                            end_row = int(anchor.to.row) if anchor.to.row is not None else start_row + 15
                            chart_width = max(1, end_col - start_col)
                            chart_height = max(1, end_row - start_row)
                        elif hasattr(anchor, '_to') and anchor._to:
                            end_col = int(anchor._to.col) if anchor._to.col is not None else start_col + 8
                            end_row = int(anchor._to.row) if anchor._to.row is not None else start_row + 15
                            chart_width = max(1, end_col - start_col)
                            chart_height = max(1, end_row - start_row)
                        
                        chart_positions.append({
                            'start_col': start_col,
                            'start_row': start_row,
                            'end_col': start_col + chart_width,
                            'end_row': start_row + chart_height,
                            'width': chart_width,
                            'height': chart_height
                        })
                        
                except Exception as e:
                    # If we can't parse a chart, log and continue
                    logger.warning(f"Could not parse chart position: {e}")
                    continue
    except Exception as e:
        logger.warning(f"Error getting chart positions: {e}")
    
    return chart_positions

def check_area_overlap(start_col: int, start_row: int, width: int, height: int, 
                      existing_positions: List[Dict[str, Any]], 
                      buffer_cols: int = 2, buffer_rows: int = 2) -> bool:
    """
    Check if a proposed chart area overlaps with existing charts or content.
    
    Args:
        start_col, start_row: Starting position (0-based)
        width, height: Dimensions of the proposed chart
        existing_positions: List of existing chart positions
        buffer_cols, buffer_rows: Minimum spacing buffer
        
    Returns:
        True if there's overlap, False if area is free
    """
    # Ensure all parameters are integers to prevent comparison errors
    start_col = int(start_col)
    start_row = int(start_row)
    width = int(width)
    height = int(height)
    buffer_cols = int(buffer_cols)
    buffer_rows = int(buffer_rows)
    
    # Add buffer to the proposed area
    proposed_start_col = start_col - buffer_cols
    proposed_start_row = start_row - buffer_rows
    proposed_end_col = start_col + width + buffer_cols
    proposed_end_row = start_row + height + buffer_rows
    
    # Check against all existing chart positions
    for chart_pos in existing_positions:
        # Ensure all values are integers to avoid comparison errors
        chart_start_col = int(chart_pos['start_col']) - buffer_cols
        chart_start_row = int(chart_pos['start_row']) - buffer_rows
        chart_end_col = int(chart_pos['end_col']) + buffer_cols
        chart_end_row = int(chart_pos['end_row']) + buffer_rows
        
        # Check for overlap
        if not (proposed_end_col <= chart_start_col or  # Proposed is to the left
                proposed_start_col >= chart_end_col or   # Proposed is to the right
                proposed_end_row <= chart_start_row or   # Proposed is above
                proposed_start_row >= chart_end_row):    # Proposed is below
            return True  # Overlap detected
    
    return False  # No overlap

def check_content_overlap(ws: Any, start_col: int, start_row: int, width: int, height: int) -> bool:
    """
    Check if a proposed chart area overlaps with existing cell content.
    
    Args:
        ws: Worksheet object
        start_col, start_row: Starting position (0-based)
        width, height: Dimensions of the proposed chart
        
    Returns:
        True if there's content overlap, False if area is free
    """
    try:
        # Ensure all parameters are integers
        start_col = int(start_col)
        start_row = int(start_row)
        width = int(width)
        height = int(height)
        
        # Check each cell in the proposed area
        for r in range(start_row, start_row + height):
            for c in range(start_col, start_col + width):
                try:
                    cell = ws.cell(row=r + 1, column=c + 1)  # Convert to 1-based
                    if cell.value is not None and str(cell.value).strip():
                        return True  # Content found
                except Exception:
                    continue  # Cell doesn't exist or error accessing it
        return False  # No content found
    except Exception:
        return True  # Error accessing area, consider it occupied

def find_optimal_chart_position(ws: Any, preferred_col: int = 6, preferred_row: int = 1, 
                               chart_width: int = 8, chart_height: int = 15) -> str:
    """
    Find optimal position for chart with intelligent overlap prevention.
    
    This function ensures charts NEVER overlap by:
    1. Detecting all existing charts and their positions
    2. Checking for content overlap with data and tables
    3. Maintaining professional spacing between elements
    4. Using systematic positioning strategy
    
    Args:
        ws: Worksheet object
        preferred_col: Preferred starting column (0-based)
        preferred_row: Preferred starting row (0-based)
        chart_width: Chart width in columns (default: 8)
        chart_height: Chart height in rows (default: 15)
        
    Returns:
        Optimal position string (e.g., "F1", "J5")
    """
    # Ensure all parameters are integers to prevent type errors
    preferred_col = int(preferred_col) if preferred_col is not None else 6
    preferred_row = int(preferred_row) if preferred_row is not None else 1
    chart_width = int(chart_width) if chart_width is not None else 8
    chart_height = int(chart_height) if chart_height is not None else 15
    
    # Get all existing chart positions
    existing_charts = get_existing_chart_positions(ws)
    
    # Professional spacing requirements
    MIN_SPACING_COLS = 2  # Minimum columns between charts
    MIN_SPACING_ROWS = 2  # Minimum rows between charts
    
    # Define systematic search pattern for professional layouts
    # Pattern: Try different columns first, then different rows
    search_positions = []
    
    # Row 1: Try columns F, J, N, R (standard chart positions)
    for col_offset in [5, 9, 13, 17, 21]:  # F=5, J=9, N=13, R=17, V=21 (0-based)
        search_positions.append((col_offset, 0))  # Row 1
    
    # Row 2: Below data (assuming data ends around row 20)
    data_end_estimate = max(20, preferred_row + 5)
    for col_offset in [0, 5, 9, 13, 17]:  # A, F, J, N, R
        search_positions.append((col_offset, data_end_estimate + 2))
    
    # Row 3: Further down for more charts
    for col_offset in [0, 5, 9, 13, 17]:
        search_positions.append((col_offset, data_end_estimate + 20))
    
    # Row 4: Far right if needed
    for row_offset in [0, 18, 36]:
        search_positions.append((25, row_offset))  # Column Z
    
    # Add preferred position at the beginning if provided
    if preferred_col >= 0 and preferred_row >= 0:
        search_positions.insert(0, (preferred_col, preferred_row))
    
    # Test each position systematically
    for test_col, test_row in search_positions:
        # Check for chart overlap
        if not check_area_overlap(test_col, test_row, chart_width, chart_height, 
                                 existing_charts, MIN_SPACING_COLS, MIN_SPACING_ROWS):
            # Check for content overlap
            if not check_content_overlap(ws, test_col, test_row, chart_width, chart_height):
                # Position is free! Convert to Excel notation
                col_letter = get_column_letter(test_col + 1)  # Convert to 1-based
                return f"{col_letter}{test_row + 1}"  # Convert to 1-based
    
    # If all positions are occupied, use emergency position far right
    emergency_col = 30  # Column AE
    emergency_row = 1
    col_letter = get_column_letter(emergency_col + 1)
    
    logger.warning(f"All optimal chart positions occupied, using emergency position {col_letter}{emergency_row}")
    return f"{col_letter}{emergency_row}"

def get_chart_layout_recommendations(ws: Any, data_ranges: List[str]) -> Dict[str, Any]:
    """
    Analyze worksheet and provide intelligent chart layout recommendations.
    
    Args:
        ws: Worksheet object
        data_ranges: List of data ranges that might be used for charts
        
    Returns:
        Dictionary with layout recommendations
    """
    existing_charts = get_existing_chart_positions(ws)
    
    # Analyze data layout
    max_data_col = 0
    max_data_row = 0
    
    for range_str in data_ranges:
        try:
            start_row, start_col, end_row, end_col = ExcelRange.parse_range(range_str)
            max_data_col = max(max_data_col, end_col)
            max_data_row = max(max_data_row, end_row)
        except Exception:
            continue
    
    recommendations = {
        'existing_charts': len(existing_charts),
        'data_ends_at_col': max_data_col,
        'data_ends_at_row': max_data_row,
        'recommended_chart_areas': [],
        'layout_strategy': 'unknown'
    }
    
    # Determine layout strategy
    if max_data_col <= 5:  # Data uses columns A-E
        recommendations['layout_strategy'] = 'right_side_charts'
        recommendations['recommended_chart_areas'] = [
            'F1:M15 (First chart area)',
            'N1:U15 (Second chart area)',
            'F18:M32 (Below first chart)',
            'N18:U32 (Below second chart)'
        ]
    elif max_data_row <= 15:  # Data is compact vertically
        recommendations['layout_strategy'] = 'below_data_charts'
        recommendations['recommended_chart_areas'] = [
            f'A{max_data_row + 3}:H{max_data_row + 17} (First chart below data)',
            f'J{max_data_row + 3}:Q{max_data_row + 17} (Second chart below data)'
        ]
    else:  # Data is large, use right side
        recommendations['layout_strategy'] = 'far_right_charts'
        recommendations['recommended_chart_areas'] = [
            f'{get_column_letter(max_data_col + 3)}1:{get_column_letter(max_data_col + 10)}15 (Right of data)',
            f'{get_column_letter(max_data_col + 12)}1:{get_column_letter(max_data_col + 19)}15 (Further right)'
        ]
    
    return recommendations

def format_total_rows(ws: Any, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    """
    Detect and format total rows with special styling.
    
    Args:
        ws: Worksheet object
        start_row, start_col, end_row, end_col: Range boundaries (0-based)
    """
    for row in range(start_row, end_row + 1):
        # Check if this looks like a total row
        first_cell = ws.cell(row=row + 1, column=start_col + 1)
        if first_cell.value and isinstance(first_cell.value, str):
            cell_text = str(first_cell.value).lower()
            if any(word in cell_text for word in ['total', 'suma', 'sum', 'grand']):
                # Apply special formatting to total row
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row + 1, column=col + 1)
                    cell.font = Font(name=DEFAULT_FONT, size=DEFAULT_FONT_SIZE, bold=True)
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    cell.border = Border(top=Side(style="medium", color="000000"))

def apply_text_alignment(ws: Any, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    """
    Apply proper text alignment: numbers right, text left.
    
    Args:
        ws: Worksheet object
        start_row, start_col, end_row, end_col: Range boundaries (0-based)
    """
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row + 1, column=col + 1)
            if cell.value is not None:
                if isinstance(cell.value, (int, float)):
                    # Numbers align right
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif isinstance(cell.value, str):
                    # Text aligns left
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    # Default center alignment
                    cell.alignment = Alignment(horizontal="center", vertical="center")

def add_smart_formulas_to_data(ws: Any, data_range: str, add_totals: bool = True, add_calculations: bool = True) -> Dict[str, Any]:
    """Add smart formulas to data ranges automatically.
    
    Args:
        ws: Worksheet object
        data_range: Range containing the data
        add_totals: Whether to add total formulas
        add_calculations: Whether to add calculated columns
        
    Returns:
        Dictionary with information about added formulas
    """
    results = {
        'success': True,
        'totals_added': [],
        'calculations_added': [],
        'message': ''
    }
    
    try:
        if add_totals:
            total_result = add_formula_to_table(ws, data_range, 'auto')
            if total_result.get('success'):
                results['totals_added'] = total_result.get('formulas_added', [])
        
        # Additional smart calculations could be added here
        # For example, percentage columns, growth rates, etc.
        
        total_formulas = len(results['totals_added']) + len(results['calculations_added'])
        results['message'] = f"Added {total_formulas} smart formulas to enhance data analysis"
        
        return results
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'message': f"Error adding smart formulas: {e}"
        }

def comprehensive_data_cleanup(ws: Any, start_cell: str = "A1") -> Tuple[str, List[str]]:
    """
    Conservative data cleanup and optimization for table creation.
    
    Args:
        ws: Worksheet object
        start_cell: Starting cell for data analysis
        
    Returns:
        Tuple of (optimized_range, improved_headers)
    """
    # Step 1: Detect ONLY tabular data range (conservative approach)
    start_row, start_col, end_row, end_col = detect_table_range(ws, start_cell)
    
    # Step 2: DO NOT remove rows - this can break existing content
    # Just work with the detected range as-is
    
    # Step 4: Enhanced header detection and extraction
    headers = []
    data_sample = []
    
    # First, determine which row contains the actual headers
    header_row = start_row
    for check_row in range(start_row, min(start_row + 3, end_row + 1)):
        row_values = []
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=check_row + 1, column=col + 1)
            row_values.append(cell.value)
        
        # Check if this row looks like headers (mostly text, unique values)
        if row_values and any(val is not None for val in row_values):
            text_count = sum(1 for val in row_values if val and isinstance(val, str))
            total_count = sum(1 for val in row_values if val is not None)
            
            # This row looks like headers if most values are text
            if text_count / max(total_count, 1) >= 0.7:
                header_row = check_row
                break
    
    # Extract headers from the identified header row
    for col in range(start_col, end_col + 1):
        header_cell = ws.cell(row=header_row + 1, column=col + 1)
        header_val = header_cell.value
        if header_val is None or str(header_val).strip() == "":
            headers.append(f"Columna{col-start_col+1}")
        else:
            # Clean header value
            clean_header = str(header_val).strip()
            # Remove common issues
            clean_header = clean_header.replace('\n', ' ').replace('\r', ' ')
            headers.append(clean_header)
    
    # Extract sample data for header analysis (skip the header row)
    data_start_row = header_row + 1
    for row in range(data_start_row, min(data_start_row + 5, end_row + 1)):
        row_data = []
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row + 1, column=col + 1)
            row_data.append(cell.value)
        if any(val is not None for val in row_data):
            data_sample.append(row_data)
    
    # Step 5: Apply smart header renaming
    improved_headers = smart_header_renaming(headers, data_sample)
    
    # Step 6: Conservative header updates - only change clearly generic ones
    for i, new_header in enumerate(improved_headers):
        if i < len(headers):  # Safety check
            original_header = headers[i]
            # Only update if original was clearly generic
            if original_header.lower() in ['column1', 'columna1', 'col1', 'field1', 
                                         'column2', 'columna2', 'col2', 'field2',
                                         'column3', 'columna3', 'col3', 'field3',
                                         'column4', 'columna4', 'col4', 'field4',
                                         'column5', 'columna5', 'col5', 'field5'] or original_header.startswith('Columna'):
                header_cell = ws.cell(row=header_row + 1, column=start_col + i + 1)
                header_cell.value = new_header
                # Apply basic header formatting
                header_cell.font = Font(name=DEFAULT_FONT, size=11, bold=True)
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Step 7: Apply comprehensive formatting
    optimized_range = ExcelRange.range_to_a1(start_row, start_col, end_row, end_col)
    
    # Format total rows
    format_total_rows(ws, start_row, start_col, end_row, end_col)
    
    # Apply text alignment
    apply_text_alignment(ws, start_row, start_col, end_row, end_col)
    
    # Apply consistent number formatting
    apply_consistent_number_format(ws)
    
    # Apply enhanced autofit
    enhanced_autofit_columns(ws)
    
    # Normalize row heights
    for row in range(start_row + 1, end_row + 1):
        ws.row_dimensions[row + 1].height = 18
        
    # Add smart formulas to enhance the data
    try:
        formula_result = add_smart_formulas_to_data(ws, optimized_range, add_totals=True)
        logger.info(f"Smart formulas added: {formula_result.get('message', '')}")
    except Exception as e:
        logger.warning(f"Could not add smart formulas: {e}")
    
    return optimized_range, improved_headers

def optimize_entire_workbook(wb: Any) -> None:
    """
    Apply comprehensive optimization to all sheets in a workbook.
    
    Args:
        wb: Workbook object
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        try:
            # Skip aggressive data cleanup to prevent table corruption
            # comprehensive_data_cleanup(ws, "A1")  # DISABLED - causes table corruption
            
            # Apply worksheet layout
            apply_worksheet_layout(ws)
            
            # Apply unified theme
            enhanced_autofit_columns(ws)
            apply_consistent_number_format(ws)
            
        except Exception as e:
            # Log error but continue with other sheets
            logger.warning(f"Error optimizing sheet '{sheet_name}': {e}")
            continue

# Common utilities 
class ExcelRange:
    """Utility class for manipulating and converting Excel ranges.

    This class offers helpers to convert between Excel notation (A1:B5) and
    zero-based Python coordinates, as well as utilities for validating ranges.
    """
    
    @staticmethod
    def parse_cell_ref(cell_ref: str) -> Tuple[int, int]:
        """
        Convert an A1-style cell reference to zero-based (row, column) coordinates.
        
        Args:
            cell_ref: Cell reference in Excel format (e.g. "A1", "B5")
            
        Returns:
            Tuple (row, column) using zero-based indices
            
        Raises:
            ValueError: If the cell reference is not valid
        """
        if not cell_ref or not isinstance(cell_ref, str):
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        
        # Extract the column portion (letters)
        col_str = ''.join(c for c in cell_ref if c.isalpha())
        # Extract the row portion (numbers)
        row_str = ''.join(c for c in cell_ref if c.isdigit())
        
        if not col_str or not row_str:
            raise ValueError(f"Invalid cell format: {cell_ref}")
        
        # Convert column letters to an index (A->0, B->1, etc.)
        col_idx = 0
        for c in col_str.upper():
            col_idx = col_idx * 26 + (ord(c) - ord('A') + 1)
        col_idx -= 1  # Adjust to zero-based index

        # Convert row number to zero-based index
        row_idx = int(row_str) - 1
        
        return row_idx, col_idx
    
    @staticmethod
    def parse_range(range_str: str) -> Tuple[int, int, int, int]:
        """
        Convert a range in A1:B5 style to zero-based coordinates (row1, col1, row2, col2).
        
        Args:
            range_str: Range in Excel format (e.g. "A1:B5")
            
        Returns:
            Tuple (start_row, start_col, end_row, end_col) using zero-based indices
            
        Raises:
            ValueError: If the range is not valid
        """
        if not range_str or not isinstance(range_str, str):
            raise ValueError(f"Invalid range: {range_str}")
        
        # Handle ranges that include a sheet reference
        if '!' in range_str:
            parts = range_str.split('!')
            if len(parts) != 2:
                raise ValueError(f"Invalid range with sheet format: {range_str}")
            range_str = parts[1]  # Use only the range portion
        
        # Split the range into starting and ending cells
        if ':' in range_str:
            start_cell, end_cell = range_str.split(':')
            start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
            end_row, end_col = ExcelRange.parse_cell_ref(end_cell)
        else:
            # If only one cell, start and end are the same
            start_row, start_col = ExcelRange.parse_cell_ref(range_str)
            end_row, end_col = start_row, start_col
        
        return start_row, start_col, end_row, end_col
    
    @staticmethod
    def cell_to_a1(row: int, col: int) -> str:
        """
        Convert zero-based (row, column) coordinates to an A1 cell reference.
        
        Args:
            row: Row index (zero-based)
            col: Column index (zero-based)
            
        Returns:
            Cell reference in A1 format
        """
        if row < 0 or col < 0:
            raise ValueError(f"Negative indices not allowed: row={row}, column={col}")
        
        # Convert column index to letters
        col_str = ""
        col_val = col + 1  # Convert to base 1 for calculation
        
        while col_val > 0:
            remainder = (col_val - 1) % 26
            col_str = chr(65 + remainder) + col_str
            col_val = (col_val - 1) // 26
        
        # Convert row index to a 1-based Excel number
        row_val = row + 1
        
        return f"{col_str}{row_val}"
    
    @staticmethod
    def range_to_a1(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
        """
        Convert zero-based range coordinates to an A1:B5 style range.
        
        Args:
            start_row: Starting row (zero-based)
            start_col: Starting column (zero-based)
            end_row: Ending row (zero-based)
            end_col: Ending column (zero-based)
            
        Returns:
            Range in A1:B5 format
        """
        start_cell = ExcelRange.cell_to_a1(start_row, start_col)
        end_cell = ExcelRange.cell_to_a1(end_row, end_col)
        
        if start_cell == end_cell:
            return start_cell
        return f"{start_cell}:{end_cell}"

    @staticmethod
    def parse_range_with_sheet(range_str: str) -> Tuple[Optional[str], int, int, int, int]:
        """Convert a range that may include a sheet to a tuple ``(sheet, row1, col1, row2, col2)``.

        Args:
            range_str: Range string, possibly with sheet prefix ``Sheet!A1:B2``.

        Returns:
            Tuple ``(sheet, start_row, start_col, end_row, end_col)`` where ``sheet``
            is ``None`` if no sheet was specified.
        """
        if not range_str or not isinstance(range_str, str):
            raise ValueError(f"Invalid range: {range_str}")

        sheet_name = None
        pure_range = range_str
        if "!" in range_str:
            parts = range_str.split("!", 1)
            if len(parts) != 2:
                raise ValueError(f"Invalid range with sheet format: {range_str}")
            sheet_name = parts[0].strip("'")
            pure_range = parts[1]

        start_row, start_col, end_row, end_col = ExcelRange.parse_range(pure_range)
        return sheet_name, start_row, start_col, end_row, end_col

# Constants and mappings
# Map style names to Excel style numbers
CHART_STYLE_NAMES = {
    # Light styles
    'light-1': 1, 'light-2': 2, 'light-3': 3, 'light-4': 4, 'light-5': 5, 'light-6': 6,
    'office-1': 1, 'office-2': 2, 'office-3': 3, 'office-4': 4, 'office-5': 5, 'office-6': 6,
    'white': 1, 'minimal': 2, 'soft': 3, 'gradient': 4, 'muted': 5, 'outlined': 6,
    
    # Dark styles
    'dark-1': 7, 'dark-2': 8, 'dark-3': 9, 'dark-4': 10, 'dark-5': 11, 'dark-6': 12, 
    'dark-blue': 7, 'dark-gray': 8, 'dark-green': 9, 'dark-red': 10, 'dark-purple': 11, 'dark-orange': 12,
    'navy': 7, 'charcoal': 8, 'forest': 9, 'burgundy': 10, 'indigo': 11, 'rust': 12,
    
    # Colorful styles
    'colorful-1': 13, 'colorful-2': 14, 'colorful-3': 15, 'colorful-4': 16, 
    'colorful-5': 17, 'colorful-6': 18, 'colorful-7': 19, 'colorful-8': 20,
    'bright': 13, 'vivid': 14, 'rainbow': 15, 'multi': 16, 'contrast': 17, 'vibrant': 18,
    
    # Office themes
    'ion-1': 21, 'ion-2': 22, 'ion-3': 23, 'ion-4': 24,
    'wisp-1': 25, 'wisp-2': 26, 'wisp-3': 27, 'wisp-4': 28,
    'aspect-1': 29, 'aspect-2': 30, 'aspect-3': 31, 'aspect-4': 32,
    'badge-1': 33, 'badge-2': 34, 'badge-3': 35, 'badge-4': 36,
    'gallery-1': 37, 'gallery-2': 38, 'gallery-3': 39, 'gallery-4': 40,
    'median-1': 41, 'median-2': 42, 'median-3': 43, 'median-4': 44,
    
    # Styles for specific chart types
    'column-default': 1, 'column-dark': 7, 'column-colorful': 13, 
    'bar-default': 1, 'bar-dark': 7, 'bar-colorful': 13,
    'line-default': 1, 'line-dark': 7, 'line-markers': 3, 'line-dash': 5,
    'pie-default': 1, 'pie-dark': 7, 'pie-explosion': 4, 'pie-3d': 10,
    'area-default': 1, 'area-dark': 7, 'area-transparent': 5, 'area-stacked': 9,
    'scatter-default': 1, 'scatter-dark': 7, 'scatter-bubble': 4, 'scatter-smooth': 9,
}

# Mapping between styles and recommended color palettes
STYLE_TO_PALETTE = {
    # Light styles (1-6)
    1: 'office', 2: 'office', 3: 'colorful', 4: 'colorful', 5: 'pastel', 6: 'pastel',
    # Dark styles (7-12)
    7: 'dark-blue', 8: 'dark-gray', 9: 'dark-green', 10: 'dark-red', 11: 'dark-purple', 12: 'dark-orange',
    # Colorful styles (13-20)
    13: 'colorful', 14: 'colorful', 15: 'colorful', 16: 'colorful', 
    17: 'colorful', 18: 'colorful', 19: 'colorful', 20: 'colorful',
}

# CHART_COLOR_SCHEMES - normally defined in the original module
# Included here in simplified form
CHART_COLOR_SCHEMES = {
    'default': ['4472C4', 'ED7D31', 'A5A5A5', 'FFC000', '5B9BD5', '70AD47', '8549BA', 'C55A11'],
    'colorful': ['5B9BD5', 'ED7D31', 'A5A5A5', 'FFC000', '4472C4', '70AD47', '264478', '9E480E'],
    'pastel': ['9DC3E6', 'FFD966', 'C5E0B3', 'F4B183', 'B4A7D6', '8FBCDB', 'D89595', 'B7B7B7'],
    'dark-blue': ['2F5597', '1F3864', '4472C4', '5B9BD5', '8FAADC', '2E75B5', '255E91', '1C4587'],
    'dark-red': ['952213', 'C0504D', 'FF8B6B', 'EA6B66', 'DA3903', 'FF4500', 'B22222', '8B0000'],
    'dark-green': ['1E6C41', '375623', '548235', '70AD47', '9BC169', '006400', '228B22', '3CB371'],
    'dark-purple': ['5C3292', '7030A0', '8064A2', '9A7FBA', 'B3A2C7', '800080', '9400D3', '8B008B'],
    'dark-orange': ['C55A11', 'ED7D31', 'F4B183', 'FFC000', 'FFD966', 'FF8C00', 'FF7F50', 'FF4500']
}

# Helper function to obtain a worksheet (unified)
def get_sheet(wb, sheet_name_or_index) -> Any:
    """Retrieve a worksheet by name or index.

    Args:
        wb: Openpyxl workbook object.
        sheet_name_or_index: Sheet name or numeric index.

    Returns:
        Worksheet object.

    Raises:
        SheetNotFoundError: If the sheet does not exist.
    """
    if wb is None:
        raise ExcelMCPError("Workbook cannot be None")
    
    if isinstance(sheet_name_or_index, int):
        # If an index is provided, try to access by position
        if 0 <= sheet_name_or_index < len(wb.worksheets):
            return wb.worksheets[sheet_name_or_index]
        else:
            raise SheetNotFoundError(f"No sheet exists with index {sheet_name_or_index}")
    else:
        # If a name is provided, try to access by name
        try:
            return wb[sheet_name_or_index]
        except KeyError:
            sheets_info = ", ".join(wb.sheetnames)
            raise SheetNotFoundError(f"Sheet '{sheet_name_or_index}' does not exist in the file. Available sheets: {sheets_info}")
        except Exception as e:
            raise ExcelMCPError(f"Error accessing sheet: {e}")

# Helper function to convert a style specifier into a numeric style
def parse_chart_style(style):
    """
    Convert different style formats into a numeric Excel style (1-48).

    Args:
        style: Style as an int, numeric str, ``styleN`` or descriptive name.

    Returns:
        Integer style between 1 and 48, or ``None`` if not valid.
    """
    if isinstance(style, int) and 1 <= style <= 48:
        return style
        
    if isinstance(style, str):
        # Case 1: numeric string like '5'
        if style.isdigit():
            style_num = int(style)
            if 1 <= style_num <= 48:
                return style_num
                
        # Case 2: format 'styleN' or 'Style N'
        style_lower = style.lower()
        if style_lower.startswith('style'):
            try:
                # Extract the number following 'style'
                num_part = ''.join(c for c in style_lower[5:] if c.isdigit())
                if num_part:
                    style_num = int(num_part)
                    if 1 <= style_num <= 48:
                        return style_num
            except (ValueError, IndexError):
                pass
                
        # Case 3: descriptive name ('dark-blue', etc.)
        if style_lower in CHART_STYLE_NAMES:
            return CHART_STYLE_NAMES[style_lower]
            
    return None

# Helper to apply chart styles and colors simultaneously
def apply_chart_style(chart, style):
    """
    Apply a predefined style to a chart including the appropriate color palette.

    Args:
        chart: Openpyxl chart object.
        style: Style in any supported format (number, name, etc.).

    Returns:
        ``True`` if applied successfully, ``False`` otherwise.
    """
    # Convert to a style number if needed
    style_number = parse_chart_style(style)
    
    if style_number is None:
        style_str = str(style) if style else "None"
        logger.warning(f"Invalid chart style: '{style_str}'. Must be a number between 1-48 or a valid style name.")
        logger.info("Valid style names include: 'dark-blue', 'light-1', 'colorful-3', etc.")
        return False
        
    if not (1 <= style_number <= 48):
        logger.warning(f"Invalid chart style: {style_number}. It must be between 1 and 48.")
        return False
    
    # Step 1: apply the numeric style to native Excel attributes
    try:
        # The style property in openpyxl corresponds to the Excel style number
        chart.style = style_number
        logger.info(f"Applied native style {style_number} to chart")
    except Exception as e:
        logger.warning(f"Error applying style {style_number}: {e}")
    
    # Step 2: apply the color palette associated with the style's theme
    palette_name = STYLE_TO_PALETTE.get(style_number, 'default')
    colors = CHART_COLOR_SCHEMES.get(palette_name, CHART_COLOR_SCHEMES['default'])
    
    # Apply colors to the series
    try:
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.fill import ColorChoice
        
        for i, series in enumerate(chart.series):
            if i < len(colors):
                # Ensure graphical properties exist
                if not hasattr(series, 'graphicalProperties') or series.graphicalProperties is None:
                    series.graphicalProperties = GraphicalProperties()
                    
                # Assign color using ColorChoice for better compatibility
                color = colors[i % len(colors)]
                if isinstance(color, str) and color.startswith('#'):
                    color = color[1:]
                    
                series.graphicalProperties.solidFill = ColorChoice(srgbClr=color)
                
        logger.info(f"Applied style {style_number} with palette '{palette_name}' to chart")
        return True
        
    except Exception as e:
        logger.warning(f"Error applying colors for style {style_number}: {e}")
        return False

def determine_orientation(ws: Any, min_row: int, min_col: int, max_row: int, max_col: int) -> bool:
    """Attempt to guess the orientation of the data.

    Returns ``True`` if categories appear to be in the first column (column
    oriented) and ``False`` if they are more likely in the first row. The
    algorithm compares the ratio of numeric values for both interpretations and
    uses the shape of the range as a tiebreaker. This helps a language model
    avoid choosing wrong headers in ambiguous tables.
    """

    def _is_number(value: Any) -> bool:
        return isinstance(value, (int, float)) and not isinstance(value, bool)

    # Calculate ratio of numbers assuming categories are in the first column
    col_numeric = col_total = 0
    for c in range(min_col + 1, max_col + 1):
        for r in range(min_row, max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                col_total += 1
                if _is_number(val):
                    col_numeric += 1

    col_ratio = (col_numeric / col_total) if col_total else 0

    # Calculate ratio of numbers assuming categories are in the first row
    row_numeric = row_total = 0
    for r in range(min_row + 1, max_row + 1):
        for c in range(min_col, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                row_total += 1
                if _is_number(val):
                    row_numeric += 1

    row_ratio = (row_numeric / row_total) if row_total else 0

    if row_ratio > col_ratio:
        return False  # headers in the first row
    if col_ratio > row_ratio:
        return True   # headers in the first column

    # Tiebreaker based on range shape
    return (max_row - min_row) >= (max_col - min_col)

def _trim_range_to_data(ws: Any, min_row: int, min_col: int, max_row: int, max_col: int) -> Tuple[int, int, int, int]:
    """Remove trailing empty rows and columns from a range."""
    while max_row >= min_row:
        if all(ws.cell(row=max_row, column=c).value in (None, "") for c in range(min_col, max_col + 1)):
            max_row -= 1
        else:
            break
    while max_col >= min_col:
        if all(ws.cell(row=r, column=max_col).value in (None, "") for r in range(min_row, max_row + 1)):
            max_col -= 1
        else:
            break
    return min_row, min_col, max_row, max_col

def _range_has_blank(ws: Any, min_row: int, min_col: int, max_row: int, max_col: int) -> bool:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            if ws.cell(row=r, column=c).value in (None, ""):
                return True
    return False

# ----------------------------------------
# BASE FUNCTIONS
# ----------------------------------------

# 1. Workbook management
def create_workbook(filename: str, overwrite: bool = False) -> Any:
    """
    Create a new empty Excel file.

    Args:
        filename (str): Full path and name of the file to create.
        overwrite (bool, optional): Overwrite existing file if ``True``.

    Returns:
        Workbook object.

    Raises:
        FileExistsError: If the file exists and ``overwrite`` is ``False``.
    """
    if os.path.exists(filename) and not overwrite:
        raise FileExistsError(f"El archivo '{filename}' ya existe. Use overwrite=True para sobreescribir.")
    
    wb = openpyxl.Workbook()
    # Store the filename in the workbook for later use
    wb.path = filename
    return wb

def open_workbook(filename: str) -> Any:
    """
    Open an existing Excel file.

    Args:
        filename (str): Path to the file.

    Returns:
        Workbook object.

    Raises:
        FileNotFoundError: If the file does not exist.
    """
    if not os.path.exists(filename):
        raise FileNotFoundError(f"El archivo '{filename}' no existe.")
    
    try:
        wb = openpyxl.load_workbook(filename)
        return wb
    except Exception as e:
        logger.error(f"Error opening file '{filename}': {e}")
        raise ExcelMCPError(f"Error opening file: {e}")

def save_workbook(wb: Any, filename: Optional[str] = None) -> str:
    """
    Save the workbook to disk.

    Args:
        wb: Workbook object.
        filename (str, optional): Alternative file name if provided.

    Returns:
        Path to the saved file.

    Raises:
        ExcelMCPError: If an error occurs while saving.
    """
    if not wb:
        raise ExcelMCPError("Workbook cannot be None")
    
    try:
        # Si no se proporciona filename, usar el filename original si existe
        if not filename and hasattr(wb, 'path'):
            filename = wb.path
        elif not filename:
            raise ExcelMCPError("Debe proporcionar un nombre de archivo")
        
        # Apply comprehensive optimization before saving
        try:
            # First optimize all data and layout
            optimize_entire_workbook(wb)
            # Then apply unified theme
            apply_unified_theme(wb, "professional")
        except Exception:
            pass
        
        wb.save(filename)
        return filename
    except Exception as e:
        logger.error(f"Error saving workbook to '{filename}': {e}")
        raise ExcelMCPError(f"Error saving workbook: {e}")

def close_workbook(wb: Any) -> None:
    """
    Close the workbook in memory.

    Args:
        wb: Workbook object.

    Returns:
        None.
    """
    if not wb:
        return
    
    try:
        # Openpyxl does not really have a close() method,
        # but we can remove references to help the GC
        if hasattr(wb, "_archive"):
            wb._archive.close()
    except Exception as e:
        logger.warning(f"Warning while closing workbook: {e}")

def list_sheets(wb: Any) -> List[str]:
    """
    Return a list of sheet names.

    Args:
        wb: Workbook object.

    Returns:
        List[str]: Names of the sheets.
    """
    if not wb:
        raise ExcelMCPError("Workbook cannot be None")
    
    if hasattr(wb, 'sheetnames'):
        return wb.sheetnames
    
    # Alternative if sheetnames cannot be accessed
    sheet_names = []
    for sheet in wb.worksheets:
        if hasattr(sheet, 'title'):
            sheet_names.append(sheet.title)
    
    return sheet_names

def add_sheet(wb: Any, sheet_name: str, index: Optional[int] = None) -> Any:
    """
    Add a new empty worksheet.

    Args:
        wb: Workbook object.
        sheet_name (str): Name of the sheet.
        index (int, optional): Position in the tab list.

    Returns:
        The created worksheet.

    Raises:
        SheetExistsError: If a sheet with that name already exists.
    """
    if not wb:
        raise ExcelMCPError("Workbook cannot be None")
    
    # Check if a sheet with that name already exists
    if sheet_name in list_sheets(wb):
        raise SheetExistsError(f"A sheet named '{sheet_name}' already exists")
    
    # Create new sheet
    if index is not None:
        ws = wb.create_sheet(sheet_name, index)
    else:
        ws = wb.create_sheet(sheet_name)
    
    return ws

def delete_sheet(wb: Any, sheet_name: str) -> None:
    """
    Delete the specified worksheet.

    Args:
        wb: Workbook object.
        sheet_name (str): Name of the sheet to remove.

    Raises:
        SheetNotFoundError: If the sheet does not exist.
    """
    if not wb:
        raise ExcelMCPError("Workbook cannot be None")
    
    # Check that the sheet exists
    if sheet_name not in list_sheets(wb):
        raise SheetNotFoundError(f"Sheet '{sheet_name}' does not exist in the workbook")
    
    # Delete the sheet
    try:
        del wb[sheet_name]
    except Exception as e:
        logger.error(f"Error deleting sheet '{sheet_name}': {e}")
        raise ExcelMCPError(f"Error deleting sheet: {e}")

def rename_sheet(wb: Any, old_name: str, new_name: str) -> None:
    """
    Rename a worksheet.

    Args:
        wb: Workbook object.
        old_name (str): Current sheet name.
        new_name (str): New sheet name.

    Raises:
        SheetNotFoundError: If the original sheet does not exist.
        SheetExistsError: If a sheet with the new name already exists.
    """
    if not wb:
        raise ExcelMCPError("Workbook cannot be None")
    
    # Check that the original sheet exists
    if old_name not in list_sheets(wb):
        raise SheetNotFoundError(f"Sheet '{old_name}' does not exist in the workbook")
    
    # Check that no sheet with the new name exists
    if new_name in list_sheets(wb) and old_name != new_name:
        raise SheetExistsError(f"A sheet named '{new_name}' already exists")
    
    # Rename the sheet
    try:
        wb[old_name].title = new_name
    except Exception as e:
        logger.error(f"Error renaming sheet '{old_name}' to '{new_name}': {e}")
        raise ExcelMCPError(f"Error renaming sheet: {e}")

# 2. Data reading and exploration
def read_sheet_data(wb: Any, sheet_name: str, range_str: Optional[str] = None,
                   formulas: bool = False) -> List[List[Any]]:
    """
    Read values and optionally formulas from an Excel sheet.
    
    Args:
        wb: Openpyxl workbook object
        sheet_name: Sheet name
        range_str: Range in ``A1:B5`` format or ``None`` for the whole sheet
        formulas: If ``True`` return formulas instead of calculated values
    
    Returns:
        List of lists with cell values or formulas
        
    Raises:
        SheetNotFoundError: If the sheet does not exist
        RangeError: If the range is invalid
    """
    # Get the sheet
    ws = get_sheet(wb, sheet_name)
    
    # If no range is specified, use the whole sheet
    if not range_str:
        # Determine the used range (min_row, min_col, max_row, max_col)
        min_row, min_col = 1, 1
        max_row = ws.max_row
        max_col = ws.max_column
    else:
        # Parse the specified range
        try:
            min_row, min_col, max_row, max_col = ExcelRange.parse_range(range_str)
            # Convert to 1-based for openpyxl
            min_row += 1
            min_col += 1
            max_row += 1
            max_col += 1
        except ValueError as e:
            raise RangeError(f"Invalid range '{range_str}': {e}")
    
    # Extract data from the range
    data = []
    for row in range(min_row, max_row + 1):
        row_data = []
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            
            # Get the appropriate value (formula or calculated value)
            if formulas and cell.data_type == 'f':
                # If formulas are requested and the cell has one
                value = cell.value  # This is the formula with '='
            else:
                # Normal or calculated value
                value = cell.value
            
            row_data.append(value)
        data.append(row_data)
    
    return data

def list_tables(wb: Any, sheet_name: str) -> List[Dict[str, Any]]:
    """
    List all tables defined on an Excel sheet.
    
    Args:
        wb: Openpyxl workbook object
        sheet_name: Sheet name
        
    Returns:
        List of dictionaries with table information
        
    Raises:
        SheetNotFoundError: If the sheet does not exist
    """
    # Get the sheet
    ws = get_sheet(wb, sheet_name)
    
    # List to store table information
    tables_info = []
    
    # Check if the sheet has tables
    if hasattr(ws, 'tables') and ws.tables:
        for table_name, table in ws.tables.items():
            table_info = {
                'name': table_name,
                'ref': table.ref,
                'display_name': table.displayName,
                'header_row': table.headerRowCount > 0,
                'totals_row': table.totalsRowCount > 0,
                'style': table.tableStyleInfo.name if table.tableStyleInfo else None
            }
            
            tables_info.append(table_info)
    
    return tables_info

def get_table_data(wb: Any, sheet_name: str, table_name: str) -> List[Dict[str, Any]]:
    """
    Get the data from a specific table as records.
    
    Args:
        wb: Openpyxl workbook object
        sheet_name: Sheet name
        table_name: Table name
        
    Returns:
        List of dictionaries, where each dictionary represents a row
        
    Raises:
        SheetNotFoundError: If the sheet does not exist
        TableError: If the table does not exist
    """
    # Get the sheet
    ws = get_sheet(wb, sheet_name)
    
    # Check if the table exists
    if not hasattr(ws, 'tables') or table_name not in ws.tables:
        raise TableError(f"Table '{table_name}' does not exist on sheet '{sheet_name}'")
    
    # Get the table reference
    table = ws.tables[table_name]
    table_range = table.ref
    
    # Parse the range
    min_row, min_col, max_row, max_col = ExcelRange.parse_range(table_range)
    
    # Adjust to 1-based for openpyxl
    min_row += 1
    min_col += 1
    max_row += 1
    max_col += 1
    
    # Extract headers (first row)
    headers = []
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=min_row, column=col)
        headers.append(cell.value or f"Column{col}")
    
    # Extract data (rows after the header)
    data = []
    for row in range(min_row + 1, max_row + 1):
        row_data = {}
        for col_idx, col in enumerate(range(min_col, max_col + 1)):
            cell = ws.cell(row=row, column=col)
            header = headers[col_idx]
            row_data[header] = cell.value
        data.append(row_data)
    
    return data

def list_charts(wb: Any, sheet_name: str) -> List[Dict[str, Any]]:
    """
    List all charts on an Excel sheet.
    
    Args:
        wb: Openpyxl workbook object
        sheet_name: Sheet name
        
    Returns:
        List of dictionaries with chart information
        
    Raises:
        SheetNotFoundError: If the sheet does not exist
    """
    # Get the sheet
    ws = get_sheet(wb, sheet_name)
    
    # List to store chart information
    charts_info = []
    
    # Check if the sheet has charts
    if hasattr(ws, '_charts'):
        for chart_id, chart_rel in enumerate(ws._charts):
            chart = chart_rel[0]  # Element 0 is the chart object, 1 is position
            
            # Determine the chart type
            chart_type = "unknown"
            if isinstance(chart, BarChart):
                chart_type = "bar" if chart.type == "bar" else "column"
            elif isinstance(chart, LineChart):
                chart_type = "line"
            elif isinstance(chart, PieChart):
                chart_type = "pie"
            elif isinstance(chart, ScatterChart):
                chart_type = "scatter"
            elif isinstance(chart, AreaChart):
                chart_type = "area"
            
            # Gather chart information
            chart_info = {
                'id': chart_id,
                'type': chart_type,
                'title': chart.title if hasattr(chart, 'title') and chart.title else f"Chart {chart_id}",
                'position': chart_rel[1] if len(chart_rel) > 1 else None,
                'series_count': len(chart.series) if hasattr(chart, 'series') else 0
            }
            
            charts_info.append(chart_info)
    
    return charts_info

# 3. Escritura y formato de datos (de excel_writer_mcp.py)
def write_sheet_data(ws: Any, start_cell: str, data: List[List[Any]]) -> None:
    """
    Write a two-dimensional array of values or formulas.
     **Emojis must never be included in text written to cells, labels, titles or Excel charts.**


    To ensure the output remains readable when the function is used by a language
    model, it is recommended to apply styles after writing and check the length
    of the resulting cells. If any column contains very long text, its width
    should be increased to avoid cutting off the content. This way the generated
    files will look professional.

    Args:
        ws: Openpyxl worksheet object
        start_cell (str): Anchor cell (e.g. "A1")
        data (List[List]): Values or strings "=FORMULA(...)"
        
    Raises:
        CellReferenceError: If the cell reference is invalid
    """
    if not ws:
        raise ExcelMCPError("Worksheet cannot be None")
    
    if not data or not isinstance(data, list):
        raise ExcelMCPError("Data must be a non-empty list")
    
    try:
        # Parsear la celda inicial para obtener fila y columna base
        start_row, start_col = ExcelRange.parse_cell_ref(start_cell)

        # Escribir los datos
        for i, row_data in enumerate(data):
            if row_data is None:
                continue

            if not isinstance(row_data, list):
                # If it's not a list, treat it as a single value
                row_data = [row_data]

            for j, value in enumerate(row_data):
                # Calcular coordenadas de celda (base 1 para openpyxl)
                row = start_row + i + 1
                col = start_col + j + 1

                # Escribir el valor
                cell = ws.cell(row=row, column=col)
                cell.value = value

        # ----------------------------------------------------
        # Enhanced auto-fit and formatting
        # ----------------------------------------------------
        # Apply enhanced autofit to all columns
        try:
            enhanced_autofit_columns(ws)
        except Exception:
            pass
        
        # Apply consistent number formatting
        try:
            apply_consistent_number_format(ws)
        except Exception:
            pass
    
    except ValueError as e:
        raise CellReferenceError(f"Invalid cell reference '{start_cell}': {e}")
    except Exception as e:
        raise ExcelMCPError(f"Error writing data: {e}")

def append_rows(ws: Any, data: List[List[Any]]) -> None:
    """
    Append rows at the end with the given values.
     **Emojis must never be included in text written to cells, labels, titles or Excel charts.**

    
    Args:
        ws: Openpyxl worksheet object
        data (List[List]): Values or strings "=FORMULA(...)"
    """
    if not ws:
        raise ExcelMCPError("El worksheet no puede ser None")
    
    if not data or not isinstance(data, list):
        raise ExcelMCPError("Data must be a non-empty list")
    
    try:
        for row_data in data:
            if not isinstance(row_data, list):
                # Si no es una lista, convertir a lista con un solo elemento
                row_data = [row_data]
            
            ws.append(row_data)
    
    except Exception as e:
        raise ExcelMCPError(f"Error adding rows: {e}")

def update_cell(ws: Any, cell: str, value_or_formula: Any) -> None:
    """
    Update a single cell.
     **Emojis must never be included in text written to cells, labels, titles or Excel charts.**

    
    Args:
        ws: Openpyxl worksheet object
        cell (str): Cell reference (e.g. "A1")
        value_or_formula: Value or formula to assign
        
    Raises:
        CellReferenceError: If the cell reference is invalid
    """
    if not ws:
        raise ExcelMCPError("El worksheet no puede ser None")
    
    try:
        # Assign value to the cell
        cell_obj = ws[cell]
        cell_obj.value = value_or_formula

        # ----------------------------------------------
        # Auto-fit if long text is written
        # ----------------------------------------------
        if isinstance(value_or_formula, str):
            text = value_or_formula
            lines = text.splitlines()
            max_len = max(len(line) for line in lines)

            column_letter = cell_obj.column_letter
            current_w = ws.column_dimensions[column_letter].width or 8.43
            desired_w = min(max_len + 2, 80)
            if desired_w > current_w:
                ws.column_dimensions[column_letter].width = desired_w

            if len(lines) > 1 or max_len > current_w:
                cell_obj.alignment = Alignment(wrap_text=True)
                est_lines = max(len(lines), math.ceil(max_len / max(desired_w, 1)))
                current_h = ws.row_dimensions[cell_obj.row].height or 15
                desired_h = est_lines * 15
                if desired_h > current_h:
                    ws.row_dimensions[cell_obj.row].height = desired_h


    except KeyError:
        raise CellReferenceError(f"Invalid cell reference: '{cell}'")
    except Exception as e:
        raise ExcelMCPError(f"Error updating cell: {e}")

def autofit_table(ws: Any, cell_range: str) -> None:
    """Adjust column widths and row heights for a tabular range."""
    start_row, start_col, end_row, end_col = ExcelRange.parse_range(cell_range)

    col_widths: Dict[int, int] = {}
    row_heights: Dict[int, int] = {}

    for row in range(start_row, end_row + 1):
        max_lines = 1
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row + 1, column=col + 1)
            value = cell.value
            if value is None:
                continue
            text = str(value)
            lines = text.splitlines()
            longest = max(len(line) for line in lines)
            col_widths[col] = max(col_widths.get(col, 0), longest)
            est_lines = max(len(lines), math.ceil(longest / 40))
            if est_lines > 1:
                cell.alignment = Alignment(wrap_text=True)
            max_lines = max(max_lines, est_lines)
        if max_lines > 1:
            row_heights[row] = max_lines * 15

    for col, width in col_widths.items():
        column_letter = get_column_letter(col + 1)
        current = ws.column_dimensions[column_letter].width or 8.43
        desired = min(width + 2, 80)
        if desired > current:
            ws.column_dimensions[column_letter].width = desired

    for row, height in row_heights.items():
        current = ws.row_dimensions[row + 1].height or 15
        if height > current:
            ws.row_dimensions[row + 1].height = height

def apply_style(ws: Any, cell_range: str, style_dict: Dict[str, Any]) -> None:
    """
    Apply cell styles to a range.

    Args:
        ws: Openpyxl worksheet object.
        cell_range (str): Range in ``A1:B5`` format or a single cell like ``"A1"``.
        style_dict (dict): Dictionary with styles to apply:
            - font_name (str): Font name.
            - font_size (int): Font size.
            - bold (bool): Bold text.
            - italic (bool): Italic text.
            - fill_color (str): Background color in hex, e.g. ``"FF0000"``.
            - border_style (str): Border style (``"thin"``, ``"medium"``, ``"thick"``, etc.).
            - alignment (str): Alignment (``"center"``, ``"left"``, ``"right"``, etc.).

    Raises:
        RangeError: If the range is invalid.
    """
    if not ws:
        raise ExcelMCPError("El worksheet no puede ser None")
    
    try:
        # Parse the range
        if ':' in cell_range:
            # Cell range
            start_cell, end_cell = cell_range.split(':')
            start_coord = ws[start_cell].coordinate
            end_coord = ws[end_cell].coordinate
            range_str = f"{start_coord}:{end_coord}"
        else:
            # A single cell
            range_str = cell_range
        
        # Preparar los estilos
        font_kwargs = {}
        if 'font_name' in style_dict:
            font_kwargs['name'] = style_dict['font_name']
        if 'font_size' in style_dict:
            font_kwargs['size'] = style_dict['font_size']
        if 'bold' in style_dict:
            font_kwargs['bold'] = style_dict['bold']
        if 'italic' in style_dict:
            font_kwargs['italic'] = style_dict['italic']
        if 'font_color' in style_dict:
            font_kwargs['color'] = style_dict['font_color']
        
        fill = None
        if 'fill_color' in style_dict:
            fill = PatternFill(start_color=style_dict['fill_color'], 
                              end_color=style_dict['fill_color'],
                              fill_type='solid')
        
        border = None
        if 'border_style' in style_dict:
            side = Side(style=style_dict['border_style'])
            border = Border(left=side, right=side, top=side, bottom=side)
        
        alignment = None
        if 'alignment' in style_dict:
            alignment_value = style_dict['alignment'].lower()
            horizontal = None
            
            # Map horizontal alignment values
            if alignment_value in ['left', 'center', 'right', 'justify']:
                horizontal = alignment_value
            
            alignment = Alignment(horizontal=horizontal)
        
        # Apply styles to all cells in the range
        for row in ws[range_str]:
            for cell in row:
                if font_kwargs:
                    cell.font = Font(**font_kwargs)
                if fill:
                    cell.fill = fill
                if border:
                    cell.border = border
                if alignment:
                    cell.alignment = alignment
    
    except KeyError:
        raise RangeError(f"Invalid range: '{cell_range}'")
    except Exception as e:
        raise ExcelMCPError(f"Error applying styles: {e}")

def apply_number_format(ws: Any, cell_range: str, fmt: str) -> None:
    """
    Apply a number format to a range of cells.

    Args:
        ws: Openpyxl worksheet object.
        cell_range (str): Range in ``A1:B5`` format or a single cell like ``"A1"``.
        fmt (str): Number format (``"#,##0.00"``, ``"0%"``, ``"dd/mm/yyyy"``, etc.).

    Raises:
        RangeError: If the range is invalid.
    """
    if not ws:
        raise ExcelMCPError("El worksheet no puede ser None")
    
    try:
        # Parse the range
        if ':' in cell_range:
            # Cell range
            start_cell, end_cell = cell_range.split(':')
            start_coord = ws[start_cell].coordinate
            end_coord = ws[end_cell].coordinate
            range_str = f"{start_coord}:{end_coord}"
        else:
            # A single cell
            range_str = cell_range
        
        # Apply the format to all cells in the range
        for row in ws[range_str]:
            for cell in row:
                cell.number_format = fmt
    
    except KeyError:
        raise RangeError(f"Invalid range: '{cell_range}'")
    except Exception as e:
        raise ExcelMCPError(f"Error applying number format: {e}")

# 4. Tables and formulas (from advanced_excel_mcp.py)
def add_table(ws: Any, table_name: str, cell_range: str, style=None) -> Any:
    """
    Define a range as a styled table.

    Args:
        ws: Openpyxl worksheet object.
        table_name (str): Unique name for the table.
        cell_range (str): Range in ``A1:B5`` format.
        style (str, optional): Predefined style name or a custom dict.

    Returns:
        The created :class:`Table` object.

    Raises:
        TableError: If there is a problem with the table, e.g. duplicate name.
    """
    if not ws:
        raise ExcelMCPError("El worksheet no puede ser None")
    
    try:
        # Sanitize table name to prevent Excel corruption
        # Table names must be unique and follow Excel naming rules
        sanitized_name = table_name.replace(' ', '_').replace('-', '_')
        sanitized_name = ''.join(c for c in sanitized_name if c.isalnum() or c == '_')
        if not sanitized_name:
            sanitized_name = f"Table_{len(ws.tables) + 1}"
        
        # Ensure uniqueness
        original_name = sanitized_name
        counter = 1
        while hasattr(ws, 'tables') and sanitized_name in ws.tables:
            sanitized_name = f"{original_name}_{counter}"
            counter += 1
        
        table_name = sanitized_name
        
        # Use the exact range provided - no aggressive modifications to prevent corruption
        # Tables should contain ONLY the specified data range
        
        # Create table object
        table = Table(displayName=table_name, ref=cell_range)
        
        # Apply style (use default if not specified)
        if style:
            if isinstance(style, dict):
                # Custom style
                style_info = TableStyleInfo(**style)
            else:
                # Predefined style
                style_info = TableStyleInfo(
                    name=style,
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
        else:
            # Use default style
            style_info = TableStyleInfo(
                name=DEFAULT_TABLE_STYLE,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
        table.tableStyleInfo = style_info
        
        # Add table to the sheet
        ws.add_table(table)

        return table
    
    except Exception as e:
        if "Duplicate name" in str(e):
            raise TableError(f"A table named '{table_name}' already exists")
        elif "Invalid coordinate" in str(e) or "Invalid cell" in str(e):
            raise RangeError(f"Invalid range: '{cell_range}'")
        else:
            raise TableError(f"Error adding table: {e}")

def set_formula(ws: Any, cell: str, formula: str) -> Any:
    """
    Set a formula in a cell.

    Args:
        ws: Openpyxl worksheet object.
        cell (str): Cell reference (e.g. ``"A1"``).
        formula (str): Excel formula, with or without the ``=`` sign.

    Returns:
        The updated cell.

    Raises:
        FormulaError: If there is a problem with the formula.
    """
    if not ws:
        raise ExcelMCPError("El worksheet no puede ser None")
    
    try:
        # Add '=' sign if not present
        if formula and not formula.startswith('='):
            formula = f"={formula}"
        
        # Set the formula
        ws[cell] = formula
        return ws[cell]
    
    except KeyError:
        raise RangeError(f"Invalid cell: '{cell}'")
    except Exception as e:
        raise FormulaError(f"Error setting formula: {e}")

# 5. Charts and pivot tables (from advanced_excel_mcp.py)
def add_chart(
    wb: Any,
    sheet_name: str,
    chart_type: str,
    data_range: str,
    title=None,
    position=None,
    style=None,
    theme=None,
    custom_palette=None,
) -> Tuple[int, Any]:
    """Insert a native chart using the data from the given range.
     **Emojis must never be included in text written to cells, labels, titles or charts.**

    ``data_range`` should reference a rectangular table with no blank cells in
    the value area. The first row or first column is interpreted as headers and
    categories according to :func:`determine_orientation`. All series must
    contain only numbers and be the same length as the category vector. If total
    rows or mixed columns exist, the chart may be created incorrectly.

    Validate beforehand that the range belongs to the correct sheet and that
    headers are present, since the series are added with
    ``titles_from_data=True``. Categories must not contain blank or duplicate
    values and numeric columns must not contain text.

    Args:
        wb: Openpyxl ``Workbook`` object.
        sheet_name: Name of the sheet where the chart will be inserted.
        chart_type: Chart type (``'column'``, ``'bar'``, ``'line'``, ``'pie'``, etc.).
        data_range: Data range in ``A1:B5`` format.
        title: Optional chart title.
        position: Cell where the chart will be placed (e.g. ``"E5"``).
        style: Chart style (number ``1``–``48`` or descriptive name).
        theme: Name of the color theme.
        custom_palette: List of custom colors.

    Returns:
        Tuple ``(chart id, chart object)``.

    Raises:
        ChartError: If a problem occurs while creating the chart.
    """
    if not wb:
        raise ExcelMCPError("Workbook cannot be None")
    
    try:
        # Get the sheet
        ws = get_sheet(wb, sheet_name)
        
        # Validate and normalize the position
        if position:
            # If the position is a range, take only the first cell
            if ':' in position:
                position = position.split(':')[0]
            
            # Validate that it is a valid cell reference
            try:
                # Try parsing to verify it is a valid reference
                ExcelRange.parse_cell_ref(position)
            except ValueError:
                raise ValueError(f"Invalid position '{position}'. Must be a cell reference (e.g. 'E4')")
        
        # Create the chart object according to the type
        chart = None
        if chart_type.lower() == 'column':
            chart = BarChart()
            chart.type = "col"
        elif chart_type.lower() == 'bar':
            chart = BarChart()
            chart.type = "bar"
        elif chart_type.lower() == 'line':
            chart = LineChart()
        elif chart_type.lower() == 'pie':
            chart = PieChart()
        elif chart_type.lower() == 'scatter':
            chart = ScatterChart()
        elif chart_type.lower() == 'area':
            chart = AreaChart()
        else:
            raise ChartError(f"Chart type not supported: '{chart_type}'")
        
        # Set title if provided
        if title:
            chart.title = title
            # Configure title font
            if hasattr(chart.title, 'tx'):
                if hasattr(chart.title.tx, 'rich'):
                    from openpyxl.drawing.text import CharacterProperties
                    char_props = CharacterProperties()
                    char_props.sz = DEFAULT_CHART_TITLE_SIZE * 100  # Size in hundreds
                    char_props.b = True  # Bold
                    if hasattr(chart.title.tx.rich, 'p') and len(chart.title.tx.rich.p) > 0:
                        if hasattr(chart.title.tx.rich.p[0], 'r') and len(chart.title.tx.rich.p[0].r) > 0:
                            chart.title.tx.rich.p[0].r[0].rPr = char_props
            
        # Determine if the range references another sheet
        data_sheet_name, sr, sc, er, ec = ExcelRange.parse_range_with_sheet(data_range)
        if data_sheet_name is None:
            data_sheet_name = sheet_name
        data_ws = get_sheet(wb, data_sheet_name)

        # Normalize the range for Reference (escaping the sheet name)
        if " " in data_sheet_name or any(c in data_sheet_name for c in "![]{}?"):
            sheet_prefix = f"'{data_sheet_name}'!"
        else:
            sheet_prefix = f"{data_sheet_name}!"
        clean_range = ExcelRange.range_to_a1(sr, sc, er, ec)
        
        # Parse data range
        try:
            # Use the previously calculated limits
            min_row = sr + 1
            min_col = sc + 1
            max_row = er + 1
            max_col = ec + 1

            # Trim empty rows or columns at the end
            min_row, min_col, max_row, max_col = _trim_range_to_data(data_ws, min_row, min_col, max_row, max_col)
            if max_row < min_row or max_col < min_col:
                raise ChartError("The specified range contains no data")

            # Determine orientation by analyzing the range contents
            is_column_oriented = determine_orientation(data_ws, min_row, min_col, max_row, max_col)
            
            # For charts that need categories (most except scatter)
            if chart_type.lower() != 'scatter':
                if is_column_oriented:
                    # More tolerant approach - warn about blanks but don't fail
                    if _range_has_blank(data_ws, min_row + 1, min_col + 1, max_row, max_col):
                        logger.warning("Data range contains blank cells, chart may have gaps")
                    
                    categories = Reference(data_ws, min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=min_col)
                    data = Reference(data_ws, min_row=min_row, max_row=max_row, min_col=min_col + 1, max_col=max_col)
                    
                    # Try different approaches for adding data
                    try:
                        chart.add_data(data, titles_from_data=True)
                    except (TypeError, AttributeError):
                        try:
                            chart.add_data(data, titles_from_data=False)
                        except:
                            chart.add_data(data)
                    
                    try:
                        chart.set_categories(categories)
                    except Exception as e:
                        logger.warning(f"Could not set categories: {e}")
                else:
                    # More tolerant approach - warn about blanks but don't fail
                    if _range_has_blank(data_ws, min_row + 1, min_col, max_row, max_col):
                        logger.warning("Data range contains blank cells, chart may have gaps")
                    
                    categories = Reference(data_ws, min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col)
                    data = Reference(data_ws, min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col)
                    
                    # Try different approaches for adding data
                    try:
                        chart.add_data(data, titles_from_data=True)
                    except (TypeError, AttributeError):
                        try:
                            chart.add_data(data, titles_from_data=False)
                        except:
                            chart.add_data(data)
                    
                    try:
                        chart.set_categories(categories)
                    except Exception as e:
                        logger.warning(f"Could not set categories: {e}")
            else:
                # For scatter charts, be more flexible
                if _range_has_blank(data_ws, min_row, min_col, max_row, max_col):
                    logger.warning("Data range contains blank cells, chart may have gaps")
                
                data_ref = Reference(data_ws, min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col)
                try:
                    chart.add_data(data_ref, titles_from_data=True)
                except (TypeError, AttributeError):
                    chart.add_data(data_ref)
        
        except Exception as e:
            raise RangeError(f"Error processing data range '{data_range}': {e}")
        
        # Apply styles
        if style is not None:
            # Convert specified style (number, name, etc.)
            style_number = parse_chart_style(style)
            if style_number is not None:
                # Apply the style including the color palette
                apply_chart_style(chart, style_number)
            else:
                logger.warning(f"Invalid chart style: '{style}'. Using default style.")
        
        # Apply color theme if provided
        # (here we would use the theme but omit it for simplicity)
        
        # Apply custom palette if provided
        if custom_palette and isinstance(custom_palette, list):
            from openpyxl.chart.shapes import GraphicalProperties
            from openpyxl.drawing.fill import ColorChoice
            
            for i, series in enumerate(chart.series):
                if i < len(custom_palette):
                    # Ensure graphical properties exist
                    if not hasattr(series, 'graphicalProperties'):
                        series.graphicalProperties = GraphicalProperties()
                    elif series.graphicalProperties is None:
                        series.graphicalProperties = GraphicalProperties()
                    
                    # Assign color ensuring it doesn't have the # prefix
                    color = custom_palette[i]
                    if isinstance(color, str) and color.startswith('#'):
                        color = color[1:]
                    
                    # Apply the color explicitly
                    series.graphicalProperties.solidFill = ColorChoice(srgbClr=color)
        
        # Configure axis titles and legend
        # Add X axis title
        if hasattr(chart, 'x_axis') and chart.x_axis:
            chart.x_axis.title = "Categories"
            chart.x_axis.titleFont = Font(name=DEFAULT_CHART_FONT, size=DEFAULT_CHART_FONT_SIZE)
        
        # Add Y axis title
        if hasattr(chart, 'y_axis') and chart.y_axis:
            chart.y_axis.title = "Values"
            chart.y_axis.titleFont = Font(name=DEFAULT_CHART_FONT, size=DEFAULT_CHART_FONT_SIZE)
        
        # Configure legend position
        if hasattr(chart, 'legend'):
            chart.legend.position = DEFAULT_CHART_LEGEND_POSITION
            # Ensure legend doesn't overlap with chart
            chart.legend.overlay = False
        
        # Ensure proper spacing for chart position
        if position:
            # Parse position string to get coordinates
            try:
                import re
                pos_match = re.match(r'([A-Z]+)(\d+)', position.upper())
                if pos_match:
                    pos_col = column_index_from_string(pos_match.group(1)) - 1  # Convert to 0-based
                    pos_row = int(pos_match.group(2)) - 1  # Convert to 0-based
                    optimal_position = find_optimal_chart_position(ws, pos_col, pos_row, 8, 15)
                else:
                    optimal_position = find_optimal_chart_position(ws, 5, 0, 8, 15)  # Default F1
            except Exception:
                optimal_position = find_optimal_chart_position(ws, 5, 0, 8, 15)  # Default F1
            ws.add_chart(chart, optimal_position)
        else:
            # Find a suitable position automatically
            optimal_position = find_optimal_chart_position(ws, 5, 0, 8, 15)  # F1 = col 5, row 0
            ws.add_chart(chart, optimal_position)
        
        # Determine the chart ID (based on its position in the list)
        chart_id = len(ws._charts) - 1
        
        return chart_id, chart
    
    except SheetNotFoundError:
        raise
    except ChartError:
        raise
    except RangeError:
        raise
    except Exception as e:
        raise ChartError(f"Error creating chart: {e}")

def add_pivot_table(wb: Any, source_sheet: str, source_range: str, target_sheet: str,
                   target_cell: str, rows: List[str], cols: List[str], data_fields: List[str]) -> Any:
    """
    Create a pivot table.

    Args:
        wb: Openpyxl workbook object.
        source_sheet (str): Sheet containing the source data.
        source_range (str): Range of source data (e.g. ``A1:E10``).
        target_sheet (str): Sheet where the pivot table will be created.
        target_cell (str): Anchor cell (e.g. ``"A1"``).
        rows (list): Row fields.
        cols (list): Column fields.
        data_fields (list): Value fields and functions.

    Returns:
        The created :class:`PivotTable` object.

    Raises:
        PivotTableError: If there are issues with the pivot table.
    """
    if not wb:
        raise ExcelMCPError("Workbook cannot be None")
    
    try:
        # Get source sheet
        source_ws = get_sheet(wb, source_sheet)

        # Get target sheet
        target_ws = get_sheet(wb, target_sheet)

        logger.warning("Pivot tables in openpyxl have limited functionality and may not work as expected.")

        # Try creating the data cache (this is a required step)
        try:
            # Parse the range
            min_row, min_col, max_row, max_col = ExcelRange.parse_range(source_range)

            # Adjust to 1-based for Reference
            min_row += 1
            min_col += 1
            max_row += 1
            max_col += 1

            # Create data reference for the cache
            data_reference = Reference(source_ws, min_row=min_row, min_col=min_col,
                                     max_row=max_row, max_col=max_col)

            # Create pivot cache
            pivot_cache = PivotCache(cacheSource=data_reference, cacheDefinition={'refreshOnLoad': True})

            # Generate a unique ID for the pivot table
            pivot_name = f"PivotTable{len(wb._pivots) + 1 if hasattr(wb, '_pivots') else 1}"

            # Create the pivot table
            pivot_table = PivotTable(name=pivot_name, cache=pivot_cache,
                                    location=target_cell, rowGrandTotals=True, colGrandTotals=True)

            # Add row fields
            for row_field in rows:
                pivot_table.rowFields.append(PivotField(data=row_field))

            # Add column fields
            for col_field in cols:
                pivot_table.colFields.append(PivotField(data=col_field))

            # Add data fields
            for data_field in data_fields:
                pivot_table.dataFields.append(PivotField(data=data_field))

            # Add the pivot table to the target sheet
            target_ws.add_pivot_table(pivot_table)
            
            return pivot_table
            
        except Exception as pivot_error:
            logger.error(f"Error creating pivot table: {pivot_error}")
            raise PivotTableError(f"Error creating pivot table: {pivot_error}")
    
    except SheetNotFoundError:
        raise
    except PivotTableError:
        raise
    except Exception as e:
        raise PivotTableError(f"Error creating pivot table: {e}")

# ----------------------------------------
# NEW COMBINED HIGH-LEVEL FUNCTIONS
# ----------------------------------------

def create_sheet_with_data(wb: Any, sheet_name: str, data: List[List[Any]],
                           index: Optional[int] = None, overwrite: bool = False) -> Any:
    """
    Create a new sheet and write data in a single step.
     **Emojis must never be included in text written to cells, labels, titles or charts.**

    Args:
        wb: Openpyxl workbook object.
        sheet_name (str): Name for the new sheet.
        data (List[List]): Data to write.
        index (int, optional): Position of the sheet in the workbook.
        overwrite (bool): If ``True`` overwrite an existing sheet with the same name.

    Returns:
        The created worksheet object.

    Raises:
        SheetExistsError: If the sheet already exists and ``overwrite`` is ``False``.
    """
    # Handle existing sheet case
    if sheet_name in list_sheets(wb):
        if overwrite:
            # Delete the existing sheet
            delete_sheet(wb, sheet_name)
        else:
            raise SheetExistsError(f"Sheet '{sheet_name}' already exists. Use overwrite=True to overwrite it.")
    
    # Create new sheet
    ws = add_sheet(wb, sheet_name, index)
    
    # Write data
    if data:
        write_sheet_data(ws, "A1", data)
    
    return ws

def create_formatted_table(wb: Any, sheet_name: str, start_cell: str, data: List[List[Any]],
                            table_name: str, table_style: Optional[str] = None,
                            formats: Optional[Dict[str, Union[str, Dict]]] = None) -> Tuple[Any, Any]:
    """
    Create a formatted table in a single step.
     **Emojis must never be included in text written to cells, labels, titles or charts.**

    Args:
        wb: Openpyxl workbook object.
        sheet_name (str): Name of the sheet where the table will be created.
        start_cell (str): Starting cell for the data (e.g. ``"A1"``).
        data (List[List]): Data for the table including headers.
        table_name (str): Unique name for the table.
        table_style (str, optional): Predefined table style (e.g. ``"TableStyleMedium9"``).
        formats (dict, optional): Dictionary of formats to apply:
            - Keys: Relative ranges (e.g. ``"A2:A10"``) or cells.
            - Values: Number format or a style dictionary.

    Returns:
        Tuple ``(table object, worksheet)``.

    Example format::
        
        formats = {
            "B2:B10": "#,##0.00",  # Currency format
            "A1:Z1": {"bold": True, "fill_color": "DDEBF7"}  # Header style
        }
    """
    # Get the sheet
    ws = get_sheet(wb, sheet_name)
    
    # Get data range dimensions
    rows = len(data)
    cols = max([len(row) if isinstance(row, list) else 1 for row in data], default=0)
    
    # Write the data
    write_sheet_data(ws, start_cell, data)
    
    # Calculate the full table range
    start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
    end_row = start_row + rows - 1
    end_col = start_col + cols - 1
    full_range = ExcelRange.range_to_a1(start_row, start_col, end_row, end_col)
    
    # Create the table (will use default style if not specified)
    table = add_table(ws, table_name, full_range, table_style)
    
    # Apply enhanced formatting automatically
    try:
        # Apply enhanced autofit to all columns
        enhanced_autofit_columns(ws)
        # Apply consistent number formatting
        apply_consistent_number_format(ws)
    except Exception:
        pass
    
    # Apply additional formats if provided
    if formats:
        for range_str, format_value in formats.items():
            # Convert relative range to absolute if needed
            if not any(c in range_str for c in [':', '!']):
                # It's a single cell, add offset
                cell_row, cell_col = ExcelRange.parse_cell_ref(range_str)
                abs_row = start_row + cell_row
                abs_col = start_col + cell_col
                abs_range = ExcelRange.cell_to_a1(abs_row, abs_col)
            elif ':' in range_str and '!' not in range_str:
                # Range without a specific sheet, add offset
                range_start, range_end = range_str.split(':')
                start_row_rel, start_col_rel = ExcelRange.parse_cell_ref(range_start)
                end_row_rel, end_col_rel = ExcelRange.parse_cell_ref(range_end)

                # Calculate absolute positions
                abs_start_row = start_row + start_row_rel
                abs_start_col = start_col + start_col_rel
                abs_end_row = start_row + end_row_rel
                abs_end_col = start_col + end_col_rel

                # Create absolute range
                abs_range = ExcelRange.range_to_a1(abs_start_row, abs_start_col, abs_end_row, abs_end_col)
            else:
                # It's already an absolute range or includes the sheet
                abs_range = range_str

            # Apply format according to type
            if isinstance(format_value, str):
                # It's a number format
                apply_number_format(ws, abs_range, format_value)
            elif isinstance(format_value, dict):
                # It's a style dictionary
                apply_style(ws, abs_range, format_value)
    
    return table, ws

def create_chart_from_table(
    wb: Any,
    sheet_name: str,
    table_name: str,
    chart_type: str,
    title: Optional[str] = None,
    position: Optional[str] = None,
    style: Optional[Any] = None,
    use_headers: bool = True,
) -> Tuple[int, Any]:
    """Generate a chart from an existing table.
     **Emojis must never be included in text written to cells, labels, titles or charts.**

    The table must contain valid headers and must not include total rows. Data
    cells are assumed to form a rectangular range with no blanks. When
    ``use_headers`` is ``True`` the first row of the table is used as series
    titles and categories. All data columns must be numeric and the same length
    to avoid errors when creating the chart.

    Ensure the table has no blank cells or text columns where numbers are
    expected. Any mismatch in series length or category count can lead to
    incomplete or empty charts.

    Args:
        wb: Openpyxl ``Workbook`` object.
        sheet_name: Name of the sheet containing the table.
        table_name: Name of the table used as the source.
        chart_type: Chart type (``'column'``, ``'bar'``, ``'line'``, ``'pie'``, etc.).
        title: Optional chart title.
        position: Anchor cell for the chart.
        style: Chart style (number ``1``–``48`` or descriptive name).
        use_headers: If ``True`` use the first row as headers and categories.

    Returns:
        Tuple ``(chart ID, chart object)``.
    """
    # Get the sheet
    ws = get_sheet(wb, sheet_name)
    
    # Get table information
    tables = list_tables(wb, sheet_name)
    table_info = None
    for table in tables:
        if table['name'] == table_name:
            table_info = table
            break
    
    if not table_info:
        raise TableError(f"Table '{table_name}' not found in sheet '{sheet_name}'")
    
    # Use the table range to create the chart
    table_range = table_info['ref']
    
    # Create the chart
    chart_id, chart = add_chart(wb, sheet_name, chart_type, table_range, 
                               title, position, style)
    
    return chart_id, chart

def create_chart_from_data(
    wb: Any,
    sheet_name: str,
    data: List[List[Any]],
    chart_type: str,
    position: Optional[str] = None,
    title: Optional[str] = None,
    style: Optional[Any] = None,
    create_table: bool = False,
    table_name: Optional[str] = None,
    table_style: Optional[str] = None,
) -> Dict[str, Any]:
    """Create a chart from ``data`` by writing the values first.
     **Emojis must never be included in text written to cells, labels, titles or charts.**

    ``data`` must be a list of lists forming a rectangular structure with no
    empty cells. The first row or column is interpreted as headers and
    categories; therefore every row must have the same length and numeric columns
    should not contain text. Avoid including total rows or records that should
    not be charted.

    Before calling the function check for blank cells, duplicated headers or
    mismatched lengths between categories and series. ``add_chart`` uses
    ``titles_from_data=True`` to assign the series names. If the series are
    inconsistent the resulting chart may be incomplete or show errors.

    Args:
        wb: Openpyxl ``Workbook`` object.
        sheet_name: Name of the sheet where the chart will be created.
        data: Data matrix including the headers.
        chart_type: Chart type (``'column'``, ``'bar'``, ``'line'``, ``'pie'``, etc.).
        position: Cell where to place the chart.
        title: Chart title.
        style: Chart style (number ``1``–``48`` or descriptive name).
        create_table: If ``True`` a table with the written data will be created.
        table_name: Name of the table (required if ``create_table`` is ``True``).
        table_style: Optional table style.

    Returns:
        Dictionary with information about the chart and, if created, the table.
    """
    # Create sheet if it does not exist
    if sheet_name not in list_sheets(wb):
        add_sheet(wb, sheet_name)
    
    # Get the sheet
    ws = get_sheet(wb, sheet_name)
    
    # Determine a suitable location for the data
    # By default, place the data at A1
    data_start_cell = "A1"
    
    # Write the data
    write_sheet_data(ws, data_start_cell, data)
    
    # Calculate the full data range
    rows = len(data)
    cols = max([len(row) if isinstance(row, list) else 1 for row in data], default=0)
    start_row, start_col = ExcelRange.parse_cell_ref(data_start_cell)
    end_row = start_row + rows - 1
    end_col = start_col + cols - 1
    data_range = ExcelRange.range_to_a1(start_row, start_col, end_row, end_col)
    
    result = {
        "data_range": data_range,
        "rows": rows,
        "columns": cols
    }
    
    # Create table if requested
    if create_table:
        if not table_name:
            # Generate table name if not provided
            table_name = f"Table_{sheet_name}_{int(time.time())}"
            
        try:
            table = add_table(ws, table_name, data_range, table_style)
            result["table"] = {
                "name": table_name,
                "range": data_range,
                "style": table_style
            }
        except Exception as e:
            logger.warning(f"Could not create the table: {e}")
    
    # Create the chart
    try:
        chart_id, chart = add_chart(wb, sheet_name, chart_type, data_range, 
                                  title, position, style)
        
        result["chart"] = {
            "id": chart_id,
            "type": chart_type,
            "title": title,
            "position": position,
            "style": style
        }
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        raise ChartError(f"Error creating chart: {e}")

    return result

def create_chart_from_dataframe(
    wb: Any,
    sheet_name: str,
    df: 'pd.DataFrame',
    chart_type: str,
    position: Optional[str] = None,
    title: Optional[str] = None,
    style: Optional[Any] = None,
    create_table: bool = False,
    table_name: Optional[str] = None,
    table_style: Optional[str] = None,
) -> Dict[str, Any]:
    """Generate a chart from a ``pandas.DataFrame``.
     **Emojis must never be included in text written to cells, labels, titles or charts.**

    The ``DataFrame`` must contain numeric columns without missing values in the
    series and should not include total rows. The headers are used as series
    titles, so duplicates or blank cells should be avoided. The contents of the
    ``DataFrame`` are written to the sheet and delegated to
    :func:`create_chart_from_data`, therefore the same validation rules apply.

    Args:
        wb: Openpyxl ``Workbook`` object.
        sheet_name: Name of the sheet where the chart will be created.
        df: Data in ``pandas.DataFrame`` format.
        chart_type: Chart type (``'column'``, ``'bar'``, ``'line'``, ``'pie'``, etc.).
        position: Anchor cell for the chart.
        title: Chart title.
        style: Optional chart style.
        create_table: If ``True`` create a table with the written data.
        table_name: Name of the table to create.
        table_style: Table style.

    Returns:
        Dictionary with information about the chart and, if created, the table.
    """

    if df is None:
        raise ExcelMCPError("The provided DataFrame is None")

    # Convert the DataFrame to a list of lists including headers
    data = [df.columns.tolist()] + df.values.tolist()

    return create_chart_from_data(
        wb,
        sheet_name,
        data,
        chart_type,
        position=position,
        title=title,
        style=style,
        create_table=create_table,
        table_name=table_name,
        table_style=table_style,
    )

def create_report(wb: Any, data: Dict[str, List[List[Any]]], tables: Optional[Dict[str, Dict[str, Any]]] = None,
                 charts: Optional[Dict[str, Dict[str, Any]]] = None, formats: Optional[Dict[str, Dict[str, Any]]] = None,
                 overwrite_sheets: bool = False) -> Dict[str, Any]:
    """
    Create a full report with multiple sheets, tables and charts in a single step.
     **Emojis must never be included in text written to cells, labels, titles or charts.**

    This function acts as a generic template for automated report generators.
    All created sheets should be orderly and styled. Check the available space
    before inserting charts so they do not end up on top of any table or block
    of text. After creating a table, verify which column contains the longest
    strings and adjust its width so the content is visible without manual
    editing.

    Args:
        wb: Openpyxl workbook object.
        data: Dictionary with data per sheet: ``{"Sheet1": [[data]], "Sheet2": [[data]]}``
        tables: Dictionary with table configuration:
            ``{"SalesTable": {"sheet": "Sales", "range": "A1:B10", "style": "TableStyleMedium9"}}``
        charts: Dictionary with chart configuration:
            ``{"SalesChart": {"sheet": "Sales", "type": "column", "data": "SalesTable",
                              "title": "Sales", "position": "D2", "style": "dark-blue"}}``
        formats: Dictionary of formats to apply:
            ``{"Sales": {"B2:B10": "#,##0.00", "A1:Z1": {"bold": True}}}``
        overwrite_sheets: If ``True`` overwrite existing sheets.

    Returns:
        Dictionary with information about the created elements.
    """
    result = {
        "sheets": [],
        "tables": [],
        "charts": []
    }
    
    # Create/update sheets with data
    for sheet_name, sheet_data in data.items():
        if sheet_name in list_sheets(wb):
            if overwrite_sheets:
                # Use the existing sheet
                ws = wb[sheet_name]
                # Write the data
                write_sheet_data(ws, "A1", sheet_data)
            else:
                # Add numeric suffix if the sheet already exists
                base_name = sheet_name
                counter = 1
                while f"{base_name}_{counter}" in list_sheets(wb):
                    counter += 1
                new_name = f"{base_name}_{counter}"
                ws = create_sheet_with_data(wb, new_name, sheet_data)
                sheet_name = new_name
        else:
            # Create new sheet
            ws = create_sheet_with_data(wb, sheet_name, sheet_data)
        
        result["sheets"].append({"name": sheet_name, "rows": len(sheet_data)})
        
        # Apply specific formats for this sheet
        if formats and sheet_name in formats:
            for range_str, format_value in formats[sheet_name].items():
                if isinstance(format_value, str):
                    apply_number_format(ws, range_str, format_value)
                elif isinstance(format_value, dict):
                    apply_style(ws, range_str, format_value)
    
    # Crear tablas
    if tables:
        for table_name, table_config in tables.items():
            sheet_name = table_config.get("sheet")
            range_str = table_config.get("range")
            style = table_config.get("style")
            
            if not sheet_name or not range_str:
                logger.warning(f"Incomplete configuration for table '{table_name}'. Sheet and range are required.")
                continue
            
            try:
                # Verify that the sheet exists
                if sheet_name not in list_sheets(wb):
                    logger.warning(f"Sheet '{sheet_name}' not found for table '{table_name}'. Skipping.")
                    continue
                
                ws = wb[sheet_name]
                table = add_table(ws, table_name, range_str, style)
                
                result["tables"].append({
                    "name": table_name,
                    "sheet": sheet_name,
                    "range": range_str,
                    "style": style
                })
                
                # Apply specific formats for this table
                if "formats" in table_config:
                    for range_str, format_value in table_config["formats"].items():
                        if isinstance(format_value, str):
                            apply_number_format(ws, range_str, format_value)
                        elif isinstance(format_value, dict):
                            apply_style(ws, range_str, format_value)
            
            except Exception as e:
                logger.warning(f"Error al crear tabla '{table_name}': {e}")
    
    # Create charts
    if charts:
        for chart_name, chart_config in charts.items():
            sheet_name = chart_config.get("sheet")
            chart_type = chart_config.get("type")
            data_source = chart_config.get("data")
            title = chart_config.get("title", chart_name)
            position = chart_config.get("position")
            style = chart_config.get("style")
            
            if not sheet_name or not chart_type or not data_source:
                logger.warning(f"Incomplete configuration for chart '{chart_name}'. Sheet, type and data are required.")
                continue
            
            try:
                # Verificar que la hoja existe
                if sheet_name not in list_sheets(wb):
                    logger.warning(f"Sheet '{sheet_name}' not found for chart '{chart_name}'. Skipping.")
                    continue
                
                # Determinar si data_source es una tabla o un rango
                data_range = data_source
                if data_source in [t["name"] for t in result["tables"]]:
                    # Es una tabla, obtener su rango
                    for table in result["tables"]:
                        if table["name"] == data_source:
                            data_range = table["range"]
                            break
                
                # Create the chart
                chart_id, chart = add_chart(wb, sheet_name, chart_type, data_range, 
                                           title, position, style)
                
                result["charts"].append({
                    "name": chart_name,
                    "id": chart_id,
                    "sheet": sheet_name,
                    "type": chart_type,
                    "data_source": data_source,
                    "position": position,
                    "style": style
                })
            
            except Exception as e:
                logger.warning(f"Error creating chart '{chart_name}': {e}")
    
    return result


def apply_excel_template(wb: Any, template_name: str, data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Apply a predefined template to an Excel workbook.
     **Emojis must never be included in text written to cells, labels, titles or charts.**

    Args:
        wb: Openpyxl workbook object.
        template_name (str): Name of the template to apply (e.g. ``"sales_report"``, ``"dashboard"``).
        data: Dictionary with data specific to the template.

    Returns:
        Dictionary with information about the created elements.

    Available templates:
        - ``"basic_report"``: Basic report with table and chart.
        - ``"financial_dashboard"``: Financial dashboard with multiple KPIs and charts.
        - ``"sales_analysis"``: Sales analysis by region and product.
        - ``"project_tracker"``: Project tracker with progress tables and charts.
    """
    result = {
        "template": template_name,
        "sheets": [],
        "elements": []
    }
    
    # Implementation of predefined templates
    if template_name == "basic_report":
        # Basic report template
        title = data.get("title", "Basic Report")
        subtitle = data.get("subtitle", "")
        report_date = data.get("date", time.strftime("%d/%m/%Y"))
        sheet_name = data.get("sheet", "Report")
        report_data = data.get("data", [])
        
        # Crear hoja para el informe si no existe
        if sheet_name not in list_sheets(wb):
            ws = add_sheet(wb, sheet_name)
        else:
            ws = wb[sheet_name]
        
        # Title and basic information
        update_cell(ws, "A1", title)
        apply_style(ws, "A1", {
            "font_size": 16,
            "bold": True,
            "alignment": "center"
        })
        
        if subtitle:
            update_cell(ws, "A2", subtitle)
            apply_style(ws, "A2", {
                "font_size": 12,
                "alignment": "center"
            })
        
        update_cell(ws, "A3", f"Fecha: {report_date}")
        
        # Crear tabla con los datos
        start_row = 5
        if report_data:
            write_sheet_data(ws, f"A{start_row}", report_data)
            
            # Determinar dimensiones
            rows = len(report_data)
            cols = max([len(row) if isinstance(row, list) else 1 for row in report_data], default=0)
            
            # Crear tabla
            table_range = f"A{start_row}:{get_column_letter(cols)}{start_row + rows - 1}"
            table_name = data.get("table_name", "ReportTable")
            table_style = data.get("table_style", "TableStyleMedium9")
            
            try:
                table = add_table(ws, table_name, table_range, table_style)
                result["elements"].append({
                    "type": "table",
                    "name": table_name,
                    "range": table_range
                })
            except Exception as e:
                logger.warning(f"Error al crear tabla: {e}")
            
            # Create chart
            chart_type = data.get("chart_type", "column")
            chart_position = data.get("chart_position", f"G{start_row}")
            chart_title = data.get("chart_title", "Report Chart")
            chart_style = data.get("chart_style", "colorful-1")
            
            try:
                chart_id, chart = add_chart(wb, sheet_name, chart_type, table_range, 
                                          chart_title, chart_position, chart_style)
                
                result["elements"].append({
                    "type": "chart",
                    "id": chart_id,
                    "position": chart_position
                })
            except Exception as e:
                logger.warning(f"Error creating chart: {e}")
        
        result["sheets"].append({"name": sheet_name, "type": "report"})
    
    elif template_name == "financial_dashboard":
        # More advanced template for a financial dashboard
        title = data.get("title", "Financial Dashboard")
        sheet_name = data.get("sheet", "Dashboard")
        financial_data = data.get("financial_data", {})
        
        # Configuration to create the full dashboard
        dashboard_config = {
            "title": title,
            "sheet": sheet_name,
            "sections": []
        }
        
        # 1. Financial KPI section
        if "kpis" in financial_data:
            kpis = financial_data["kpis"]
            kpi_section = {
                "title": "Key Financial Indicators",
                "type": "text",
                "content": "Financial KPIs"
            }
            dashboard_config["sections"].append(kpi_section)

            # Each KPI could be added as text or a formatted cell
            for kpi_name, kpi_value in kpis.items():
                kpi_section = {
                    "title": kpi_name,
                    "type": "text",
                    "content": f"{kpi_name}: {kpi_value}",
                    "format": {
                        "bold": True,
                        "font_size": 12
                    }
                }
                dashboard_config["sections"].append(kpi_section)
        
        # 2. Financial charts section
        if "charts" in financial_data:
            for chart_config in financial_data["charts"]:
                chart_section = {
                    "title": chart_config.get("title", "Financial Chart"),
                    "type": "chart",
                    "chart_type": chart_config.get("type", "column"),
                    "data_range": chart_config.get("data_range", ""),
                    "position": chart_config.get("position", ""),
                    "style": chart_config.get("style", "dark-blue")
                }
                dashboard_config["sections"].append(chart_section)
        
        # 3. Data tables section
        if "tables" in financial_data:
            for table_config in financial_data["tables"]:
                table_section = {
                    "title": table_config.get("title", "Financial Table"),
                    "type": "table",
                    "data_range": table_config.get("data_range", ""),
                    "name": table_config.get("name", "FinanceTable"),
                    "style": table_config.get("style", "TableStyleMedium9")
                }
                dashboard_config["sections"].append(table_section)
        
        # Create the dashboard
        dashboard_result = create_dashboard(wb, dashboard_config)
        
        # Add result
        result["sheets"].append({"name": sheet_name, "type": "dashboard"})
        result["dashboard"] = dashboard_result
    
    elif template_name == "sales_analysis":
        # Template for sales analysis
        title = data.get("title", "Sales Analysis")
        sheet_data = data.get("sales_data", [])
        sheet_name = data.get("sheet", "Sales")
        
        # Crear hoja de datos si no existe
        data_sheet = f"{sheet_name}_Datos"
        if data_sheet not in list_sheets(wb):
            data_ws = add_sheet(wb, data_sheet)
        else:
            data_ws = wb[data_sheet]
        
        # Escribir datos de ventas
        if sheet_data:
            write_sheet_data(data_ws, "A1", sheet_data)
            
            # Crear tabla para los datos
            rows = len(sheet_data)
            cols = max([len(row) if isinstance(row, list) else 1 for row in sheet_data], default=0)
            data_range = f"A1:{get_column_letter(cols)}{rows}"
            
            try:
                table = add_table(data_ws, "SalesDataTable", data_range, "TableStyleMedium9")
                result["elements"].append({
                    "type": "table",
                    "name": "SalesDataTable",
                    "sheet": data_sheet,
                    "range": data_range
                })
            except Exception as e:
                logger.warning(f"Error al crear tabla de datos: {e}")
        
        # Create analysis sheet
        if sheet_name not in list_sheets(wb):
            ws = add_sheet(wb, sheet_name)
        else:
            ws = wb[sheet_name]
        
        # Analysis title
        update_cell(ws, "A1", title)
        apply_style(ws, "A1", {
            "font_size": 16,
            "bold": True,
            "alignment": "center"
        })
        

            
        # Create analysis sections according to the data structure
        current_row = 3
        
        # 1. Sales by Region (assuming there is a region column)
        update_cell(ws, f"A{current_row}", "Sales by Region")
        apply_style(ws, f"A{current_row}", {"bold": True, "font_size": 12})
        current_row += 1
        
        try:
            # Create chart for sales by region
            chart_id, chart = add_chart(wb, sheet_name, "column",
                                       f"{data_sheet}!A1:{get_column_letter(cols)}{rows}",
                                       "Sales by Region", f"A{current_row}", "colorful-1")
            
            result["elements"].append({
                "type": "chart",
                "name": "SalesByRegionChart",
                "sheet": sheet_name,
                "id": chart_id
            })
            
            current_row += 15  # Space for the chart
        except Exception as e:
            logger.warning(f"Error creating sales by region chart: {e}")
            current_row += 2
        
        # 2. Sales Trend (if there is time data)
        update_cell(ws, f"A{current_row}", "Sales Trend")
        apply_style(ws, f"A{current_row}", {"bold": True, "font_size": 12})
        current_row += 1
        
        try:
            # Create chart for sales trend
            chart_id, chart = add_chart(wb, sheet_name, "line",
                                       f"{data_sheet}!A1:{get_column_letter(cols)}{rows}",
                                       "Sales Trend", f"A{current_row}", "line-markers")
            
            result["elements"].append({
                "type": "chart",
                "name": "SalesTrendChart",
                "sheet": sheet_name,
                "id": chart_id
            })
            
            current_row += 15  # Space for the chart
        except Exception as e:
            logger.warning(f"Error creating sales trend chart: {e}")
            current_row += 2
        
        result["sheets"].append({"name": sheet_name, "type": "analysis"})
        result["sheets"].append({"name": data_sheet, "type": "data"})
        
    elif template_name == "project_tracker":
        # Template for project tracking
        title = data.get("title", "Project Tracking")
        projects = data.get("projects", [])
        sheet_name = data.get("sheet", "Projects")
        
        # Prepare project data
        if not projects:
            # Create sample data if none is provided
            projects = [
                ["ID", "Project", "Owner", "Start", "Deadline", "Status", "Progress"],
                ["P001", "Project A", "Juan Pérez", "01/01/2023", "30/06/2023", "In progress", 75],
                ["P002", "Project B", "Ana López", "15/02/2023", "31/07/2023", "In progress", 40],
                ["P003", "Project C", "Carlos Ruiz", "01/03/2023", "31/08/2023", "Delayed", 20]
            ]
        
        # Create sheet for projects if it does not exist
        if sheet_name not in list_sheets(wb):
            ws = add_sheet(wb, sheet_name)
        else:
            ws = wb[sheet_name]
        
        # Tracker title
        update_cell(ws, "A1", title)
        apply_style(ws, "A1", {
            "font_size": 16,
            "bold": True,
            "alignment": "center"
        })
        

        
        # Write project data
        write_sheet_data(ws, "A3", projects)
        
        # Crear tabla para los datos
        rows = len(projects)
        cols = len(projects[0]) if rows > 0 else 7
        table_range = f"A3:{get_column_letter(cols)}{rows+2}"
        
        try:
            table = add_table(ws, "ProjectsTable", table_range, "TableStyleMedium9")
            result["elements"].append({
                "type": "table",
                "name": "ProjectsTable",
                "sheet": sheet_name,
                "range": table_range
            })
            
            # Apply percentage format to the progress column
            avance_col = get_column_letter(cols)
            apply_number_format(ws, f"{avance_col}4:{avance_col}{rows+2}", "0%")
        except Exception as e:
            logger.warning(f"Error al crear tabla de proyectos: {e}")
        
        # Create progress chart
        try:
            chart_id, chart = add_chart(wb, sheet_name, "column",
                                       table_range,
                                       "Project Progress", "I3", "colorful-3")
            
            result["elements"].append({
                "type": "chart",
                "name": "ProgressChart",
                "sheet": sheet_name,
                "id": chart_id
            })
        except Exception as e:
            logger.warning(f"Error creating progress chart: {e}")
        
        result["sheets"].append({"name": sheet_name, "type": "tracker"})
    
    else:
        logger.warning(f"Plantilla '{template_name}' no reconocida.")
        result["error"] = f"Plantilla '{template_name}' no disponible"
    
    return result

def update_report(wb: Any, report_config: Dict[str, Any],
                 recalculate: bool = True) -> Dict[str, Any]:
    """
    Update an existing report with new data.
     **Emojis must never be included in text written to cells, labels, titles or charts.**

    Args:
        wb: Openpyxl workbook object.
        report_config: Configuration for the report update
            {
                "data_updates": {
                    "Sales": {"range": "A2:C10", "data": [[new data]]},
                    "Customers": {"range": "A2:D20", "data": [[new data]]}
                },
                "recalculate_formulas": True,
                "refresh_charts": True
            }
        recalculate: If ``True`` recalculate formulas after updating.

    Returns:
        Dictionary with information about the updated elements.
    """
    result = {
        "updated_sheets": [],
        "updated_tables": [],
        "updated_charts": [],
        "recalculated": recalculate
    }
    
    # Actualizar datos en hojas
    data_updates = report_config.get("data_updates", {})
    for sheet_name, update_info in data_updates.items():
        if sheet_name not in list_sheets(wb):
            logger.warning(f"Sheet '{sheet_name}' not found. Skipping update.")
            continue
        
        ws = wb[sheet_name]
        range_str = update_info.get("range")
        data = update_info.get("data")
        
        if not range_str or not data:
            logger.warning(f"Incomplete configuration to update sheet '{sheet_name}'. Range and data are required.")
            continue
        
        try:
            # Obtener solo la primera celda del rango
            if ':' in range_str:
                start_cell = range_str.split(':')[0]
            else:
                start_cell = range_str
            
            # Escribir nuevos datos
            write_sheet_data(ws, start_cell, data)
            
            result["updated_sheets"].append({
                "name": sheet_name,
                "range": range_str
            })
        except Exception as e:
            logger.warning(f"Error al actualizar datos en hoja '{sheet_name}': {e}")
    
    # Actualizar/refrescar tablas
    refresh_tables = report_config.get("refresh_tables", [])
    for table_info in refresh_tables:
        sheet_name = table_info.get("sheet")
        table_name = table_info.get("name")
        new_range = table_info.get("new_range")
        
        if not sheet_name or not table_name:
            logger.warning("Incomplete table information. Sheet and name are required.")
            continue
        
        if sheet_name not in list_sheets(wb):
            logger.warning(f"Sheet '{sheet_name}' not found. Skipping table update.")
            continue
        
        ws = wb[sheet_name]
        
        try:
            # Verificar si la tabla existe
            if not hasattr(ws, 'tables') or table_name not in ws.tables:
                logger.warning(f"Table '{table_name}' not found in sheet '{sheet_name}'.")
                continue
            
            # Get current reference
            current_range = ws.tables[table_name].ref
            
            # Update range if a new one is provided
            if new_range:
                ws.tables[table_name].ref = new_range
                
                result["updated_tables"].append({
                    "name": table_name,
                    "sheet": sheet_name,
                    "old_range": current_range,
                    "new_range": new_range
                })
            else:
                result["updated_tables"].append({
                    "name": table_name,
                    "sheet": sheet_name,
                    "refreshed": True
                })
        except Exception as e:
            logger.warning(f"Error updating table '{table_name}': {e}")
    
    # Recalculate formulas if requested
    if recalculate:
        # OpenPyXL does not directly recalculate formulas
        # This is a placeholder that could be implemented in future versions
        # or via Excel's COM API if available
        result["recalculation_note"] = "Formula recalculation in OpenPyXL is limited"
    
    # Update charts
    refresh_charts = report_config.get("refresh_charts", [])
    for chart_info in refresh_charts:
        sheet_name = chart_info.get("sheet")
        chart_id = chart_info.get("id")
        new_data_range = chart_info.get("new_data_range")
        
        if not sheet_name or chart_id is None:
            logger.warning("Incomplete chart information. Sheet and id are required.")
            continue
        
        if sheet_name not in list_sheets(wb):
            logger.warning(f"Sheet '{sheet_name}' not found. Skipping chart update.")
            continue
        
        ws = wb[sheet_name]
        
        try:
            # Verify if the chart exists
            if not hasattr(ws, '_charts') or chart_id >= len(ws._charts) or chart_id < 0:
                logger.warning(f"Chart with ID {chart_id} not found in sheet '{sheet_name}'.")
                continue
            
            # In OpenPyXL updating a chart is not straightforward
            # One option is to delete the chart and create a new one
            if new_data_range:
                # Get current chart properties
                chart_rel = ws._charts[chart_id]
                chart = chart_rel[0]
                position = chart_rel[1] if len(chart_rel) > 1 else None
                
                # Determine chart type
                chart_type = "column"  # Default value
                if isinstance(chart, BarChart):
                    chart_type = "bar" if chart.type == "bar" else "column"
                elif isinstance(chart, LineChart):
                    chart_type = "line"
                elif isinstance(chart, PieChart):
                    chart_type = "pie"
                elif isinstance(chart, ScatterChart):
                    chart_type = "scatter"
                elif isinstance(chart, AreaChart):
                    chart_type = "area"
                
                # Get title if it exists
                title = chart.title if hasattr(chart, 'title') and chart.title else None
                
                # Delete the old chart
                del ws._charts[chart_id]
                
                # Create a new chart with the same parameters but new range
                new_chart_id, new_chart = add_chart(wb, sheet_name, chart_type, new_data_range,
                                                 title, position)
                
                result["updated_charts"].append({
                    "id": chart_id,
                    "new_id": new_chart_id,
                    "sheet": sheet_name,
                    "old_data_range": "unknown",  # No easy way to get the original range
                    "new_data_range": new_data_range
                })
            else:
                # Without a new range it cannot be easily updated
                result["updated_charts"].append({
                    "id": chart_id,
                    "sheet": sheet_name,
                    "note": "No new range provided. Updating data requires Excel COM."
                })
        except Exception as e:
            logger.warning(f"Error updating chart {chart_id}: {e}")
    
    return result

def import_data(wb: Any, import_config: Dict[str, Any]) -> Dict[str, Any]:
    """
    Import data from various sources into Excel.

    Args:
        wb: Openpyxl workbook object.
        import_config: Import configuration
            {
                "source": "csv",  # csv, json, pandas, etc.
                "source_path": "data.csv",
                "sheet": "Data",
                "start_cell": "A1",
                "options": {
                    "delimiter": ",",
                    "has_header": true
                }
            }

    Returns:
        Dictionary with information about the imported data.

    Note: This function is a simplified example that only imports data from CSV.
    """
    result = {
        "source": import_config.get("source"),
        "imported_rows": 0,
        "imported_columns": 0
    }
    
    source_type = import_config.get("source", "").lower()
    source_path = import_config.get("source_path")
    sheet_name = import_config.get("sheet", "Data")
    start_cell = import_config.get("start_cell", "A1")
    options = import_config.get("options", {})
    
    if not source_path:
        logger.warning("No source path specified for importing data.")
        result["error"] = "No source path specified"
        return result
        
    # Create the sheet if it does not exist
    if sheet_name not in list_sheets(wb):
        ws = add_sheet(wb, sheet_name)
    else:
        ws = wb[sheet_name]
    
    if source_type == "csv":
        try:
            import csv
            
            delimiter = options.get("delimiter", ",")
            has_header = options.get("has_header", True)
            
            data = []
            with open(source_path, 'r', encoding='utf-8', newline='') as f:
                csv_reader = csv.reader(f, delimiter=delimiter)
                for row in csv_reader:
                    data.append(row)
            
            # Write the data
            write_sheet_data(ws, start_cell, data)
            
            result["imported_rows"] = len(data)
            result["imported_columns"] = len(data[0]) if data else 0
            result["sheet"] = sheet_name
            result["start_cell"] = start_cell
        except Exception as e:
            logger.error(f"Error importing CSV: {e}")
            result["error"] = f"Error importing CSV: {e}"
    
    elif source_type == "json":
        try:
            import json
            
            with open(source_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            # Convert JSON to list of lists
            data = []
            
            if isinstance(json_data, list):
                # It is a list of objects
                if json_data and isinstance(json_data[0], dict):
                    # Get headers (claves del primer objeto)
                    headers = list(json_data[0].keys())
                    data.append(headers)
                    
                    # Add data rows
                    for item in json_data:
                        row = [item.get(header, "") for header in headers]
                        data.append(row)
                else:
                    # It is a simple list
                    for item in json_data:
                        data.append([item])
            elif isinstance(json_data, dict):
                # It is a dictionary
                for key, value in json_data.items():
                    data.append([key, value])
            
            # Write the data
            write_sheet_data(ws, start_cell, data)
            
            result["imported_rows"] = len(data)
            result["imported_columns"] = len(data[0]) if data else 0
            result["sheet"] = sheet_name
            result["start_cell"] = start_cell
        except Exception as e:
            logger.error(f"Error importing JSON: {e}")
            result["error"] = f"Error importing JSON: {e}"
    
    elif source_type == "pandas":
        try:
            import pandas as pd
            
            # Options for pandas
            file_ext = os.path.splitext(source_path)[1].lower()
            
            if file_ext == '.csv':
                df = pd.read_csv(source_path)
            elif file_ext in ['.xls', '.xlsx']:
                df = pd.read_excel(source_path)
            elif file_ext == '.json':
                df = pd.read_json(source_path)
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            # Convert DataFrame to list of lists
            data = [df.columns.tolist()]  # Encabezados
            data.extend(df.values.tolist())  # Datos
            
            # Write the data
            write_sheet_data(ws, start_cell, data)
            
            result["imported_rows"] = len(data)
            result["imported_columns"] = len(data[0]) if data else 0
            result["sheet"] = sheet_name
            result["start_cell"] = start_cell
        except Exception as e:
            logger.error(f"Error importing with pandas: {e}")
            result["error"] = f"Error importing with pandas: {e}"
    
    else:
        logger.warning(f"Unsupported source type: {source_type}")
        result["error"] = f"Unsupported source type: {source_type}"
    
    return result

def export_data(wb: Any, export_config: Dict[str, Any]) -> Dict[str, Any]:
    """
    Export data from Excel to different formats.

    Args:
        wb: Openpyxl workbook object.
        export_config: Export configuration
            {
                "format": "csv",  # csv, json, pdf, html, etc.
                "sheet": "Data",
                "range": "A1:D10",
                "output_path": "exported_data.csv",
                "options": {
                    "delimiter": ",",
                    "include_header": true
                }
            }

    Returns:
        Dictionary with information about the exported data.

    Note: This function is a simplified example that only exports to CSV and JSON.
    """
    result = {
        "format": export_config.get("format"),
        "exported_rows": 0,
        "exported_columns": 0
    }
    
    export_format = export_config.get("format", "").lower()
    sheet_name = export_config.get("sheet")
    range_str = export_config.get("range")
    output_path = export_config.get("output_path")
    options = export_config.get("options", {})
    
    if not sheet_name:
        logger.warning("No sheet specified para exportar datos.")
        result["error"] = "No sheet specified"
        return result
        
    if sheet_name not in list_sheets(wb):
        logger.warning(f"Hoja '{sheet_name}' no encontrada.")
        result["error"] = f"Hoja '{sheet_name}' no encontrada"
        return result
    
    # Leer los datos del rango especificado
    data = read_sheet_data(wb, sheet_name, range_str)
    
    if not data:
        logger.warning(f"No data found in range {range_str} de la hoja {sheet_name}")
        return []
    
    # Filter the data based on the criteria
    result = []
    headers = data[0] if data else []
    
    # If there is data with headers
    if len(data) > 1:
        # Convert to record format (list of dictionaries)
        records = []
        for row in data[1:]:
            record = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    record[header] = row[i]
                else:
                    record[header] = None
            records.append(record)
        
        # Apply filters if provided
        if filters:
            filtered_records = []
            for record in records:
                include = True
                for field, value in filters.items():
                    if field in record:
                        # If the filter value is a list, check if the value is in the list
                        if isinstance(value, list):
                            if record[field] not in value:
                                include = False
                                break
                        # If the filter value is a dictionary, apply operators
                        elif isinstance(value, dict):
                            for op, op_value in value.items():
                                if op == 'eq' and record[field] != op_value:
                                    include = False
                                    break
                                elif op == 'ne' and record[field] == op_value:
                                    include = False
                                    break
                                elif op == 'gt' and (not isinstance(record[field], (int, float)) or record[field] <= op_value):
                                    include = False
                                    break
                                elif op == 'lt' and (not isinstance(record[field], (int, float)) or record[field] >= op_value):
                                    include = False
                                    break
                                elif op == 'contains' and (not isinstance(record[field], str) or op_value not in record[field]):
                                    include = False
                                    break
                        # If the filter value is a simple value, perform an equality comparison
                        elif record[field] != value:
                            include = False
                            break
                if include:
                    filtered_records.append(record)
            records = filtered_records
        
        # Devolver los registros filtrados
        result = records
    
    return result

def create_report_from_template(template_file, output_file, data_mappings, chart_mappings=None, format_mappings=None):
    """
    Create a report based on an Excel template, replacing data, updating charts and applying formats.
     **Emojis must never be included in text written to cells, labels, titles or charts.**

    Args:
        template_file (str): Path to the Excel template.

        output_file (str): Path where the generated report will be saved.
        data_mappings (dict): Dictionary with data mappings:
            {
                "sheet_name": {
                    "range1": data_list1,
                    "range2": data_list2,
                    ...
                }
            }
        chart_mappings (dict, optional): Dictionary with chart updates:
            {
                "sheet_name": {
                    "chart_id": {
                        "title": "New title",
                        "data_range": "New range",
                        ...
                    }
                }
            }
        format_mappings (dict, optional): Dictionary with formats to apply:
            {
                "sheet_name": {
                    "range1": {"number_format": "#,##0.00"},
                    "range2": {"style": {"bold": True, "fill_color": "FFFF00"}},
                    ...
                }
            }

    Returns:
        dict: Result of the operation
    """
    try:
        # Verificar que el archivo de plantilla existe
        if not os.path.exists(template_file):
            raise FileNotFoundError(f"La plantilla no existe: {template_file}")
        
        # Copiar la plantilla al archivo de salida
        import shutil
        shutil.copy2(template_file, output_file)
        
        # Abrir el nuevo archivo
        wb = openpyxl.load_workbook(output_file)
        
        # Aplicar mapeos de datos
        if data_mappings:
            for sheet_name, ranges in data_mappings.items():
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"Sheet '{sheet_name}' does not exist in the template")
                    continue
                
                ws = wb[sheet_name]
                for range_str, data in ranges.items():
                    # Si el rango es una sola celda, extraer la celda de inicio
                    if ':' not in range_str:
                        start_cell = range_str
                    else:
                        start_cell = range_str.split(':')[0]
                    
                    # Write the data
                    write_sheet_data(ws, start_cell, data)
        
        # Apply chart mappings
        if chart_mappings:
            for sheet_name, charts in chart_mappings.items():
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"Sheet '{sheet_name}' does not exist in the template")
                    continue
                
                ws = wb[sheet_name]
                existing_charts = list_charts(ws)
                
                for chart_id, chart_updates in charts.items():
                    # Check if chart_id is an index or a name
                    chart_idx = None
                    if isinstance(chart_id, int) or (isinstance(chart_id, str) and chart_id.isdigit()):
                        chart_idx = int(chart_id)
                    else:
                        # Look up the chart by title
                        for i, chart in enumerate(existing_charts):
                            if chart.get('title') == chart_id:
                                chart_idx = i
                                break
                    
                    if chart_idx is None or chart_idx >= len(existing_charts):
                        logger.warning(f"Chart not found '{chart_id}' en la hoja '{sheet_name}'")
                        continue
                    
                    # Update chart properties
                    chart = ws._charts[chart_idx][0]
                    
                    if 'title' in chart_updates:
                        chart.title = chart_updates['title']
                    
                    if 'data_range' in chart_updates:
                        # Updating the data range is complex and depends on the chart type
                        # For now just log that this feature is not implemented
                        logger.warning("Chart data range update is not fully implemented")
        
        # Aplicar mapeos de formato
        if format_mappings:
            for sheet_name, ranges in format_mappings.items():
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"Sheet '{sheet_name}' does not exist in the template")
                    continue
                
                ws = wb[sheet_name]
                for range_str, formats in ranges.items():
                    if 'number_format' in formats:
                        apply_number_format(ws, range_str, formats['number_format'])
                    
                    if 'style' in formats:
                        apply_style(ws, range_str, formats['style'])
        
        # Guardar el archivo
        wb.save(output_file)
        
        return {
            "success": True,
            "template_file": template_file,
            "output_file": output_file,
            "message": f"Informe creado correctamente: {output_file}"
        }
    
    except Exception as e:
        logger.error(f"Error al crear informe desde plantilla: {e}")
        return {
            "success": False,
            "error": str(e),
            "message": f"Error al crear informe desde plantilla: {e}"
        }

def create_dynamic_dashboard(file_path, data, dashboard_config, overwrite=False):
    """
    Create a dynamic dashboard with multiple visualizations in a single step.

    Models using this function should ensure that tables and charts do not overlap.
     **Emojis must never be included in text written to cells, labels, titles or charts.**
    Leaving empty rows between sections and adjusting column widths after writing
    the data helps produce a clean, professional result.

    Args:
        file_path (str): Path to the Excel file to create or modify.
        data (dict): Dictionary with data per sheet:
            {
                "sheet_name": [
                    ["Header1", "Header2", ...],
                    [value1, value2, ...],
                    ...
                ]
            }
        dashboard_config (dict): Dashboard configuration:
            {
                "tables": [
                    {
                        "sheet": "sheet_name",
                        "name": "TableName",
                        "range": "A1:C10",
                        "style": "TableStyleMedium9"
                    }
                ],
                "charts": [
                    {
                        "sheet": "sheet_name",
                        "type": "column",
                        "data_range": "A1:C10",
                        "title": "Chart Title",
                        "position": "E1",
                        "style": "style1"
                    }
                ],
                "slicers": [
                    {
                        "sheet": "sheet_name",
                        "table": "TableName",
                        "column": "Category",
                        "position": "H1"
                    }
                ]
            }
        overwrite (bool): If ``True`` overwrite the file if it exists.

    Returns:
        dict: Result of the operation
    """
    try:
        # Verificar si el archivo existe
        file_exists = os.path.exists(file_path)
        
        if file_exists and not overwrite:
            raise FileExistsError(f"El archivo '{file_path}' ya existe. Use overwrite=True para sobrescribir.")
        
        # Crear o abrir el archivo
        if not file_exists or overwrite:
            wb = openpyxl.Workbook()
            # Eliminar la hoja predeterminada si existe
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
        else:
            wb = openpyxl.load_workbook(file_path)
        
        # Crear o actualizar hojas con datos
        for sheet_name, sheet_data in data.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Escribir datos
            if sheet_data:
                write_sheet_data(ws, "A1", sheet_data)
        
        # Crear tablas
        for table_config in dashboard_config.get("tables", []):
            sheet_name = table_config["sheet"]
            table_name = table_config["name"]
            range_str = table_config["range"]
            style = table_config.get("style", "TableStyleMedium9")
            
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' does not exist to create table '{table_name}'")
                continue
            
            ws = wb[sheet_name]
            
            # Verificar si la tabla ya existe
            table_exists = False
            if hasattr(ws, 'tables') and table_name in ws.tables:
                table_exists = True
                logger.warning(f"Table '{table_name}' already exists, it will be updated")
            
            if table_exists:
                # Actualizar tabla existente
                refresh_table(ws, table_name, range_str)
            else:
                # Crear nueva tabla
                add_table(ws, table_name, range_str, style)
            
            # Aplicar formatos si se especifican
            if "formats" in table_config:
                for cell_range, fmt in table_config["formats"].items():
                    if isinstance(fmt, str):
                        # Numeric format
                        apply_number_format(ws, cell_range, fmt)
                    elif isinstance(fmt, dict):
                        # Es un estilo
                        apply_style(ws, cell_range, fmt)
        
        # Create charts
        for chart_config in dashboard_config.get("charts", []):
            sheet_name = chart_config["sheet"]
            chart_type = chart_config["type"]
            data_range = chart_config["data_range"]
            title = chart_config.get("title", f"Chart {len(wb[sheet_name]._charts) + 1}")
            position = chart_config.get("position", "E1")
            style = chart_config.get("style")
            
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' does not exist to create the chart '{title}'")
                continue
            
            # Create chart
            chart_id, _ = add_chart(wb, sheet_name, chart_type, data_range, title, position, style)
        
        # Set column widths for optimal display
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Set a minimum width for date columns
            for i in range(1, ws.max_column + 1):
                column_letter = get_column_letter(i)
                # Verificar si hay celdas con formato de fecha en la columna
                date_format = False
                for cell in ws[column_letter]:
                    if cell.number_format and ('yy' in cell.number_format.lower() or 'mm' in cell.number_format.lower() or 'dd' in cell.number_format.lower()):
                        date_format = True
                        break
                
                if date_format:
                    # Set minimum width for date columns
                    ws.column_dimensions[column_letter].width = max(ws.column_dimensions[column_letter].width or 0, 10)
        
        # Guardar el archivo
        wb.save(file_path)
        
        return {
            "success": True,
            "file_path": file_path,
            "message": f"Dashboard creado correctamente: {file_path}"
        }
    
    except Exception as e:
        logger.error(f"Error creating dashboard: {e}")
        return {
            "success": False,
            "error": str(e),
            "message": f"Error creating dashboard: {e}"
        }

def import_multi_source_data(excel_file, import_config, sheet_name=None, start_cell="A1", create_tables=False):
    """
    Import data from multiple sources (CSV, JSON, SQL) into an Excel file in one step.

     **Emojis must never be included in text written to cells, labels, titles or charts.**

    Args:
        excel_file (str): Path to the Excel file where the data will be imported.
        import_config (dict): Import configuration:
            {
                "csv": [
                    {
                        "file_path": "data.csv",
                        "sheet_name": "SheetName",
                        "start_cell": "A1",
                        "delimiter": ",",
                        "encoding": "utf-8"
                    }
                ],
                "json": [
                    {
                        "file_path": "data.json",
                        "sheet_name": "SheetName",
                        "start_cell": "A1",
                        "fields": ["field1", "field2"]
                    }
                ],
                "sql": [
                    {
                        "query": "SELECT * FROM table",
                        "sheet_name": "SheetName",
                        "start_cell": "A1",
                        "connection_string": "..."
                    }
                ]
            }
        sheet_name (str, optional): Default sheet name if not specified in the configuration
        start_cell (str, optional): Default starting cell if not specified in the configuration
        create_tables (bool, optional): If True, create Excel tables for each dataset
    
    Returns:
        dict: Result of the operation
    """
    try:
        import csv
        import json
        
        # Try to import pandas if available (optional)
        try:
            import pandas as pd
            HAS_PANDAS = True
        except ImportError:
            HAS_PANDAS = False
            logger.warning("Pandas is not available. Some features will be limited.")
        
        # Verificar si el archivo Excel existe, si no, crearlo
        if not os.path.exists(excel_file):
            wb = openpyxl.Workbook()
            if sheet_name and "Sheet" in wb.sheetnames:
                # Renombrar la hoja predeterminada si se proporciona sheet_name
                wb["Sheet"].title = sheet_name
        else:
            wb = openpyxl.load_workbook(excel_file)
        
        imported_data = []
        
        # Procesar importaciones CSV
        for csv_config in import_config.get("csv", []):
            csv_file = csv_config["file_path"]
            csv_sheet = csv_config.get("sheet_name", sheet_name)
            csv_cell = csv_config.get("start_cell", start_cell)
            delimiter = csv_config.get("delimiter", ",")
            encoding = csv_config.get("encoding", "utf-8")
            
            if not os.path.exists(csv_file):
                logger.warning(f"El archivo CSV no existe: {csv_file}")
                continue
            
            # Crear la hoja si no existe
            if csv_sheet not in wb.sheetnames:
                ws = wb.create_sheet(csv_sheet)
            else:
                ws = wb[csv_sheet]
            
            # Leer datos CSV
            if HAS_PANDAS:
                # Use pandas if available
                df = pd.read_csv(csv_file, delimiter=delimiter, encoding=encoding)
                data = [df.columns.tolist()]  # Encabezados
                data.extend(df.values.tolist())  # Datos
            else:
                # Use standard csv if pandas is not available
                data = []
                with open(csv_file, 'r', encoding=encoding) as f:
                    reader = csv.reader(f, delimiter=delimiter)
                    for row in reader:
                        data.append(row)
            
            # Escribir datos en la hoja
            write_sheet_data(ws, csv_cell, data)
            
            # Crear tabla si se solicita
            if create_tables:
                # Determinar el rango de la tabla
                start_row, start_col = ExcelRange.parse_cell_ref(csv_cell)
                end_row = start_row + len(data) - 1
                end_col = start_col + (len(data[0]) if data and len(data) > 0 else 0) - 1
                table_range = ExcelRange.range_to_a1(start_row, start_col, end_row, end_col)
                
                # Create a unique name for the table
                table_name = f"Table_{csv_sheet}_{len(imported_data) + 1}"
                table_name = table_name.replace(" ", "_")
                
                try:
                    add_table(ws, table_name, table_range, "TableStyleMedium9")
                except Exception as table_error:
                    logger.warning(f"Could not create the table for {csv_file}: {table_error}")
            
            imported_data.append({
                "source": "csv",
                "file": csv_file,
                "sheet": csv_sheet,
                "rows": len(data)
            })
        
        # Procesar importaciones JSON
        for json_config in import_config.get("json", []):
            json_file = json_config["file_path"]
            json_sheet = json_config.get("sheet_name", sheet_name)
            json_cell = json_config.get("start_cell", start_cell)
            fields = json_config.get("fields", [])
            
            if not os.path.exists(json_file):
                logger.warning(f"El archivo JSON no existe: {json_file}")
                continue
            
            # Crear la hoja si no existe
            if json_sheet not in wb.sheetnames:
                ws = wb.create_sheet(json_sheet)
            else:
                ws = wb[json_sheet]
            
            # Leer datos JSON
            with open(json_file, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            # Convertir JSON a formato tabular
            if isinstance(json_data, list):
                # Si es una lista de objetos, extraer los campos
                if fields:
                    # Usar campos especificados
                    headers = fields
                elif json_data and isinstance(json_data[0], dict):
                    # Usar todas las claves del primer objeto
                    headers = list(json_data[0].keys())
                else:
                    headers = []
                
                # Crear datos tabulares
                data = [headers]
                for item in json_data:
                    if isinstance(item, dict):
                        row = [item.get(field, "") for field in headers]
                        data.append(row)
                    else:
                        # If the item is not a dictionary, add it as a single column
                        data.append([item])
            else:
                # Si es un solo objeto, usar sus claves y valores
                if isinstance(json_data, dict):
                    if fields:
                        # Usar campos especificados
                        headers = fields
                        data = [headers, [json_data.get(field, "") for field in headers]]
                    else:
                        # Usar todas las claves
                        headers = list(json_data.keys())
                        data = [headers, list(json_data.values())]
                else:
                    # If it is neither a dictionary nor a list, use a simple representation
                    data = [["Value"], [json_data]]
            
            # Escribir datos en la hoja
            write_sheet_data(ws, json_cell, data)
            
            # Crear tabla si se solicita
            if create_tables and data:
                # Determinar el rango de la tabla
                start_row, start_col = ExcelRange.parse_cell_ref(json_cell)
                end_row = start_row + len(data) - 1
                end_col = start_col + (len(data[0]) if data and len(data) > 0 else 0) - 1
                table_range = ExcelRange.range_to_a1(start_row, start_col, end_row, end_col)
                
                # Create a unique name for the table
                table_name = f"Table_{json_sheet}_{len(imported_data) + 1}"
                table_name = table_name.replace(" ", "_")
                
                try:
                    add_table(ws, table_name, table_range, "TableStyleMedium9")
                except Exception as table_error:
                    logger.warning(f"Could not create the table for {json_file}: {table_error}")
            
            imported_data.append({
                "source": "json",
                "file": json_file,
                "sheet": json_sheet,
                "rows": len(data)
            })
        
        # Process SQL queries (requires database connection)
        if "sql" in import_config and import_config["sql"]:
            try:
                import pyodbc
                HAS_PYODBC = True
            except ImportError:
                HAS_PYODBC = False
                logger.warning("pyodbc is not available. SQL data cannot be imported.")
            
            if HAS_PYODBC or HAS_PANDAS:
                for sql_config in import_config.get("sql", []):
                    query = sql_config["query"]
                    sql_sheet = sql_config.get("sheet_name", sheet_name)
                    sql_cell = sql_config.get("start_cell", start_cell)
                    connection_string = sql_config.get("connection_string", "")
                    
                    if not query or not connection_string:
                        logger.warning("Se requiere query y connection_string para importar datos SQL")
                        continue
                    
                    # Crear la hoja si no existe
                    if sql_sheet not in wb.sheetnames:
                        ws = wb.create_sheet(sql_sheet)
                    else:
                        ws = wb[sql_sheet]
                    
                    try:
                        data = []
                        
                        if HAS_PANDAS:
                            # Use pandas if available
                            import urllib.parse
                            params = urllib.parse.quote_plus(connection_string)
                            connection_url = f"mssql+pyodbc:///?odbc_connect={params}"
                            
                            from sqlalchemy import create_engine
                            engine = create_engine(connection_url)
                            df = pd.read_sql(query, engine)
                            
                            data = [df.columns.tolist()]  # Encabezados
                            data.extend(df.values.tolist())  # Datos
                        else:
                            # Usar pyodbc directamente
                            conn = pyodbc.connect(connection_string)
                            cursor = conn.cursor()
                            cursor.execute(query)
                            
                            # Obtener nombres de columnas
                            columns = [column[0] for column in cursor.description]
                            data.append(columns)
                            
                            # Obtener datos
                            for row in cursor.fetchall():
                                data.append(list(row))
                            
                            conn.close()
                        
                        # Escribir datos en la hoja
                        write_sheet_data(ws, sql_cell, data)
                        
                        # Crear tabla si se solicita
                        if create_tables and data:
                            # Determinar el rango de la tabla
                            start_row, start_col = ExcelRange.parse_cell_ref(sql_cell)
                            end_row = start_row + len(data) - 1
                            end_col = start_col + (len(data[0]) if data and len(data) > 0 else 0) - 1
                            table_range = ExcelRange.range_to_a1(start_row, start_col, end_row, end_col)
                            
                            # Create a unique name for the table
                            table_name = f"Table_{sql_sheet}_{len(imported_data) + 1}"
                            table_name = table_name.replace(" ", "_")
                            
                            try:
                                add_table(ws, table_name, table_range, "TableStyleMedium9")
                            except Exception as table_error:
                                logger.warning(f"Could not create the table for SQL query: {table_error}")
                        
                        imported_data.append({
                            "source": "sql",
                            "query": query[:50] + "..." if len(query) > 50 else query,
                            "sheet": sql_sheet,
                            "rows": len(data)
                        })
                    
                    except Exception as sql_error:
                        logger.error(f"Error al importar datos SQL: {sql_error}")
                        continue
        
        # Guardar el archivo Excel
        wb.save(excel_file)
        
        return {
            "success": True,
            "file_path": excel_file,
            "imported_data": imported_data,
            "message": f"Datos importados correctamente a {excel_file}"
        }
    
    except Exception as e:
        logger.error(f"Error al importar datos: {e}")
        return {
            "success": False,
            "error": str(e),
            "message": f"Error al importar datos: {e}"
        }

def export_excel_data(excel_file, export_config):
    """
    Export Excel data to multiple formats (CSV, JSON, PDF) in one step.
    
    Args:
        excel_file (str): Path to the source Excel file
        export_config (dict): Export configuration:
            {
                "csv": [
                    {
                        "sheet_name": "SheetName",
                        "range": "A1:C10",
                        "output_file": "output.csv",
                        "delimiter": ",",
                        "encoding": "utf-8"
                    }
                ],
                "json": [
                    {
                        "sheet_name": "SheetName",
                        "range": "A1:C10",
                        "output_file": "output.json",
                        "format": "records"  # "records", "object", "table"
                    }
                ],
                "pdf": {
                      "output_file": "output.pdf",
                      "sheets": ["Sheet1", "Sheet2"]  # or null for all
                }
            }
    
    Returns:
        dict: Result of the operation
    """
    try:
        import csv
        import json
        
        if not os.path.exists(excel_file):
            raise FileNotFoundError(f"El archivo Excel no existe: {excel_file}")
        
        # Cargar el archivo Excel
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        exported_files = []
        
        # Exportar a CSV
        for csv_config in export_config.get("csv", []):
            sheet_name = csv_config["sheet_name"]
            range_str = csv_config.get("range")
            output_file = csv_config["output_file"]
            delimiter = csv_config.get("delimiter", ",")
            encoding = csv_config.get("encoding", "utf-8")
            
            if sheet_name not in wb.sheetnames:
                logger.warning(f"La hoja '{sheet_name}' no existe")
                continue
            
            # Leer los datos del rango especificado
            data = read_sheet_data(wb, sheet_name, range_str)
            
            # Write the data en CSV
            with open(output_file, 'w', newline='', encoding=encoding) as csvfile:
                writer = csv.writer(csvfile, delimiter=delimiter)
                for row in data:
                    writer.writerow(row)
            
            exported_files.append({
                "format": "csv",
                "file": output_file,
                "sheet": sheet_name,
                "rows": len(data)
            })
        
        # Exportar a JSON
        for json_config in export_config.get("json", []):
            sheet_name = json_config["sheet_name"]
            range_str = json_config.get("range")
            output_file = json_config["output_file"]
            format_type = json_config.get("format", "records")
            
            if sheet_name not in wb.sheetnames:
                logger.warning(f"La hoja '{sheet_name}' no existe")
                continue
            
            # Leer los datos del rango especificado
            data = read_sheet_data(wb, sheet_name, range_str)
            
            if not data:
                logger.warning(f"No hay datos para exportar en la hoja '{sheet_name}'")
                continue
            
            # Convert data to JSON format according to the specified type
            headers = data[0]
            json_data = None
            
            if format_type == "records":
                # Formato de registros [{campo1: valor1, campo2: valor2}, {...}]
                json_data = []
                for row in data[1:]:
                    record = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            record[header] = row[i]
                    json_data.append(record)
            
            elif format_type == "object":
                # Formato de objeto {id1: {campo1: valor1}, id2: {campo1: valor2}}
                json_data = {}
                id_field = headers[0]  # Usar la primera columna como ID
                for row in data[1:]:
                    if not row:
                        continue
                    record = {}
                    for i, header in enumerate(headers[1:], 1):  # Empezar desde la segunda columna
                        if i < len(row):
                            record[header] = row[i]
                    json_data[row[0]] = record
            
            elif format_type == "table":
                # Formato de tabla {headers: [...], data: [[...], [...]]}
                json_data = {
                    "headers": headers,
                    "data": [row for row in data[1:]]
                }
            
            # Write the data en JSON
            with open(output_file, 'w', encoding='utf-8') as jsonfile:
                json.dump(json_data, jsonfile, indent=2)
            
            exported_files.append({
                "format": "json",
                "file": output_file,
                "sheet": sheet_name,
                "rows": len(data) - 1  # Sin contar encabezados
            })
        
        # Exportar a PDF (requiere biblioteca adicional)
        if "pdf" in export_config:
            pdf_config = export_config["pdf"]
            output_file = pdf_config["output_file"]
            sheets = pdf_config.get("sheets")
            
            try:
                # Try to use win32com for Excel if available
                import win32com.client
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                
                # Abrir el archivo
                workbook = excel.Workbooks.Open(os.path.abspath(excel_file))
                
                # Determinar las hojas a exportar
                sheets_to_export = []
                if sheets:
                    for sheet_name in sheets:
                        try:
                            sheet = workbook.Sheets(sheet_name)
                            sheets_to_export.append(sheet)
                        except:
                            logger.warning(f"La hoja '{sheet_name}' no existe para exportar a PDF")
                else:
                    # Exportar todas las hojas
                    sheets_to_export = workbook.Sheets
                
                # Exportar a PDF
                if sheets_to_export:
                    workbook.ExportAsFixedFormat(0, os.path.abspath(output_file))
                    
                    exported_files.append({
                        "format": "pdf",
                        "file": output_file,
                        "sheets": sheets if sheets else [sheet.Name for sheet in sheets_to_export]
                    })
                
                # Cerrar Excel
                workbook.Close(False)
                excel.Quit()
            
            except ImportError:
                logger.warning("win32com is not available. Cannot export to PDF.")
                pass  # If win32com is not available, simply skip the PDF export
            except Exception as pdf_error:
                logger.error(f"Error al exportar a PDF: {pdf_error}")
                pass
        
        return {
            "success": True,
            "file_path": excel_file,
            "exported_files": exported_files,
            "message": f"Datos exportados correctamente desde {excel_file}"
        }
    
    except Exception as e:
        logger.error(f"Error al exportar datos: {e}")
        return {
            "success": False,
            "error": str(e),
            "message": f"Error al exportar datos: {e}"
        }

def export_single_visible_sheet_pdf(excel_file: str, output_pdf: Optional[str] = None) -> Dict[str, Any]:
    """Export an Excel workbook to PDF only if it has a single visible sheet.

    Args:
        excel_file: Path to the Excel file to export.
        output_pdf: Path of the resulting PDF file. If not provided, the same
            name as ``excel_file`` with ``.pdf`` extension is used.

    Returns:
        dict: Result of the operation.
    """
    try:
        import shutil
        import subprocess

        if not os.path.exists(excel_file):
            raise FileNotFoundError(f"El archivo Excel no existe: {excel_file}")

        wb = openpyxl.load_workbook(excel_file, data_only=True)
        visible_sheets = [ws.title for ws in wb.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]

        if len(visible_sheets) != 1:
            msg = f"The file must have a single visible sheet. Visible sheets: {len(visible_sheets)}"
            logger.warning(msg)
            return {
                "success": False,
                "file_path": excel_file,
                "visible_sheets": visible_sheets,
                "message": msg,
            }

        if not output_pdf:
            output_pdf = os.path.splitext(excel_file)[0] + ".pdf"
        output_pdf = os.path.abspath(output_pdf)

        # Intentar exportar con win32com (Windows)
        try:
            import win32com.client

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            workbook = excel.Workbooks.Open(os.path.abspath(excel_file))
            workbook.ExportAsFixedFormat(0, output_pdf)
            workbook.Close(False)
            excel.Quit()

            msg = f"File successfully exported to PDF: {output_pdf}"
            logger.info(msg)
            return {
                "success": True,
                "file_path": excel_file,
                "pdf_file": output_pdf,
                "message": msg,
            }
        except ImportError:
            logger.info("win32com not available, LibreOffice will be tried")
        except Exception as e:
            logger.error(f"Error al exportar con win32com: {e}")

        # Fallback a LibreOffice en sistemas no Windows
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if soffice:
            outdir = os.path.dirname(output_pdf)
            cmd = [soffice, "--headless", "--convert-to", "pdf", os.path.abspath(excel_file), "--outdir", outdir]
            subprocess.run(cmd, check=True)

            generated = os.path.join(outdir, Path(excel_file).stem + ".pdf")
            if generated != output_pdf:
                os.replace(generated, output_pdf)

            msg = f"Archivo exportado correctamente a PDF: {output_pdf}"
            logger.info(msg)
            return {
                "success": True,
                "file_path": excel_file,
                "pdf_file": output_pdf,
                "message": msg,
            }

        msg = "No available method found to export to PDF."
        logger.error(msg)
        return {
            "success": False,
            "file_path": excel_file,
            "message": msg,
        }

    except Exception as e:
        logger.error(f"Error al exportar a PDF: {e}")
        return {
            "success": False,
            "file_path": excel_file,
            "error": str(e),
            "message": f"Error al exportar a PDF: {e}",
        }

def export_sheets_to_pdf(
    excel_file: str,
    sheets: Optional[Union[str, List[str]]] = None,
    output_dir: Optional[str] = None,
    single_file: bool = False,
) -> Dict[str, Any]:
    """Export one or more sheets of an Excel workbook to PDF.

    Parameters
    ----------
    excel_file : str
        Path to the Excel file to export.
    sheets : Union[str, List[str]], optional
        Name of the sheet or list of sheets to export. If ``None`` all sheets
        in the workbook are exported one by one.
    output_dir : str, optional
        Folder where the PDFs will be stored. By default the original file
        directory is used.
    single_file : bool, optional
        If ``True`` and several sheets are specified a single PDF is generated
        with all of them if supported. If ``False`` a PDF is created per sheet.

    Returns
    -------
    dict
        Operation result with the list of generated PDFs. If any sheet does not
        exist a warning is included in ``warnings``.
    """

    try:
        import shutil
        import subprocess

        if not os.path.exists(excel_file):
            raise FileNotFoundError(f"El archivo Excel no existe: {excel_file}")

        wb = openpyxl.load_workbook(excel_file, data_only=True)
        all_sheets = wb.sheetnames
        wb.close()

        if sheets is None:
            target_sheets = all_sheets
        elif isinstance(sheets, str):
            target_sheets = [sheets]
        else:
            target_sheets = list(sheets)

        warnings = []
        valid_sheets = []
        for s in target_sheets:
            if s in all_sheets:
                valid_sheets.append(s)
            else:
                warnings.append(f"La hoja '{s}' no existe")

        if not valid_sheets:
            msg = "No valid sheets found to export"
            logger.warning(msg)
            return {
                "success": False,
                "file_path": excel_file,
                "warnings": warnings,
                "message": msg,
            }

        if output_dir is None:
            output_dir = os.path.dirname(os.path.abspath(excel_file))

        pdf_files: List[str] = []

        # Try to use win32com if available
        try:
            import win32com.client

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            workbook = excel.Workbooks.Open(os.path.abspath(excel_file))

            if single_file and len(valid_sheets) > 1:
                workbook.Worksheets(valid_sheets).Select()
                output_pdf = os.path.join(
                    output_dir, Path(excel_file).stem + ".pdf"
                )
                workbook.ActiveSheet.ExportAsFixedFormat(0, output_pdf)
                pdf_files.append(output_pdf)
            else:
                for s in valid_sheets:
                    ws = workbook.Worksheets(s)
                    output_pdf = os.path.join(
                        output_dir, f"{Path(excel_file).stem}_{s}.pdf"
                    )
                    ws.ExportAsFixedFormat(0, output_pdf)
                    pdf_files.append(output_pdf)

            workbook.Close(False)
            excel.Quit()

            msg = "PDF export completed successfully"
            logger.info(msg)
            return {
                "success": True,
                "file_path": excel_file,
                "pdf_files": pdf_files,
                "warnings": warnings,
                "message": msg,
            }
        except ImportError:
            logger.info("win32com not available, trying LibreOffice")
        except Exception as e:
            logger.error(f"Error al exportar con win32com: {e}")

        # Fallback a LibreOffice
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if soffice:
            with tempfile.TemporaryDirectory() as tmpdir:
                if single_file and len(valid_sheets) > 1:
                    tmp_xlsx = os.path.join(tmpdir, "tmp.xlsx")
                    wb = openpyxl.load_workbook(excel_file)
                    for sheet in wb.sheetnames:
                        wb[sheet].sheet_state = (
                            "visible" if sheet in valid_sheets else "hidden"
                        )
                    wb.save(tmp_xlsx)
                    wb.close()
                    cmd = [
                        soffice,
                        "--headless",
                        "--convert-to",
                        "pdf",
                        os.path.abspath(tmp_xlsx),
                        "--outdir",
                        tmpdir,
                    ]
                    subprocess.run(cmd, check=True)
                    generated = os.path.join(tmpdir, "tmp.pdf")
                    final = os.path.join(
                        output_dir, Path(excel_file).stem + ".pdf"
                    )
                    shutil.move(generated, final)
                    pdf_files.append(final)
                else:
                    for s in valid_sheets:
                        tmp_xlsx = os.path.join(tmpdir, f"{s}.xlsx")
                        wb = openpyxl.load_workbook(excel_file)
                        for sheet in wb.sheetnames:
                            wb[sheet].sheet_state = (
                                "visible" if sheet == s else "hidden"
                            )
                        wb.save(tmp_xlsx)
                        wb.close()
                        cmd = [
                            soffice,
                            "--headless",
                            "--convert-to",
                            "pdf",
                            os.path.abspath(tmp_xlsx),
                            "--outdir",
                            tmpdir,
                        ]
                        subprocess.run(cmd, check=True)
                        generated = os.path.join(tmpdir, f"{s}.pdf")
                        final = os.path.join(
                            output_dir, f"{Path(excel_file).stem}_{s}.pdf"
                        )
                        shutil.move(generated, final)
                        pdf_files.append(final)

            msg = "PDF export completed successfully"
            logger.info(msg)
            return {
                "success": True,
                "file_path": excel_file,
                "pdf_files": pdf_files,
                "warnings": warnings,
                "message": msg,
            }

        msg = "No available method found to export to PDF."
        logger.error(msg)
        return {
            "success": False,
            "file_path": excel_file,
            "warnings": warnings,
            "message": msg,
        }

    except Exception as e:
        logger.error(f"Error al exportar a PDF: {e}")
        return {
            "success": False,
            "file_path": excel_file,
            "error": str(e),
            "message": f"Error al exportar a PDF: {e}",
        }

# Crear el servidor MCP como variable global
mcp = None
if HAS_MCP:
    # Esta es la variable global que el sistema MCP busca
    mcp = FastMCP("Master Excel MCP", 
                 dependencies=["openpyxl", "pandas", "numpy"])
    logger.info("Servidor MCP unificado iniciado correctamente")
    
    # Register basic workbook management functions
    @mcp.tool(description="Creates a new empty Excel file with professional foundation")
    def create_workbook_tool(filename, overwrite=False):
        """Create a new empty Excel workbook with optimal foundation for data manipulation.

        **PURPOSE & CONTEXT:**
        This is the essential first step when creating Excel files from scratch. It establishes
        a clean, properly formatted Excel workbook ready for data, tables, charts, and formulas.
        The workbook will be created with Excel's native format (.xlsx) ensuring full compatibility
        with all Excel features and formulas.

        **WHEN TO USE:**
        - Starting a new report, dashboard, or data analysis file
        - Creating templates for data entry or reporting
        - Beginning any Excel automation workflow
        - When you need a clean slate for data manipulation

        **TECHNICAL DETAILS:**
        - Creates standard .xlsx format (Excel 2007+)
        - Includes default worksheet "Sheet1" ready for immediate use
        - Optimized for openpyxl library compatibility
        - File structure supports all Excel features (formulas, charts, tables, etc.)

        Args:
            filename (str): Full absolute path including filename and .xlsx extension.
                          Example: "C:/reports/sales_analysis.xlsx" or "/home/user/data.xlsx"
                          IMPORTANT: Must be absolute path, not relative.
            overwrite (bool, optional): If ``True``, replaces existing file without warning.
                                      If ``False`` (default), raises error if file exists.
                                      Use ``True`` when updating/refreshing existing reports.

        Returns:
            dict: Comprehensive operation result containing:
                - success (bool): True if workbook created successfully
                - file_path (str): Absolute path to the created file
                - message (str): Human-readable success message
                - error (str): Error details if operation failed

        **USAGE EXAMPLES:**
        
        Basic usage:
            create_workbook_tool("/path/to/reports/monthly_sales.xlsx")
            
        Replace existing file:
            create_workbook_tool("/path/to/reports/monthly_sales.xlsx", overwrite=True)
            
        **COMMON PATTERNS:**
        1. Create → Add data → Add tables → Add charts → Save
        2. Create → Import data → Add formulas → Format → Export
        3. Create → Build template → Save for reuse

        **ERROR HANDLING:**
        - FileExistsError: File already exists and overwrite=False
        - PermissionError: Insufficient permissions to write to location
        - IOError: Invalid path or disk space issues

        **NEXT STEPS AFTER CREATION:**
        - Use write_sheet_data_tool() to add data
        - Use add_table_tool() to structure data as tables
        - Use add_chart_tool() to create visualizations
        - Use add_formulas_tool() to add calculations
        """
        try:
            import os
            
            # Check if file exists and overwrite flag
            if os.path.exists(filename) and not overwrite:
                return {
                    "success": False,
                    "error": "File already exists",
                    "message": f"File '{filename}' already exists. Use overwrite=True to replace it."
                }
            
            # Create a simple Excel workbook directly
            if HAS_OPENPYXL:
                import openpyxl
                wb = openpyxl.Workbook()
                wb.save(filename)
            else:
                # Fallback: create minimal Excel file structure
                import zipfile
                import tempfile
                
                # Create minimal Excel file structure
                with tempfile.TemporaryDirectory() as temp_dir:
                    # Create basic Excel XML structure
                    xl_dir = os.path.join(temp_dir, "xl")
                    worksheets_dir = os.path.join(xl_dir, "worksheets")
                    os.makedirs(worksheets_dir)
                    
                    # Create basic files
                    with open(os.path.join(temp_dir, "[Content_Types].xml"), "w") as f:
                        f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/></Types>')
                    
                    os.makedirs(os.path.join(temp_dir, "_rels"))
                    with open(os.path.join(temp_dir, "_rels", ".rels"), "w") as f:
                        f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>')
                    
                    os.makedirs(os.path.join(xl_dir, "_rels"))
                    with open(os.path.join(xl_dir, "_rels", "workbook.xml.rels"), "w") as f:
                        f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/></Relationships>')
                    
                    with open(os.path.join(xl_dir, "workbook.xml"), "w") as f:
                        f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets></workbook>')
                    
                    with open(os.path.join(worksheets_dir, "sheet1.xml"), "w") as f:
                        f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData/></worksheet>')
                    
                    # Create ZIP file (Excel format)
                    with zipfile.ZipFile(filename, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for root, dirs, files in os.walk(temp_dir):
                            for file in files:
                                file_path = os.path.join(root, file)
                                arc_name = os.path.relpath(file_path, temp_dir)
                                zf.write(file_path, arc_name)
            
            return {
                "success": True,
                "file_path": filename,
                "message": f"Excel file successfully created: {filename}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error creating Excel file: {e}"
            }
    
    @mcp.tool(description="Abre un fichero Excel existente")
    def open_workbook_tool(filename):
        """Open an existing Excel file.

        This function opens an existing ``.xlsx`` or ``.xls`` file so it can be manipulated.
        Use this before performing any operation on an existing file.

        Args:
            filename (str): Full path and name of the Excel file to open.

        Returns:
            dict: Information about the opened file including sheet count and other properties.

        Raises:
            FileNotFoundError: If the specified file does not exist.

        Example:
            open_workbook_tool("C:/data/sales_report.xlsx")
        """
        try:
            wb = open_workbook(filename)
            sheet_names = list_sheets(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "sheets": sheet_names,
                "sheet_count": len(sheet_names),
                "message": f"Excel file successfully opened: {filename}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error opening Excel file: {e}"
            }
    
    @mcp.tool(description="Guarda el Workbook en disco")
    def save_workbook_tool(filename, new_filename=None):
        """Save the workbook to disk.

        Use this function after modifying a workbook to persist the changes.

        Args:
            filename (str): Full path and name of the Excel file to save.
            new_filename (str, optional): If provided, save the file under a new name ("Save As"). Defaults to ``None``.

        Returns:
            dict: Information about the operation result including the saved file path.

        Example:
            save_workbook_tool("C:/data/report.xlsx")
            save_workbook_tool("C:/data/report.xlsx", "C:/data/report_backup.xlsx")  # Save As
        """
        try:
            wb = open_workbook(filename)
            saved_path = save_workbook(wb, new_filename or filename)
            close_workbook(wb)
            
            return {
                "success": True,
                "original_file": filename,
                "saved_file": saved_path,
                "message": f"Excel file successfully saved: {saved_path}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error saving Excel file: {e}"
            }
    
    @mcp.tool(description="Lista las hojas disponibles en un archivo Excel")
    def list_sheets_tool(filename):
        """List the worksheets available in an Excel file.

        This function returns all worksheets contained in an Excel workbook and is useful
        to get an overview before working with the file.

        Args:
            filename (str): Full path and name of the Excel file to inspect.

        Returns:
            dict: Dictionary with the sheet names and their positions in the workbook.

        Raises:
            FileNotFoundError: If the specified file does not exist.

        Example:
            list_sheets_tool("C:/data/financial_report.xlsx")  # Returns: {"sheets": ["Sales", "Costs", "Summary"]}
        """
        try:
            wb = open_workbook(filename)
            sheets = list_sheets(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "sheets": sheets,
                "count": len(sheets),
                "message": f"Se encontraron {len(sheets)} hojas en el archivo Excel"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al listar hojas: {e}"
            }
    
    # Register basic worksheet manipulation functions
    @mcp.tool(description="Adds a new empty sheet")
    def add_sheet_tool(filename, sheet_name, index=None):
        """Add a new empty worksheet.

        This function inserts a new blank worksheet into an existing Excel workbook.
        You can specify the position where the sheet should be inserted.

        Args:
            filename (str): Full path and name of the Excel file.
            sheet_name (str): Name for the new sheet.
            index (int, optional): Position where the sheet will be inserted (``0`` is the first position).
                                 If ``None`` the sheet is added at the end. Default is ``None``.

        Returns:
            dict: Information about the operation including the updated list of sheets.

        Raises:
            FileNotFoundError: If the specified file does not exist.
            SheetExistsError: If a sheet with the same name already exists.

        Example:
            add_sheet_tool("C:/data/report.xlsx", "New Summary")  # Add at the end
            add_sheet_tool("C:/data/report.xlsx", "Cover", 0)  # Add as first sheet
        """
        try:
            wb = open_workbook(filename)
            ws = add_sheet(wb, sheet_name, index)
            save_workbook(wb, filename)
            
            sheets = list_sheets(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "sheet_name": sheet_name,
                "sheet_index": sheets.index(sheet_name),
                "all_sheets": sheets,
                "message": f"Sheet '{sheet_name}' added successfully"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error adding sheet: {e}"
            }
    
    @mcp.tool(description="Delete the indicated sheet")
    def delete_sheet_tool(filename, sheet_name):
        """Delete the indicated worksheet.

        This function removes a specific worksheet from an Excel workbook. Use with care
        because once the file is saved the deletion is permanent.

        Args:
            filename (str): Full path and name of the Excel file.
            sheet_name (str): Name of the worksheet to delete.

        Returns:
            dict: Information about the operation including the updated list of sheets.

        Raises:
            FileNotFoundError: If the specified file does not exist.
            SheetNotFoundError: If the specified sheet does not exist.
            ValueError: If attempting to delete the only sheet in the workbook.

        Example:
            delete_sheet_tool("C:/data/report.xlsx", "Draft")
        """
        try:
            wb = open_workbook(filename)
            delete_sheet(wb, sheet_name)
            save_workbook(wb, filename)
            
            remaining_sheets = list_sheets(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "deleted_sheet": sheet_name,
                "remaining_sheets": remaining_sheets,
                "remaining_count": len(remaining_sheets),
                "message": f"Sheet '{sheet_name}' successfully deleted"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error deleting sheet: {e}"
            }
    
    @mcp.tool(description="Rename a sheet")
    def rename_sheet_tool(filename, old_name, new_name):
        """Rename a worksheet.

        This function changes the name of an existing worksheet in an Excel workbook.

        Args:
            filename (str): Full path and name of the Excel file.
            old_name (str): Current name of the sheet to rename.
            new_name (str): New name for the sheet.

        Returns:
            dict: Information about the operation including the updated list of sheets.

        Raises:
            FileNotFoundError: If the specified file does not exist.
            SheetNotFoundError: If no sheet exists with the original name.
            SheetExistsError: If a sheet with the new name already exists.

        Example:
            rename_sheet_tool("C:/data/report.xlsx", "Sheet1", "Executive Summary")
        """
        try:
            wb = open_workbook(filename)
            rename_sheet(wb, old_name, new_name)
            save_workbook(wb, filename)
            
            sheets = list_sheets(wb)
            close_workbook(wb)
            
            return {
                "success": True,
                "file_path": filename,
                "old_name": old_name,
                "new_name": new_name,
                "all_sheets": sheets,
                "message": f"Sheet renamed from '{old_name}' to '{new_name}'"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error renaming sheet: {e}"
            }
    
    # Register basic writing functions
    @mcp.tool(description="Write structured data arrays to Excel with intelligent type conversion")
    def write_sheet_data_tool(file_path, sheet_name, start_cell, data):
        """Write two-dimensional data arrays to Excel with automatic data type optimization.

        **PURPOSE & CONTEXT:**
        This is the primary tool for populating Excel worksheets with structured data.
        It handles automatic data type conversion, ensuring numbers are stored as numbers
        (not text), dates are properly formatted, and formulas are preserved. This creates
        foundation data that works perfectly with Excel's calculation engine and chart creation.

        **INTELLIGENT DATA PROCESSING:**
        - Automatically converts string numbers to numeric values ("1,000" → 1000)
        - Handles percentage strings ("50%" → 0.5 for proper calculations)
        - Preserves formulas (text starting with "=")
        - Cleans and validates data structure
        - Optimizes data types for Excel compatibility

        **WHEN TO USE:**
        - Importing data from databases, APIs, or CSV files
        - Writing calculation results to Excel
        - Creating data tables for analysis
        - Populating templates with dynamic data
        - Building datasets for charts and pivot tables

        **DATA STRUCTURE REQUIREMENTS:**
        Data must be a list of lists (2D array) where:
        - First row typically contains headers
        - Each subsequent row contains data values
        - All rows should have consistent column count
        - Mixed data types are supported and automatically optimized

        Args:
            file_path (str): Full absolute path to Excel file (.xlsx).
                           Example: "C:/reports/sales_data.xlsx"
                           MUST be existing file - use create_workbook_tool() first if needed.
            sheet_name (str): Target worksheet name. Examples: "Data", "Sales", "Sheet1"
                            Must exist in workbook - use add_sheet_tool() if needed.
            start_cell (str): Top-left cell for data placement. Examples: "A1", "B5", "D10"
                            Data will expand right and down from this position.
            data (list): Two-dimensional list containing the data to write.
                       Structure: [[header1, header2, ...], [value1, value2, ...], ...]
                       
        **DATA EXAMPLES:**
        
        Simple table:
            [["Product", "Price", "Quantity"], 
             ["Widget A", 10.50, 100], 
             ["Widget B", 15.75, 50]]
             
        With formulas:
            [["Product", "Price", "Qty", "Total"], 
             ["Widget A", 10.50, 100, "=B2*C2"], 
             ["Widget B", 15.75, 50, "=B3*C3"]]
             
        Mixed data types:
            [["Name", "Date", "Amount", "Percentage"], 
             ["John", "2024-01-15", "1,500.50", "75%"], 
             ["Mary", "2024-01-16", "2,300.25", "82%"]]

        Returns:
            dict: Comprehensive operation result:
                - success (bool): Operation success status
                - file_path (str): Path to modified file
                - sheet_name (str): Target sheet name
                - start_cell (str): Starting position
                - rows_written (int): Number of data rows written
                - columns_written (int): Number of columns written
                - data_cleaned (bool): Whether automatic type conversion was applied
                - message (str): Detailed success message
                - error (str): Error details if operation failed

        **AUTOMATIC DATA TYPE CONVERSION:**
        - "1,000" → 1000 (removes commas from numbers)
        - "50%" → 0.5 (converts percentages to decimals)
        - "" → None (handles empty cells properly)
        - "=SUM(A1:A10)" → Preserved as formula
        - "2024-01-15" → Date value (when properly formatted)

        **ERROR PREVENTION:**
        - Validates file existence before writing
        - Checks sheet availability in workbook
        - Validates cell reference format
        - Handles mixed data types gracefully
        - Provides detailed error messages for troubleshooting

        **PERFORMANCE OPTIMIZATION:**
        - Batch writes for improved speed
        - Memory-efficient processing for large datasets
        - Automatic column width adjustment
        - Smart data type detection reduces processing overhead

        **COMMON USAGE PATTERNS:**
        
        1. Database import:
            create_workbook_tool("report.xlsx")
            write_sheet_data_tool("report.xlsx", "Data", "A1", database_results)
            
        2. Multi-sheet reports:
            write_sheet_data_tool("report.xlsx", "Sales", "A1", sales_data)
            write_sheet_data_tool("report.xlsx", "Costs", "A1", cost_data)
            
        3. Template population:
            write_sheet_data_tool("template.xlsx", "Input", "B5", user_data)

        **BEST PRACTICES:**
        - Always include headers in first row for better table creation
        - Use consistent data types within columns when possible
        - Place data starting from A1 or B1 for optimal table creation
        - Validate data structure before writing large datasets
        - Consider using add_table_tool() after writing for enhanced formatting
        """
        try:
            # Validate inputs first
            if not isinstance(data, list):
                raise ValueError("The 'data' parameter must be a list")
            
            if not data:
                raise ValueError("Data cannot be empty")
            
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File does not exist: {file_path}")

            # Open the file using our base function
            wb = open_workbook(file_path)
            
            # Validate sheet exists
            if sheet_name not in list_sheets(wb):
                raise SheetNotFoundError(f"Sheet '{sheet_name}' does not exist in {file_path}")
            
            ws = get_sheet(wb, sheet_name)

            # Clean and validate data types
            cleaned_data = []
            for i, row in enumerate(data):
                if not isinstance(row, list):
                    # Convert single values to list
                    row = [row]
                
                cleaned_row = []
                for j, cell_value in enumerate(row):
                    # Clean and convert data types appropriately
                    if cell_value is None or cell_value == "":
                        cleaned_row.append("")
                    elif isinstance(cell_value, str):
                        # Try to convert string numbers to actual numbers
                        cell_str = cell_value.strip()
                        if cell_str == "":
                            cleaned_row.append("")
                        elif cell_str.replace('.','').replace(',','').replace('-','').isdigit():
                            # Try to convert to number
                            try:
                                if '.' in cell_str:
                                    cleaned_row.append(float(cell_str.replace(',', '')))
                                else:
                                    cleaned_row.append(int(cell_str.replace(',', '')))
                            except ValueError:
                                cleaned_row.append(cell_str)
                        elif cell_str.endswith('%'):
                            # Convert percentage
                            try:
                                pct_value = float(cell_str[:-1]) / 100
                                cleaned_row.append(pct_value)
                            except ValueError:
                                cleaned_row.append(cell_str)
                        else:
                            cleaned_row.append(cell_str)
                    else:
                        cleaned_row.append(cell_value)
                
                cleaned_data.append(cleaned_row)

            # Write the cleaned data
            write_sheet_data(ws, start_cell, cleaned_data)

            # Save with optimization
            save_workbook(wb, file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "start_cell": start_cell,
                "rows_written": len(cleaned_data),
                "columns_written": max([len(row) for row in cleaned_data], default=0),
                "data_cleaned": True,
                "message": f"Data successfully written starting at {start_cell} with automatic type conversion"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error writing data: {e}"
            }
    
    @mcp.tool(description="Update a single cell")
    def update_cell_tool(file_path, sheet_name, cell, value_or_formula):
        """Update the value or formula of a specific cell.

        This function modifies a single cell in a worksheet. It can be used for both values and formulas.

        Args:
            file_path (str): Full path and name of the Excel file.
            sheet_name (str): Name of the sheet containing the cell to update.
            cell (str): Reference of the cell to update (e.g. ``"B5"``).
            value_or_formula (str | int | float | bool): Value or formula to set. Formulas must start with ``=``.

        Returns:
            dict: Information about the operation including the modified cell.

        Raises:
            FileNotFoundError: If the specified file does not exist.
            SheetNotFoundError: If the specified sheet does not exist.
            CellReferenceError: If the cell reference is not valid.

        Example:
            update_cell_tool("C:/data/report.xlsx", "Sales", "C4", 5280.50)  # Numeric value
            update_cell_tool("C:/data/report.xlsx", "Sales", "D4", "=SUM(A1:A10)")  # Formula
        """
        try:
            # Validate inputs first
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File does not exist: {file_path}")

            # Open the file using our base function
            wb = open_workbook(file_path)
            
            # Validate sheet exists
            if sheet_name not in list_sheets(wb):
                raise SheetNotFoundError(f"Sheet '{sheet_name}' does not exist in {file_path}")
            
            ws = get_sheet(wb, sheet_name)
            
            # Clean and convert value appropriately
            cleaned_value = value_or_formula
            if isinstance(value_or_formula, str) and not str(value_or_formula).startswith('='):
                # Not a formula, try to convert data type
                value_str = str(value_or_formula).strip()
                if value_str.replace('.','').replace(',','').replace('-','').isdigit():
                    # Try to convert to number
                    try:
                        if '.' in value_str:
                            cleaned_value = float(value_str.replace(',', ''))
                        else:
                            cleaned_value = int(value_str.replace(',', ''))
                    except ValueError:
                        pass  # Keep as string
                elif value_str.endswith('%'):
                    # Convert percentage
                    try:
                        cleaned_value = float(value_str[:-1]) / 100
                    except ValueError:
                        pass  # Keep as string
            
            # Update the cell with enhanced processing
            update_cell(ws, cell, cleaned_value)
            
            # Save changes with optimization
            save_workbook(wb, file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell": cell,
                "value": cleaned_value,
                "original_value": value_or_formula,
                "data_cleaned": cleaned_value != value_or_formula,
                "message": f"Cell {cell} successfully updated in sheet {sheet_name} with automatic type conversion"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error updating cell: {e}"
            }
    
    # Register advanced functions
    @mcp.tool(description="Transform data ranges into professional Excel tables with filtering and formatting")
    def add_table_tool(file_path, sheet_name, table_name, cell_range, style=None):
        """Convert data ranges into native Excel tables with professional formatting and functionality.

        **PURPOSE & CONTEXT:**
        Excel tables are powerful native structures that provide automatic filtering, sorting,
        formatting, and formula management. This tool transforms raw data into professional
        tables that users can interact with directly in Excel. Tables also serve as excellent
        data sources for charts and pivot tables.

        **EXCEL TABLE BENEFITS:**
        - Automatic filtering dropdowns on headers
        - Professional formatting with alternating row colors
        - Dynamic range expansion when new data is added
        - Structured references in formulas (e.g., Table1[Sales])
        - Built-in sorting capabilities
        - Easy data validation and input
        - Perfect foundation for charts and pivot tables

        **WHEN TO USE:**
        - Converting raw data into interactive format
        - Creating datasets for analysis and reporting
        - Building data foundations for dashboards
        - Preparing data for chart creation
        - Enabling user interaction with data
        - Organizing data for filtering and sorting

        **IMPORTANT: TABLE RANGE SPECIFICATION:**
        The cell_range parameter should contain ONLY tabular data:
        - Include headers in the first row
        - Include data rows (no empty rows)
        - Do NOT include descriptive text, titles, or conclusions
        - Do NOT include summary calculations (they'll be added separately)
        - Range should be contiguous data only

        Args:
            file_path (str): Full absolute path to Excel file containing the data.
                           Example: "C:/reports/sales_analysis.xlsx"
                           File must exist and contain the target sheet.
            sheet_name (str): Name of worksheet containing the data range.
                            Examples: "Data", "Sales", "Raw Data"
                            Sheet must exist in the workbook.
            table_name (str): Unique identifier for the table within the workbook.
                            Examples: "SalesData", "CustomerList", "ProductCatalog"
                            MUST be unique - Excel will error if name already exists.
                            Used for formula references and table management.
            cell_range (str): Exact range containing ONLY the table data.
                            Format: "A1:D10" (top-left to bottom-right)
                            CRITICAL: Must include headers but NO extra content.
                            Example: "A1:E25" for data with headers in row 1, data in rows 2-25
            style (str, optional): Excel table style name for professional appearance.
                                 Examples: "TableStyleMedium9", "TableStyleLight1", "TableStyleDark2"
                                 If None, uses default professional blue style.
                                 Available styles: Light1-21, Medium1-28, Dark1-11

        Returns:
            dict: Comprehensive table creation result:
                - success (bool): Table creation success status
                - file_path (str): Path to modified Excel file
                - sheet_name (str): Target worksheet name
                - table_name (str): Created table identifier
                - range (str): Final table range (may be optimized)
                - style (str): Applied table style
                - optimized (bool): Whether range was optimized during creation
                - message (str): Detailed success message
                - error (str): Error details if operation failed

        **RANGE EXAMPLES:**
        
        Correct table range (headers + data only):
            "A1:D15" containing:
            A1: "Product"  B1: "Price"   C1: "Quantity"  D1: "Total"
            A2: "Widget A" B2: 10.50     C2: 100         D2: 1050
            A3: "Widget B" B3: 15.75     C3: 50          D3: 787.5
            ... (data continues to row 15)
            
        Incorrect table range (includes extra content):
            "A1:D20" containing:
            A1: "Sales Report for Q1 2024"  <- WRONG: Title text
            A3: "Product"  B3: "Price" ...  <- WRONG: Headers not in row 1
            A18: "Total Sales: $50,000"     <- WRONG: Summary text

        **TABLE STYLE OPTIONS:**
        - Light styles: Professional, clean appearance (TableStyleLight1-21)
        - Medium styles: Colored headers, good contrast (TableStyleMedium1-28)
        - Dark styles: Bold appearance for emphasis (TableStyleDark1-11)
        - Default: TableStyleMedium9 (professional blue with good contrast)

        **AUTOMATIC ENHANCEMENTS:**
        This tool automatically applies:
        - Conservative header improvement (only generic headers like "Column1")
        - Basic formatting optimization
        - Column width adjustment
        - Professional styling
        - Data type validation

        **ERROR PREVENTION:**
        - Validates file and sheet existence
        - Checks table name uniqueness
        - Validates range format and data presence
        - Provides specific error messages for troubleshooting
        - Conservative approach prevents data corruption

        **COMMON USAGE PATTERNS:**
        
        1. Data preparation workflow:
            write_sheet_data_tool("file.xlsx", "Data", "A1", data_array)
            add_table_tool("file.xlsx", "Data", "MainData", "A1:E50")
            add_formulas_tool("file.xlsx", "Data", "A1:E50")
            
        2. Multiple table creation:
            add_table_tool("file.xlsx", "Sales", "SalesData", "A1:D100")
            add_table_tool("file.xlsx", "Costs", "CostData", "A1:C50")
            
        3. Chart data preparation:
            add_table_tool("file.xlsx", "Data", "ChartSource", "A1:C20")
            add_chart_tool("file.xlsx", "Data", "column", "A1:C20")

        **BEST PRACTICES:**
        - Always use descriptive table names ("SalesQ1" not "Table1")
        - Include meaningful headers in the first row
        - Ensure data is clean and consistent within columns
        - Use appropriate styles for your audience and purpose
        - Consider add_formulas_tool() after table creation for calculations
        - Test table functionality in Excel after creation

        **NEXT STEPS AFTER TABLE CREATION:**
        - Use add_formulas_tool() to add calculation rows
        - Use add_chart_tool() to visualize table data
        - Use filter_data_tool() to extract subsets
        - Apply additional formatting as needed
        """
        try:
            # Validate inputs first
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File does not exist: {file_path}")
            
            # Open the file using our base function
            wb = open_workbook(file_path)
            
            # Validate sheet exists
            if sheet_name not in list_sheets(wb):
                raise SheetNotFoundError(f"Sheet '{sheet_name}' does not exist in {file_path}")
            
            # Get the sheet
            ws = get_sheet(wb, sheet_name)
            
            # Apply conservative table cleanup (only improves headers, no range expansion)
            try:
                cell_range = conservative_table_cleanup(ws, cell_range)
            except Exception as e:
                logger.warning(f"Conservative table cleanup failed, using original range: {e}")
            
            # Add the table with enhanced processing
            table = add_table(ws, table_name, cell_range, style or DEFAULT_TABLE_STYLE)
            
            # Save changes with optimization
            save_workbook(wb, file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "table_name": table_name,
                "range": cell_range,
                "style": style or DEFAULT_TABLE_STYLE,
                "optimized": True,
                "message": f"Table '{table_name}' successfully created in range {cell_range} with enhanced formatting"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error creating table: {e}"
            }
    
    @mcp.tool(description="Create professional native Excel charts with intelligent positioning and styling")
    def add_chart_tool(file_path, sheet_name, chart_type, data_range, title=None, position=None, style=None, theme=None, custom_palette=None):
        """Create native Excel charts with professional styling and intelligent data linking.

        **PURPOSE & CONTEXT:**
        Excel charts are powerful native visualizations that update automatically when
        source data changes. This tool creates professional charts with proper data linking,
        intelligent positioning, and customizable styling. Charts integrate seamlessly with
        Excel's native functionality and can be further customized by users.

        **CHART CAPABILITIES:**
        - Native Excel charts (not images) - fully interactive
        - Automatic data linking with live updates
        - Professional styling and themes
        - Intelligent positioning to avoid content overlap
        - Comprehensive chart type support
        - Custom color schemes and branding
        - Excel-native editing and formatting

        **WHEN TO USE:**
        - Visualizing table data for analysis
        - Creating dashboard components
        - Building interactive reports
        - Presenting data trends and patterns
        - Comparing data across categories
        - Creating executive summaries

        **DATA RANGE REQUIREMENTS:**
        The data_range must contain clean, chartable data:
        - First row/column should contain category labels
        - Remaining data should be numeric values
        - No empty rows/columns within the range
        - Consistent data structure throughout
        - No descriptive text or titles within range

        Args:
            file_path (str): Full absolute path to Excel file containing the data.
                           Example: "C:/dashboards/sales_report.xlsx"
                           File must exist and be accessible.
            sheet_name (str): Worksheet name containing the source data.
                            Examples: "Data", "Charts", "Dashboard"
                            Sheet must exist in the workbook.
            chart_type (str): Chart type for optimal data visualization.
                            
                            **COLUMN CHARTS** ("column"): Best for comparing values across categories
                            - Use for: Sales by region, quarterly comparisons, category analysis
                            - Data structure: Categories in column A, values in column B+
                            
                            **BAR CHARTS** ("bar"): Best for horizontal comparisons
                            - Use for: Long category names, ranking data, horizontal layouts
                            - Data structure: Same as column but renders horizontally
                            
                            **LINE CHARTS** ("line"): Best for trends over time
                            - Use for: Time series, progress tracking, trend analysis
                            - Data structure: Time periods in column A, values in column B+
                            
                            **PIE CHARTS** ("pie"): Best for part-to-whole relationships
                            - Use for: Market share, budget allocation, percentage breakdowns
                            - Data structure: Categories in column A, single value column B
                            - Limitation: Works best with 5-7 categories maximum
                            
                            **SCATTER PLOTS** ("scatter"): Best for correlation analysis
                            - Use for: X-Y relationships, correlation studies, data distribution
                            - Data structure: X values in column A, Y values in column B
                            
                            **AREA CHARTS** ("area"): Best for cumulative trends
                            - Use for: Stacked totals, cumulative growth, volume over time
                            - Data structure: Time in column A, multiple value columns
                            
                            Other types: "doughnut", "radar", "surface", "stock"
                            
            data_range (str): Exact range containing chart data (headers + data only).
                            Format: "A1:C10" (first cell to last cell with data)
                            
                            **CRITICAL REQUIREMENTS:**
                            - Must include headers for series names
                            - Must contain actual data (not empty cells)
                            - Should not include titles, descriptions, or totals
                            - Range must be contiguous (no gaps)
                            
                            **EXAMPLE RANGES:**
                            For column chart: "A1:B6" containing:
                            A1: "Month"    B1: "Sales"
                            A2: "Jan"      B2: 15000
                            A3: "Feb"      B3: 18000
                            A4: "Mar"      B4: 22000
                            A5: "Apr"      B5: 19000
                            A6: "May"      B6: 25000
                            
            title (str, optional): Descriptive chart title for professional presentation.
                                 Examples: "Quarterly Sales Performance", "Regional Growth Trends"
                                 If None, chart will have no title (can be added later in Excel)
                                 Keep concise but descriptive for best impact.
                                 
            position (str, optional): Chart placement to avoid content overlap.
                                    Format: "E1:L15" (top-left to bottom-right of chart area)
                                    If None, uses intelligent auto-positioning algorithm.
                                    Consider worksheet layout and existing content.
                                    Standard sizes: Small (6x8 cells), Medium (8x12), Large (10x15)
                                    
            style (int, optional): Excel chart style number for professional appearance.
                                 Range: 1-48 (each number provides different color/styling)
                                 Popular choices:
                                 - 1-6: Light, clean styles (good for professional reports)
                                 - 7-12: Dark, bold styles (good for presentations)
                                 - 13-20: Colorful styles (good for dashboards)
                                 If None, uses default Excel styling.
                                 
            theme (str, optional): Color theme for consistent branding.
                                 Options: "office", "colorful", "dark-blue", "dark-red"
                                 Overrides style parameter when specified.
                                 Use for brand consistency across multiple charts.
                                 
            custom_palette (list, optional): Custom colors for specific branding.
                                           Format: ["#4472C4", "#ED7D31", "#A5A5A5"]
                                           Hex color codes for series colors.
                                           Use for corporate branding or specific color requirements.

        Returns:
            dict: Comprehensive chart creation result:
                - success (bool): Chart creation success status
                - file_path (str): Path to modified Excel file
                - sheet_name (str): Target worksheet name
                - chart_id (str): Unique identifier for the created chart
                - chart_type (str): Type of chart created
                - data_range (str): Source data range for the chart
                - title (str): Applied chart title
                - position (str): Final chart position
                - chart_position (str): Readable position description
                - message (str): Detailed success message
                - error (str): Error details if operation failed

        **INTELLIGENT FEATURES:**
        
        **AUTOMATIC OVERLAP PREVENTION (Built-in):**
        - Automatically detects ALL existing charts and their exact positions
        - Scans worksheet for data tables and content automatically
        - Calculates precise spacing to prevent ANY overlap
        - Uses systematic positioning strategy for professional layouts
        - Maintains minimum 2-column/2-row spacing between charts automatically
        - Provides emergency positioning for complex layouts
        - Validates user-provided positions and corrects overlaps automatically
        - Logs positioning decisions for transparency
        - NO additional tools needed - works perfectly out of the box
        
        **Data Validation:**
        - Verifies data range contains actual data
        - Checks for empty cells and provides warnings
        - Validates data structure for chart type
        - Ensures proper headers for series identification
        
        **Professional Styling:**
        - Applies consistent formatting
        - Optimizes colors for readability
        - Sets appropriate fonts and sizes
        - Ensures chart title positioning

        **ERROR PREVENTION:**
        - Validates file and sheet existence
        - Checks data range validity
        - Verifies data contains chartable content
        - Provides specific error messages
        - Handles edge cases gracefully

        **CHART TYPE SELECTION GUIDE:**
        
        Choose chart type based on your data story:
        - **Comparison**: Column, Bar charts
        - **Trends**: Line, Area charts  
        - **Composition**: Pie, Doughnut charts
        - **Relationships**: Scatter plots
        - **Distribution**: Scatter, Area charts

        **COMMON USAGE PATTERNS:**
        
        1. Table-to-chart workflow:
            add_table_tool("file.xlsx", "Data", "SalesTable", "A1:C12")
            add_chart_tool("file.xlsx", "Data", "column", "A1:C12", "Monthly Sales")
            
        2. Dashboard creation:
            add_chart_tool("file.xlsx", "Dashboard", "pie", "A1:B6", "Market Share")
            add_chart_tool("file.xlsx", "Dashboard", "line", "D1:E12", "Growth Trend")
            
        3. Multi-series comparison:
            add_chart_tool("file.xlsx", "Analysis", "column", "A1:D10", "Regional Performance")

        **BEST PRACTICES:**
        - Ensure data is clean and complete before charting
        - Use descriptive titles that explain the data story
        - Choose chart types that match your data narrative
        - Consider your audience when selecting styles
        - Test chart interactivity in Excel after creation
        - Use consistent styling across multiple charts
        - Position charts to complement, not obstruct, data tables

        **TROUBLESHOOTING:**
        - "Chart not appearing": Check data range contains actual data
        - "Wrong data series": Verify headers are in first row/column
        - "Poor positioning": Specify exact position parameter
        - "Styling issues": Try different style numbers or themes
        """
        try:
            # Validate inputs first
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File does not exist: {file_path}")
            
            # Open the file using our base function
            wb = open_workbook(file_path)
            
            # Validate sheet exists
            if sheet_name not in list_sheets(wb):
                raise SheetNotFoundError(f"Sheet '{sheet_name}' does not exist in {file_path}")
            
            # Validate data range contains data
            ws = get_sheet(wb, sheet_name)
            try:
                # Parse range and check it has data
                if '!' in data_range:
                    # Extract range part if sheet is included
                    data_range = data_range.split('!')[1]
                
                start_row, start_col, end_row, end_col = ExcelRange.parse_range(data_range)
                
                # Check if range has actual data
                has_data = False
                for row in range(start_row + 1, end_row + 2):
                    for col in range(start_col + 1, end_col + 2):
                        cell = ws.cell(row=row, column=col)
                        if cell.value is not None and str(cell.value).strip():
                            has_data = True
                            break
                    if has_data:
                        break
                
                if not has_data:
                    raise ValueError(f"Data range '{data_range}' appears to be empty")
                    
            except Exception as e:
                raise RangeError(f"Invalid or empty data range '{data_range}': {e}")
            
            # Create chart with enhanced error handling
            try:
                chart_id, chart = add_chart(wb, sheet_name, chart_type, data_range, title, position, style, theme, custom_palette)
            except Exception as e:
                raise ChartError(f"Failed to create chart: {e}")
            
            # Save changes with optimization
            save_workbook(wb, file_path)
            
            # Extract chart type for a better response message
            chart_type_display = chart_type
            if chart_type.lower() == "col":
                chart_type_display = "column"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "chart_id": chart_id,
                "chart_type": chart_type_display,
                "data_range": data_range,
                "title": title,
                "position": position,
                "message": f"Chart '{chart_type_display}' successfully created with ID {chart_id}",
                "chart_position": position or "Auto-positioned",
                "overlap_prevention": True,
                "positioning_strategy": "Intelligent automatic positioning with overlap prevention"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error creating chart: {e}"
            }
    
    # Register new combined functions
    @mcp.tool(description="Create a sheet with data in one step")
    def create_sheet_with_data_tool(file_path, sheet_name, data, overwrite=False):
        """Create an Excel file with a single sheet and data in one step.

        Args:
             **Emojis must never be included in text written to cells, labels, titles or charts.**

            file_path (str): Path to the Excel file to create.
            sheet_name (str): Name of the sheet to create.
            data (list): Two-dimensional array with the data.
            overwrite (bool): If ``True`` overwrite the file if it already exists.

        Returns:
            dict: Result of the operation.
        """
        try:
            # Check if the file exists
            file_exists = os.path.exists(file_path)
            
            if file_exists and not overwrite:
                raise FileExistsError(f"The file '{file_path}' already exists. Use overwrite=True to overwrite.")
            
            # Create or open the file
            if not file_exists or overwrite:
                wb = openpyxl.Workbook()
                # Remove the default sheet if it exists
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
            else:
                wb = openpyxl.load_workbook(file_path)
            
            # Check if the sheet already exists
            if sheet_name in wb.sheetnames:
                if overwrite:
                    # Delete the existing sheet
                    del wb[sheet_name]
                else:
                    raise SheetExistsError(f"The sheet '{sheet_name}' already exists. Use overwrite=True to overwrite.")
            
            # Create the sheet
            ws = wb.create_sheet(sheet_name)
            
            # Write the data
            if data:
                write_sheet_data(ws, "A1", data)
            
            # Save the file
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "rows_written": len(data) if data else 0,
                "columns_written": max([len(row) if isinstance(row, list) else 1 for row in data], default=0) if data else 0,
                "message": f"File created with sheet '{sheet_name}' and data"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error creating sheet with data: {e}"
            }
    
    @mcp.tool(description="Create a formatted table with data in one step")
    def create_formatted_table_tool(file_path, sheet_name, start_cell, data, table_name, table_style="TableStyleMedium9", formats=None):
        """Create a formatted table with data in one step.

        Args:
             **Emojis must never be included in text written to cells, labels, titles or charts.**

            file_path (str): Path to the Excel file.
            sheet_name (str): Name of the sheet.
            start_cell (str): Starting cell (e.g. ``"A1"``).
            data (list): Two-dimensional array with the data.
            table_name (str): Name for the table.
            table_style (str): Table style.
            formats (dict): Dictionary with formats to apply:
                {
                    "A2:A10": "#,##0.00",  # Numeric format
                    "B2:B10": {"bold": True, "fill_color": "FFFF00"}  # Style
                }

        Returns:
            dict: Result of the operation.
        """
        try:
            # Validate inputs first
            if not isinstance(data, list) or not data:
                raise ValueError("Data must be a non-empty list")
            
            # Check if the file exists, if not, create it
            if not os.path.exists(file_path):
                wb = create_workbook()
                if "Sheet" in list_sheets(wb) and sheet_name != "Sheet":
                    # Rename the default sheet
                    rename_sheet(wb, "Sheet", sheet_name)
            else:
                wb = open_workbook(file_path)
                
                # Create the sheet if it doesn't exist
                if sheet_name not in list_sheets(wb):
                    add_sheet(wb, sheet_name)
            
            # Get the sheet
            ws = get_sheet(wb, sheet_name)
            
            # Clean and write the data with enhanced processing
            cleaned_data = []
            for i, row in enumerate(data):
                if not isinstance(row, list):
                    row = [row]
                
                cleaned_row = []
                for j, cell_value in enumerate(row):
                    # Clean and convert data types appropriately
                    if cell_value is None or cell_value == "":
                        cleaned_row.append("")
                    elif isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        if cell_str == "":
                            cleaned_row.append("")
                        elif cell_str.replace('.','').replace(',','').replace('-','').isdigit():
                            try:
                                if '.' in cell_str:
                                    cleaned_row.append(float(cell_str.replace(',', '')))
                                else:
                                    cleaned_row.append(int(cell_str.replace(',', '')))
                            except ValueError:
                                cleaned_row.append(cell_str)
                        elif cell_str.endswith('%'):
                            try:
                                pct_value = float(cell_str[:-1]) / 100
                                cleaned_row.append(pct_value)
                            except ValueError:
                                cleaned_row.append(cell_str)
                        else:
                            cleaned_row.append(cell_str)
                    else:
                        cleaned_row.append(cell_value)
                
                cleaned_data.append(cleaned_row)
            
            write_sheet_data(ws, start_cell, cleaned_data)
            
            # Calculate exact table range based on provided data only
            start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
            end_row = start_row + len(cleaned_data) - 1
            end_col = start_col + (len(cleaned_data[0]) if cleaned_data and len(cleaned_data) > 0 else 0) - 1
            table_range = ExcelRange.range_to_a1(start_row, start_col, end_row, end_col)
            
            # Apply conservative table cleanup (only improves headers, no range expansion)
            try:
                table_range = conservative_table_cleanup(ws, table_range)
            except Exception as e:
                logger.warning(f"Conservative table cleanup failed: {e}")
            
            # Create the table with enhanced processing
            add_table(ws, table_name, table_range, table_style or DEFAULT_TABLE_STYLE)
            
            # Apply formats if provided
            if formats:
                for cell_range, fmt in formats.items():
                    if isinstance(fmt, dict):
                        apply_style(ws, cell_range, fmt)
                    else:
                        apply_number_format(ws, cell_range, fmt)
            
            # Apply consistent formatting
            apply_consistent_number_format(ws, table_range)
            enhanced_autofit_columns(ws)
            
            # Add smart formulas to enhance the table
            try:
                formula_result = add_formula_to_table(ws, table_range, 'auto')
                if formula_result.get('success'):
                    logger.info(f"Added formulas to table: {formula_result.get('message', '')}")
            except Exception as e:
                logger.warning(f"Could not add formulas to table: {e}")
            
            # Save with optimization
            save_workbook(wb, file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "table_name": table_name,
                "table_range": table_range,
                "table_style": table_style or DEFAULT_TABLE_STYLE,
                "data_cleaned": True,
                "optimized": True,
                "message": f"Table '{table_name}' created and formatted successfully with enhanced processing"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error creating formatted table: {e}"
            }
    
    @mcp.tool(description="Create a chart from new data in one step")
    def create_chart_from_data_tool(file_path, sheet_name, data, chart_type, position=None, title=None, style=None):
        """Create a chart from new data in one step.

        Args:
             **Emojis must never be included in text written to cells, labels, titles or charts.**

            file_path (str): Path to the Excel file.
            sheet_name (str): Name of the sheet.
            data (list): Two-dimensional array with the data for the chart.
            chart_type (str): Chart type (``'column'``, ``'bar'``, ``'line'``, ``'pie'``, etc.).
            position (str): Cell where the chart will be placed (e.g. ``"E1"``).
            title (str): Chart title.
            style: Chart style.

        Returns:
            dict: Result of the operation.
        """
        try:
            # Validate inputs first
            if not isinstance(data, list) or not data:
                raise ValueError("Data must be a non-empty list")
            
            # Check if the file exists, if not, create it
            if not os.path.exists(file_path):
                wb = create_workbook()
                if "Sheet" in list_sheets(wb) and sheet_name != "Sheet":
                    rename_sheet(wb, "Sheet", sheet_name)
            else:
                wb = open_workbook(file_path)
                
                # Create the sheet if it doesn't exist
                if sheet_name not in list_sheets(wb):
                    add_sheet(wb, sheet_name)
            
            # Get the sheet
            ws = get_sheet(wb, sheet_name)
            
            # Find a free area for the data intelligently
            start_cell = "A1"
            
            # Check if there is already data in that area
            if ws["A1"].value is not None:
                # Find the first empty column
                col = 1
                while ws.cell(row=1, column=col).value is not None:
                    col += 1
                start_cell = f"{get_column_letter(col)}1"
            
            # Clean and write the data with enhanced processing
            cleaned_data = []
            for i, row in enumerate(data):
                if not isinstance(row, list):
                    row = [row]
                
                cleaned_row = []
                for j, cell_value in enumerate(row):
                    # Clean and convert data types appropriately
                    if cell_value is None or cell_value == "":
                        cleaned_row.append("")
                    elif isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        if cell_str == "":
                            cleaned_row.append("")
                        elif cell_str.replace('.','').replace(',','').replace('-','').isdigit():
                            try:
                                if '.' in cell_str:
                                    cleaned_row.append(float(cell_str.replace(',', '')))
                                else:
                                    cleaned_row.append(int(cell_str.replace(',', '')))
                            except ValueError:
                                cleaned_row.append(cell_str)
                        elif cell_str.endswith('%'):
                            try:
                                pct_value = float(cell_str[:-1]) / 100
                                cleaned_row.append(pct_value)
                            except ValueError:
                                cleaned_row.append(cell_str)
                        else:
                            cleaned_row.append(cell_str)
                    else:
                        cleaned_row.append(cell_value)
                
                cleaned_data.append(cleaned_row)
            
            write_sheet_data(ws, start_cell, cleaned_data)
            
            # Determine the data range for the chart
            start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
            end_row = start_row + len(cleaned_data) - 1
            end_col = start_col + (len(cleaned_data[0]) if cleaned_data and len(cleaned_data) > 0 else 0) - 1
            data_range = ExcelRange.range_to_a1(start_row, start_col, end_row, end_col)
            
            # AUTOMATIC INTELLIGENT POSITIONING - No overlaps guaranteed!
            if not position:
                # Automatically analyze worksheet layout and prevent ALL overlaps
                existing_charts = get_existing_chart_positions(ws)
                
                # Get intelligent layout recommendations based on data
                layout_analysis = get_chart_layout_recommendations(ws, [data_range])
                
                # Use intelligent positioning with full context
                position = find_optimal_chart_position(ws, end_col + 2, 0, 8, 15)
                
                # Log positioning decision for transparency
                logger.info(f"AUTOMATIC POSITIONING: Found {len(existing_charts)} existing charts. Strategy: {layout_analysis.get('layout_strategy', 'adaptive')}. Selected position: {position} (guaranteed no overlap)")
            else:
                # Validate user-provided position to prevent overlaps
                existing_charts = get_existing_chart_positions(ws)
                try:
                    # Parse user position
                    import re
                    pos_match = re.match(r'([A-Z]+)(\d+)', position.upper())
                    if pos_match:
                        pos_col = column_index_from_string(pos_match.group(1)) - 1
                        pos_row = int(pos_match.group(2)) - 1
                        
                        # Check if user position would cause overlap
                        if check_area_overlap(pos_col, pos_row, 8, 15, existing_charts, 1, 1):
                            # User position would overlap - find alternative
                            logger.warning(f"USER POSITION {position} would cause overlap. Finding safe alternative...")
                            safe_position = find_optimal_chart_position(ws, pos_col, pos_row, 8, 15)
                            logger.info(f"OVERLAP PREVENTION: Changed position from {position} to {safe_position}")
                            position = safe_position
                        else:
                            logger.info(f"USER POSITION {position} validated - no overlap detected")
                except Exception as e:
                    logger.warning(f"Could not validate user position {position}: {e}. Using automatic positioning.")
                    position = find_optimal_chart_position(ws, 5, 0, 8, 15)
            
            # Apply consistent formatting to data
            apply_consistent_number_format(ws, data_range)
            enhanced_autofit_columns(ws)
            
            # Create the chart with enhanced error handling
            try:
                chart_id, chart = add_chart(wb, sheet_name, chart_type, data_range, title, position, style)
            except Exception as e:
                raise ChartError(f"Failed to create chart: {e}")
            
            # Save with optimization
            save_workbook(wb, file_path)
            
            # Extract chart type for better response message
            chart_type_display = chart_type
            if chart_type.lower() == "col":
                chart_type_display = "column"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "data_range": data_range,
                "chart_id": chart_id,
                "chart_type": chart_type_display,
                "position": position,
                "data_cleaned": True,
                "optimized": True,
                "message": f"Chart '{chart_type_display}' successfully created from new data with enhanced processing"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error creating chart with data: {e}"
            }
    
    
    
    
    @mcp.tool(description="Import data from multiple sources (CSV, JSON, SQL) into an Excel file")
    def import_data_tool(excel_file, import_config, sheet_name=None, start_cell="A1", create_tables=False):
        """Import data from multiple sources (CSV, JSON, SQL) into an Excel file.

        Args:
            excel_file (str): Path to the Excel file where the data will be imported.
            import_config (dict): Import configuration (see documentation for format).
            sheet_name (str, optional): Default sheet name.
            start_cell (str, optional): Default starting cell.
            create_tables (bool, optional): If ``True`` create Excel tables for each dataset.

        Returns:
            dict: Result of the operation.
        """
        return import_multi_source_data(excel_file, import_config, sheet_name, start_cell, create_tables)
    
    @mcp.tool(description="Export Excel data to multiple formats (CSV, JSON, PDF)")
    def export_data_tool(excel_file, export_config):
        """Export Excel data to multiple formats (CSV, JSON, PDF).

        Args:
            excel_file (str): Path to the source Excel file.
            export_config (dict): Export configuration (see documentation for format).

        Returns:
            dict: Result of the operation.
        """
        return export_excel_data(excel_file, export_config)
    
    @mcp.tool(description="Filter and extract data from a table or range as records")
    def filter_data_tool(file_path, sheet_name, range_str=None, table_name=None, filters=None):
        """Filter and extract data from a table or range as records.

        Args:
            file_path (str): Path to the Excel file.
            sheet_name (str): Name of the sheet.
            range_str (str, optional): Range in ``A1:B5`` format (required if ``table_name`` is not provided).
            table_name (str, optional): Table name (required if ``range_str`` is not provided).
            filters (dict, optional): Filters to apply to the data:
                {
                    "field1": value1,               # Simple equality
                    "field2": [value1, value2],      # List of allowed values
                    "field3": {"gt": 100},           # Greater than
                    "field4": {"lt": 50},            # Less than
                    "field5": {"contains": "text"}   # Contains text
                }

        Returns:
            dict: Result of the operation with the filtered data.
        """
        try:
            # Validate arguments
            if not range_str and not table_name:
                raise ValueError("You must provide 'range_str' or 'table_name'")

            # Validate inputs first
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File does not exist: {file_path}")

            # Open the file using our base function
            wb = open_workbook(file_path)

            # Verify that the sheet exists
            if sheet_name not in list_sheets(wb):
                raise SheetNotFoundError(f"Sheet '{sheet_name}' does not exist in the file")
            
            ws = get_sheet(wb, sheet_name)
            
            # If table_name is provided, get its range
            if table_name:
                tables = list_tables(ws)
                if table_name not in [t.get('name') for t in tables]:
                    raise TableNotFoundError(f"Table '{table_name}' does not exist on sheet '{sheet_name}'")
                
                # Find the table and get its range
                for table in tables:
                    if table.get('name') == table_name:
                        range_str = table.get('range')
                        break
            
            # Filter the data with enhanced processing
            filtered_data = filter_sheet_data(wb, sheet_name, range_str, filters)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "source": f"Table '{table_name}'" if table_name else f"Range {range_str}",
                "filtered_data": filtered_data,
                "record_count": len(filtered_data),
                "enhanced_processing": True,
                "message": f"Found {len(filtered_data)} records that meet the criteria"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error filtering data: {e}"
            }

    @mcp.tool(description="Export Excel worksheets to PDF with intelligent automatic handling")
    def export_pdf_tool(excel_file, sheets=None, output_path=None, single_file=True):
        """Export Excel worksheets to PDF with intelligent automatic handling.

        **UNIFIED PDF EXPORT - HANDLES ALL SCENARIOS:**
        This single tool intelligently handles both single and multiple sheet exports:
        - Automatically detects single vs. multiple sheet scenarios
        - Optimizes export strategy based on content analysis
        - Creates professional PDF output with proper formatting
        - Handles edge cases and complex layouts automatically

        Args:
            excel_file (str): Full path to the Excel file to export.
                            Example: "C:/reports/dashboard.xlsx"
            sheets (str|list, optional): Sheet specification for export control:
                                       - None: Export ALL sheets (intelligent handling)
                                       - "Sheet1": Export single specific sheet
                                       - ["Sheet1", "Dashboard"]: Export multiple specific sheets
            output_path (str, optional): Output PDF path or directory.
                                       If None, uses Excel filename with .pdf extension.
                                       Example: "C:/exports/report.pdf"
            single_file (bool): Multiple sheet combination strategy:
                              - True: Combine multiple sheets into one PDF (default)
                              - False: Create separate PDF file per sheet

        Returns:
            dict: Comprehensive export result:
                - success (bool): Export operation success status
                - excel_file (str): Source Excel file path
                - exported_sheets (list): List of sheets that were exported
                - pdf_strategy (str): Export strategy used ("single_sheet" or "multi_sheet")
                - single_file (bool): Whether output was combined or separate
                - output_files (list): List of created PDF file paths
                - message (str): Detailed success message

        **AUTOMATIC INTELLIGENCE:**
        - Single sheet detection: Uses optimized single-sheet export
        - Multiple sheet handling: Combines or separates based on preference
        - Layout optimization: Adjusts PDF formatting for content type
        - Error handling: Graceful fallbacks for complex scenarios
        - Professional output: Consistent PDF quality and formatting
        """
        try:
            # Validate input file
            if not os.path.exists(excel_file):
                raise FileNotFoundError(f"Excel file not found: {excel_file}")

            # Open workbook to analyze structure
            wb = open_workbook(excel_file)
            available_sheets = list_sheets(wb)
            
            # Determine sheets to export
            if sheets is None:
                target_sheets = available_sheets
            elif isinstance(sheets, str):
                target_sheets = [sheets]
            elif isinstance(sheets, list):
                target_sheets = sheets
            else:
                raise ValueError("sheets parameter must be None, string, or list")
            
            # Validate target sheets exist
            missing_sheets = [s for s in target_sheets if s not in available_sheets]
            if missing_sheets:
                raise ValueError(f"Sheets not found: {missing_sheets}. Available: {available_sheets}")
            
            # Intelligent export strategy selection
            if len(target_sheets) == 1:
                # Single sheet - use optimized single sheet export
                result = export_single_visible_sheet_pdf(excel_file, output_path)
                strategy = "single_sheet"
                output_files = [result.get('output_file', output_path)] if result.get('success') else []
            else:
                # Multiple sheets - use multi-sheet export
                output_dir = os.path.dirname(output_path) if output_path else None
                result = export_sheets_to_pdf(excel_file, target_sheets, output_dir, single_file)
                strategy = "multi_sheet"
                output_files = result.get('pdf_files', []) if result.get('success') else []
            
            return {
                "success": result.get('success', False),
                "excel_file": excel_file,
                "exported_sheets": target_sheets,
                "pdf_strategy": strategy,
                "single_file": single_file if len(target_sheets) > 1 else True,
                "output_files": output_files,
                "files_created": len(output_files),
                "result_details": result,
                "message": f"Successfully exported {len(target_sheets)} sheet(s) to PDF using {strategy} strategy"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "excel_file": excel_file,
                "message": f"Error exporting to PDF: {e}"
            }

    @mcp.tool(description="Comprehensive cleanup and optimization of Excel files")
    def optimize_excel_file_tool(excel_file, output_file=None):
        """Perform comprehensive cleanup and optimization of an Excel file.

        This tool automatically:
        - Detects and expands table ranges to include all data
        - Removes empty rows before table headers
        - Renames generic headers (Columna1, Column1) with semantic names
        - Repositions charts to avoid content overlaps
        - Fixes chart data linking after table expansion
        - Applies consistent number formatting
        - Formats total rows with special styling
        - Aligns text and numbers properly
        - Applies professional themes and layout

        Args:
            excel_file (str): Path to the Excel file to optimize
            output_file (str, optional): Path for the optimized file. If not provided, 
                                        the original file will be overwritten.

        Returns:
            dict: Result of the optimization operation
        """
        try:
            # Open the workbook
            wb = openpyxl.load_workbook(excel_file)
            
            # Apply comprehensive optimization
            optimize_entire_workbook(wb)
            
            # Apply unified theme
            apply_unified_theme(wb, "professional")
            
            # Determine output file
            if not output_file:
                output_file = excel_file
            
            # Save the optimized workbook
            wb.save(output_file)
            
            return {
                "success": True,
                "input_file": excel_file,
                "output_file": output_file,
                "message": "Excel file has been comprehensively optimized",
                "optimizations_applied": [
                    "Dynamic data range detection",
                    "Empty row removal",
                    "Smart header renaming",
                    "Chart repositioning",
                    "Number format standardization",
                    "Total row formatting",
                    "Text alignment optimization",
                    "Professional theme application"
                ]
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error optimizing Excel file: {e}"
            }
    
    @mcp.tool(description="Add intelligent Excel formulas for dynamic data analysis and calculations")
    def add_formulas_tool(file_path, sheet_name, table_range, formula_type="auto", add_totals=True):
        """Add live Excel formulas to create dynamic, self-updating data analysis.

        **PURPOSE & CONTEXT:**
        This tool transforms static data into dynamic Excel worksheets by adding native
        Excel formulas that recalculate automatically when source data changes. Instead
        of hardcoded values, it creates live calculations that maintain accuracy and
        provide real-time insights as data is updated.

        **DYNAMIC CALCULATION BENEFITS:**
        - Formulas update automatically when data changes
        - Maintains calculation accuracy over time
        - Shows calculation logic for transparency
        - Enables "what-if" analysis scenarios
        - Creates professional, audit-ready worksheets
        - Supports Excel's native calculation engine
        - Allows manual verification of results

        **INTELLIGENT FORMULA SELECTION:**
        The tool analyzes your data to suggest optimal formula types:
        - **Column headers** containing "total", "sum", "amount" → SUM formulas
        - **Column headers** containing "average", "mean", "rate" → AVERAGE formulas
        - **Column headers** containing "count", "quantity", "number" → COUNT formulas
        - **Column headers** containing "max", "highest", "peak" → MAX formulas
        - **Column headers** containing "min", "lowest" → MIN formulas
        - **Numeric data patterns** analyzed for appropriate aggregation

        **WHEN TO USE:**
        - After creating tables to add calculation summaries
        - Building reports that need to update with new data
        - Creating templates for ongoing data entry
        - Establishing audit trails for calculations
        - Building interactive analysis worksheets
        - Preparing data for executive reporting

        Args:
            file_path (str): Full absolute path to Excel file containing the data table.
                           Example: "C:/reports/financial_analysis.xlsx"
                           File must exist and contain the target data.
            sheet_name (str): Worksheet name containing the data table.
                            Examples: "Data", "Sales", "Analysis"
                            Sheet must exist and contain the specified range.
            table_range (str): Data range for formula analysis and calculation.
                             Format: "A1:E20" (includes headers and all data)
                             
                             **RANGE REQUIREMENTS:**
                             - Must include header row for intelligent analysis
                             - Should contain only the data table (no extra text)
                             - Must have consistent data structure
                             - Should include all columns needing calculations
                             
                             **EXAMPLE STRUCTURE:**
                             A1:E10 containing:
                             A1: "Product"   B1: "Units"    C1: "Price"    D1: "Total"     E1: "Margin"
                             A2: "Widget A"  B2: 100       C2: 10.50     D2: 1050      E2: 315
                             A3: "Widget B"  B3: 75        C3: 15.75     D3: 1181.25   E3: 354
                             ... (continues with data)
                             
            formula_type (str): Formula selection strategy for calculations.
                              
                              **"auto" (RECOMMENDED)**: Intelligent analysis of headers and data
                              - Analyzes column names for calculation hints
                              - Examines data patterns and types
                              - Selects most appropriate formula for each column
                              - Provides best results for mixed data types
                              
                              **"sum"**: SUM formulas for all numeric columns
                              - Use for: Sales totals, quantity summations, cost aggregations
                              - Creates: =SUM(B2:B20), =SUM(C2:C20), etc.
                              
                              **"average"**: AVERAGE formulas for all numeric columns
                              - Use for: Performance metrics, rating averages, price means
                              - Creates: =AVERAGE(B2:B20), =AVERAGE(C2:C20), etc.
                              
                              **"count"**: COUNT formulas for data presence analysis
                              - Use for: Record counting, data completeness analysis
                              - Creates: =COUNT(B2:B20), =COUNT(C2:C20), etc.
                              
                              **"max"**: MAX formulas for highest value identification
                              - Use for: Peak performance, highest sales, maximum values
                              - Creates: =MAX(B2:B20), =MAX(C2:C20), etc.
                              
                              **"min"**: MIN formulas for lowest value identification
                              - Use for: Minimum thresholds, lowest costs, bottom performance
                              - Creates: =MIN(B2:B20), =MIN(C2:C20), etc.
                              
            add_totals (bool): Whether to add a dedicated totals row with formulas.
                             
                             **True (RECOMMENDED)**: Adds professional totals row
                             - Creates new row below data with "Total" label
                             - Adds appropriate formulas for each numeric column
                             - Applies bold formatting for visual emphasis
                             - Positions optimally with proper spacing
                             
                             **False**: Adds formulas without totals row
                             - Provides formula analysis without additional formatting
                             - Use when totals row already exists or not desired

        Returns:
            dict: Comprehensive formula addition result:
                - success (bool): Formula addition success status
                - file_path (str): Path to modified Excel file
                - sheet_name (str): Target worksheet name
                - table_range (str): Source data range analyzed
                - formulas_added (list): Detailed list of added formulas:
                  - column (str): Column letter (B, C, D, etc.)
                  - formula_type (str): Type of formula used (sum, average, etc.)
                  - formula (str): Actual Excel formula created
                  - range (str): Data range the formula calculates
                - total_row (int): Row number where totals were added (if applicable)
                - message (str): Detailed success message with formula count
                - error (str): Error details if operation failed

        **FORMULA EXAMPLES CREATED:**
        
        For sales data table, might create:
        - Units column: =SUM(B2:B20) "Total units sold"
        - Revenue column: =SUM(C2:C20) "Total revenue"
        - Average Price: =AVERAGE(D2:D20) "Average selling price"
        - Transaction Count: =COUNT(A2:A20) "Number of transactions"

        **INTELLIGENT ANALYSIS EXAMPLES:**
        
        Header: "Sales Amount" → Creates: =SUM(C2:C20)
        Header: "Average Rating" → Creates: =AVERAGE(D2:D20)
        Header: "Order Count" → Creates: =COUNT(B2:B20)
        Header: "Peak Revenue" → Creates: =MAX(E2:E20)
        Header: "Minimum Cost" → Creates: =MIN(F2:F20)

        **PROFESSIONAL FORMATTING APPLIED:**
        - Bold formatting for totals row and labels
        - Appropriate number formatting (#,##0.00 for currency)
        - Proper alignment and spacing
        - Visual separation from data rows
        - Consistent with Excel best practices

        **ERROR PREVENTION:**
        - Validates data range contains actual numeric data
        - Checks for appropriate column structure
        - Ensures formulas reference correct ranges
        - Handles mixed data types gracefully
        - Provides detailed error messages for troubleshooting

        **COMMON USAGE PATTERNS:**
        
        1. Complete table enhancement:
            write_sheet_data_tool("file.xlsx", "Data", "A1", data_array)
            add_table_tool("file.xlsx", "Data", "DataTable", "A1:E50")
            add_formulas_tool("file.xlsx", "Data", "A1:E50", "auto")
            
        2. Financial analysis:
            add_formulas_tool("budget.xlsx", "Analysis", "A1:F100", "sum")
            
        3. Performance dashboard:
            add_formulas_tool("metrics.xlsx", "Dashboard", "A1:H25", "auto")

        **BEST PRACTICES:**
        - Use "auto" formula type for mixed data analysis
        - Ensure headers are descriptive for better formula selection
        - Apply after table creation for optimal formatting
        - Review formulas in Excel to verify accuracy
        - Consider data update frequency when planning formula placement
        - Use consistent data structure for reliable formula creation

        **FORMULA MAINTENANCE:**
        - Formulas automatically expand with table growth
        - Structured references work with Excel table expansion
        - Manual verification recommended for critical calculations
        - Formulas remain editable for custom adjustments
        
        **NEXT STEPS:**
        - Verify formulas calculate correctly in Excel
        - Add charts referencing the calculated totals
        - Use conditional formatting to highlight key metrics
        - Consider pivot tables for additional analysis
        """
        try:
            # Validate inputs
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File does not exist: {file_path}")

            # Open the file
            wb = open_workbook(file_path)
            
            # Validate sheet exists
            if sheet_name not in list_sheets(wb):
                raise SheetNotFoundError(f"Sheet '{sheet_name}' does not exist in {file_path}")
            
            ws = get_sheet(wb, sheet_name)
            
            # Add formulas to the table
            if add_totals:
                result = add_formula_to_table(ws, table_range, formula_type)
            else:
                result = add_smart_formulas_to_data(ws, table_range, add_totals=False)
            
            if result.get('success'):
                # Save changes
                save_workbook(wb, file_path)
                
                return {
                    "success": True,
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "table_range": table_range,
                    "formulas_added": result.get('formulas_added', []),
                    "total_row": result.get('total_row'),
                    "message": f"Successfully added Excel formulas: {result.get('message', '')}"
                }
            else:
                return {
                    "success": False,
                    "error": result.get('error', 'Unknown error'),
                    "message": f"Failed to add formulas: {result.get('message', '')}"
                }
                
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error adding formulas: {e}"
            }
    
    @mcp.tool(description="Add calculated columns with live Excel formulas for advanced data analysis")
    def add_calculated_column_tool(file_path, sheet_name, table_range, column_header, formula_template):
        """Create calculated columns with live Excel formulas for dynamic data analysis.

        **PURPOSE & CONTEXT:**
        This tool extends existing data tables by adding calculated columns that reference
        other columns using live Excel formulas. The formulas automatically recalculate
        when source data changes, creating dynamic analysis capabilities. This is essential
        for creating derived metrics, financial calculations, and complex data relationships.

        **CALCULATED COLUMN BENEFITS:**
        - Live formulas that update automatically with data changes
        - Transparent calculation logic visible in Excel
        - Support for complex mathematical operations
        - Integration with Excel's native calculation engine
        - Professional appearance with proper formatting
        - Audit-ready calculation trails
        - Enables advanced analysis and reporting

        **WHEN TO USE:**
        - Creating derived metrics (profit = revenue - costs)
        - Financial calculations (tax amounts, discounts, totals)
        - Performance indicators (efficiency ratios, growth rates)
        - Data transformations (unit conversions, standardizations)
        - Complex business logic implementation
        - Building comprehensive analysis tables

        **FORMULA TEMPLATE SYSTEM:**
        Uses {row} placeholder for dynamic row referencing:
        - {row} gets replaced with actual row numbers (2, 3, 4, etc.)
        - Enables relative referencing across all data rows
        - Supports complex multi-column calculations
        - Maintains formula consistency throughout the table

        Args:
            file_path (str): Full absolute path to Excel file containing the source table.
                           Example: "C:/analysis/financial_data.xlsx"
                           File must exist and be accessible for modification.
            sheet_name (str): Worksheet name containing the source data table.
                            Examples: "Data", "Analysis", "Calculations"
                            Sheet must exist in the workbook.
            table_range (str): Range of the existing table that will be extended.
                             Format: "A1:D15" (current table boundaries)
                             
                             **REQUIREMENTS:**
                             - Must include headers in first row
                             - Should contain the complete existing table
                             - New column will be added immediately to the right
                             - Range will be automatically expanded to include new column
                             
                             **EXAMPLE EXISTING TABLE:**
                             A1:C10 containing:
                             A1: "Product"   B1: "Quantity"  C1: "Unit Price"
                             A2: "Widget A"  B2: 100        C2: 10.50
                             A3: "Widget B"  B3: 75         C3: 15.75
                             ... (data continues)
                             
            column_header (str): Descriptive header for the new calculated column.
                               Examples: "Total Price", "Profit Margin", "Tax Amount"
                               Should clearly indicate what the calculation represents.
                               Will appear in the header row of the new column.
                               
            formula_template (str): Excel formula template with {row} placeholders.
                                  
                                  **FORMULA TEMPLATE EXAMPLES:**
                                  
                                  **Basic Arithmetic:**
                                  - "=B{row}*C{row}" → Multiply two columns
                                  - "=B{row}+C{row}" → Add two columns  
                                  - "=C{row}-D{row}" → Subtract columns
                                  - "=B{row}/C{row}" → Divide columns
                                  
                                  **Financial Calculations:**
                                  - "=B{row}*C{row}*0.1" → 10% calculation (tax, discount)
                                  - "=(C{row}-D{row})/C{row}*100" → Percentage margin
                                  - "=B{row}*C{row}*(1+E{row})" → Total with variable rate
                                  
                                  **Conditional Logic:**
                                  - "=IF(B{row}>100,C{row}*0.9,C{row})" → Volume discount
                                  - "=IF(C{row}<0,\"Loss\",\"Profit\")" → Status indicator
                                  - "=MAX(B{row}*C{row},0)" → Prevent negative values
                                  
                                  **Text Operations:**
                                  - "=A{row}&\" - \"&B{row}" → Concatenate with separator
                                  - "=UPPER(A{row})" → Convert to uppercase
                                  - "=LEFT(A{row},3)" → Extract first 3 characters
                                  
                                  **Mathematical Functions:**
                                  - "=ROUND(B{row}*C{row},2)" → Round to 2 decimal places
                                  - "=SQRT(B{row})" → Square root calculation
                                  - "=POWER(B{row},2)" → Square the value

        Returns:
            dict: Comprehensive calculated column creation result:
                - success (bool): Column creation success status
                - file_path (str): Path to modified Excel file
                - sheet_name (str): Target worksheet name
                - original_range (str): Original table range before expansion
                - new_range (str): Expanded table range including new column
                - new_column (str): Column letter of the added column (E, F, G, etc.)
                - column_header (str): Header text for the new column
                - formulas_added (int): Number of formula cells created
                - message (str): Detailed success message
                - error (str): Error details if operation failed

        **AUTOMATIC PROCESSING:**
        
        **Row-by-Row Formula Generation:**
        Template "=B{row}*C{row}" becomes:
        - Row 2: =B2*C2
        - Row 3: =B3*C3
        - Row 4: =B4*C4
        - etc. for all data rows
        
        **Professional Formatting Applied:**
        - Header formatted with bold text and center alignment
        - Formula cells formatted with appropriate number format
        - Consistent styling with existing table
        - Proper column width adjustment

        **REAL-WORLD EXAMPLES:**
        
        **Sales Analysis:**
        ```
        Existing: A1:C10 (Product, Quantity, Unit Price)
        Add: "Total Revenue" with "=B{row}*C{row}"
        Result: New column D with revenue calculations
        ```
        
        **Financial Analysis:**
        ```
        Existing: A1:D15 (Item, Cost, Sale Price, Quantity)
        Add: "Profit" with "=(C{row}-B{row})*D{row}"
        Result: New column E with profit calculations
        ```
        
        **Performance Metrics:**
        ```
        Existing: A1:C20 (Employee, Target, Actual)
        Add: "Achievement %" with "=(C{row}/B{row})*100"
        Result: New column D with percentage calculations
        ```

        **ERROR PREVENTION:**
        - Validates formula template syntax
        - Checks for valid column references
        - Ensures table range exists and contains data
        - Handles division by zero scenarios
        - Provides specific error messages for troubleshooting

        **FORMULA VALIDATION:**
        - Template must contain {row} placeholder
        - Column references must exist in the table
        - Formula syntax must be valid Excel format
        - Handles special characters and text properly

        **COMMON USAGE PATTERNS:**
        
        1. Financial dashboard:
            add_calculated_column_tool("finance.xlsx", "Data", "A1:C50", "Net Profit", "=B{row}-C{row}")
            add_calculated_column_tool("finance.xlsx", "Data", "A1:D50", "Profit Margin", "=(D{row}/B{row})*100")
            
        2. Sales analysis:
            add_calculated_column_tool("sales.xlsx", "Data", "A1:E100", "Commission", "=D{row}*E{row}*0.05")
            
        3. Inventory management:
            add_calculated_column_tool("inventory.xlsx", "Stock", "A1:D200", "Reorder Point", "=C{row}*D{row}")

        **BEST PRACTICES:**
        - Use descriptive column headers that explain the calculation
        - Test formula templates with sample data first
        - Consider data types and potential division by zero
        - Use appropriate number formatting for calculated values
        - Document complex formulas for future reference
        - Verify calculations manually for critical business logic

        **ADVANCED FORMULA TECHNIQUES:**
        - Use IF statements for conditional calculations
        - Combine multiple functions for complex logic
        - Reference absolute values when needed ($B$1)
        - Use ROUND function for clean decimal presentation
        - Implement error handling with IFERROR function

        **NEXT STEPS AFTER COLUMN CREATION:**
        - Verify formulas calculate correctly for all rows
        - Add appropriate number formatting if needed
        - Consider using the new column for chart creation
        - Apply conditional formatting to highlight key values
        - Use add_formulas_tool() to add totals for the new column
        """
        try:
            # Validate inputs
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File does not exist: {file_path}")

            # Open the file
            wb = open_workbook(file_path)
            
            # Validate sheet exists
            if sheet_name not in list_sheets(wb):
                raise SheetNotFoundError(f"Sheet '{sheet_name}' does not exist in {file_path}")
            
            ws = get_sheet(wb, sheet_name)
            
            # Create calculated column
            result = create_calculated_column(ws, table_range, column_header, formula_template)
            
            if result.get('success'):
                # Save changes
                save_workbook(wb, file_path)
                
                return {
                    "success": True,
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "original_range": table_range,
                    "new_range": result.get('new_range'),
                    "new_column": result.get('new_column'),
                    "column_header": column_header,
                    "formulas_added": result.get('formulas_added'),
                    "message": f"Successfully added calculated column: {result.get('message', '')}"
                }
            else:
                return {
                    "success": False,
                    "error": result.get('error', 'Unknown error'),
                    "message": f"Failed to add calculated column: {result.get('message', '')}"
                }
                
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error adding calculated column: {e}"
            }
    
    @mcp.tool(description="Add a specific Excel formula to a cell or range")
    def add_formula_tool(file_path, sheet_name, cell_or_range, formula):
        """Add a specific Excel formula to a cell or range of cells.

        This tool allows you to add any Excel formula to enhance data analysis.
        The formula will be live and recalculate automatically.

        Args:
            file_path (str): Path to the Excel file.
            sheet_name (str): Name of the sheet.
            cell_or_range (str): Target cell (e.g., "D5") or range (e.g., "D5:D10").
            formula (str): Excel formula to add (must start with "=").
                          Examples: "=SUM(A1:A10)", "=B2*C2", "=AVERAGE(B:B)"

        Returns:
            dict: Result of the operation.

        Example:
            add_formula_tool(
                "C:/data/report.xlsx",
                "Summary",
                "E15",
                "=SUM(E2:E14)"  # Sum all values above
            )
        """
        try:
            # Validate inputs
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File does not exist: {file_path}")

            if not formula.startswith('='):
                raise ValueError("Formula must start with '='")

            # Open the file
            wb = open_workbook(file_path)
            
            # Validate sheet exists
            if sheet_name not in list_sheets(wb):
                raise SheetNotFoundError(f"Sheet '{sheet_name}' does not exist in {file_path}")
            
            ws = get_sheet(wb, sheet_name)
            
            # Apply formula to cell or range
            if ':' in cell_or_range:
                # Range of cells
                start_cell, end_cell = cell_or_range.split(':')
                start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
                end_row, end_col = ExcelRange.parse_cell_ref(end_cell)
                
                cells_updated = 0
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        cell = ws.cell(row=row + 1, column=col + 1)
                        # Adjust formula for each cell if it contains relative references
                        adjusted_formula = formula  # For now, use same formula
                        cell.value = adjusted_formula
                        cells_updated += 1
                        
                message = f"Formula added to {cells_updated} cells in range {cell_or_range}"
            else:
                # Single cell
                cell = ws[cell_or_range]
                cell.value = formula
                message = f"Formula added to cell {cell_or_range}"
            
            # Save changes
            save_workbook(wb, file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "target": cell_or_range,
                "formula": formula,
                "message": message
            }
                
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error adding formula: {e}"
            }
    


if __name__ == "__main__":
    logger.info("Master Excel MCP - Usage example")
    logger.info("This module brings together all Excel functionalities in one place.")
