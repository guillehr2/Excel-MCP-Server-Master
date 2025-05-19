#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel MCP (Multi-purpose Connector for Python with Excel)
-------------------------------------------------------
Biblioteca para manipular archivos Excel con funcionalidades avanzadas:
- Lectura y extracción de datos, fórmulas, tablas y gráficos
- Edición y modificación de archivos
- Creación de gráficos nativos de Excel
- Importación y exportación de datos

Author: MCP Team
Version: 2.0
"""

import os
import sys
import json
import logging
import tempfile
import time
from pathlib import Path
from typing import List, Dict, Union, Optional, Tuple, Any, Callable

# Configuración de logging
logger = logging.getLogger("excel_mcp")
logger.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stderr)
handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
logger.addHandler(handler)

# Importar MCP
try:
    from mcp.server.fastmcp import FastMCP
    HAS_MCP = True
except ImportError:
    logger.warning("No se pudo importar FastMCP. Las funcionalidades de servidor MCP no estarán disponibles.")
    HAS_MCP = False

# Intentar importar las bibliotecas necesarias
try:
    import pandas as pd
    import numpy as np
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.table import Table
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, ScatterChart, 
        Reference, Series
    )
    HAS_OPENPYXL = True
except ImportError as e:
    logger.warning(f"Error al importar bibliotecas esenciales: {e}")
    logger.warning("Es posible que algunas funcionalidades no estén disponibles")
    HAS_OPENPYXL = False

# Excepciones personalizadas
class ExcelMCPError(Exception):
    """Excepción base para todos los errores de Excel MCP."""
    pass

class FileNotFoundError(ExcelMCPError):
    """Se lanza cuando no se encuentra un archivo Excel."""
    pass

class SheetNotFoundError(ExcelMCPError):
    """Se lanza cuando no se encuentra una hoja en el archivo Excel."""
    pass

class RangeError(ExcelMCPError):
    """Se lanza cuando hay un problema con un rango de celdas."""
    pass

class TableNotFoundError(ExcelMCPError):
    """Se lanza cuando no se encuentra una tabla en Excel."""
    pass

class ChartNotFoundError(ExcelMCPError):
    """Se lanza cuando no se encuentra un gráfico en Excel."""
    pass

# Clase auxiliar para gestionar rangos de Excel
class ExcelRange:
    """
    Clase para manipular y convertir rangos de Excel.
    
    Esta clase proporciona métodos para convertir entre notación de Excel (A1:B5)
    y coordenadas de Python (0-based), además de validar rangos.
    """
    
    @staticmethod
    def parse_cell_ref(cell_ref: str) -> Tuple[int, int]:
        """
        Convierte una referencia de celda en estilo A1 a coordenadas (fila, columna) 0-based.
        
        Args:
            cell_ref: Referencia de celda en formato Excel (ej: 'A1', 'B5')
            
        Returns:
            Tupla (fila, columna) con índices base 0
            
        Raises:
            ValueError: Si la referencia de celda no es válida
        """
        if not cell_ref or not isinstance(cell_ref, str):
            raise ValueError(f"Referencia de celda inválida: {cell_ref}")
        
        # Extraer la parte de columna (letras)
        col_str = ''.join(c for c in cell_ref if c.isalpha())
        # Extraer la parte de fila (números)
        row_str = ''.join(c for c in cell_ref if c.isdigit())
        
        if not col_str or not row_str:
            raise ValueError(f"Formato de celda inválido: {cell_ref}")
        
        # Convertir columna a índice (A->0, B->1, etc.)
        col_idx = 0
        for c in col_str.upper():
            col_idx = col_idx * 26 + (ord(c) - ord('A') + 1)
        col_idx -= 1  # Ajustar a base 0
        
        # Convertir fila a índice (base 0)
        row_idx = int(row_str) - 1
        
        return row_idx, col_idx
    
    @staticmethod
    def parse_range(range_str: str) -> Tuple[int, int, int, int]:
        """
        Convierte un rango en estilo A1:B5 a coordenadas (row1, col1, row2, col2) 0-based.
        
        Args:
            range_str: Rango en formato Excel (ej: 'A1:B5')
            
        Returns:
            Tupla (fila_inicio, col_inicio, fila_fin, col_fin) con índices base 0
            
        Raises:
            ValueError: Si el rango no es válido
        """
        if not range_str or not isinstance(range_str, str):
            raise ValueError(f"Rango inválido: {range_str}")
        
        # Manejar rangos con referencia a hoja
        if '!' in range_str:
            parts = range_str.split('!')
            if len(parts) != 2:
                raise ValueError(f"Formato de rango con hoja inválido: {range_str}")
            range_str = parts[1]  # Usar solo la parte del rango
        
        # Dividir el rango en celdas de inicio y fin
        if ':' in range_str:
            start_cell, end_cell = range_str.split(':')
            start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
            end_row, end_col = ExcelRange.parse_cell_ref(end_cell)
        else:
            # Si es una sola celda, inicio y fin son iguales
            start_row, start_col = ExcelRange.parse_cell_ref(range_str)
            end_row, end_col = start_row, start_col
        
        return start_row, start_col, end_row, end_col
    
    @staticmethod
    def cell_to_a1(row: int, col: int) -> str:
        """
        Convierte coordenadas (fila, columna) 0-based a referencia de celda A1.
        
        Args:
            row: Índice de fila (base 0)
            col: Índice de columna (base 0)
            
        Returns:
            Referencia de celda en formato A1
        """
        if row < 0 or col < 0:
            raise ValueError(f"Índices negativos no válidos: fila={row}, columna={col}")
        
        # Convertir columna a letras
        col_str = ""
        col_val = col + 1  # Convertir a base 1 para cálculo
        
        while col_val > 0:
            remainder = (col_val - 1) % 26
            col_str = chr(65 + remainder) + col_str
            col_val = (col_val - 1) // 26
        
        # Convertir fila a número (base 1 para Excel)
        row_val = row + 1
        
        return f"{col_str}{row_val}"
    
    @staticmethod
    def range_to_a1(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
        """
        Convierte coordenadas de rango 0-based a rango A1:B5.
        
        Args:
            start_row: Fila inicial (base 0)
            start_col: Columna inicial (base 0)
            end_row: Fila final (base 0)
            end_col: Columna final (base 0)
            
        Returns:
            Rango en formato A1:B5
        """
        start_cell = ExcelRange.cell_to_a1(start_row, start_col)
        end_cell = ExcelRange.cell_to_a1(end_row, end_col)
        
        return f"{start_cell}:{end_cell}"


# Clase principal para leer datos de Excel
class ExcelReader:
    """
    Clase para leer y extraer datos de archivos Excel.
    
    Proporciona métodos para leer valores, fórmulas, tablas y gráficos.
    """
    
    def __init__(self, file_path: str):
        """
        Inicializa un lector de Excel.
        
        Args:
            file_path: Ruta al archivo Excel
            
        Raises:
            FileNotFoundError: Si el archivo no existe
        """
        self.file_path = self._resolve_file_path(file_path)
        
        # Verificar que el archivo existe
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"El archivo no existe: {self.file_path}")
        
        # Cargar el archivo con openpyxl para acceso avanzado
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)
            logger.info(f"Archivo Excel cargado: {self.file_path}")
            logger.info(f"Hojas disponibles: {self.workbook.sheetnames}")
        except Exception as e:
            logger.error(f"Error al cargar archivo Excel {self.file_path}: {e}")
            raise ExcelMCPError(f"Error al cargar archivo Excel: {e}")
        
        # También cargar una versión con valores calculados (sin fórmulas)
        try:
            self.workbook_values = openpyxl.load_workbook(self.file_path, data_only=True)
        except Exception as e:
            logger.warning(f"Error al cargar valores calculados: {e}")
            self.workbook_values = self.workbook  # Usar el mismo workbook como fallback
    
    def _resolve_file_path(self, file_path: str) -> str:
        """
        Resuelve una ruta de archivo, buscando en directorios permitidos si es necesario.
        
        Args:
            file_path: Ruta al archivo (puede ser relativa o absoluta)
            
        Returns:
            Ruta absoluta al archivo si se encuentra, o la ruta original si no
        """
        # Si la ruta es absoluta y existe, usarla directamente
        if os.path.isabs(file_path) and os.path.exists(file_path):
            return file_path
        
        # Lista de posibles directorios donde buscar
        search_dirs = [
            os.getcwd(),  # Directorio actual
            os.path.expanduser("~"),  # Directorio del usuario
            os.path.join(os.path.expanduser("~"), "Downloads"),  # Carpeta de descargas
            os.path.join(os.path.expanduser("~"), "Documents"),  # Carpeta de documentos
        ]
        
        # Si la ruta tiene separadores de directorio, extraer el nombre base
        file_name = os.path.basename(file_path)
        
        # Buscar en posibles ubicaciones
        for search_dir in search_dirs:
            # Comprobar la ruta completa
            full_path = os.path.join(search_dir, file_path)
            if os.path.exists(full_path):
                return full_path
            
            # Comprobar sólo con el nombre del archivo
            name_only_path = os.path.join(search_dir, file_name)
            if os.path.exists(name_only_path):
                return name_only_path
        
        # No se encontró el archivo, devolver la ruta original
        return file_path
    
    def get_sheet_names(self) -> List[str]:
        """
        Obtiene la lista de nombres de hojas disponibles en el archivo Excel.
        
        Returns:
            Lista de nombres de hojas
        """
        return self.workbook.sheetnames
    
    def read_sheet_data(self, sheet_name: str, range_str: Optional[str] = None, 
                       formulas: bool = False) -> List[List[Any]]:
        """
        Lee valores y, opcionalmente, fórmulas de una hoja de Excel.
        
        Args:
            sheet_name: Nombre de la hoja
            range_str: Rango en formato A1:B5, o None para toda la hoja
            formulas: Si es True, devuelve fórmulas en lugar de valores calculados
        
        Returns:
            Lista de listas con los valores o fórmulas de las celdas
            
        Raises:
            SheetNotFoundError: Si la hoja no existe
            RangeError: Si el rango es inválido
        """
        # Seleccionar el libro adecuado según se quieren fórmulas o valores
        wb = self.workbook if formulas else self.workbook_values
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            sheets_info = ", ".join(wb.sheetnames)
            raise SheetNotFoundError(f"La hoja '{sheet_name}' no existe en el archivo. Hojas disponibles: {sheets_info}")
        
        # Obtener la hoja
        ws = wb[sheet_name]
        
        # Si no se especifica rango, usar toda la hoja
        if not range_str:
            # Determinar el rango usado (min_row, min_col, max_row, max_col)
            min_row, min_col = 1, 1
            max_row = ws.max_row
            max_col = ws.max_column
        else:
            # Parsear el rango especificado
            try:
                min_row, min_col, max_row, max_col = ExcelRange.parse_range(range_str)
                # Convertir a base 1 para openpyxl
                min_row += 1
                min_col += 1
                max_row += 1
                max_col += 1
            except ValueError as e:
                raise RangeError(f"Rango inválido '{range_str}': {e}")
        
        # Extraer datos del rango
        data = []
        for row in range(min_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                
                # Obtener el valor adecuado (fórmula o valor calculado)
                if formulas and cell.data_type == 'f':
                    # Si queremos fórmulas y la celda tiene una fórmula
                    value = cell.value  # Esto es la fórmula con '='
                else:
                    # Valor normal o calculado
                    value = cell.value
                
                row_data.append(value)
            data.append(row_data)
        
        return data
    
    def list_tables(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Lista todas las tablas definidas en una hoja de Excel.
        
        Args:
            sheet_name: Nombre de la hoja
            
        Returns:
            Lista de diccionarios con información de las tablas:
            [{'name': 'Tabla1', 'ref': 'A1:D10', 'header_row': True, 'totals_row': False}, ...]
            
        Raises:
            SheetNotFoundError: Si la hoja no existe
        """
        # Verificar que la hoja existe
        if sheet_name not in self.workbook.sheetnames:
            raise SheetNotFoundError(f"La hoja '{sheet_name}' no existe en el archivo")
        
        # Obtener la hoja
        ws = self.workbook[sheet_name]
        
        # Lista para almacenar información de las tablas
        tables_info = []
        
        # Verificar si la hoja tiene tablas
        if hasattr(ws, 'tables') and ws.tables:
            for table_name, table in ws.tables.items():
                table_info = {
                    'name': table_name,
                    'ref': table.ref,
                    'display_name': table.displayName,
                    'header_row': table.headerRowCount > 0,
                    'totals_row': table.totalsRowCount > 0,
                    'style': table.tableStyleInfo.name if table.tableStyleInfo else None
                }
                
                tables_info.append(table_info)
        
        return tables_info
    
    def get_table_data(self, sheet_name: str, table_name: str) -> List[Dict[str, Any]]:
        """
        Obtiene los datos de una tabla específica en formato de registros.
        
        Args:
            sheet_name: Nombre de la hoja
            table_name: Nombre de la tabla
            
        Returns:
            Lista de diccionarios, donde cada diccionario representa una fila
            
        Raises:
            SheetNotFoundError: Si la hoja no existe
            TableNotFoundError: Si la tabla no existe
        """
        # Verificar que la hoja existe
        if sheet_name not in self.workbook_values.sheetnames:  # Usar workbook_values para obtener valores calculados
            raise SheetNotFoundError(f"La hoja '{sheet_name}' no existe en el archivo")
        
        # Obtener la hoja
        ws = self.workbook_values[sheet_name]
        
        # Verificar si la tabla existe
        if not hasattr(ws, 'tables') or table_name not in ws.tables:
            raise TableNotFoundError(f"La tabla '{table_name}' no existe en la hoja '{sheet_name}'")
        
        # Obtener la referencia de la tabla
        table = ws.tables[table_name]
        table_range = table.ref
        
        # Parsear el rango
        min_row, min_col, max_row, max_col = ExcelRange.parse_range(table_range)
        
        # Ajustar a base 1 para openpyxl
        min_row += 1
        min_col += 1
        max_row += 1
        max_col += 1
        
        # Extraer encabezados (primera fila)
        headers = []
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=min_row, column=col)
            headers.append(cell.value or f"Column{col}")
        
        # Extraer datos (filas después del encabezado)
        data = []
        for row in range(min_row + 1, max_row + 1):
            row_data = {}
            for col_idx, col in enumerate(range(min_col, max_col + 1)):
                cell = ws.cell(row=row, column=col)
                header = headers[col_idx]
                row_data[header] = cell.value
            data.append(row_data)
        
        return data
    
    def list_charts(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Lista todos los gráficos en una hoja de Excel.
        
        Args:
            sheet_name: Nombre de la hoja
            
        Returns:
            Lista de diccionarios con información de los gráficos:
            [{'id': 1, 'type': 'bar', 'title': 'Ventas por región'}, ...]
            
        Raises:
            SheetNotFoundError: Si la hoja no existe
        """
        # Verificar que la hoja existe
        if sheet_name not in self.workbook.sheetnames:
            raise SheetNotFoundError(f"La hoja '{sheet_name}' no existe en el archivo")
        
        # Obtener la hoja
        ws = self.workbook[sheet_name]
        
        # Lista para almacenar información de los gráficos
        charts_info = []
        
        # Verificar si la hoja tiene gráficos
        chart_id = 0
        for chart_rel in ws._charts:
            chart = chart_rel[0]  # El elemento 0 es el objeto chart, el 1 es position
            
            # Determinar el tipo de gráfico
            chart_type = "desconocido"
            if isinstance(chart, BarChart):
                chart_type = "bar" if chart.type == "bar" else "column"
            elif isinstance(chart, LineChart):
                chart_type = "line"
            elif isinstance(chart, PieChart):
                chart_type = "pie"
            elif isinstance(chart, ScatterChart):
                chart_type = "scatter"
            
            # Recopilar información del gráfico
            chart_info = {
                'id': chart_id,
                'type': chart_type,
                'title': chart.title if hasattr(chart, 'title') and chart.title else f"Chart {chart_id}",
                'position': chart_rel[1] if len(chart_rel) > 1 else None,
                'series_count': len(chart.series) if hasattr(chart, 'series') else 0
            }
            
            charts_info.append(chart_info)
            chart_id += 1
        
        return charts_info
    
    def get_chart_info(self, sheet_name: str, chart_id: int) -> Dict[str, Any]:
        """
        Obtiene información detallada de un gráfico específico.
        
        Args:
            sheet_name: Nombre de la hoja
            chart_id: ID del gráfico
            
        Returns:
            Diccionario con información detallada del gráfico
            
        Raises:
            SheetNotFoundError: Si la hoja no existe
            ChartNotFoundError: Si el gráfico no existe
        """
        # Verificar que la hoja existe
        if sheet_name not in self.workbook.sheetnames:
            raise SheetNotFoundError(f"La hoja '{sheet_name}' no existe en el archivo")
        
        # Obtener la hoja
        ws = self.workbook[sheet_name]
        
        # Verificar si el chart_id es válido
        if not hasattr(ws, '_charts') or chart_id >= len(ws._charts):
            raise ChartNotFoundError(f"El gráfico con ID {chart_id} no existe en la hoja '{sheet_name}'")
        
        # Obtener el gráfico
        chart_rel = ws._charts[chart_id]
        chart = chart_rel[0]
        
        # Determinar el tipo de gráfico
        chart_type = "desconocido"
        if isinstance(chart, BarChart):
            chart_type = "bar" if chart.type == "bar" else "column"
        elif isinstance(chart, LineChart):
            chart_type = "line"
        elif isinstance(chart, PieChart):
            chart_type = "pie"
        elif isinstance(chart, ScatterChart):
            chart_type = "scatter"
        
        # Extraer información sobre series
        series_info = []
        if hasattr(chart, 'series'):
            for idx, series in enumerate(chart.series):
                series_data = {
                    'index': idx,
                    'title': series.tx.v if hasattr(series, 'tx') and hasattr(series.tx, 'v') else f"Series {idx}"
                }
                
                # Intentar obtener el rango de datos si está disponible
                if hasattr(series, 'val') and hasattr(series.val, 'reference'):
                    series_data['data_range'] = str(series.val.reference)
                
                series_info.append(series_data)
        
        # Recopilar información completa del gráfico
        chart_info = {
            'id': chart_id,
            'type': chart_type,
            'title': chart.title if hasattr(chart, 'title') and chart.title else f"Chart {chart_id}",
            'position': chart_rel[1] if len(chart_rel) > 1 else None,
            'series': series_info,
            'has_legend': hasattr(chart, 'legend') and chart.legend is not None,
            'x_axis_title': chart.x_axis.title if hasattr(chart, 'x_axis') and hasattr(chart.x_axis, 'title') else None,
            'y_axis_title': chart.y_axis.title if hasattr(chart, 'y_axis') and hasattr(chart.y_axis, 'title') else None
        }
        
        return chart_info
    
    def get_named_ranges(self) -> Dict[str, str]:
        """
        Recupera todos los rangos con nombre definidos en el libro de Excel.
        
        Returns:
            Diccionario con nombres de rangos y sus referencias
        """
        named_ranges = {}
        
        # Iterar sobre todos los rangos con nombre en el libro
        try:
            if hasattr(self.workbook, 'defined_names'):
                # En versiones nuevas de openpyxl
                for name, defined_name in self.workbook.defined_names.items():
                    try:
                        # Intentar obtener destinos
                        destinations = list(defined_name.destinations)
                        if destinations:
                            # Usar el primer destino
                            sheet, coord = destinations[0]
                            named_ranges[name] = f"'{sheet}'!{coord}"
                        else:
                            # Si no hay destinos, usar attr_text si está disponible
                            if hasattr(defined_name, 'attr_text'):
                                named_ranges[name] = defined_name.attr_text
                            else:
                                named_ranges[name] = str(defined_name)
                    except Exception as e:
                        logger.warning(f"Error al procesar rango con nombre '{name}': {e}")
                        named_ranges[name] = "Error al procesar"
        except Exception as e:
            logger.warning(f"Error al acceder a los rangos con nombre: {e}")
            # Intento alternativo
            try:
                # Para compatibilidad con diferentes versiones
                for name in self.workbook.get_named_ranges():
                    named_ranges[name.name] = name.value
            except:
                pass
        
        return named_ranges
    
    def close(self):
        """Cierra el archivo Excel y libera recursos."""
        if hasattr(self, 'workbook'):
            self.workbook.close()
        if hasattr(self, 'workbook_values') and self.workbook_values != self.workbook:
            self.workbook_values.close()


# Función auxiliar para validar argumentos
def validate_workbook(wb) -> None:
    """
    Valida que el objeto workbook sea válido.
    
    Args:
        wb: Objeto workbook a validar
        
    Raises:
        ValueError: Si el workbook no es válido
    """
    if wb is None:
        raise ValueError("El workbook no puede ser None")
    
    if not hasattr(wb, 'sheetnames'):
        raise ValueError("El objeto proporcionado no parece ser un workbook válido")


# Funciones de utilidad para la interfaz
def read_sheet_data(wb, sheet_name: str, range_str: Optional[str] = None, 
                   formulas: bool = False) -> List[List[Any]]:
    """
    Lee valores y, opcionalmente, fórmulas de una hoja de Excel.
    
    Args:
        wb: Objeto workbook de openpyxl
        sheet_name: Nombre de la hoja
        range_str: Rango en formato A1:B5, o None para toda la hoja
        formulas: Si es True, devuelve fórmulas en lugar de valores calculados
    
    Returns:
        Lista de listas con los valores o fórmulas de las celdas
        
    Raises:
        ValueError: Si los argumentos no son válidos
        SheetNotFoundError: Si la hoja no existe
        RangeError: Si el rango es inválido
    """
    # Validar workbook
    validate_workbook(wb)
    
    # Crear un ExcelReader temporal con un archivo existente
    # Nota: Es un poco artificioso pero permite reutilizar el código
    # En una implementación real, habría que refactorizar para no depender del archivo
    if hasattr(wb, '_path'):
        temp_reader = ExcelReader(wb._path)
    else:
        # Guardar temporalmente el workbook en un archivo
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        wb.save(temp_file.name)
        temp_reader = ExcelReader(temp_file.name)
    
    try:
        # Usar el método de la clase
        return temp_reader.read_sheet_data(sheet_name, range_str, formulas)
    finally:
        # Limpiar recursos
        temp_reader.close()
        if not hasattr(wb, '_path'):
            os.unlink(temp_file.name)


def list_tables(ws) -> List[Dict[str, Any]]:
    """
    Lista todas las tablas definidas en una hoja de Excel.
    
    Args:
        ws: Objeto worksheet de openpyxl
        
    Returns:
        Lista de diccionarios con información de las tablas
    """
    # Verificar que ws es un objeto worksheet válido
    if ws is None:
        raise ValueError("El worksheet no puede ser None")
    
    # Lista para almacenar información de las tablas
    tables_info = []
    
    # Verificar si la hoja tiene tablas
    if hasattr(ws, 'tables') and ws.tables:
        for table_name, table in ws.tables.items():
            table_info = {
                'name': table_name,
                'ref': table.ref,
                'display_name': table.displayName,
                'header_row': table.headerRowCount > 0,
                'totals_row': table.totalsRowCount > 0,
                'style': table.tableStyleInfo.name if table.tableStyleInfo else None
            }
            
            tables_info.append(table_info)
    
    return tables_info


def get_table_data(ws, table_name: str) -> List[Dict[str, Any]]:
    """
    Obtiene los datos de una tabla específica en formato de registros.
    
    Args:
        ws: Objeto worksheet de openpyxl
        table_name: Nombre de la tabla
        
    Returns:
        Lista de diccionarios, donde cada diccionario representa una fila
        
    Raises:
        ValueError: Si los argumentos no son válidos
        TableNotFoundError: Si la tabla no existe
    """
    # Verificar que ws es un objeto worksheet válido
    if ws is None:
        raise ValueError("El worksheet no puede ser None")
    
    # Verificar si la tabla existe
    if not hasattr(ws, 'tables') or table_name not in ws.tables:
        raise TableNotFoundError(f"La tabla '{table_name}' no existe en la hoja")
    
    # Obtener la referencia de la tabla
    table = ws.tables[table_name]
    table_range = table.ref
    
    # Parsear el rango
    min_row, min_col, max_row, max_col = ExcelRange.parse_range(table_range)
    
    # Ajustar a base 1 para openpyxl
    min_row += 1
    min_col += 1
    max_row += 1
    max_col += 1
    
    # Extraer encabezados (primera fila)
    headers = []
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=min_row, column=col)
        headers.append(cell.value or f"Column{col}")
    
    # Extraer datos (filas después del encabezado)
    data = []
    for row in range(min_row + 1, max_row + 1):
        row_data = {}
        for col_idx, col in enumerate(range(min_col, max_col + 1)):
            cell = ws.cell(row=row, column=col)
            header = headers[col_idx]
            row_data[header] = cell.value
        data.append(row_data)
    
    return data


def list_charts(ws) -> List[Dict[str, Any]]:
    """
    Lista todos los gráficos en una hoja de Excel.
    
    Args:
        ws: Objeto worksheet de openpyxl
        
    Returns:
        Lista de diccionarios con información de los gráficos
    """
    # Verificar que ws es un objeto worksheet válido
    if ws is None:
        raise ValueError("El worksheet no puede ser None")
    
    # Lista para almacenar información de los gráficos
    charts_info = []
    
    # Verificar si la hoja tiene gráficos
    if hasattr(ws, '_charts'):
        for chart_id, chart_rel in enumerate(ws._charts):
            chart = chart_rel[0]  # El elemento 0 es el objeto chart, el 1 es position
            
            # Determinar el tipo de gráfico
            chart_type = "desconocido"
            if isinstance(chart, BarChart):
                chart_type = "bar" if chart.type == "bar" else "column"
            elif isinstance(chart, LineChart):
                chart_type = "line"
            elif isinstance(chart, PieChart):
                chart_type = "pie"
            elif isinstance(chart, ScatterChart):
                chart_type = "scatter"
            
            # Recopilar información del gráfico
            chart_info = {
                'id': chart_id,
                'type': chart_type,
                'title': chart.title if hasattr(chart, 'title') and chart.title else f"Chart {chart_id}",
                'position': chart_rel[1] if len(chart_rel) > 1 else None,
                'series_count': len(chart.series) if hasattr(chart, 'series') else 0
            }
            
            charts_info.append(chart_info)
    
    return charts_info


def get_chart_info(ws, chart_id: int) -> Dict[str, Any]:
    """
    Obtiene información detallada de un gráfico específico.
    
    Args:
        ws: Objeto worksheet de openpyxl
        chart_id: ID del gráfico
        
    Returns:
        Diccionario con información detallada del gráfico
        
    Raises:
        ValueError: Si los argumentos no son válidos
        ChartNotFoundError: Si el gráfico no existe
    """
    # Verificar que ws es un objeto worksheet válido
    if ws is None:
        raise ValueError("El worksheet no puede ser None")
    
    # Verificar si el chart_id es válido
    if not hasattr(ws, '_charts') or chart_id >= len(ws._charts):
        raise ChartNotFoundError(f"El gráfico con ID {chart_id} no existe en la hoja")
    
    # Obtener el gráfico
    chart_rel = ws._charts[chart_id]
    chart = chart_rel[0]
    
    # Determinar el tipo de gráfico
    chart_type = "desconocido"
    if isinstance(chart, BarChart):
        chart_type = "bar" if chart.type == "bar" else "column"
    elif isinstance(chart, LineChart):
        chart_type = "line"
    elif isinstance(chart, PieChart):
        chart_type = "pie"
    elif isinstance(chart, ScatterChart):
        chart_type = "scatter"
    
    # Extraer información sobre series
    series_info = []
    if hasattr(chart, 'series'):
        for idx, series in enumerate(chart.series):
            series_data = {
                'index': idx,
                'title': series.tx.v if hasattr(series, 'tx') and hasattr(series.tx, 'v') else f"Series {idx}"
            }
            
            # Intentar obtener el rango de datos si está disponible
            if hasattr(series, 'val') and hasattr(series.val, 'reference'):
                series_data['data_range'] = str(series.val.reference)
            
            series_info.append(series_data)
    
    # Recopilar información completa del gráfico
    chart_info = {
        'id': chart_id,
        'type': chart_type,
        'title': chart.title if hasattr(chart, 'title') and chart.title else f"Chart {chart_id}",
        'position': chart_rel[1] if len(chart_rel) > 1 else None,
        'series': series_info,
        'has_legend': hasattr(chart, 'legend') and chart.legend is not None,
        'x_axis_title': chart.x_axis.title if hasattr(chart, 'x_axis') and hasattr(chart.x_axis, 'title') else None,
        'y_axis_title': chart.y_axis.title if hasattr(chart, 'y_axis') and hasattr(chart.y_axis, 'title') else None
    }
    
    return chart_info


def get_named_ranges(wb) -> Dict[str, str]:
    """
    Recupera todos los rangos con nombre definidos en el libro de Excel.
    
    Args:
        wb: Objeto workbook de openpyxl
        
    Returns:
        Diccionario con nombres de rangos y sus referencias
    """
    # Validar workbook
    validate_workbook(wb)
    
    named_ranges = {}
    
    # Iterar sobre todos los rangos con nombre en el libro
    if hasattr(wb, 'defined_names') and hasattr(wb.defined_names, 'definedName'):
        for named_range in wb.defined_names.definedName:
            named_ranges[named_range.name] = named_range.attr_text
    
    return named_ranges


# Crear el servidor MCP como variable global
mcp = None
if HAS_MCP:
    # Esta es la variable global que el sistema MCP busca
    mcp = FastMCP("Excel MCP", 
                 dependencies=["pandas", "numpy", "openpyxl"])
    logger.info("Servidor MCP iniciado correctamente")
    
    @mcp.tool(description="Explora un archivo Excel y devuelve su estructura completa")
    def explore_excel_tool(file_path):
        """Explora un archivo Excel y devuelve información completa sobre su estructura"""
        try:
            reader = ExcelReader(file_path)
            result = {}
            
            # Obtener información básica del archivo
            result["file_path"] = reader.file_path
            result["file_name"] = os.path.basename(reader.file_path)
            result["file_size"] = os.path.getsize(reader.file_path)
            result["last_modified"] = time.ctime(os.path.getmtime(reader.file_path))
            
            # Obtener lista de hojas
            sheets = reader.get_sheet_names()
            result["sheets"] = sheets
            
            # Recopilar información detallada de cada hoja
            sheets_info = {}
            for sheet_name in sheets:
                sheet_info = {}
                
                try:
                    # Dimensiones de la hoja
                    ws = reader.workbook[sheet_name]
                    sheet_info["dimensions"] = {
                        "rows": ws.max_row,
                        "columns": ws.max_column
                    }
                    
                    # Muestra de datos (primeras filas)
                    try:
                        sample_data = reader.read_sheet_data(sheet_name, f"A1:{get_column_letter(min(10, ws.max_column))}5")
                        sheet_info["sample_data"] = sample_data
                    except Exception as e:
                        sheet_info["sample_data_error"] = str(e)
                    
                    # Buscar tablas
                    try:
                        tables = reader.list_tables(sheet_name)
                        if tables:
                            sheet_info["tables"] = tables
                    except Exception as e:
                        sheet_info["tables_error"] = str(e)
                    
                    # Buscar gráficos
                    try:
                        charts = reader.list_charts(sheet_name)
                        if charts:
                            sheet_info["charts"] = charts
                    except Exception as e:
                        sheet_info["charts_error"] = str(e)
                    
                except Exception as sheet_e:
                    sheet_info["error"] = str(sheet_e)
                
                sheets_info[sheet_name] = sheet_info
            
            result["sheets_info"] = sheets_info
            
            # Obtener rangos con nombre
            try:
                named_ranges = reader.get_named_ranges()
                if named_ranges:
                    result["named_ranges"] = named_ranges
            except Exception as e:
                result["named_ranges_error"] = str(e)
            
            reader.close()
            
            return {
                "success": True,
                "file": result["file_name"],
                "path": result["file_path"],
                "sheet_count": len(sheets),
                "structure": result,
                "message": f"Archivo Excel explorado correctamente: {len(sheets)} hojas encontradas"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al explorar archivo Excel: {e}"
            }
    
    @mcp.tool(description="Lista las hojas disponibles en un archivo Excel")
    def list_sheets_tool(file_path):
        """Lista todas las hojas disponibles en un archivo Excel"""
        try:
            reader = ExcelReader(file_path)
            sheets = reader.get_sheet_names()
            reader.close()
            
            return {
                "success": True,
                "file_path": reader.file_path,
                "sheets": sheets,
                "count": len(sheets),
                "message": f"Se encontraron {len(sheets)} hojas en el archivo Excel"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al listar hojas: {e}"
            }
    
    # Función auxiliar para buscar hojas similares
    def _find_similar_sheet(reader, sheet_name):
        """Busca una hoja similar si la especificada no existe"""
        if sheet_name in reader.get_sheet_names():
            return sheet_name
            
        sheets = reader.get_sheet_names()
        suggestion = None
        
        # Buscar coincidencia parcial
        for sheet in sheets:
            if sheet_name.lower() in sheet.lower() or sheet.lower() in sheet_name.lower():
                suggestion = sheet
                break
        
        if suggestion:
            logger.info(f"Hoja '{sheet_name}' no encontrada, usando '{suggestion}' como alternativa")
            return suggestion
        elif sheets:  # Si hay hojas disponibles, usar la primera
            first_sheet = sheets[0]
            logger.info(f"Hoja '{sheet_name}' no encontrada, usando la primera hoja disponible: '{first_sheet}'")
            return first_sheet
        else:
            raise SheetNotFoundError(f"No se encontró la hoja '{sheet_name}' ni alternativas")
    @mcp.tool(description="Lee valores y, opcionalmente, fórmulas de una hoja de Excel")
    def read_sheet_data_tool(file_path, sheet_name, range_str=None, formulas=False):
        """Interfaz MCP para read_sheet_data"""
        try:
            reader = ExcelReader(file_path)
            sheet_name = _find_similar_sheet(reader, sheet_name)
            
            result = reader.read_sheet_data(sheet_name, range_str, formulas)
            reader.close()
            return {
                "success": True,
                "sheet_name": sheet_name,
                "data": result,
                "rows": len(result),
                "columns": len(result[0]) if result and len(result) > 0 else 0,
                "message": f"Datos leídos correctamente de {sheet_name}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al leer datos: {e}"
            }
    
    @mcp.tool(description="Lista todas las tablas definidas en una hoja de Excel")
    def list_tables_tool(file_path, sheet_name):
        """Interfaz MCP para list_tables"""
        try:
            reader = ExcelReader(file_path)
            sheet_name = _find_similar_sheet(reader, sheet_name)
            
            result = reader.list_tables(sheet_name)
            reader.close()
            return {
                "success": True,
                "sheet_name": sheet_name,
                "file_path": reader.file_path,
                "tables": result,
                "count": len(result),
                "message": f"Se encontraron {len(result)} tablas en la hoja {sheet_name}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al listar tablas: {e}"
            }
    
    @mcp.tool(description="Obtiene los datos de una tabla específica en formato de registros")
    def get_table_data_tool(file_path, sheet_name, table_name):
        """Interfaz MCP para get_table_data"""
        try:
            reader = ExcelReader(file_path)
            sheet_name = _find_similar_sheet(reader, sheet_name)
            
            result = reader.get_table_data(sheet_name, table_name)
            reader.close()
            return {
                "success": True,
                "sheet_name": sheet_name,
                "table_name": table_name,
                "data": result,
                "row_count": len(result),
                "message": f"Datos de tabla '{table_name}' leídos correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al obtener datos de tabla: {e}"
            }
    
    @mcp.tool(description="Lista todos los gráficos en una hoja de Excel")
    def list_charts_tool(file_path, sheet_name):
        """Interfaz MCP para list_charts"""
        try:
            reader = ExcelReader(file_path)
            sheet_name = _find_similar_sheet(reader, sheet_name)
            
            result = reader.list_charts(sheet_name)
            reader.close()
            return {
                "success": True,
                "sheet_name": sheet_name,
                "file_path": reader.file_path,
                "charts": result,
                "count": len(result),
                "message": f"Se encontraron {len(result)} gráficos en la hoja {sheet_name}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al listar gráficos: {e}"
            }
    
    @mcp.tool(description="Obtiene información detallada de un gráfico específico")
    def get_chart_info_tool(file_path, sheet_name, chart_id):
        """Interfaz MCP para get_chart_info"""
        try:
            reader = ExcelReader(file_path)
            sheet_name = _find_similar_sheet(reader, sheet_name)
            
            # Convertir chart_id a entero si es posible
            try:
                chart_id = int(chart_id)
            except (ValueError, TypeError):
                # Si no es un entero válido, dejar el valor original
                pass
                
            result = reader.get_chart_info(sheet_name, chart_id)
            reader.close()
            return {
                "success": True,
                "sheet_name": sheet_name,
                "chart_id": chart_id,
                "chart": result,
                "message": f"Información del gráfico {chart_id} obtenida correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al obtener información del gráfico: {e}"
            }
    
    @mcp.tool(description="Recupera todos los rangos con nombre definidos en el libro de Excel")
    def get_named_ranges_tool(file_path):
        """Interfaz MCP para get_named_ranges"""
        try:
            reader = ExcelReader(file_path)
            result = reader.get_named_ranges()
            
            # Obtener información adicional sobre el archivo
            file_info = {
                "file_path": reader.file_path,
                "file_name": os.path.basename(reader.file_path),
                "sheet_count": len(reader.get_sheet_names()),
                "sheets": reader.get_sheet_names()
            }
            
            reader.close()
            return {
                "success": True,
                "file_info": file_info,
                "ranges": result,
                "count": len(result),
                "message": f"Se encontraron {len(result)} rangos con nombre"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al obtener rangos con nombre: {e}"
            }

if __name__ == "__main__":
    # Código de ejemplo de uso
    logger.info("Excel MCP - Ejemplo de uso")
    
    # Verificar argumentos
    if len(sys.argv) > 1:
        archivo = sys.argv[1]
        logger.info(f"Leyendo archivo: {archivo}")
        
        try:
            reader = ExcelReader(archivo)
            
            # Listar hojas
            logger.info(f"Hojas en el archivo: {reader.workbook.sheetnames}")
            
            # Leer datos de la primera hoja
            primera_hoja = reader.workbook.sheetnames[0]
            datos = reader.read_sheet_data(primera_hoja)
            logger.info(f"Primeras 5 filas de datos: {datos[:5]}")
            
            # Listar tablas en la primera hoja
            tablas = reader.list_tables(primera_hoja)
            logger.info(f"Tablas en la hoja {primera_hoja}: {tablas}")
            
            # Listar gráficos en la primera hoja
            graficos = reader.list_charts(primera_hoja)
            logger.info(f"Gráficos en la hoja {primera_hoja}: {graficos}")
            
            # Listar rangos con nombre
            rangos = reader.get_named_ranges()
            logger.info(f"Rangos con nombre: {rangos}")
            
            reader.close()
            
        except Exception as e:
            logger.error(f"Error al procesar el archivo: {e}")
    else:
        logger.info("Uso: python excel_mcp_complete.py archivo.xlsx")
