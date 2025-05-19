#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel Writer MCP (Multi-purpose Connector for Python with Excel)
-------------------------------------------------------
Biblioteca para escribir y dar formato a archivos Excel:
- Escritura y edición de datos
- Formateo y estilo de celdas
- Manipulación de filas y columnas
- Gestión de celdas combinadas

Author: MCP Team
Version: 1.0
"""

import os
import sys
import logging
from typing import List, Dict, Union, Optional, Tuple, Any, Callable

# Configuración de logging
logger = logging.getLogger("excel_writer_mcp")
logger.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stderr)
handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
logger.addHandler(handler)

# Importar MCP
try:
    from mcp.server.fastmcp import FastMCP
    HAS_MCP = True
except ImportError:
    logger.warning("No se pudo importar FastMCP. Las funcionalidades de servidor MCP no estarán disponibles.")
    HAS_MCP = False

# Intentar importar las bibliotecas necesarias
try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment, 
        NamedStyle, Protection, Color, colors
    )
    HAS_OPENPYXL = True
except ImportError as e:
    logger.warning(f"Error al importar bibliotecas esenciales: {e}")
    logger.warning("Es posible que algunas funcionalidades no estén disponibles")
    HAS_OPENPYXL = False

# Excepciones personalizadas
class ExcelWriterError(Exception):
    """Excepción base para todos los errores de Excel Writer."""
    pass

class FileNotFoundError(ExcelWriterError):
    """Se lanza cuando no se encuentra un archivo Excel."""
    pass

class SheetNotFoundError(ExcelWriterError):
    """Se lanza cuando no se encuentra una hoja en el archivo Excel."""
    pass

class CellReferenceError(ExcelWriterError):
    """Se lanza cuando hay un problema con una referencia de celda."""
    pass

class RangeError(ExcelWriterError):
    """Se lanza cuando hay un problema con un rango de celdas."""
    pass


# Clase auxiliar para gestionar rangos de Excel
class ExcelRange:
    """
    Clase para manipular y convertir rangos de Excel.
    
    Esta clase proporciona métodos para convertir entre notación de Excel (A1:B5)
    y coordenadas de Python (0-based), además de validar rangos.
    """
    
    @staticmethod
    def parse_cell_ref(cell_ref: str) -> Tuple[int, int]:
        """
        Convierte una referencia de celda en estilo A1 a coordenadas (fila, columna) 0-based.
        
        Args:
            cell_ref: Referencia de celda en formato Excel (ej: 'A1', 'B5')
            
        Returns:
            Tupla (fila, columna) con índices base 0
            
        Raises:
            ValueError: Si la referencia de celda no es válida
        """
        if not cell_ref or not isinstance(cell_ref, str):
            raise ValueError(f"Referencia de celda inválida: {cell_ref}")
        
        # Extraer la parte de columna (letras)
        col_str = ''.join(c for c in cell_ref if c.isalpha())
        # Extraer la parte de fila (números)
        row_str = ''.join(c for c in cell_ref if c.isdigit())
        
        if not col_str or not row_str:
            raise ValueError(f"Formato de celda inválido: {cell_ref}")
        
        # Convertir columna a índice (A->0, B->1, etc.)
        col_idx = 0
        for c in col_str.upper():
            col_idx = col_idx * 26 + (ord(c) - ord('A') + 1)
        col_idx -= 1  # Ajustar a base 0
        
        # Convertir fila a índice (base 0)
        row_idx = int(row_str) - 1
        
        return row_idx, col_idx
    
    @staticmethod
    def parse_range(range_str: str) -> Tuple[int, int, int, int]:
        """
        Convierte un rango en estilo A1:B5 a coordenadas (row1, col1, row2, col2) 0-based.
        
        Args:
            range_str: Rango en formato Excel (ej: 'A1:B5')
            
        Returns:
            Tupla (fila_inicio, col_inicio, fila_fin, col_fin) con índices base 0
            
        Raises:
            ValueError: Si el rango no es válido
        """
        if not range_str or not isinstance(range_str, str):
            raise ValueError(f"Rango inválido: {range_str}")
        
        # Manejar rangos con referencia a hoja
        if '!' in range_str:
            parts = range_str.split('!')
            if len(parts) != 2:
                raise ValueError(f"Formato de rango con hoja inválido: {range_str}")
            range_str = parts[1]  # Usar solo la parte del rango
        
        # Dividir el rango en celdas de inicio y fin
        if ':' in range_str:
            start_cell, end_cell = range_str.split(':')
            start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
            end_row, end_col = ExcelRange.parse_cell_ref(end_cell)
        else:
            # Si es una sola celda, inicio y fin son iguales
            start_row, start_col = ExcelRange.parse_cell_ref(range_str)
            end_row, end_col = start_row, start_col
        
        return start_row, start_col, end_row, end_col
    
    @staticmethod
    def cell_to_a1(row: int, col: int) -> str:
        """
        Convierte coordenadas (fila, columna) 0-based a referencia de celda A1.
        
        Args:
            row: Índice de fila (base 0)
            col: Índice de columna (base 0)
            
        Returns:
            Referencia de celda en formato A1
        """
        if row < 0 or col < 0:
            raise ValueError(f"Índices negativos no válidos: fila={row}, columna={col}")
        
        # Convertir columna a letras
        col_str = ""
        col_val = col + 1  # Convertir a base 1 para cálculo
        
        while col_val > 0:
            remainder = (col_val - 1) % 26
            col_str = chr(65 + remainder) + col_str
            col_val = (col_val - 1) // 26
        
        # Convertir fila a número (base 1 para Excel)
        row_val = row + 1
        
        return f"{col_str}{row_val}"
    
    @staticmethod
    def range_to_a1(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
        """
        Convierte coordenadas de rango 0-based a rango A1:B5.
        
        Args:
            start_row: Fila inicial (base 0)
            start_col: Columna inicial (base 0)
            end_row: Fila final (base 0)
            end_col: Columna final (base 0)
            
        Returns:
            Rango en formato A1:B5
        """
        start_cell = ExcelRange.cell_to_a1(start_row, start_col)
        end_cell = ExcelRange.cell_to_a1(end_row, end_col)
        
        if start_cell == end_cell:
            return start_cell
        return f"{start_cell}:{end_cell}"

# Funciones de utilidad
def get_sheet(wb, sheet_name_or_index) -> Any:
    """
    Obtiene una hoja de Excel por nombre o índice.
    
    Args:
        wb: Objeto workbook de openpyxl
        sheet_name_or_index: Nombre o índice de la hoja
        
    Returns:
        Objeto worksheet
        
    Raises:
        SheetNotFoundError: Si la hoja no existe
    """
    if wb is None:
        raise ExcelWriterError("El workbook no puede ser None")
    
    if isinstance(sheet_name_or_index, int):
        # Si es un índice, intentar acceder por posición
        if 0 <= sheet_name_or_index < len(wb.worksheets):
            return wb.worksheets[sheet_name_or_index]
        else:
            raise SheetNotFoundError(f"No existe una hoja con el índice {sheet_name_or_index}")
    else:
        # Si es un nombre, intentar acceder por nombre
        try:
            return wb[sheet_name_or_index]
        except KeyError:
            raise SheetNotFoundError(f"No existe una hoja con el nombre '{sheet_name_or_index}'")
        except Exception as e:
            raise ExcelWriterError(f"Error al acceder a la hoja: {e}")


# 4. Escritura y Edición de Datos
def write_sheet_data(ws, start_cell, data):
    """
    Escribe un array bidimensional de valores o fórmulas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        start_cell (str): Celda de anclaje (e.j. "A1")
        data (List[List]): Valores o cadenas "=FÓRMULA(...)"
        
    Raises:
        CellReferenceError: Si la referencia de celda es inválida
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    if not data or not isinstance(data, list):
        raise ExcelWriterError("Los datos deben ser una lista no vacía")
    
    try:
        # Parsear la celda inicial para obtener fila y columna base
        start_row, start_col = ExcelRange.parse_cell_ref(start_cell)
        
        # Escribir los datos
        for i, row_data in enumerate(data):
            if row_data is None:
                continue
            
            if not isinstance(row_data, list):
                # Si no es una lista, tratar como valor único
                row_data = [row_data]
                
            for j, value in enumerate(row_data):
                # Calcular coordenadas de celda (base 1 para openpyxl)
                row = start_row + i + 1
                col = start_col + j + 1
                
                # Escribir el valor
                cell = ws.cell(row=row, column=col)
                cell.value = value
    
    except ValueError as e:
        raise CellReferenceError(f"Referencia de celda inválida '{start_cell}': {e}")
    except Exception as e:
        raise ExcelWriterError(f"Error al escribir datos: {e}")

def append_rows(ws, data):
    """
    Añade filas al final con los valores dados.
    
    Args:
        ws: Objeto worksheet de openpyxl
        data (List[List]): Valores o cadenas "=FÓRMULA(...)"
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    if not data or not isinstance(data, list):
        raise ExcelWriterError("Los datos deben ser una lista no vacía")
    
    try:
        for row_data in data:
            if not isinstance(row_data, list):
                # Si no es una lista, convertir a lista con un solo elemento
                row_data = [row_data]
            
            ws.append(row_data)
    
    except Exception as e:
        raise ExcelWriterError(f"Error al añadir filas: {e}")

def update_cell(ws, cell, value_or_formula):
    """
    Actualiza individualmente una celda.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell (str): Referencia de celda (e.j. "A1")
        value_or_formula: Valor o fórmula a asignar
        
    Raises:
        CellReferenceError: Si la referencia de celda es inválida
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        # Asignar valor a la celda
        ws[cell] = value_or_formula
    
    except KeyError:
        raise CellReferenceError(f"Referencia de celda inválida: '{cell}'")
    except Exception as e:
        raise ExcelWriterError(f"Error al actualizar celda: {e}")

def delete_rows(ws, min_row, max_row=None):
    """
    Elimina filas del worksheet.
    
    Args:
        ws: Objeto worksheet de openpyxl
        min_row (int): Índice de primera fila a eliminar (base 1)
        max_row (int, opcional): Índice de última fila a eliminar (base 1)
            Si no se especifica, solo se elimina min_row
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        # Si max_row no está especificado, eliminar solo min_row
        if max_row is None:
            max_row = min_row
            
        # Validar que los índices sean enteros positivos
        if not isinstance(min_row, int) or min_row <= 0:
            raise ValueError(f"El índice de fila debe ser un entero positivo: {min_row}")
        if not isinstance(max_row, int) or max_row <= 0:
            raise ValueError(f"El índice de fila debe ser un entero positivo: {max_row}")
        
        # Verificar que min_row <= max_row
        if min_row > max_row:
            min_row, max_row = max_row, min_row
        
        # Eliminar las filas desde max_row hasta min_row (en orden inverso)
        ws.delete_rows(min_row, max_row - min_row + 1)
    
    except ValueError as e:
        raise ExcelWriterError(f"Error en los índices de fila: {e}")
    except Exception as e:
        raise ExcelWriterError(f"Error al eliminar filas: {e}")

def delete_cols(ws, min_col, max_col=None):
    """
    Elimina columnas del worksheet.
    
    Args:
        ws: Objeto worksheet de openpyxl
        min_col (int or str): Índice de primera columna a eliminar (base 1 o letra)
        max_col (int or str, opcional): Índice de última columna a eliminar (base 1 o letra)
            Si no se especifica, solo se elimina min_col
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        # Convertir letras de columna a índices si es necesario
        if isinstance(min_col, str):
            min_col = column_index_from_string(min_col)
        if isinstance(max_col, str):
            max_col = column_index_from_string(max_col)
            
        # Si max_col no está especificado, eliminar solo min_col
        if max_col is None:
            max_col = min_col
            
        # Validar que los índices sean enteros positivos
        if not isinstance(min_col, int) or min_col <= 0:
            raise ValueError(f"El índice de columna debe ser un entero positivo: {min_col}")
        if not isinstance(max_col, int) or max_col <= 0:
            raise ValueError(f"El índice de columna debe ser un entero positivo: {max_col}")
        
        # Verificar que min_col <= max_col
        if min_col > max_col:
            min_col, max_col = max_col, min_col
        
        # Eliminar las columnas desde max_col hasta min_col (en orden inverso)
        ws.delete_cols(min_col, max_col - min_col + 1)
    
    except ValueError as e:
        raise ExcelWriterError(f"Error en los índices de columna: {e}")
    except Exception as e:
        raise ExcelWriterError(f"Error al eliminar columnas: {e}")

def clear_range(ws, cell_range):
    """
    Borra valores, fórmulas y formatos en un rango especificado.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell_range (str): Rango en formato A1:B5, o una sola celda (e.j. "A1")
        
    Raises:
        RangeError: Si el rango es inválido
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        # Parsear el rango
        if ':' in cell_range:
            # Rango de celdas
            start_cell, end_cell = cell_range.split(':')
            start_coord = ws[start_cell].coordinate
            end_coord = ws[end_cell].coordinate
            range_str = f"{start_coord}:{end_coord}"
        else:
            # Una sola celda
            range_str = cell_range
        
        # Obtener todas las celdas del rango
        for row in ws[range_str]:
            for cell in row:
                # Limpiar valor y estilo
                cell.value = None
                cell.style = 'Normal'
                # Reiniciar formatos
                cell.font = Font()
                cell.fill = PatternFill()
                cell.border = Border()
                cell.alignment = Alignment()
                cell.number_format = 'General'
    
    except KeyError:
        raise RangeError(f"Rango inválido: '{cell_range}'")
    except Exception as e:
        raise ExcelWriterError(f"Error al limpiar rango: {e}")


# 5. Formatos y Estilos
def apply_number_format(ws, cell_range, fmt):
    """
    Aplica formato numérico a un rango de celdas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell_range (str): Rango en formato A1:B5, o una sola celda (e.j. "A1")
        fmt (str): Formato numérico ("#,##0.00", "0%", "dd/mm/yyyy", etc.)
        
    Raises:
        RangeError: Si el rango es inválido
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        # Parsear el rango
        if ':' in cell_range:
            # Rango de celdas
            start_cell, end_cell = cell_range.split(':')
            start_coord = ws[start_cell].coordinate
            end_coord = ws[end_cell].coordinate
            range_str = f"{start_coord}:{end_coord}"
        else:
            # Una sola celda
            range_str = cell_range
        
        # Aplicar formato a todas las celdas del rango
        for row in ws[range_str]:
            for cell in row:
                cell.number_format = fmt
    
    except KeyError:
        raise RangeError(f"Rango inválido: '{cell_range}'")
    except Exception as e:
        raise ExcelWriterError(f"Error al aplicar formato numérico: {e}")

def apply_style(ws, cell_range, style_dict):
    """
    Aplica estilos de celda a un rango.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell_range (str): Rango en formato A1:B5, o una sola celda (e.j. "A1")
        style_dict (dict): Diccionario con estilos a aplicar:
            - font_name (str): Nombre de la fuente
            - font_size (int): Tamaño de la fuente
            - bold (bool): Negrita
            - italic (bool): Cursiva
            - fill_color (str): Color de fondo (formato hex: "FF0000")
            - border_style (str): Estilo de borde ('thin', 'medium', 'thick', etc.)
            - alignment (str): Alineación ('center', 'left', 'right', etc.)
            
    Raises:
        RangeError: Si el rango es inválido
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        # Parsear el rango
        if ':' in cell_range:
            # Rango de celdas
            start_cell, end_cell = cell_range.split(':')
            start_coord = ws[start_cell].coordinate
            end_coord = ws[end_cell].coordinate
            range_str = f"{start_coord}:{end_coord}"
        else:
            # Una sola celda
            range_str = cell_range
        
        # Preparar los estilos
        font_kwargs = {}
        if 'font_name' in style_dict:
            font_kwargs['name'] = style_dict['font_name']
        if 'font_size' in style_dict:
            font_kwargs['size'] = style_dict['font_size']
        if 'bold' in style_dict:
            font_kwargs['bold'] = style_dict['bold']
        if 'italic' in style_dict:
            font_kwargs['italic'] = style_dict['italic']
        if 'font_color' in style_dict:
            font_kwargs['color'] = style_dict['font_color']
        
        fill = None
        if 'fill_color' in style_dict:
            fill = PatternFill(start_color=style_dict['fill_color'], 
                              end_color=style_dict['fill_color'],
                              fill_type='solid')
        
        border = None
        if 'border_style' in style_dict:
            side = Side(style=style_dict['border_style'])
            border = Border(left=side, right=side, top=side, bottom=side)
        
        alignment = None
        if 'alignment' in style_dict:
            alignment_value = style_dict['alignment'].lower()
            horizontal = None
            
            # Mapear valores de alineación horizontal
            if alignment_value in ['left', 'center', 'right', 'justify']:
                horizontal = alignment_value
            
            alignment = Alignment(horizontal=horizontal)
        
        # Aplicar estilos a todas las celdas del rango
        for row in ws[range_str]:
            for cell in row:
                if font_kwargs:
                    cell.font = Font(**font_kwargs)
                if fill:
                    cell.fill = fill
                if border:
                    cell.border = border
                if alignment:
                    cell.alignment = alignment
    
    except KeyError:
        raise RangeError(f"Rango inválido: '{cell_range}'")
    except Exception as e:
        raise ExcelWriterError(f"Error al aplicar estilos: {e}")

def merge_cells(ws, cell_range):
    """
    Combina celdas en un rango especificado.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell_range (str): Rango en formato A1:B5
        
    Raises:
        RangeError: Si el rango es inválido
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        ws.merge_cells(cell_range)
    except Exception as e:
        raise RangeError(f"Error al combinar celdas '{cell_range}': {e}")

def unmerge_cells(ws, cell_range):
    """
    Separa celdas previamente combinadas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cell_range (str): Rango en formato A1:B5
        
    Raises:
        RangeError: Si el rango es inválido
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        ws.unmerge_cells(cell_range)
    except Exception as e:
        raise RangeError(f"Error al separar celdas '{cell_range}': {e}")

def set_row_height(ws, row, height):
    """
    Establece la altura de una fila.
    
    Args:
        ws: Objeto worksheet de openpyxl
        row (int): Índice de fila (base 1)
        height (float): Altura en puntos
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        # Validar que el índice sea entero positivo
        if not isinstance(row, int) or row <= 0:
            raise ValueError(f"El índice de fila debe ser un entero positivo: {row}")
        
        # Establecer la altura
        ws.row_dimensions[row].height = height
    
    except ValueError as e:
        raise ExcelWriterError(f"Error en el índice de fila: {e}")
    except Exception as e:
        raise ExcelWriterError(f"Error al establecer altura de fila: {e}")

def set_column_width(ws, col, width):
    """
    Establece el ancho de una columna.
    
    Args:
        ws: Objeto worksheet de openpyxl
        col (int or str): Índice (base 1) o letra de columna
        width (float): Ancho en caracteres
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        # Convertir índice a letra si es necesario
        if isinstance(col, int):
            col_letter = get_column_letter(col)
        else:
            col_letter = col
        
        # Establecer el ancho
        ws.column_dimensions[col_letter].width = width
    
    except ValueError as e:
        raise ExcelWriterError(f"Error en el índice de columna: {e}")
    except Exception as e:
        raise ExcelWriterError(f"Error al establecer ancho de columna: {e}")

def hide_rows(ws, rows):
    """
    Oculta filas especificadas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        rows (int or list): Índice de fila o lista de índices (base 1)
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        # Convertir a lista si es un solo índice
        if not isinstance(rows, list):
            rows = [rows]
        
        # Ocultar cada fila
        for row in rows:
            # Validar que el índice sea entero positivo
            if not isinstance(row, int) or row <= 0:
                raise ValueError(f"El índice de fila debe ser un entero positivo: {row}")
            
            ws.row_dimensions[row].hidden = True
    
    except ValueError as e:
        raise ExcelWriterError(f"Error en el índice de fila: {e}")
    except Exception as e:
        raise ExcelWriterError(f"Error al ocultar filas: {e}")

def hide_columns(ws, cols):
    """
    Oculta columnas especificadas.
    
    Args:
        ws: Objeto worksheet de openpyxl
        cols (int or str or list): Índice, letra o lista de columnas
    """
    if not ws:
        raise ExcelWriterError("El worksheet no puede ser None")
    
    try:
        # Convertir a lista si es un solo índice o letra
        if not isinstance(cols, list):
            cols = [cols]
        
        # Ocultar cada columna
        for col in cols:
            # Convertir índice a letra si es necesario
            if isinstance(col, int):
                col_letter = get_column_letter(col)
            else:
                col_letter = col
            
            ws.column_dimensions[col_letter].hidden = True
    
    except ValueError as e:
        raise ExcelWriterError(f"Error en el índice de columna: {e}")
    except Exception as e:
        raise ExcelWriterError(f"Error al ocultar columnas: {e}")


# Crear el servidor MCP como variable global
mcp = None
if HAS_MCP:
    # Esta es la variable global que el sistema MCP busca
    mcp = FastMCP("Excel Writer MCP", 
                 dependencies=["openpyxl"])
    logger.info("Servidor MCP iniciado correctamente")
    
    # 4. Funciones de Escritura y Edición de Datos
    @mcp.tool(description="Escribe un array bidimensional de valores o fórmulas")
    def write_sheet_data_tool(file_path, sheet_name, start_cell, data):
        """Escribe un array bidimensional en una hoja de Excel"""
        try:
            # Validar argumentos
            if not isinstance(data, list):
                raise ValueError("El parámetro 'data' debe ser una lista")
            
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Escribir los datos
            write_sheet_data(ws, start_cell, data)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "start_cell": start_cell,
                "rows_written": len(data),
                "columns_written": max([len(row) if isinstance(row, list) else 1 for row in data], default=0),
                "message": f"Datos escritos correctamente desde {start_cell}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al escribir datos: {e}"
            }
    
    @mcp.tool(description="Añade filas al final de la hoja con los valores dados")
    def append_rows_tool(file_path, sheet_name, data):
        """Añade filas al final de una hoja de Excel"""
        try:
            # Validar argumentos
            if not isinstance(data, list):
                raise ValueError("El parámetro 'data' debe ser una lista")
            
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Añadir las filas
            append_rows(ws, data)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "rows_appended": len(data),
                "message": f"Se añadieron {len(data)} filas a la hoja {sheet_name}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al añadir filas: {e}"
            }
    
    @mcp.tool(description="Actualiza individualmente una celda")
    def update_cell_tool(file_path, sheet_name, cell, value_or_formula):
        """Actualiza individualmente una celda en una hoja de Excel"""
        try:
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Actualizar la celda
            update_cell(ws, cell, value_or_formula)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell": cell,
                "value": value_or_formula,
                "message": f"Celda {cell} actualizada correctamente en la hoja {sheet_name}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al actualizar celda: {e}"
            }
    
    @mcp.tool(description="Elimina filas de una hoja de Excel")
    def delete_rows_tool(file_path, sheet_name, min_row, max_row=None):
        """Elimina filas de una hoja de Excel"""
        try:
            # Convertir a enteros si son cadenas
            min_row = int(min_row)
            max_row = int(max_row) if max_row is not None else None
            
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Eliminar las filas
            delete_rows(ws, min_row, max_row)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            # Calcular mensaje según si se eliminó una o varias filas
            if max_row is None or min_row == max_row:
                message = f"Fila {min_row} eliminada correctamente de la hoja {sheet_name}"
            else:
                message = f"Filas {min_row} a {max_row} eliminadas correctamente de la hoja {sheet_name}"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "min_row": min_row,
                "max_row": max_row,
                "message": message
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al eliminar filas: {e}"
            }
    
    @mcp.tool(description="Elimina columnas de una hoja de Excel")
    def delete_cols_tool(file_path, sheet_name, min_col, max_col=None):
        """Elimina columnas de una hoja de Excel"""
        try:
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Eliminar las columnas
            delete_cols(ws, min_col, max_col)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            # Convertir a letras si son números para el mensaje
            if isinstance(min_col, int):
                min_col_str = get_column_letter(min_col)
            else:
                min_col_str = min_col
                
            if max_col is None:
                max_col_str = min_col_str
            elif isinstance(max_col, int):
                max_col_str = get_column_letter(max_col)
            else:
                max_col_str = max_col
            
            # Calcular mensaje según si se eliminó una o varias columnas
            if max_col is None or min_col_str == max_col_str:
                message = f"Columna {min_col_str} eliminada correctamente de la hoja {sheet_name}"
            else:
                message = f"Columnas {min_col_str} a {max_col_str} eliminadas correctamente de la hoja {sheet_name}"
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "min_col": min_col_str,
                "max_col": max_col_str,
                "message": message
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al eliminar columnas: {e}"
            }
    
    @mcp.tool(description="Borra valores, fórmulas y formatos en un rango especificado")
    def clear_range_tool(file_path, sheet_name, cell_range):
        """Borra valores, fórmulas y formatos en un rango especificado"""
        try:
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Limpiar el rango
            clear_range(ws, cell_range)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "range": cell_range,
                "message": f"Rango {cell_range} limpiado correctamente en la hoja {sheet_name}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al limpiar rango: {e}"
            }
    
    # 5. Funciones de Formato y Estilo
    @mcp.tool(description="Aplica formato numérico a un rango de celdas")
    def apply_number_format_tool(file_path, sheet_name, cell_range, fmt):
        """Aplica formato numérico a un rango de celdas"""
        try:
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Aplicar el formato
            apply_number_format(ws, cell_range, fmt)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "range": cell_range,
                "format": fmt,
                "message": f"Formato numérico aplicado correctamente al rango {cell_range}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al aplicar formato numérico: {e}"
            }
    
    @mcp.tool(description="Aplica estilos de celda a un rango")
    def apply_style_tool(file_path, sheet_name, cell_range, style_dict):
        """Aplica estilos de celda a un rango"""
        try:
            # Validar argumentos
            if not isinstance(style_dict, dict):
                # Intentar convertir si es una cadena JSON
                import json
                if isinstance(style_dict, str):
                    style_dict = json.loads(style_dict)
                else:
                    raise ValueError("El parámetro 'style_dict' debe ser un diccionario o una cadena JSON")
            
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Aplicar los estilos
            apply_style(ws, cell_range, style_dict)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "range": cell_range,
                "styles": style_dict,
                "message": f"Estilos aplicados correctamente al rango {cell_range}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al aplicar estilos: {e}"
            }
    
    @mcp.tool(description="Combina celdas en un rango especificado")
    def merge_cells_tool(file_path, sheet_name, cell_range):
        """Combina celdas en un rango especificado"""
        try:
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Combinar celdas
            merge_cells(ws, cell_range)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "range": cell_range,
                "message": f"Celdas combinadas correctamente en el rango {cell_range}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al combinar celdas: {e}"
            }
    
    @mcp.tool(description="Separa celdas previamente combinadas")
    def unmerge_cells_tool(file_path, sheet_name, cell_range):
        """Separa celdas previamente combinadas"""
        try:
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Separar celdas
            unmerge_cells(ws, cell_range)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "range": cell_range,
                "message": f"Celdas separadas correctamente en el rango {cell_range}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al separar celdas: {e}"
            }
    
    @mcp.tool(description="Establece la altura de una fila")
    def set_row_height_tool(file_path, sheet_name, row, height):
        """Establece la altura de una fila"""
        try:
            # Convertir a enteros si son cadenas
            row = int(row)
            height = float(height)
            
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Establecer la altura de la fila
            set_row_height(ws, row, height)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "row": row,
                "height": height,
                "message": f"Altura de fila {row} establecida a {height} puntos"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al establecer altura de fila: {e}"
            }
    
    @mcp.tool(description="Establece el ancho de una columna")
    def set_column_width_tool(file_path, sheet_name, col, width):
        """Establece el ancho de una columna"""
        try:
            # Convertir a número si es posible
            if isinstance(col, str) and col.isdigit():
                col = int(col)
            
            # Convertir ancho a float
            width = float(width)
            
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Establecer el ancho de la columna
            set_column_width(ws, col, width)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            # Formular mensaje con letra de columna
            if isinstance(col, int):
                col_str = get_column_letter(col)
            else:
                col_str = col
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "column": col_str,
                "width": width,
                "message": f"Ancho de columna {col_str} establecido a {width} caracteres"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al establecer ancho de columna: {e}"
            }
    
    @mcp.tool(description="Oculta filas especificadas")
    def hide_rows_tool(file_path, sheet_name, rows):
        """Oculta filas especificadas"""
        try:
            # Convertir a lista si es una cadena (ej. "1,2,3")
            if isinstance(rows, str):
                rows = [int(r.strip()) for r in rows.split(',')]
            # Convertir a lista si es un solo número
            elif isinstance(rows, (int, float)):
                rows = [int(rows)]
            
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Ocultar las filas
            hide_rows(ws, rows)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "rows": rows,
                "message": f"Filas {rows} ocultadas correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al ocultar filas: {e}"
            }
    
    @mcp.tool(description="Oculta columnas especificadas")
    def hide_columns_tool(file_path, sheet_name, cols):
        """Oculta columnas especificadas"""
        try:
            # Convertir a lista si es una cadena (ej. "A,B,C" o "1,2,3")
            if isinstance(cols, str):
                if ',' in cols:
                    cols = [col.strip() for col in cols.split(',')]
                else:
                    cols = [cols]
            # Convertir a lista si es un solo número o letra
            elif isinstance(cols, (int, float)):
                cols = [int(cols)]
            
            # Abrir el archivo y obtener la hoja
            wb = openpyxl.load_workbook(file_path)
            ws = get_sheet(wb, sheet_name)
            
            # Ocultar las columnas
            hide_columns(ws, cols)
            
            # Guardar y cerrar
            wb.save(file_path)
            
            # Convertir a letras para el mensaje si son números
            cols_str = []
            for col in cols:
                if isinstance(col, int):
                    cols_str.append(get_column_letter(col))
                else:
                    cols_str.append(col)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "columns": cols_str,
                "message": f"Columnas {cols_str} ocultadas correctamente"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al ocultar columnas: {e}"
            }


if __name__ == "__main__":
    # Código de ejemplo de uso
    logger.info("Excel Writer MCP - Ejemplo de uso")
    
    import sys
    
    if len(sys.argv) > 1:
        opciones = ", ".join(["escribir", "formato", "combinar", "altura", "ocultar"])
        logger.info(f"Opciones disponibles: {opciones}")
        
        # Test write_sheet_data
        if sys.argv[1].lower() == "escribir" and len(sys.argv) > 2:
            archivo = sys.argv[2]
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                
                # Ejemplo de datos
                datos = [
                    ["Nombre", "Edad", "Ciudad"],
                    ["Ana", 25, "Madrid"],
                    ["Carlos", 30, "Barcelona"],
                    ["Elena", 22, "Valencia"]
                ]
                
                write_sheet_data(ws, "A1", datos)
                wb.save(archivo)
                logger.info(f"Datos escritos correctamente en {archivo}")
                
            except Exception as e:
                logger.error(f"Error: {e}")
        
        # Test apply_style
        elif sys.argv[1].lower() == "formato" and len(sys.argv) > 2:
            archivo = sys.argv[2]
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                
                # Aplicar formato a encabezados
                apply_style(ws, "A1:C1", {
                    'bold': True,
                    'fill_color': 'FFFF00',
                    'border_style': 'thin',
                    'alignment': 'center'
                })
                
                wb.save(archivo)
                logger.info(f"Formato aplicado correctamente en {archivo}")
                
            except Exception as e:
                logger.error(f"Error: {e}")
                
        else:
            logger.info(f"Comando no reconocido o faltan argumentos")
    else:
        logger.info("Uso: python excel_writer_mcp.py [escribir|formato|combinar|altura|ocultar] archivo.xlsx")
